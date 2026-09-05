"""The blank snag-list workbook the import wizard hands out.

One endpoint, one shape: a single worksheet holding the five header labels the
wizard's column guess understands, over `TEMPLATE_ROW_COUNT` ready-to-fill rows.
Every label here is already in the `services/snag_parser/guess.py` vocabulary --
`Area` / `Category` / `Description` / `Remarks` claim their roles and `S.No` claims
the serial -- so a workbook filled from this template maps itself on upload with no
hand-mapping.

There is deliberately NO Status column. The import does not read one: every snag
starts at `Pending` and its status is moved in the app afterwards, so a Status column
here would invite a consultant to fill in a value that is then silently discarded.

The ruled rows ship EMPTY -- every cell, S.No included. Nothing here writes a value
a person did not: a number we printed would look like the consultant's own the moment
the sheet came back, and the import cannot tell the two apart. A row the consultant
leaves unnumbered is numbered by its position at ingest (`import_wizard._serials_for`),
which is where a number we invent belongs -- attributable to the import, not to them.

Styling mirrors the "Project Snag" print format the consultant gets back -- grey
header, grey Area column -- so the sheet they fill in and the PDF it prints to
read as the same document.

Wire contract: none. This returns a FILE (`frappe.local.response.type =
"download"`), not a payload in `frontend/src/pages/SnagList/types.ts`.
"""

from __future__ import annotations

import io

import frappe

from nirmaan_stack.api.snags import require_import_access

#: Column order IS the template. Changing it changes what a consultant sends back.
TEMPLATE_HEADERS: "tuple[str, ...]" = (
    "S.No",
    "Area",
    "Category",
    "Description",
    "Remarks",
)

#: Display width per header, in the same order. Description is the one that holds
#: a sentence; the rest hold a phrase.
_COLUMN_WIDTHS: "tuple[int, ...]" = (8, 24, 22, 60, 32)

#: Blank rows ruled below the header. A consultant with more snags than this adds
#: rows as usual -- the number is a starting point, not a limit.
TEMPLATE_ROW_COUNT = 50

_FILENAME = "Snag List Template.xlsx"
_SHEET_TITLE = "Snag List"

# Grey shared by the header row and the Area column, matching the print format's
# table chrome (tailwind gray-200 / gray-300, the app's own palette).
_FILL_HEX = "FFE5E7EB"
_BORDER_HEX = "FFD1D5DB"


@frappe.whitelist()
def download_snag_template():
    """Return the blank template as an .xlsx download.

    Gated on IMPORT access rather than READ: the template exists to be filled in
    and imported, so the people who may import are exactly the people it is for.
    """
    require_import_access("download the snag list template")

    frappe.local.response.filename = _FILENAME
    frappe.local.response.filecontent = build_template_bytes()
    frappe.local.response.type = "download"


def build_template_bytes() -> bytes:
    """Build the workbook in memory. The only openpyxl touch in this module."""
    import openpyxl  # noqa: PLC0415
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: PLC0415
    from openpyxl.utils import get_column_letter  # noqa: PLC0415

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = _SHEET_TITLE

    grey = PatternFill("solid", fgColor=_FILL_HEX)
    edge = Side(style="thin", color=_BORDER_HEX)
    box = Border(left=edge, right=edge, top=edge, bottom=edge)
    header_font = Font(bold=True)
    area_font = Font(bold=True)
    left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    centre_top = Alignment(horizontal="center", vertical="top")

    sheet.append(list(TEMPLATE_HEADERS))
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = grey
        cell.border = box
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # The ruled rows carry FORMAT ONLY -- borders, alignment, the Area column's grey.
    # Not one cell is given a value.
    for offset in range(TEMPLATE_ROW_COUNT):
        row = 2 + offset
        for column in range(1, len(TEMPLATE_HEADERS) + 1):
            cell = sheet.cell(row=row, column=column)
            cell.border = box
            cell.alignment = centre_top if column == 1 else left_top
        area = sheet.cell(row=row, column=2)
        area.font = area_font
        area.fill = grey

    for index, width in enumerate(_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    # So the header stays put while the consultant types down the sheet.
    sheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
