# Copyright (c) 2026, Nirmaan (Stratos Infra Technologies Pvt. Ltd.) and Contributors
# See license.txt

"""Tests for the INTERNAL priced workbook -- the client export PLUS the cost block.

TWO THINGS ARE UNDER TEST HERE, and the second matters as much as the first:

  1. THE BLOCK IS RIGHT -- the cost columns land where the remark column would have, carry
     the grid's own headers, hold values read straight off `BoQ Row BCS Rate`, and the Total
     column holds an EXCEL FORMULA that blanks exactly where the screen blanks.
  2. THE SEPARATION IS REAL -- the client export is untouched, still leaks nothing, and the
     two exports over the SAME fixture differ in exactly the intended way. A guard saying
     "the client file has no BCS" is only worth something beside one saying "and this file
     does", or the first could be passing because there was nothing to find.

Run:
    bench --site localhost run-tests --module \\
        nirmaan_stack.api.boq.wizard.test_export_bcs_writeback
"""
import base64
import io
import json
import os
import tempfile

import frappe
import openpyxl
from frappe.tests.utils import FrappeTestCase
from openpyxl import load_workbook

from nirmaan_stack.api.boq.wizard.export_bcs_writeback import (
    _BCS_FILLED_HEX,
    _amount_body,
    _builtin_total,
    _margin_body,
    _next_empty_col,
    _generate_internal_workbook,
    _guarded,
    _next_empty_col,
    _ref_to_excel,
    _tree_to_excel,
    export_priced_workbook_with_bcs,
)
from nirmaan_stack.api.boq.wizard.export_writeback import _generate_priced_workbook
from nirmaan_stack.api.boq.wizard.test_export_writeback import _seed_committed_sheet
from nirmaan_stack.api.boq.wizard.test_review_screen import _cleanup_project, _make_project

_ROLE_MAP = {
    "A": {"role": "description", "area": None},
    "B": {"role": "unit", "area": None},
    "C": {"role": "qty_total", "area": None},
    "D": {"role": "rate_supply", "area": None},
    "E": {"role": "rate_install", "area": None},
    "F": {"role": "amount_total", "area": None},
}
# Deliberately absurd sentinels, so a hit in the produced workbook can only be ours.
_SUPPLY = 987654.21
_INSTALL = 123456.78


def _ref(value_field):
    return {"ref": {"value_field": value_field, "value_key": None, "rate_subkey": None}}


def _declared_total_tree():
    """`(bcs_supply + bcs_install) * qty_total` -- the exact shape all four declared
    `bcs_total` formulas on the live bench carry (measured 2026-08-19)."""
    return {"op": "*", "operands": [
        {"op": "+", "operands": [_ref("bcs_supply"), _ref("bcs_install")]},
        _ref("qty_total"),
    ]}


# ===========================================================================
# Group 1: the pure Excel-emitting helpers -- no site data at all
# ===========================================================================
class TestExcelEmitters(FrappeTestCase):
    """The formula writer, at its own seam. Nothing here needs a BoQ."""

    _COST = {"supply": "G", "install": "H"}
    _ROLE = _ROLE_MAP

    def test_a_cost_operand_resolves_to_the_column_we_wrote(self):
        body, cells = _ref_to_excel(_ref("bcs_supply")["ref"], 5, self._COST, ["C"], self._ROLE)
        self.assertEqual(body, "G5")
        self.assertEqual(cells, [], "a cost cell WE wrote is never a sheet cell to guard on")

    def test_the_qty_operand_sums_the_sheets_quantity_cells(self):
        """`bcs_qty` has no column of its own -- it IS the derived quantity columns, summed,
        which is what `bcsRowQuantity` does on screen."""
        body, cells = _ref_to_excel(_ref("bcs_qty")["ref"], 5, self._COST, ["C", "K"], self._ROLE)
        self.assertEqual(body, "(C5+K5)")
        self.assertEqual(cells, ["C5", "K5"])

    def test_a_sheet_column_ref_goes_through_the_shared_resolver(self):
        """A sheet-column ref must mean the SAME thing in both exports, so it resolves through
        `export_template_workbook.resolve_ref_col` rather than a second role-map walk."""
        body, cells = _ref_to_excel(_ref("qty_total")["ref"], 9, self._COST, ["C"], self._ROLE)
        self.assertEqual(body, "C9")
        self.assertEqual(cells, ["C9"], "a CLIENT cell can be blank, so it must be guarded on")

    def test_the_bcs_total_operand_resolves_to_the_total_column(self):
        """`bcs_total` is BOTH a target and an operand -- choosing "BCS Total Amount" in a
        margin numerator must mean "whatever that column currently computes". As an Excel
        REFERENCE it means exactly that, and stays live."""
        body, cells = _ref_to_excel(
            _ref("bcs_total")["ref"], 5, self._COST, ["C"], self._ROLE, total_col="I")
        self.assertEqual(body, "I5")
        self.assertEqual(cells, [], "a formula WE wrote is already guarded; it is not a "
                                    "client cell that might be blank")

    def test_the_bcs_total_operand_fails_safe_when_there_is_no_total_column(self):
        """A numerator naming the Total on a sheet that got none. Blank is the honest answer:
        there is no such column to divide by, and any number here would be invented."""
        body, cells = _ref_to_excel(
            _ref("bcs_total")["ref"], 5, self._COST, ["C"], self._ROLE, total_col=None)
        self.assertIsNone(body)
        self.assertEqual(cells, [])

    def test_the_default_denominator_sums_the_amount_columns_and_guards_every_one(self):
        """Each amount cell belongs to the CLIENT's workbook and may legitimately be blank, so
        every one is returned for guarding -- unlike a cost cell we wrote ourselves."""
        body, cells = _amount_body(7, ["F"], self._ROLE)
        self.assertEqual(body, "F7", "one column needs no parentheses")
        self.assertEqual(cells, ["F7"])
        body, cells = _amount_body(7, ["F", "K"], self._ROLE)
        self.assertEqual(body, "(F7+K7)")
        self.assertEqual(cells, ["F7", "K7"])
        self.assertEqual(_amount_body(7, [], self._ROLE), (None, []))

    def test_the_margin_body_divides_by_the_amount_and_never_the_other_way_round(self):
        """★ THE DIRECTION IS OWNER-SETTLED AND WAS ONCE RELAYED BACKWARDS. Dividing by the
        AMOUNT is what makes a one-sided sheet read LOWER and go sharply negative once cost
        passes amount; cost-over-amount and mark-up-on-cost both read HIGHER on exactly the
        sheets that need a warning."""
        self.assertEqual(_margin_body(2, "I2", "F2"), "((F2-I2)/F2)*100")

    def test_the_margin_body_is_the_owners_formula_rearranged(self):
        """The owner states it as `(1 - cost/amount) x 100`; this emits `(amount - cost) /
        amount x 100`. The same expression, and this form needs no literal 1. Checked
        numerically rather than by eye, because the MISREAD of the owner's form --
        `1 - (c/a) x 100` -- returns -59 where the answer is +40, and reads plausible."""
        amount, cost = 100.0, 60.0
        emitted = ((amount - cost) / amount) * 100
        self.assertAlmostEqual(emitted, (1 - cost / amount) * 100)
        self.assertAlmostEqual(emitted, 40.0)
        self.assertNotAlmostEqual(emitted, 1 - (cost / amount) * 100)

    def test_a_guard_with_an_extra_test_ors_them_into_one_if(self):
        """One builder emits every guard in this module, so there is no second place a guard
        could be written differently. A single test skips the `OR`, which is what keeps the
        ordinary Total byte-identical to what it emitted before the parameter existed."""
        self.assertEqual(_guarded("X", ["C2"], False), '=IF(COUNT(C2)=0,"",X)')
        self.assertEqual(_guarded("X", ["C2"], False, extra_tests=["F2<=0"]),
                         '=IF(OR(COUNT(C2)=0,F2<=0),"",X)')
        self.assertEqual(_guarded("X", [], False, extra_tests=["F2<=0"]),
                         '=IF(F2<=0,"",X)', "no cells to count, so no COUNT term")
        self.assertEqual(_guarded("X", [], False), "=X", "nothing to guard at all")

    def test_an_unresolvable_ref_fails_safe_to_blank(self):
        for ref in (
            _ref("make_model")["ref"],          # a real field, not mapped on this sheet
            _ref("not_a_field_at_all")["ref"],
            {"value_field": "bcs_combined"},    # a kind this sheet has no column for
        ):
            body, cells = _ref_to_excel(ref, 5, self._COST, ["C"], self._ROLE)
            self.assertIsNone(body, ref)
            self.assertEqual(cells, [])

    def test_the_qty_operand_with_no_quantity_columns_fails_safe(self):
        body, _ = _ref_to_excel(_ref("bcs_qty")["ref"], 5, self._COST, [], self._ROLE)
        self.assertIsNone(body)

    def test_a_declared_tree_becomes_the_expected_excel(self):
        body, cells = _tree_to_excel(_declared_total_tree(), 5, self._COST, ["C"], self._ROLE)
        self.assertEqual(body, "((G5+H5)*C5)")
        self.assertEqual(cells, ["C5"])

    def test_every_operator_node_is_bracketed_so_the_fold_survives(self):
        """`-` and `/` fold LEFT TO RIGHT from operands[0], so the list order IS the
        arithmetic. Bracketing every node is what makes Excel read an n-ary `(A-B-C)` the way
        `foldOperands` folds it."""
        tree = {"op": "-", "operands": [_ref("bcs_supply"), _ref("bcs_install"),
                                        _ref("qty_total")]}
        body, _ = _tree_to_excel(tree, 2, self._COST, ["C"], self._ROLE)
        self.assertEqual(body, "(G2-H2-C2)")

    def test_an_unsupported_operator_blanks_the_whole_cell(self):
        tree = {"op": "^", "operands": [_ref("bcs_supply"), _ref("bcs_install")]}
        self.assertEqual(_tree_to_excel(tree, 2, self._COST, ["C"], self._ROLE), (None, []))

    def test_one_unresolvable_operand_blanks_the_whole_cell(self):
        """Fail-safe, never partial: half a cost formula is worse than none."""
        tree = {"op": "*", "operands": [_ref("bcs_supply"), _ref("make_model")]}
        self.assertEqual(_tree_to_excel(tree, 2, self._COST, ["C"], self._ROLE), (None, []))

    def test_a_malformed_node_blanks_rather_than_raising(self):
        for node in (None, [], {}, {"operands": []}, {"op": "*", "operands": []}):
            self.assertEqual(_tree_to_excel(node, 2, self._COST, ["C"], self._ROLE), (None, []))

    def test_the_builtin_rule_is_costs_summed_times_quantity_summed(self):
        body, cells = _builtin_total(7, ["supply", "install"], self._COST, ["C"])
        self.assertEqual(body, "(G7+H7)*(C7)")
        self.assertEqual(cells, ["C7"])

    def test_the_builtin_rule_needs_both_sides(self):
        self.assertEqual(_builtin_total(7, ["supply"], self._COST, []), (None, []))
        self.assertEqual(_builtin_total(7, [], self._COST, ["C"]), (None, []))

    # -- the guard, which is where a blank stays a blank -------------------
    def test_the_builtin_guard_blanks_only_when_NOTHING_is_numeric(self):
        """`bcsRowQuantity` sums whatever resolves and returns null only when nothing does, so
        the built-in guard is `COUNT = 0`. A genuine 0 quantity still computes and still reads
        0 -- which is what the screen shows for it."""
        self.assertEqual(
            _guarded("(G5+H5)*(C5)", ["C5"], False), '=IF(COUNT(C5)=0,"",(G5+H5)*(C5))'
        )

    def test_the_declared_guard_blanks_when_ANY_operand_is_missing(self):
        """`evaluateBcsTotalFormula` returns on the FIRST unresolved ref, so a declared formula
        blanks when any of its sheet operands is non-numeric -- `COUNT < n`, not `COUNT = 0`.
        The two shapes are not interchangeable."""
        self.assertEqual(
            _guarded("((G5+H5)*C5)", ["C5", "K5"], True),
            '=IF(COUNT(C5,K5)<2,"",((G5+H5)*C5))',
        )

    def test_a_formula_reading_no_client_cell_needs_no_guard(self):
        """Every cell it reads is one WE wrote, and those exist whenever the row is costed."""
        self.assertEqual(_guarded("(G5+H5)", [], False), "=(G5+H5)")

    def test_the_column_scan_steps_past_anything_that_carries_data(self):
        """★ THE SAFETY PROPERTY OF THE WHOLE PLACEMENT RULE, and it had no test until now.
        The block starts at the TRUE data edge and scans RIGHT past any column that carries
        real content -- a stray estimator note, or a trailing column the role map never
        covered. Without the scan the block would silently overwrite it."""
        ws = openpyxl.Workbook().active
        ws["G2"] = "an estimator note one past the mapped edge"
        ws["H5"] = 0                      # a genuine 0 is CONTENT, not emptiness
        ws["I3"] = ""                     # an empty string is NOT content
        self.assertEqual(_next_empty_col(ws, 7), 9, "G and H are taken; I is free")
        self.assertEqual(_next_empty_col(ws, 9), 9, "an already-empty start is returned as is")
        self.assertEqual(_next_empty_col(ws, 20), 20)

    def test_every_emitted_formula_starts_with_an_equals_sign(self):
        """openpyxl types a cell as a formula from the leading `=`. Without it the workbook
        would carry a string that LOOKS like a formula and computes nothing -- and the fidelity
        delta would be wrong too, since the cell would not count as a formula."""
        for out in (
            _guarded("(G5+H5)*(C5)", ["C5"], False),
            _guarded("((G5+H5)*C5)", ["C5"], True),
            _guarded("(G5+H5)", [], False),
        ):
            self.assertTrue(out.startswith("="), out)


# ===========================================================================
# Group 2 + 3: end to end on a synthetic workbook, and the separation
# ===========================================================================
class TestInternalWorkbookEndToEnd(FrappeTestCase):
    """One upload-origin BoQ, three sheets, exported BOTH ways from the SAME fixture.

    The workbook is synthetic and `_generate_internal_workbook` is called directly -- the same
    seam `export_writeback` created so a test can inject a workbook and bypass the S3 fetch.
    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls._tmps = []
        cls.project = _make_project()
        boq = frappe.new_doc("BOQs")
        boq.project = cls.project.name
        boq.boq_name = "Internal Export BoQ"
        boq.tax_treatment = "Pre-tax"
        boq.source_file_url = "/files/synthetic.xlsx"  # never read: we inject the path
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        cls.boq = boq.name
        now = frappe.utils.now()

        grid = [{"row_number": r, "cells": json.dumps({})} for r in range(1, 7)]
        for sheet in ("Elec ", "HVAC "):
            _seed_committed_sheet(
                cls.boq, sheet, "data", 1, _ROLE_MAP, [], 1, sheet.strip(), now,
                grid,
                [{"excel_row": 2, "col_letter": "D", "rate": 25}],
                None,
            )
            frappe.db.set_value("BoQ Sheet",
                                {"boq": cls.boq, "sheet_name": sheet, "is_current": 1},
                                "bcs_enabled", 1, update_modified=False)
        # A grid-only general-specs sheet: no rates, no nodes, no costs.
        _seed_committed_sheet(cls.boq, "Specs ", "master_preamble", 1, _ROLE_MAP, [], 3,
                              "Specs", now, grid, [], None, sheet_disposition="grid_only")

        # HVAC declares a bcs_total formula; Elec does not, so it takes the built-in rule.
        af = frappe.new_doc("BoQ Cell Amount Formula")
        af.boq = cls.boq
        af.sheet_name = "HVAC "
        af.committed_version = 1
        af.target_value_field = "bcs_total"
        af.target_col = None
        af.formula = json.dumps(_declared_total_tree())
        af.formula_version = 1
        af.is_current = 1
        af.defined_at = now
        af.insert(ignore_permissions=True)

        # Costs on rows 2 and 3 of both sheets. Row 4 is UNCOSTED and row 5 is costed but has
        # NO quantity -- the two rows that must come out blank, for different reasons.
        for sheet in ("Elec ", "HVAC "):
            for excel_row in (2, 3, 5):
                doc = frappe.new_doc("BoQ Row BCS Rate")
                doc.boq = cls.boq
                doc.sheet_name = sheet
                doc.excel_row = excel_row
                doc.committed_version = 1
                doc.supply_rate = _SUPPLY
                doc.install_rate = _INSTALL
                doc.combined_rate = 0.0
                doc.is_filled = 1
                doc.bcs_version = 1
                doc.is_current = 1
                doc.bcs_rated_at = now
                doc.insert(ignore_permissions=True)

        # A remark, so the remark column's placement relative to the block is under test.
        rm = frappe.new_doc("BoQ Cell Remark")
        rm.boq = cls.boq
        rm.sheet_name = "Elec "
        rm.excel_row = 2
        rm.committed_version = 1
        rm.remark = "a remark"
        rm.remark_version = 1
        rm.is_current = 1
        rm.remarked_at = now
        rm.insert(ignore_permissions=True)
        frappe.db.commit()

        cls.sheets = ["Elec ", "HVAC ", "Specs "]
        cls.result = _generate_internal_workbook(
            cls.boq, cls.sheets, cls._synthetic_workbook(), "Internal Export BoQ"
        )
        cls.wb = load_workbook(
            io.BytesIO(base64.b64decode(cls.result["content_base64"])), data_only=False
        )
        # ⚠️ CAPTURED BEFORE THE CLIENT EXPORT RUNS. The client export STAMPS
        # `last_exported_at` and commits -- that is its correct behaviour -- so reading the
        # field afterwards would tell us nothing about ours. This is the only moment the
        # question can be asked honestly.
        cls.stamp_after_internal = frappe.db.get_value(
            "BoQ Sheet", {"boq": cls.boq, "sheet_name": "HVAC ", "is_current": 1},
            "last_exported_at")

        # The CLIENT export over the very same fixture, for the separation guards. Called at
        # its own injectable seam (the one its docstring says exists so a test can supply a
        # synthetic workbook) -- the public endpoint would go to S3 for the source file.
        cls.client_result = _generate_priced_workbook(
            cls.boq, cls.sheets, cls._synthetic_workbook(), "Internal Export BoQ"
        )

    @classmethod
    def _synthetic_workbook(cls):
        """A minimal three-sheet workbook shaped like _ROLE_MAP, written to a tempfile the
        caller stamps. F holds a real formula, so the fidelity guard has something to protect."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for title in ("Elec ", "HVAC ", "Specs "):
            ws = wb.create_sheet(title=title)
            ws["A1"], ws["B1"], ws["C1"] = "Description", "Unit", "Qty"
            ws["D1"], ws["E1"], ws["F1"] = "Rate (S)", "Rate (I)", "Amount"
            for r in (2, 3, 4):
                ws[f"A{r}"] = f"item {r}"
                ws[f"C{r}"] = 10 * r
                ws[f"F{r}"] = f"=C{r}*(D{r}+E{r})"
            ws["A5"] = "no quantity on this row"   # C5 deliberately EMPTY
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb.save(path)
        cls._tmps.append(path)  # every copy is tracked, so none is leaked
        return path

    @classmethod
    def tearDownClass(cls):
        for path in cls._tmps:
            try:
                os.unlink(path)
            except OSError:
                pass
        for dt in ("BoQ Row BCS Rate", "BoQ Cell Remark", "BoQ Cell Amount Formula",
                   "BoQ Cell Pricing", "BoQ Committed Sheet Grid", "BoQ Sheet"):
            frappe.db.delete(dt, {"boq": cls.boq})
        frappe.db.delete("BOQs", {"name": cls.boq})
        frappe.db.commit()
        _cleanup_project(cls.project.name)
        super().tearDownClass()

    # -- placement ---------------------------------------------------------
    def test_the_block_lands_after_the_remark_column_at_the_true_data_edge(self):
        """F is the rightmost MAPPED column and `Nirmaan Remarks` now takes G, so the block
        starts at H -- the same scan-right rule, one column further along. Deliberately NOT
        openpyxl's max_column, and deliberately not an offset from F either.

        ⚠️ REVERSED AT THE LIVE CHECK (owner 2026-08-19). The block used to come first. All
        that changed is which of the two placers is CALLED first; neither knows about the
        other, and no arithmetic anywhere was adjusted."""
        block = self.result["cost_blocks"]["Elec "]
        self.assertEqual(block["cost_columns"], {"supply": "H", "install": "I"})
        self.assertEqual(block["total_column"], "J")

    def test_the_headers_read_exactly_as_the_grids_columns_read(self):
        ws = self.wb["Elec "]
        self.assertEqual(ws["H1"].value, "BCS Cost (Supply)")
        self.assertEqual(ws["I1"].value, "BCS Cost (Installation)")
        self.assertEqual(ws["J1"].value, "BCS Total Amount")

    def test_the_remark_column_comes_BEFORE_the_block(self):
        """★ OWNER RULING 2026-08-19, after the live check -- this REVERSES the original
        "block first, remarks last". `Nirmaan Remarks` keeps the position it holds in the
        CLIENT export, at the right-hand edge of the client's own data, and everything
        INTERNAL sits beyond it: costs, then the Total, then the margin.

        The whole layout falls out of the CALL ORDER. Both placers scan rightward past any
        occupied column, so whichever runs first claims the nearer one -- which is why this
        reversal touched no arithmetic and why the full left-to-right order is asserted here
        rather than each column in isolation."""
        ws = self.wb["Elec "]
        self.assertEqual(self.result["remark_columns"]["Elec "], "G")
        self.assertEqual(
            [ws[f"{c}1"].value for c in ("G", "H", "I", "J", "K")],
            ["Nirmaan Remarks", "BCS Cost (Supply)", "BCS Cost (Installation)",
             "BCS Total Amount", "% Margin"],
            "the internal block must sit entirely to the RIGHT of the shared remark column",
        )

    # -- values + formulas -------------------------------------------------
    def test_the_cost_cells_hold_the_stored_numbers_verbatim(self):
        ws = self.wb["Elec "]
        self.assertEqual(ws["H2"].value, _SUPPLY)
        self.assertEqual(ws["I2"].value, _INSTALL)
        self.assertEqual(ws["H3"].value, _SUPPLY)

    def test_the_total_is_a_formula_and_not_a_number(self):
        """The whole design: no server code computes a BCS number."""
        cell = self.wb["Elec "]["J2"]
        self.assertEqual(cell.data_type, "f")
        self.assertEqual(cell.value, '=IF(COUNT(C2)=0,"",(H2+I2)*(C2))')

    def test_a_declared_formula_is_translated_rather_than_replaced_by_the_builtin(self):
        """HVAC declares `(bcs_supply + bcs_install) * qty_total`, so its Total must read that
        shape -- and take the ANY-operand-missing guard, not the built-in's NOTHING-numeric
        one."""
        # ⚠️ HVAC CARRIES NO REMARKS, so its block still starts at G and its Total is I --
        # one column left of Elec's. That the two sheets differ is the placement rule working:
        # each sheet's layout follows ITS OWN occupied columns, never a workbook-wide offset.
        self.assertEqual(self.wb["HVAC "]["I2"].value, '=IF(COUNT(C2)<1,"",((G2+H2)*C2))')

    def test_an_uncosted_row_gets_no_cost_and_no_total(self):
        """Row 4 has no `BoQ Row BCS Rate` record, so the screen shows a blank. `(blank+blank)
        x qty` would be a confident 0 -- a claim that the row costs nothing."""
        ws = self.wb["Elec "]
        for addr in ("H4", "I4", "J4", "K4"):
            self.assertIsNone(ws[addr].value, addr)

    def test_a_costed_row_with_no_quantity_blanks_through_the_guard(self):
        """Row 5 IS costed but C5 is empty. Excel reads an empty cell as 0, so without the
        guard the Total would render 0 where the screen renders blank + `no quantity`."""
        ws = self.wb["Elec "]
        self.assertEqual(ws["H5"].value, _SUPPLY, "the cost itself is still written")
        self.assertEqual(ws["J5"].value, '=IF(COUNT(C5)=0,"",(H5+I5)*(C5))')

    # -- the % Margin column (slice 6) -------------------------------------
    def test_the_margin_column_lands_after_the_total_and_is_headed_plainly(self):
        """Column order is cost boxes, Total, margin -- the same order the grid presents them
        in, so a pricer reading the file left to right reads the same story they read on
        screen."""
        block = self.result["cost_blocks"]["Elec "]
        self.assertEqual(block["total_column"], "J")
        self.assertEqual(block["margin_column"], "K")
        self.assertIsNone(block["margin_skipped"])
        self.assertEqual(self.wb["Elec "]["K1"].value, "% Margin")

    def test_the_margin_is_a_formula_and_carries_the_sign_guard(self):
        """★ THE CENTRAL CASE. `(amount - cost) / amount x 100`, dividing by the AMOUNT cell
        and by the Total column this module wrote -- and wrapped in BOTH guards.

        ⚠️ THE `<=0` IS THE WHOLE REASON THIS COLUMN SHIPS instead of being left to a user to
        add by hand. A NEGATIVE denominator flips the inequality, so an amount of -100 against
        a cost of 50 computes +150%: a loss displayed as a profit. A hand-typed
        `=(F2-I2)/F2*100` computes the identical number on every ordinary row and gets that
        one catastrophically wrong, silently. The zero case rides the same test.

        The COUNT guard is the second half: Excel reads an empty amount cell as 0, which WOULD
        reach `<=0` and blank correctly -- but by the wrong route, and indistinguishably from
        a genuine zero."""
        cell = self.wb["Elec "]["K2"]
        self.assertEqual(cell.data_type, "f")
        self.assertEqual(cell.value, '=IF(OR(COUNT(F2)=0,F2<=0),"",((F2-J2)/F2)*100)')

    def test_the_margin_names_the_denominator_identically_in_the_guard_and_the_body(self):
        """The guard and the division must test the SAME expression. Written separately they
        could drift -- guarding one column while dividing by another is a blank that arrives
        for a reason unrelated to the number on the row, which is worse than no guard because
        it looks deliberate."""
        import re

        value = self.wb["Elec "]["K2"].value
        guard = re.search(r"OR\(COUNT\(([^)]+)\)=0,([^<]+)<=0\)", value)
        self.assertIsNotNone(guard, value)
        self.assertEqual(guard.group(1), guard.group(2))
        self.assertIn(f"({guard.group(2)}-", value)
        self.assertIn(f")/{guard.group(2)})", value)

    def test_the_margins_numerator_defaults_to_the_total_column_by_reference(self):
        """NOT a copy of the Total's rule -- a REFERENCE to its cell. That is what keeps the
        margin live: edit a cost in the workbook and both the Total and the margin follow. A
        second copy of `(G+H)*C` inlined here would compute the same number today and drift
        the moment anyone touched the Total."""
        self.assertIn("J2", self.wb["Elec "]["K2"].value)
        self.assertNotIn("H2+I2", self.wb["Elec "]["K2"].value)

    def test_a_declared_bcs_total_flows_into_the_margin_through_the_same_reference(self):
        """HVAC's Total is a DECLARED formula, and its margin still just points at the Total
        column. So the numerator inherits whatever that sheet declared, with no second
        translation and nothing to keep in step."""
        # HVAC has no remark column, so its margin is J and it divides by the Total in I.
        self.assertEqual(self.wb["HVAC "]["J2"].value,
                         '=IF(OR(COUNT(F2)=0,F2<=0),"",((F2-I2)/F2)*100)')

    def test_an_uncosted_row_gets_no_margin_either(self):
        """Row 4 has no cost record, so it gets no cost, no Total and no margin. A margin on
        an uncosted row would read 100% -- 'this row is pure profit' -- which is the most
        confidently wrong number this column could produce."""
        self.assertIsNone(self.wb["Elec "]["K4"].value)

    def test_a_costed_row_with_no_quantity_still_gets_a_margin_cell(self):
        """⚠️ DELIBERATE, and it is not an oversight. Row 5 has no QUANTITY, so its Total
        blanks -- but the margin's guard is about the AMOUNT, and F5 is a different question.
        The formula is written and Excel resolves it: with a real amount and a blank Total it
        reads 100%, and with no amount the COUNT guard blanks it. The row is not special-cased
        here because the margin does not depend on quantity; inventing a dependency would make
        the column lie about which fact it is missing."""
        self.assertEqual(self.wb["Elec "]["K5"].value,
                         '=IF(OR(COUNT(F5)=0,F5<=0),"",((F5-J5)/F5)*100)')

    # -- the light-blue fill on filled BCS cells (owner 2026-08-19) --------
    def _fill_of(self, sheet, addr):
        f = self.wb[sheet][addr].fill
        return (f.fgColor.rgb or "")[-6:] if f and f.fill_type == "solid" else None

    def test_every_filled_bcs_cell_carries_the_light_blue_fill(self):
        """★ OWNER RULING after the live check: a BCS cell holding a figure is marked, the way
        a stamped rate cell is marked -- so a reader can see at a glance which rows were costed
        without reading the numbers. Costs, the Total and the margin all count as figures."""
        for addr in ("H2", "I2", "J2", "K2", "H3", "I3", "J3", "K3"):
            self.assertEqual(self._fill_of("Elec ", addr), _BCS_FILLED_HEX, addr)

    def test_an_uncosted_row_is_left_UNFILLED(self):
        """⚠️ THE FILL MARKS CELLS, NOT COLUMNS, and this is the case that proves it. Row 4 has
        no cost record, so its BCS cells are empty and stay unfilled. Filling the column's full
        height would say "every row is costed" -- the same false claim in colour that the COUNT
        guard exists to stop the Total making in numbers."""
        for addr in ("H4", "I4", "J4", "K4"):
            self.assertIsNone(self._fill_of("Elec ", addr), addr)

    def test_the_header_row_is_not_filled(self):
        """A header is a label, not a figure -- and the rate highlight this mirrors marks no
        header either."""
        for addr in ("H1", "I1", "J1", "K1"):
            self.assertIsNone(self._fill_of("Elec ", addr), addr)

    def test_a_costed_row_with_no_quantity_is_still_filled_where_it_has_figures(self):
        """Row 5 is costed but has no quantity. Its COSTS are real figures and are filled; its
        Total and margin are formulas we wrote, so they are filled too -- what those formulas
        RESOLVE to is Excel's business and is not something the fill can or should predict."""
        self.assertEqual(self._fill_of("Elec ", "H5"), _BCS_FILLED_HEX)
        self.assertEqual(self._fill_of("Elec ", "J5"), _BCS_FILLED_HEX)

    def test_the_fill_never_reaches_the_clients_own_columns(self):
        """The mark belongs to the internal block. A client column -- including the shared
        remark column now sitting immediately left of it -- must be untouched by it."""
        for addr in ("A2", "C2", "F2"):
            self.assertNotEqual(self._fill_of("Elec ", addr), _BCS_FILLED_HEX, addr)
        self.assertNotEqual(self._fill_of("Elec ", "G2"), _BCS_FILLED_HEX,
                            "the remark column is shared, not internal")

    def test_the_fill_leaves_the_value_alone(self):
        """A fill sets `.fill` and nothing else -- the number and the formula both survive it.
        Asserted because a styling pass that quietly rewrote a cell would be invisible until a
        pricer opened the file."""
        ws = self.wb["Elec "]
        self.assertEqual(ws["H2"].value, _SUPPLY)
        self.assertEqual(ws["J2"].data_type, "f")

    # -- skips, reported not silent ---------------------------------------
    def test_a_grid_only_sheet_gets_no_block_and_says_why(self):
        self.assertNotIn("Specs ", self.result["cost_blocks"])
        self.assertIn("general-specs", self.result["cost_skipped"]["Specs "])

    def test_a_bcs_disabled_sheet_is_exported_plain_and_says_why(self):
        """Owner ruling: a sheet with cost tracking off is still exportable, just without the
        block -- you often want the whole book with cost on the sheets that have it."""
        frappe.db.set_value("BoQ Sheet",
                            {"boq": self.boq, "sheet_name": "HVAC ", "is_current": 1},
                            "bcs_enabled", 0, update_modified=False)
        frappe.db.commit()
        try:
            out = _generate_internal_workbook(self.boq, ["HVAC "], self._synthetic_workbook(),
                                              "x")
            self.assertNotIn("HVAC ", out["cost_blocks"])
            self.assertIn("switched off", out["cost_skipped"]["HVAC "])
            self.assertEqual(out["exported_sheets"], ["HVAC "])
        finally:
            frappe.db.set_value("BoQ Sheet",
                                {"boq": self.boq, "sheet_name": "HVAC ", "is_current": 1},
                                "bcs_enabled", 1, update_modified=False)
            frappe.db.commit()

    # -- the client export is untouched -----------------------------------
    def test_the_client_export_still_carries_no_cost_from_the_same_fixture(self):
        """★ THE PAIRING. The standing guard in test_export_writeback says the client file has
        no BCS in it; on its own that could pass because there was nothing to find. Here the
        SAME BoQ produces both files, and only one of them carries the cost."""
        client = load_workbook(
            io.BytesIO(base64.b64decode(self.client_result["content_base64"])), data_only=False
        )
        for sentinel in (_SUPPLY, _INSTALL):
            rendered = repr(sentinel)
            hits = [str(c.value) for ws in client.worksheets for row in ws.iter_rows()
                    for c in row if c.value is not None and rendered in str(c.value)]
            self.assertEqual(hits, [], f"{sentinel} leaked into the CLIENT workbook")

    def test_the_internal_export_really_does_carry_it(self):
        """The anti-vacuity half of the pairing."""
        hits = [str(c.value) for ws in self.wb.worksheets for row in ws.iter_rows()
                for c in row if c.value is not None and repr(_SUPPLY) in str(c.value)]
        self.assertTrue(hits, "the internal workbook must carry the cost")

    def test_the_client_export_has_no_extra_columns(self):
        """G/H/I exist only in the internal file. The client's remark column still lands at G,
        which is where it landed before this module existed."""
        client = load_workbook(
            io.BytesIO(base64.b64decode(self.client_result["content_base64"])), data_only=False
        )
        self.assertEqual(self.client_result["remark_columns"]["Elec "], "G")
        self.assertIsNone(client["Elec "]["H1"].value)

    def test_this_export_writes_nothing_to_the_database(self):
        """Owner ruling: `last_exported_at` means 'when the CLIENT last got this sheet'. An
        internal download stamping it would make the changed-since-export chip claim the client
        holds something they have never been sent. The payload does not even carry the key."""
        self.assertNotIn("last_exported_at", self.result)
        self.assertIsNone(self.stamp_after_internal)

    def test_the_fidelity_delta_is_really_exercised(self):
        """★ THE GUARD IS ONLY WORTH SOMETHING IF THIS EXPORT ADDS FORMULAS, and it does: the
        source carries 9 (three amount formulas on each of three sheets) and the product
        carries 21: the six Totals AND the six % Margins on the two costed sheets. Without
        this, `before + 0` would be the plain equality guard and the delta path would be
        untested -- a guard that looks adjusted and is not. Miscounting fails LOUDLY: dropping
        the count aborts the export with `formulas: 9 -> 21` and produces no file, which is
        the reject-mutates-nothing behaviour the client export already has.

        ⚠️ 15 -> 21 AT SLICE 6. The split below is asserted per KIND, not as one total, so a
        slice that silently stopped writing Totals while adding Margins would keep the sum and
        still go red here."""
        produced = [c for ws in self.wb.worksheets for row in ws.iter_rows()
                    for c in row if c.data_type == "f"]
        self.assertEqual(len(produced), 21)
        ours = [c for c in produced if str(c.value).startswith("=IF(")]
        self.assertEqual(len(ours), 12, "a Total and a Margin on three costed rows, twice")
        margins = [c for c in ours if str(c.value).endswith(")*100)")]
        self.assertEqual(len(margins), 6, "three costed rows on each of the two costed sheets")
        self.assertEqual(len(ours) - len(margins), 6, "and a Total beside each one")

    def test_the_filename_says_the_file_is_internal(self):
        self.assertIn("priced_bcs_internal", self.result["filename"])
        self.assertTrue(self.result["filename"].endswith(".xlsx"))

    def test_the_clients_own_formulas_survive_the_round_trip(self):
        """The fidelity guard's real job. F still holds the client's amount formula, and the
        rate stamped into D is there beside it."""
        ws = self.wb["Elec "]
        self.assertEqual(ws["F2"].value, "=C2*(D2+E2)")
        self.assertEqual(ws["D2"].value, 25)


# ===========================================================================
# Group 4: the refusals, and the construction that keeps the wall standing
# ===========================================================================
class TestRefusalsAndConstruction(FrappeTestCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.project = _make_project()
        boq = frappe.new_doc("BOQs")
        boq.project = cls.project.name
        boq.boq_name = "Template Origin BoQ"
        boq.tax_treatment = "Pre-tax"
        boq.origin = "template"
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        cls.boq = boq.name

    @classmethod
    def tearDownClass(cls):
        frappe.db.delete("BOQs", {"name": cls.boq})
        frappe.db.commit()
        _cleanup_project(cls.project.name)
        super().tearDownClass()

    def test_a_template_origin_boq_is_refused_by_name(self):
        """A template BoQ has no source workbook -- its priced export is BUILT FROM SCRATCH by
        a different generator. v1 says so out loud rather than failing obscurely on a missing
        source_file_url."""
        with self.assertRaises(frappe.ValidationError) as ctx:
            export_priced_workbook_with_bcs(boq_name=self.boq,
                                            sheet_names=json.dumps(["Any Sheet"]))
        self.assertIn("master template", str(ctx.exception))

    def test_the_gate_runs_before_anything_is_looked_up(self):
        """A user outside admin + estimation is refused, and refused FIRST -- the BoQ is never
        even resolved, so the refusal cannot leak whether it exists."""
        frappe.set_user("Guest")
        try:
            with self.assertRaises(frappe.PermissionError):
                export_priced_workbook_with_bcs(boq_name="does-not-exist",
                                                sheet_names=json.dumps(["x"]))
        finally:
            frappe.set_user("Administrator")

    def test_the_client_export_module_is_untouched_and_still_names_no_cost(self):
        """★ THE SEPARATION, asserted from this side too. `test_export_writeback` greps its own
        module; this pins the same property from the module that DOES name BCS, so the pair
        cannot both be deleted by someone merging the two exports."""
        import inspect

        from nirmaan_stack.api.boq.wizard import export_bcs_writeback, export_writeback

        client_src = inspect.getsource(export_writeback).lower()
        for token in ("boq row bcs rate", "supply_rate", "install_rate", "combined_rate", "bcs"):
            self.assertNotIn(token, client_src,
                             f"the CLIENT export must never name {token!r}")
        ours = inspect.getsource(export_bcs_writeback).lower()
        self.assertIn("boq row bcs rate", ours,
                      "and this module must -- or the guard above is vacuous")

    def test_this_module_never_stamps_the_client_export_marker(self):
        """`last_exported_at` is the client's staleness signal. Grepping for the assignment is
        cheap insurance against it being reintroduced by a copy-paste from the client path."""
        import inspect

        from nirmaan_stack.api.boq.wizard import export_bcs_writeback

        src = inspect.getsource(export_bcs_writeback)
        self.assertNotIn('"last_exported_at"', src)
        self.assertNotIn("frappe.db.commit()", src,
                         "this export writes nothing, so it needs no commit")


# ===========================================================================
# Group 5: placement against real content, the combined-rate shape, and every
#          skip branch
# ===========================================================================
_COMBINED_ROLE_MAP = {
    "A": {"role": "description", "area": None},
    "C": {"role": "qty_total", "area": None},
    "D": {"role": "rate_combined", "area": None},
    "F": {"role": "amount_total", "area": None},
}
_NO_RATE_ROLE_MAP = {
    "A": {"role": "description", "area": None},
    "C": {"role": "qty_total", "area": None},
    "F": {"role": "amount_total", "area": None},
}
_NO_QTY_ROLE_MAP = {
    "A": {"role": "description", "area": None},
    "D": {"role": "rate_supply", "area": None},
    "E": {"role": "rate_install", "area": None},
    "F": {"role": "amount_total", "area": None},
}
# ⚠️ A ROLE MAP IN WHICH NO KEY IS AN EXCEL LETTER. `_build_column_descriptors` accepts ANY key
# (it never parses one), while `_rightmost_mapped_col_index` skips what it cannot parse -- so
# this is the ONE shape that reaches the "no mapped columns" branch with rate columns present.
#
# ⚠️ EVERY key has to be unparseable, and the first draft of this fixture got that wrong: it
# also mapped a description on "A", which parses, so the edge resolved to 1, the block started
# at B, and the branch never fired. The test caught it. A single valid column anywhere in the
# map is enough to define an edge.
_BAD_LETTER_ROLE_MAP = {
    "??": {"role": "rate_supply", "area": None},
}
# Rates, a quantity and therefore a Total -- but NO amount column anywhere, so there is
# nothing to measure a margin against. The sheet is otherwise entirely healthy, which is the
# point: it gets its full cost block and its Total, and only the margin is absent.
_NO_AMOUNT_ROLE_MAP = {
    "A": {"role": "description", "area": None},
    "C": {"role": "qty_total", "area": None},
    "D": {"role": "rate_supply", "area": None},
    "E": {"role": "rate_install", "area": None},
}


class TestPlacementCombinedRateAndEverySkip(FrappeTestCase):
    """The branches Group 2 does not reach: the block landing beside PRE-EXISTING content, the
    single-box combined-rate sheet, and each of the four remaining skip reasons.

    Every skip carries a REASON to the results modal, and each is asserted by its own case --
    a shared "some reason was returned" assertion would let two branches swap without a test
    noticing, and the reason is the whole difference between "this sheet costs nothing" and
    "this sheet was never costed"."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls._tmps = []
        cls.project = _make_project()
        boq = frappe.new_doc("BOQs")
        boq.project = cls.project.name
        boq.boq_name = "Placement And Skips BoQ"
        boq.tax_treatment = "Pre-tax"
        boq.source_file_url = "/files/synthetic.xlsx"
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        cls.boq = boq.name
        now = frappe.utils.now()
        grid = [{"row_number": r, "cells": json.dumps({})} for r in range(1, 7)]

        for sheet, role_map, costed in (
            ("Stray ", _ROLE_MAP, True),          # a workbook column already holds a note
            ("Combined ", _COMBINED_ROLE_MAP, True),
            ("NoRate ", _NO_RATE_ROLE_MAP, True),
            ("NoCosts ", _ROLE_MAP, False),
            ("NoQty ", _NO_QTY_ROLE_MAP, True),
            ("BadLetter ", _BAD_LETTER_ROLE_MAP, True),
            ("NoAmount ", _NO_AMOUNT_ROLE_MAP, True),
            ("BadAmount ", _ROLE_MAP, True),
            ("DeclAmt ", _ROLE_MAP, True),
            ("DeclCost ", _ROLE_MAP, True),
        ):
            _seed_committed_sheet(cls.boq, sheet, "data", 1, role_map, [], 1, sheet.strip(),
                                  now, grid, [], None)
            frappe.db.set_value("BoQ Sheet",
                                {"boq": cls.boq, "sheet_name": sheet, "is_current": 1},
                                "bcs_enabled", 1, update_modified=False)
            if not costed:
                continue
            for excel_row in (2, 3):
                doc = frappe.new_doc("BoQ Row BCS Rate")
                doc.boq = cls.boq
                doc.sheet_name = sheet
                doc.excel_row = excel_row
                doc.committed_version = 1
                doc.supply_rate = _SUPPLY
                doc.install_rate = _INSTALL
                doc.combined_rate = _SUPPLY + _INSTALL
                doc.is_filled = 1
                doc.bcs_version = 1
                doc.is_current = 1
                doc.bcs_rated_at = now
                doc.insert(ignore_permissions=True)
        frappe.db.commit()

        # BadAmount declares a `boq_total` naming `amount_supply`, which its role map does NOT
        # carry -- so the denominator tree resolves on no row and the margin is skipped for a
        # reason distinct from "this sheet maps no amount column". The sheet DOES map
        # amount_total, so it never reaches that earlier branch.
        for sheet, target, tree in (
            # A denominator that cannot resolve -- its own skip reason.
            ("BadAmount ", "boq_total", _ref("amount_supply")),
            # A denominator that CAN: the same column the default would have found, declared
            # explicitly, so the two paths are told apart by the GUARD they take rather than
            # by the cells they name.
            ("DeclAmt ", "boq_total", _ref("amount_total")),
            # A numerator naming the raw cost boxes instead of the Total column.
            ("DeclCost ", "bcs_margin_cost",
             {"op": "+", "operands": [_ref("bcs_supply"), _ref("bcs_install")]}),
        ):
            af = frappe.new_doc("BoQ Cell Amount Formula")
            af.boq = cls.boq
            af.sheet_name = sheet
            af.committed_version = 1
            af.target_value_field = target
            af.target_col = None
            af.formula = json.dumps(tree)
            af.formula_version = 1
            af.is_current = 1
            af.defined_at = frappe.utils.now()
            af.insert(ignore_permissions=True)
        frappe.db.commit()

        cls.sheets = ["Stray ", "Combined ", "NoRate ", "NoCosts ", "NoQty ", "BadLetter ",
                      "NoAmount ", "BadAmount ", "DeclAmt ", "DeclCost "]
        cls.result = _generate_internal_workbook(
            cls.boq, cls.sheets, cls._workbook(), "Placement And Skips BoQ"
        )
        cls.wb = load_workbook(
            io.BytesIO(base64.b64decode(cls.result["content_base64"])), data_only=False
        )

    @classmethod
    def _workbook(cls):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for title in ("Stray ", "Combined ", "NoRate ", "NoCosts ", "NoQty ", "BadLetter ",
                      "NoAmount ", "BadAmount ", "DeclAmt ", "DeclCost "):
            ws = wb.create_sheet(title=title)
            for r in (2, 3):
                ws[f"A{r}"] = f"item {r}"
                ws[f"C{r}"] = 10 * r
        # ONE sheet already carries a note in the first column past the mapped edge.
        wb["Stray "]["G2"] = "estimator note -- must survive"
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb.save(path)
        cls._tmps.append(path)
        return path

    @classmethod
    def tearDownClass(cls):
        for path in cls._tmps:
            try:
                os.unlink(path)
            except OSError:
                pass
        for dt in ("BoQ Row BCS Rate", "BoQ Cell Amount Formula", "BoQ Cell Pricing",
                   "BoQ Committed Sheet Grid", "BoQ Sheet"):
            frappe.db.delete(dt, {"boq": cls.boq})
        frappe.db.delete("BOQs", {"name": cls.boq})
        frappe.db.commit()
        _cleanup_project(cls.project.name)
        super().tearDownClass()

    # -- placement against real content -----------------------------------
    def test_the_block_steps_past_a_pre_existing_column_and_never_overwrites_it(self):
        """★ The mapped edge is F, so the block WOULD start at G -- but G already holds an
        estimator note. It must step to H, and the note must still be there afterwards."""
        ws = self.wb["Stray "]
        self.assertEqual(ws["G2"].value, "estimator note -- must survive")
        self.assertEqual(self.result["cost_blocks"]["Stray "]["cost_columns"],
                         {"supply": "H", "install": "I"})
        self.assertEqual(self.result["cost_blocks"]["Stray "]["total_column"], "J")
        self.assertEqual(ws["H1"].value, "BCS Cost (Supply)")

    def test_the_total_formula_follows_the_shifted_columns(self):
        """The shift has to reach the FORMULA too, or the Total would read the note's column."""
        self.assertEqual(self.wb["Stray "]["J2"].value, '=IF(COUNT(C2)=0,"",(H2+I2)*(C2))')

    # -- the combined-rate shape ------------------------------------------
    def test_a_combined_rate_sheet_gets_ONE_undifferentiated_box(self):
        """`live_rate_kinds` returns ["combined"] for a sheet that quotes one figure, so the
        block is two columns, not three, and the box is headed `BCS Cost` with no qualifier."""
        block = self.result["cost_blocks"]["Combined "]
        self.assertEqual(block["cost_columns"], {"combined": "G"})
        self.assertEqual(block["total_column"], "H")
        self.assertEqual(self.wb["Combined "]["G1"].value, "BCS Cost")

    def test_the_combined_total_multiplies_the_ONE_box(self):
        """⚠️ And it must never also add the halves. `bcs.py` forbids summing combined_rate
        with them -- the fixture stores all three, so a formula naming more than G would be
        double-counting numbers that are genuinely present in the row."""
        self.assertEqual(self.wb["Combined "]["H2"].value, '=IF(COUNT(C2)=0,"",(G2)*(C2))')
        self.assertEqual(self.wb["Combined "]["G2"].value, _SUPPLY + _INSTALL)

    # -- every remaining skip branch, each by its own reason ---------------
    def test_a_sheet_with_no_rate_column_is_skipped_by_name(self):
        self.assertNotIn("NoRate ", self.result["cost_blocks"])
        self.assertIn("maps no rate column", self.result["cost_skipped"]["NoRate "])

    def test_a_sheet_nobody_has_costed_is_skipped_by_name(self):
        """Distinct from 'no rate column': this sheet COULD be costed and simply has not
        been. Collapsing the two would tell a pricer to fix the wrong thing."""
        self.assertNotIn("NoCosts ", self.result["cost_blocks"])
        self.assertIn("no costs have been entered", self.result["cost_skipped"]["NoCosts "])

    def test_a_sheet_whose_quantity_cannot_be_resolved_still_gets_its_COST_columns(self):
        """★ A PARTIAL BLOCK, DELIBERATELY. The costs are facts we hold and can write; only
        the Total needs a quantity. Withholding the cost columns too would discard information
        that is perfectly good, and the reason says exactly what is missing."""
        block = self.result["cost_blocks"]["NoQty "]
        self.assertEqual(block["cost_columns"], {"supply": "G", "install": "H"})
        self.assertIsNone(block["total_column"], "no quantity, so no Total")
        self.assertIn("quantity columns could not be resolved",
                      self.result["cost_skipped"]["NoQty "])
        self.assertEqual(self.wb["NoQty "]["G2"].value, _SUPPLY)

    def test_an_unparseable_column_key_reaches_the_undefined_data_edge_branch(self):
        """The ONE shape that gets here: `_build_column_descriptors` accepts any role-map key
        and never parses it, while `_rightmost_mapped_col_index` skips what it cannot parse.
        So rate columns exist while the data edge does not, and the block has nowhere to
        start. Pinned rather than assumed -- the branch reads as dead code otherwise.

        ⚠️ IT NEEDS EVERY KEY TO BE UNPARSEABLE. One valid column anywhere in the map defines
        an edge and the block places normally -- which is what the first version of this
        fixture did, mapping a description on "A" beside the bad key and never reaching the
        branch at all."""
        self.assertNotIn("BadLetter ", self.result["cost_blocks"])
        self.assertIn("no mapped columns", self.result["cost_skipped"]["BadLetter "])

    # -- the % Margin column's DECLARED-formula paths (slice 6) -------------
    def test_a_declared_denominator_takes_the_all_operands_guard(self):
        """DeclAmt declares `boq_total = amount_total` -- the same column the DEFAULT path
        finds. The cells are therefore identical and the GUARD is what differs: a declared
        formula blanks when ANY operand is missing (`COUNT < n`), the default when NOTHING
        resolves (`COUNT = 0`). Same shape as the Total's two paths, for the same reason --
        `evaluateBcsTotalFormula` returns on the first unresolved ref while `bcsRowAmount`
        sums whatever it finds."""
        self.assertEqual(self.wb["DeclAmt "]["J2"].value,
                         '=IF(OR(COUNT(F2)<1,F2<=0),"",((F2-I2)/F2)*100)')

    def test_a_declared_numerator_replaces_the_reference_to_the_total_column(self):
        """DeclCost measures against the raw cost boxes rather than the Total. So the
        numerator reads `(G+H)` and the Total column is not mentioned at all -- the sheet said
        which cost figure it means, and it is not the Total."""
        value = self.wb["DeclCost "]["J2"].value
        self.assertEqual(value, '=IF(OR(COUNT(F2)=0,F2<=0),"",((F2-(G2+H2))/F2)*100)')
        self.assertNotIn("I2", value, "the Total column is deliberately not read here")
        self.assertEqual(self.wb["DeclCost "]["I1"].value, "BCS Total Amount",
                         "and the Total column still exists -- it is simply not the numerator")

    def test_the_combined_rate_sheet_gets_a_margin_over_its_one_box(self):
        """A single undifferentiated cost box, so the Total multiplies only G -- and the
        margin divides that Total into the amount exactly as anywhere else. The margin needs
        no combined-rate special case, because it never touches the boxes."""
        self.assertEqual(self.wb["Combined "]["I2"].value,
                         '=IF(OR(COUNT(F2)=0,F2<=0),"",((F2-H2)/F2)*100)')
        self.assertEqual(self.result["cost_blocks"]["Combined "]["margin_column"], "I")

    # -- the % Margin column's OWN skips (slice 6) -------------------------
    def test_a_sheet_with_no_amount_column_gets_its_costs_and_total_but_no_margin(self):
        """★ THE MARGIN SKIPS ON ITS OWN, and this sheet is why the reason had to be separate
        rather than folded into the block's. Rates, a quantity, costs and a Total -- nothing
        about it is wrong -- and simply no column saying what we CHARGE. Reporting "no Total"
        here would be false, and reporting nothing would be the silent absence this whole
        module refuses."""
        block = self.result["cost_blocks"]["NoAmount "]
        self.assertEqual(block["cost_columns"], {"supply": "F", "install": "G"})
        self.assertEqual(block["total_column"], "H", "the Total is unaffected")
        self.assertIsNone(block["margin_column"])
        self.assertIn("maps no amount column", block["margin_skipped"])

    def test_a_declared_denominator_that_resolves_on_no_row_is_its_own_reason(self):
        """A `boq_total` naming `amount_supply` on a sheet that maps only `amount_total`. The
        sheet DOES have an amount column, so it never reaches the earlier branch -- the
        formula simply cannot be resolved, which is a different fact and gets a different
        sentence.

        ⚠️ IT FAIL-SAFES TO BLANK RATHER THAN FALLING BACK to the sheet's own amount columns.
        A declared formula is a statement about which figure the margin measures against;
        quietly substituting a different one would produce a percentage nobody asked for, and
        it would look right."""
        block = self.result["cost_blocks"]["BadAmount "]
        self.assertEqual(block["total_column"], "I", "the Total still lands")
        self.assertIsNone(block["margin_column"])
        self.assertIn("could not be resolved on any", block["margin_skipped"])

    def test_a_sheet_with_no_total_column_has_no_margin_to_measure(self):
        """The margin's DEFAULT numerator IS the Total column, so a sheet that got none has no
        cost figure to divide. Not a zero -- a zero would claim the row costs nothing."""
        block = self.result["cost_blocks"]["NoQty "]
        self.assertIsNone(block["total_column"])
        self.assertIsNone(block["margin_column"])
        self.assertIn("no BCS Total Amount column", block["margin_skipped"])

    def test_every_sheet_is_still_reported_as_exported(self):
        """A skipped BLOCK is not a skipped SHEET -- each of these still went through the
        client-facing passes and still belongs in the file."""
        self.assertEqual(sorted(self.result["exported_sheets"]), sorted(self.sheets))

    def test_every_skip_reason_in_the_module_is_covered_by_one_of_these_cases(self):
        """★ ANTI-DRIFT. A branch added later with a new reason must come with a case, or this
        goes red -- otherwise a silent skip reappears exactly where the whole point is that a
        missing cost block must never be silent. `general-specs` and `switched off` are pinned
        in Group 2, so the union of both groups is what is compared."""
        import inspect
        import re

        from nirmaan_stack.api.boq.wizard import export_bcs_writeback

        src = inspect.getsource(export_bcs_writeback)
        reasons = set(re.findall(r'"reason": "([^"]+)', src))
        reasons |= set(re.findall(r'"reason": None if bodies else "([^"]+)', src))
        covered = set(self.result["cost_skipped"].values()) | {
            "this is a general-specs sheet, which carries no priced rows",
            "cost tracking is switched off for this sheet",
        }
        # ⚠️ A MARGIN SKIP IS REPORTED ON THE BLOCK, NOT IN `cost_skipped` -- the sheet DID get
        # a cost block, so it is not a skipped block. Slice 6 added three such reasons, and
        # without this line the guard would have gone on comparing only the block's own set
        # and called itself complete: an anti-drift guard that quietly stopped covering a
        # whole family is worse than none, because it still reads as green.
        covered |= {b["margin_skipped"] for b in self.result["cost_blocks"].values()
                    if b.get("margin_skipped")}
        for reason in reasons:
            with self.subTest(reason=reason):
                self.assertTrue(any(c.startswith(reason[:40]) for c in covered),
                                f"no test exercises the skip reason {reason!r}")
