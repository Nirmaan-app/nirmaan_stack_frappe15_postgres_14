"""
Preview endpoint: return a window of raw cell values from a BoQ sheet.

Performance rationale: openpyxl.load_workbook(path, data_only=True, read_only=True)
takes ~0.56 s on a 7.65 MB workbook.  BoqReader takes ~27 s on the same file because
it opens the workbook TWICE (data_only + formula pass) and pre-scans merged ranges.
This endpoint does NOT use BoqReader -- raw openpyxl read_only is the correct path for
a synchronous preview.

S3 safety: BOQs.source_file_url is an S3 API redirect URL after the frappe_s3_attachment
plugin moves the file.  frappe.get_doc("File", ...).get_content() does not work because
the plugin deletes the local file after upload.  Bytes are fetched via S3Operations and
written to a NamedTemporaryFile that is always unlinked in a finally block.
"""
from __future__ import annotations

import datetime
import os
import tempfile
import urllib.parse

import frappe
import openpyxl
from openpyxl.utils import get_column_letter

_PREVIEW_MAX_ROWS = 200  # hard cap on a single preview window


def _derive_s3_key(source_file_url: str) -> str:
    """Return the S3 object key from a frappe_s3_attachment private file URL.

    Primary: parse the 'key' query param.  Private-file URL format (controller.py line 142):
      /api/method/frappe_s3_attachment.controller.generate_file?key=<KEY>&file_name=<encoded>
    Fallback: the plugin stores the S3 key in the File doctype's content_hash column
    (controller.py line 148), so a direct DB lookup works if the URL lacks the param.
    """
    parsed = urllib.parse.urlparse(source_file_url)
    params = urllib.parse.parse_qs(parsed.query)
    key_list = params.get("key")
    if key_list:
        return key_list[0]

    key = frappe.db.get_value("File", {"file_url": source_file_url}, "content_hash")
    if key:
        return key

    frappe.throw(
        f"Cannot derive S3 key from source_file_url: {source_file_url!r}. "
        "Expected /api/method/frappe_s3_attachment.controller.generate_file?key=<KEY>&...",
        title="S3 key not found",
    )


def _fetch_boq_file_to_tempfile(source_file_url: str) -> str:
    """Download the BoQ file from S3 to a NamedTemporaryFile; return the path.

    Caller is responsible for os.unlink in a finally block.
    Raises a frappe error if the key cannot be derived or the S3 fetch fails.
    The tempfile is only created after bytes are successfully downloaded, so a
    failed fetch never leaves an orphaned file.
    """
    from frappe_s3_attachment.controller import S3Operations  # noqa: PLC0415

    key = _derive_s3_key(source_file_url)

    try:
        s3 = S3Operations()
        response = s3.read_file_from_s3(key)
        file_bytes = response["Body"].read()
    except Exception as exc:
        frappe.throw(
            f"Failed to fetch BoQ file from S3 (key={key!r}): {exc}",
            title="S3 fetch failed",
        )

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    try:
        tmp.write(file_bytes)
    finally:
        tmp.close()

    return tmp.name


def _to_json_serializable(value):
    """Coerce any openpyxl cell value to a JSON-primitive or None."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        return value
    if isinstance(value, datetime.datetime):
        return value.isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, datetime.timedelta):
        return str(value)
    return str(value)


@frappe.whitelist()
def get_sheet_preview(boq_name=None, sheet_name=None, start_row=1, end_row=40):
    """Return a windowed preview of raw cell values for one sheet of a BoQ workbook.

    Values-only (data_only=True) + streaming (read_only=True) open -- fast enough for
    synchronous use.  Does NOT use BoqReader (see module docstring for timing rationale).

    Params (Frappe passes query-string values as strings; all are coerced internally):
      boq_name   -- required; must match an existing BOQs document name.
      sheet_name -- required; VERBATIM match against workbook sheet names (no strip).
      start_row  -- 1-indexed, inclusive (default 1).
      end_row    -- 1-indexed, inclusive (default 40).  Window capped at _PREVIEW_MAX_ROWS.

    Returns:
      {
        "sheet_name": str,           # echoed verbatim
        "start_row": int,
        "end_row_requested": int,    # after clamping
        "rows": [{"row_number": int, "cells": {col_letter: value}}, ...],
        "returned_count": int,
        "has_more": bool,            # True when the sheet has rows beyond end_row
      }

    URL: /api/method/nirmaan_stack.api.boq.wizard.sheet_preview.get_sheet_preview
    """
    if not boq_name:
        frappe.throw("boq_name is required.", title="Missing field: boq_name")
    if not sheet_name:
        frappe.throw("sheet_name is required.", title="Missing field: sheet_name")

    try:
        start_row = int(start_row)
        end_row = int(end_row)
    except (ValueError, TypeError):
        frappe.throw("start_row and end_row must be integers.", title="Invalid params")

    if start_row < 1:
        frappe.throw("start_row must be >= 1.", title="Invalid params")
    if end_row < start_row:
        frappe.throw("end_row must be >= start_row.", title="Invalid params")

    # Clamp silently; no error -- frontend paginator should never exceed the cap anyway.
    if end_row - start_row + 1 > _PREVIEW_MAX_ROWS:
        end_row = start_row + _PREVIEW_MAX_ROWS - 1

    if not frappe.db.exists("BOQs", boq_name):
        frappe.throw(f"BOQs '{boq_name}' not found.", title="Not found")

    source_file_url = frappe.db.get_value("BOQs", boq_name, "source_file_url")
    if not source_file_url:
        frappe.throw(
            f"BOQs '{boq_name}' has no source_file_url set.",
            title="Missing source file",
        )

    tempfile_path = None
    wb = None
    try:
        tempfile_path = _fetch_boq_file_to_tempfile(source_file_url)
        wb = openpyxl.load_workbook(tempfile_path, data_only=True, read_only=True)

        if sheet_name not in wb.sheetnames:
            frappe.throw(
                f"Sheet '{sheet_name}' not found in the workbook. "
                f"Available sheets: {wb.sheetnames}",
                title="Sheet not found",
            )

        ws = wb[sheet_name]
        sheet_max_row = ws.max_row  # from dimension metadata; may be None for empty sheets

        rows_out = []
        for row_cells in ws.iter_rows(min_row=start_row, max_row=end_row):
            if not row_cells:
                continue
            # In read_only mode openpyxl returns EmptyCell objects for padding positions
            # (up to the sheet's max_column). EmptyCell has .value=None but no .row/.column.
            # Skip entirely-empty padding rows; skip individual EmptyCell within real rows.
            row_num = next((c.row for c in row_cells if hasattr(c, "row")), None)
            if row_num is None:
                continue
            cells = {
                get_column_letter(cell.column): _to_json_serializable(cell.value)
                for cell in row_cells
                if hasattr(cell, "column")
            }
            rows_out.append({"row_number": row_num, "cells": cells})

        returned_count = len(rows_out)

        # has_more: True if the sheet has content rows beyond end_row.
        # ws.max_row is read from the <dimension> tag in read_only mode (reliable for
        # well-formed xlsx).  Fallback: proxy from returned_count when max_row is None
        # (e.g. empty/malformed sheets).
        if sheet_max_row is not None:
            has_more = sheet_max_row > end_row
        else:
            # If we got a full window the sheet might continue; if fewer rows came back
            # we're at or past the end.
            has_more = returned_count == (end_row - start_row + 1)

        return {
            "sheet_name": sheet_name,
            "start_row": start_row,
            "end_row_requested": end_row,
            "rows": rows_out,
            "returned_count": returned_count,
            "has_more": has_more,
        }
    finally:
        if wb is not None:
            wb.close()
        if tempfile_path is not None:
            try:
                os.unlink(tempfile_path)
            except OSError:
                pass
