# Copyright (c) 2026, Nirmaan (Stratos Infra Technologies Pvt. Ltd.) and Contributors
# See license.txt

"""Tests for the BoQ pricing layer (Phase 5 Slice 1b): the BoQ Cell Pricing doctype +
the save_cell_price / get_sheet_pricing endpoints.

Coverage matrix:
  POS  save -> a current, is_filled=1 pricing record at the Excel-address key, pricing_version=1.
  POS  re-save the same cell -> prior frozen (is_current=0), new current pricing_version=2
       (freeze-and-supersede), exactly ONE current (the invariant).
  POS  multi-area -> two prices at the same excel_row, different col_letter, both current + distinct.
  POS  read -> get_sheet_pricing returns the current set for (boq, sheet, committed_version).
  NEG  save to a non-existent cell / sheet / committed_version -> throws (no price for a non-cell).
  NEG  missing required args -> throws.

The committed-node fixture (build_committed_sheet_fixture) is the SHARED hermetic builder
(defined in test_review_screen.py) -- the committed tier is read-only here; the pricing layer
sits on top of it. sheet_name carries a trailing space (#152) throughout.
"""
import json

import frappe
from frappe.tests.utils import FrappeTestCase

from nirmaan_stack.api.boq.wizard.pricing import (
    _sheet_formulas_complete,
    get_committed_sheet_grid,
    get_priced_rows,
    get_sheet_amount_formulas,
    get_sheet_colors,
    get_sheet_dismissals,
    get_sheet_pricing,
    get_sheet_reconciliation_choices,
    get_sheet_remarks,
    save_amount_formula,
    save_cell_color,
    save_cell_dismissal,
    save_cell_price,
    save_cell_reconciliation_choice,
    save_row_remark,
)
from nirmaan_stack.api.boq.wizard import pricing_lock
from nirmaan_stack.api.boq.wizard.pricing_lock import (
    LOCK_STALE_SECONDS,
    _LOCK_HELD_MARKER,
    _lock_identity,
    acquire_or_refresh,
    read_lock_info,
)
from nirmaan_stack.api.boq.wizard.review_screen import get_committed_rows
from nirmaan_stack.api.boq.wizard.test_review_screen import (
    _cleanup_project,
    _make_project,
    build_committed_sheet_fixture,
    cleanup_committed_fixture,
)

_PRICING = "BoQ Cell Pricing"
_LOCK_DT = "BoQ Sheet Pricing Lock"
_REMARK_DT = "BoQ Cell Remark"
_COLOR_DT = "BoQ Cell Color"
_FORMULA_DT = "BoQ Cell Amount Formula"
_DISMISSAL_DT = "BoQ Cell Dismissal"
_CHOICE_DT = "BoQ Cell Reconciliation Choice"


# A minimal structurally-valid amount formula (a single leaf -- presence is what coverage needs;
# F1 validates structure only, never the ref against descriptors).
_FIXTURE_FORMULA_LEAF = json.dumps(
    {"ref": {"value_field": "qty_by_area", "value_key": None, "rate_subkey": None}}
)


def _declare_fixture_amount_formulas(boq_name, sheet_name, commit_version):
    """Make the SHARED per-area committed fixture (build_committed_sheet_fixture: amount cols
    F [Phase 1, rate_subkey "total"] + I [Phase 2, rate_subkey "install"]) formula-COMPLETE for
    the MANDATORY amount-formula gate (save_cell_price -> _sheet_formulas_complete rejects ANY
    rate write until every amount column has a declared formula).

    The two amount columns carry DIFFERENT rate_subkeys (total + install), so a wildcard-DEFAULT
    formula is declared for EACH (one (amount_by_area, value_key=None, "total") + one
    (..., "install")); together they cover both columns via pickFormula's area-wildcard
    resolution. This is exactly the spec's "declare formulas in-test to flip completeness", kept
    LOCAL to test_pricing.py so neither the shared fixture nor any other test file is touched.
    sheet_name VERBATIM (#152). save_amount_formula commits internally."""
    for kind in ("total", "install"):
        save_amount_formula(
            boq_name=boq_name, sheet_name=sheet_name, committed_version=commit_version,
            target_value_field="amount_by_area", target_value_key=None,
            target_rate_subkey=kind, formula=_FIXTURE_FORMULA_LEAF,
        )


class TestCellPricing(FrappeTestCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.test_project = _make_project()
        boq = frappe.new_doc("BOQs")
        boq.project = cls.test_project.name
        boq.boq_name = "Pricing Test BoQ"
        boq.tax_treatment = "Pre-tax"
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        cls.boq = boq.name
        cls.sheet = "Pricing Fix "  # VERBATIM trailing space (#152)
        cls.cv = 1
        cls.fixture = build_committed_sheet_fixture(cls.boq, cls.sheet, commit_version=cls.cv)
        # fixture line items: source_row 34 (li[0]) + 35 (li[1]); rate cols E (Phase 1) / H (Phase 2)
        cls.li_34 = cls.fixture["line_items"][0]
        # MANDATORY amount-formula gate: the fixture has amount cols F/I -> declare covering
        # formulas so save_cell_price is not rejected (see _declare_fixture_amount_formulas).
        _declare_fixture_amount_formulas(cls.boq, cls.sheet, cls.cv)

    @classmethod
    def tearDownClass(cls):
        cleanup_committed_fixture(cls.boq)
        _cleanup_project(cls.test_project.name)
        super().tearDownClass()

    def setUp(self):
        # Each test starts with a clean pricing layer + lock (the committed fixture persists).
        # save_cell_price now acquires a single-editor lock, so clear it for isolation.
        frappe.db.delete(_PRICING, {"boq": self.boq})
        frappe.db.delete(_LOCK_DT, {"boq": self.boq})
        frappe.db.commit()

    # -- helpers ------------------------------------------------------------

    def _current(self, excel_row, col_letter):
        return frappe.get_all(
            _PRICING,
            filters={"boq": self.boq, "sheet_name": self.sheet, "excel_row": excel_row,
                     "col_letter": col_letter, "committed_version": self.cv, "is_current": 1},
            fields=["name", "rate", "is_filled", "pricing_version", "node", "area", "rate_kind"],
        )

    def _all_versions(self, excel_row, col_letter):
        return frappe.get_all(
            _PRICING,
            filters={"boq": self.boq, "sheet_name": self.sheet, "excel_row": excel_row,
                     "col_letter": col_letter, "committed_version": self.cv},
            fields=["pricing_version", "is_current", "rate"],
            order_by="pricing_version asc",
        )

    # -- POSITIVE -----------------------------------------------------------

    def test_save_creates_current_filled_record_v1(self):
        res = save_cell_price(
            boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
            committed_version=self.cv, rate=150.0, area="Phase 1", rate_kind="combined",
            description="cable 1.1.2",
        )
        self.assertTrue(res["ok"])
        self.assertEqual(res["pricing_version"], 1)
        self.assertEqual(res["froze_prior"], 0, "first save freezes nothing")
        cur = self._current(34, "E")
        self.assertEqual(len(cur), 1, "exactly one current pricing record")
        self.assertEqual(cur[0]["rate"], 150.0)
        self.assertEqual(cur[0]["is_filled"], 1, "the layer's own filled-flag is set (node 0.0 is not it)")
        self.assertEqual(cur[0]["node"], self.li_34, "the resolved committed node is stored as the pointer")
        self.assertEqual(cur[0]["area"], "Phase 1")

    def test_resave_freezes_prior_and_supersedes(self):
        save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
                        committed_version=self.cv, rate=150.0)
        res2 = save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
                               committed_version=self.cv, rate=175.0)
        self.assertEqual(res2["pricing_version"], 2)
        self.assertEqual(res2["froze_prior"], 1, "the re-save froze the prior current")
        cur = self._current(34, "E")
        self.assertEqual(len(cur), 1, "exactly ONE current after re-save (freeze-and-supersede)")
        self.assertEqual(cur[0]["pricing_version"], 2)
        self.assertEqual(cur[0]["rate"], 175.0)
        versions = self._all_versions(34, "E")
        self.assertEqual([v["pricing_version"] for v in versions], [1, 2])
        self.assertEqual([v["is_current"] for v in versions], [0, 1], "v1 frozen, v2 current")

    def test_one_current_invariant_after_multiple_resaves(self):
        for r in (10.0, 20.0, 30.0):
            save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=35, col_letter="H",
                            committed_version=self.cv, rate=r)
        cur = self._current(35, "H")
        self.assertEqual(len(cur), 1, "exactly one current per cell identity after N re-saves")
        self.assertEqual(cur[0]["pricing_version"], 3)
        self.assertEqual(cur[0]["rate"], 30.0)

    def test_multi_area_two_cells_same_row_distinct(self):
        # Same excel_row 34, different col_letter (E = Phase 1, H = Phase 2) -> two distinct
        # current records that do NOT freeze each other.
        save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
                        committed_version=self.cv, rate=100.0, area="Phase 1")
        save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="H",
                        committed_version=self.cv, rate=200.0, area="Phase 2")
        self.assertEqual(len(self._current(34, "E")), 1)
        self.assertEqual(len(self._current(34, "H")), 1)
        self.assertEqual(self._current(34, "E")[0]["rate"], 100.0)
        self.assertEqual(self._current(34, "H")[0]["rate"], 200.0)

    def test_get_sheet_pricing_returns_current_set(self):
        save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
                        committed_version=self.cv, rate=100.0)
        save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="H",
                        committed_version=self.cv, rate=200.0)
        # a re-save of one cell -> still only the CURRENT shows in the read
        save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
                        committed_version=self.cv, rate=125.0)
        res = get_sheet_pricing(boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv)
        prices = res["pricing"]
        self.assertEqual(len(prices), 2, "two current cells (E re-saved once, H once)")
        by_col = {p["col_letter"]: p for p in prices}
        self.assertEqual(by_col["E"]["rate"], 125.0)
        self.assertEqual(by_col["E"]["pricing_version"], 2)
        self.assertEqual(by_col["H"]["rate"], 200.0)
        for p in prices:
            self.assertEqual(p["is_current"], 1)

    def test_get_sheet_pricing_empty_when_unpriced(self):
        res = get_sheet_pricing(boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv)
        self.assertEqual(res, {"pricing": []})

    # -- NEGATIVE -----------------------------------------------------------

    def test_save_to_nonexistent_excel_row_throws(self):
        with self.assertRaises(frappe.ValidationError):
            save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=9999,
                            col_letter="E", committed_version=self.cv, rate=10.0)
        self.assertEqual(self._current(9999, "E"), [], "no price created for a non-existent cell")

    def test_save_to_nonexistent_version_throws(self):
        with self.assertRaises(frappe.ValidationError):
            save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                            col_letter="E", committed_version=99, rate=10.0)

    def test_save_to_nonexistent_sheet_throws(self):
        with self.assertRaises(frappe.ValidationError):
            save_cell_price(boq_name=self.boq, sheet_name="No Such Sheet ", excel_row=34,
                            col_letter="E", committed_version=self.cv, rate=10.0)

    def test_save_nonexistent_boq_throws(self):
        with self.assertRaises(frappe.ValidationError):
            save_cell_price(boq_name="NOPE-DOES-NOT-EXIST", sheet_name=self.sheet, excel_row=34,
                            col_letter="E", committed_version=self.cv, rate=10.0)

    def test_save_missing_args_throw(self):
        with self.assertRaises(frappe.ValidationError):
            save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                            committed_version=self.cv, rate=10.0)  # no col_letter
        with self.assertRaises(frappe.ValidationError):
            save_cell_price(boq_name=self.boq, sheet_name=self.sheet, col_letter="E",
                            committed_version=self.cv, rate=10.0)  # no excel_row

    def test_read_missing_args_throw(self):
        with self.assertRaises(frappe.ValidationError):
            get_sheet_pricing(boq_name=self.boq, sheet_name=self.sheet)  # no committed_version


def _build_scalar_rate_committed_sheet(boq_name: str, sheet_name: str, commit_version: int = 1) -> str:
    """Insert a minimal committed sheet whose rate column is SCALAR (area=None) -- the branch
    the per-area shared fixture (all rate cols are *_by_area) does NOT exercise. 1 BoQ Sheet
    (role D = scalar rate_combined) + 1 Preamble root + 1 Line Item child carrying a scalar
    combined_rate at Excel row 10. Returns the line-item node name. sheet_name VERBATIM (#152).
    Cleaned up by cleanup_committed_fixture (deletes all committed rows for the boq)."""
    now = frappe.utils.now()
    bs = frappe.new_doc("BoQ Sheet")
    bs.boq = boq_name
    bs.sheet_name = sheet_name  # VERBATIM (#152)
    bs.sheet_order = 2
    bs.treat_as = "data"
    bs.header_row = 1
    bs.header_row_count = 1
    bs.column_role_map = {
        "A": {"role": "sl_no", "area": None},
        "B": {"role": "description", "area": None},
        "C": {"role": "qty_total", "area": None},
        "D": {"role": "rate_combined", "area": None},  # SCALAR rate (area=None)
    }
    bs.column_headers = {}
    bs.area_dimensions = json.dumps([])
    bs.commit_version = commit_version
    bs.is_current = 1
    bs.committed_at = now
    bs.insert(ignore_permissions=True)

    pre = frappe.new_doc("BOQ Nodes")
    pre.sheet = bs.name
    pre.node_type = "Preamble"
    pre.row_class = "preamble"
    pre.level = 1
    pre.description = "SCALAR PREAMBLE"
    pre.code = "S.0"
    pre.sort_order = 0
    pre.source_row_number = 8
    pre.commit_version = commit_version
    pre.is_current = 1
    pre.committed_at = now
    pre.insert(ignore_permissions=True)

    li = frappe.new_doc("BOQ Nodes")
    li.sheet = bs.name
    li.node_type = "Line Item"
    li.row_class = "line_item"
    li.description = "scalar item"
    li.code = "S.1"
    li.parent_node = pre.name
    li.qty = 5.0
    li.unit = "Nos"
    li.combined_rate = 999.0  # the committed scalar rate (the un-priced baseline)
    li.source_row_number = 10
    li.sort_order = 1
    li.commit_version = commit_version
    li.is_current = 1
    li.committed_at = now
    li.insert(ignore_permissions=True)

    frappe.db.commit()
    return li.name


class TestGetPricedRows(FrappeTestCase):
    """get_priced_rows -- the pricing-overlay read (committed rows + current prices merged).

    Reuses the shared per-area committed fixture (build_committed_sheet_fixture: line items
    at excel_row 34/35; rate cols E [Phase 1] / H [Phase 2]) + a local SCALAR-rate committed
    sheet (rate col D, area=None) so both descriptor branches are covered. sheet_name carries
    a trailing space (#152) throughout. The committed tier is read-only here; each test starts
    with a clean pricing layer.
    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.test_project = _make_project()
        boq = frappe.new_doc("BOQs")
        boq.project = cls.test_project.name
        boq.boq_name = "Priced-Rows Test BoQ"
        boq.tax_treatment = "Pre-tax"
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        cls.boq = boq.name
        cls.cv = 1
        cls.sheet = "Priced Fix "  # VERBATIM trailing space (#152)
        cls.fixture = build_committed_sheet_fixture(cls.boq, cls.sheet, commit_version=cls.cv)
        cls.scalar_sheet = "Scalar Fix "  # VERBATIM trailing space (#152)
        cls.scalar_node = _build_scalar_rate_committed_sheet(cls.boq, cls.scalar_sheet, cls.cv)
        # MANDATORY amount-formula gate: the per-area fixture has amount cols F/I -> declare
        # covering formulas so save_cell_price succeeds. The scalar sheet has zero amount cols
        # (trivially complete), so it needs none.
        _declare_fixture_amount_formulas(cls.boq, cls.sheet, cls.cv)

    @classmethod
    def tearDownClass(cls):
        cleanup_committed_fixture(cls.boq)
        _cleanup_project(cls.test_project.name)
        super().tearDownClass()

    def setUp(self):
        frappe.db.delete(_PRICING, {"boq": self.boq})
        frappe.db.delete(_LOCK_DT, {"boq": self.boq})  # save_cell_price acquires a lock now
        frappe.db.commit()

    # -- helpers ------------------------------------------------------------

    def _row(self, res, excel_row):
        return next((r for r in res["rows"] if r.get("source_row_number") == excel_row), None)

    # -- POSITIVE -----------------------------------------------------------

    def test_priced_and_unpriced_in_one_row(self):
        # Price col E (Phase 1 combined); leave col H (Phase 2 combined) un-priced.
        save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
                        committed_version=self.cv, rate=150.0, area="Phase 1", rate_kind="combined")
        res = get_priced_rows(boq_name=self.boq, sheet_name=self.sheet)
        row = self._row(res, 34)
        self.assertIsNotNone(row, "row at excel_row 34 must be present")
        # E priced: saved rate stamped + marked priced.
        self.assertEqual(row["rate_by_area"]["Phase 1"]["combined_rate"], 150.0)
        self.assertTrue(row["priced_by_area"]["Phase 1"]["combined_rate"], "E is marked priced")
        # H un-priced: committed value (0.0) untouched + NO priced marker.
        self.assertEqual(row["rate_by_area"]["Phase 2"]["combined_rate"], 0.0,
                         "un-priced H keeps its committed value")
        self.assertNotIn("combined_rate", row.get("priced_by_area", {}).get("Phase 2", {}),
                         "H carries no priced marker")

    def test_zero_rate_is_priced_not_unpriced(self):
        # The load-bearing correctness case: a saved rate of 0.0 is PRICED (record present +
        # is_filled), NOT un-priced. A zero-check marker would fail this.
        save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
                        committed_version=self.cv, rate=0.0, area="Phase 1", rate_kind="combined")
        res = get_priced_rows(boq_name=self.boq, sheet_name=self.sheet)
        row = self._row(res, 34)
        self.assertEqual(row["rate_by_area"]["Phase 1"]["combined_rate"], 0.0)
        self.assertTrue(row["priced_by_area"]["Phase 1"]["combined_rate"],
                        "a 0.0 saved rate is PRICED (presence + is_filled), not un-priced")

    def test_multi_area_independence(self):
        # Price BOTH E and H on row 34 with different rates -> each stamped to its own cell,
        # both marked priced, neither bleeds into the other.
        save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
                        committed_version=self.cv, rate=111.0, area="Phase 1", rate_kind="combined")
        save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="H",
                        committed_version=self.cv, rate=222.0, area="Phase 2", rate_kind="combined")
        res = get_priced_rows(boq_name=self.boq, sheet_name=self.sheet)
        row = self._row(res, 34)
        self.assertEqual(row["rate_by_area"]["Phase 1"]["combined_rate"], 111.0)
        self.assertEqual(row["rate_by_area"]["Phase 2"]["combined_rate"], 222.0)
        self.assertTrue(row["priced_by_area"]["Phase 1"]["combined_rate"])
        self.assertTrue(row["priced_by_area"]["Phase 2"]["combined_rate"])

    def test_commit_version_passthrough(self):
        save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
                        committed_version=self.cv, rate=150.0, area="Phase 1", rate_kind="combined")
        res = get_priced_rows(boq_name=self.boq, sheet_name=self.sheet)
        self.assertEqual(res["commit_version"], self.cv,
                         "the overlay surfaces the committed version it priced")
        # And the price from THAT version is what merged.
        self.assertEqual(self._row(res, 34)["rate_by_area"]["Phase 1"]["combined_rate"], 150.0)

    def test_free_sheet_editable_and_no_lock(self):
        # A committed sheet with NO lock (no one has edited it) -> free: editable True,
        # lock_info None. (Slice A: lock_info is no longer ALWAYS None -- see
        # TestSingleEditorLock for the locked shapes.)
        res = get_priced_rows(boq_name=self.boq, sheet_name=self.sheet)
        self.assertIs(res["editable"], True, "a free sheet is editable")
        self.assertIsNone(res["lock_info"], "a free sheet (no lock) has lock_info None")

    def test_scalar_rate_cell_priced(self):
        # Un-priced baseline: the committed scalar rate (999.0) shows, no marker.
        res0 = get_priced_rows(boq_name=self.boq, sheet_name=self.scalar_sheet)
        row0 = self._row(res0, 10)
        self.assertIsNotNone(row0, "scalar sheet must return its row")
        self.assertEqual(row0["rate_combined"], 999.0)
        self.assertNotIn("priced_rate_combined", row0, "un-priced scalar carries no marker")
        # Price the scalar rate cell (col D, area=None).
        save_cell_price(boq_name=self.boq, sheet_name=self.scalar_sheet, excel_row=10,
                        col_letter="D", committed_version=self.cv, rate=750.0, rate_kind="combined")
        res1 = get_priced_rows(boq_name=self.boq, sheet_name=self.scalar_sheet)
        row1 = self._row(res1, 10)
        self.assertEqual(row1["rate_combined"], 750.0, "scalar saved rate stamped in place")
        self.assertTrue(row1["priced_rate_combined"], "scalar priced cell is marked")

    def test_descriptors_pass_through_unchanged(self):
        committed = get_committed_rows(boq_name=self.boq, sheet_name=self.sheet)
        res = get_priced_rows(boq_name=self.boq, sheet_name=self.sheet)
        self.assertEqual(res["column_descriptors"], committed["column_descriptors"],
                         "column_descriptors pass through from get_committed_rows unchanged")

    def test_no_amount_or_qty_stamping(self):
        # A saved price must land ONLY on a rate cell -- never amount / qty.
        baseline = get_committed_rows(boq_name=self.boq, sheet_name=self.sheet)
        base_row = self._row({"rows": baseline["rows"]}, 34)
        base_amount = base_row["amount_by_area"]
        base_qty = base_row["qty_by_area"]
        save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
                        committed_version=self.cv, rate=150.0, area="Phase 1", rate_kind="combined")
        res = get_priced_rows(boq_name=self.boq, sheet_name=self.sheet)
        row = self._row(res, 34)
        self.assertEqual(row["amount_by_area"], base_amount, "amount cells untouched")
        self.assertEqual(row["qty_by_area"], base_qty, "qty cells untouched")
        self.assertNotIn("priced_amount_total", row)
        self.assertNotIn("priced_qty_total", row)
        # priced_by_area carries ONLY rate kinds (never amount kinds supply/install/total).
        for area, kinds in row.get("priced_by_area", {}).items():
            self.assertTrue(
                set(kinds).issubset({"supply_rate", "install_rate", "combined_rate"}),
                "priced_by_area carries only rate kinds",
            )

    # -- NEGATIVE / edge ----------------------------------------------------

    def test_uncommitted_sheet_returns_empty_merged_shape(self):
        res = get_priced_rows(boq_name=self.boq, sheet_name="No Such Sheet ZZ ")
        self.assertEqual(res["rows"], [])
        self.assertEqual(res["column_descriptors"], [])
        self.assertIsNone(res["commit_version"])
        self.assertIs(res["editable"], True)
        self.assertIsNone(res["lock_info"])

    def test_missing_args_and_unknown_boq_throw(self):
        with self.assertRaises(frappe.ValidationError):
            get_priced_rows(sheet_name=self.sheet)  # no boq_name
        with self.assertRaises(frappe.ValidationError):
            get_priced_rows(boq_name=self.boq)  # no sheet_name
        with self.assertRaises(frappe.ValidationError):
            get_priced_rows(boq_name="NOPE-DOES-NOT-EXIST", sheet_name="X")


class TestSingleEditorLock(FrappeTestCase):
    """Single-editor pricing lock (slice A): acquire-on-first-edit, refresh, reject,
    stale takeover, the ATOMIC exactly-one-winner guarantee, and the get_priced_rows
    lock_info shapes. Reuses the shared per-area committed fixture so save_cell_price's
    committed-cell check passes. sheet_name carries a trailing space (#152).

    `me` = the session user (the saver in save_cell_price); `other` = a DIFFERENT real
    User used to stand up a competing lock holder.
    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.test_project = _make_project()
        boq = frappe.new_doc("BOQs")
        boq.project = cls.test_project.name
        boq.boq_name = "Pricing Lock Test BoQ"
        boq.tax_treatment = "Pre-tax"
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        cls.boq = boq.name
        cls.sheet = "Lock Fix "  # VERBATIM trailing space (#152)
        cls.cv = 1
        cls.fixture = build_committed_sheet_fixture(cls.boq, cls.sheet, commit_version=cls.cv)
        # MANDATORY amount-formula gate: declare covering formulas so save_cell_price reaches the
        # lock logic (the gate fires BEFORE the lock acquire). setUp clears the lock this leaves.
        _declare_fixture_amount_formulas(cls.boq, cls.sheet, cls.cv)
        cls.me = frappe.session.user
        cls.other = frappe.db.get_value(
            "User", {"name": ["not in", [cls.me, "Guest"]], "enabled": 1}, "name"
        )
        assert cls.other, "need a second real User to play the competing lock holder"

    @classmethod
    def tearDownClass(cls):
        frappe.db.delete(_LOCK_DT, {"boq": cls.boq})
        cleanup_committed_fixture(cls.boq)
        _cleanup_project(cls.test_project.name)
        super().tearDownClass()

    def setUp(self):
        frappe.db.delete(_PRICING, {"boq": self.boq})
        frappe.db.delete(_LOCK_DT, {"boq": self.boq})
        frappe.db.commit()

    # -- helpers ------------------------------------------------------------

    def _lock_row(self):
        name = _lock_identity(self.boq, self.sheet, self.cv)
        return frappe.db.get_value(
            _LOCK_DT, name, ["locked_by", "last_edit_at"], as_dict=True
        )

    def _lock_count(self):
        return frappe.db.count(
            _LOCK_DT,
            {"boq": self.boq, "sheet_name": self.sheet, "committed_version": self.cv},
        )

    def _seed_lock(self, user, last_edit_at):
        """Stand up a lock row owned by `user` with a controlled last_edit_at."""
        acquire_or_refresh(self.boq, self.sheet, self.cv, user, frappe.utils.now_datetime())
        name = _lock_identity(self.boq, self.sheet, self.cv)
        frappe.db.set_value(_LOCK_DT, name, "last_edit_at", last_edit_at, update_modified=False)
        frappe.db.commit()

    # -- acquire / refresh --------------------------------------------------

    def test_first_save_acquires_lock(self):
        save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
                        committed_version=self.cv, rate=150.0)
        lock = self._lock_row()
        self.assertIsNotNone(lock, "the first save acquires a lock")
        self.assertEqual(lock["locked_by"], self.me, "holder is the saver (session user)")
        self.assertIsNotNone(lock["last_edit_at"], "last_edit_at is stamped")
        self.assertEqual(self._lock_count(), 1)

    def test_second_save_by_holder_refreshes_and_keeps_holder(self):
        save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
                        committed_version=self.cv, rate=150.0)
        name = _lock_identity(self.boq, self.sheet, self.cv)
        # Backdate so the refresh is observable.
        old = frappe.utils.add_to_date(frappe.utils.now_datetime(), minutes=-2)
        frappe.db.set_value(_LOCK_DT, name, "last_edit_at", old, update_modified=False)
        frappe.db.commit()
        save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
                        committed_version=self.cv, rate=175.0)
        lock = self._lock_row()
        self.assertEqual(lock["locked_by"], self.me, "holder unchanged on a refresh")
        self.assertGreater(
            frappe.utils.get_datetime(lock["last_edit_at"]),
            frappe.utils.get_datetime(old),
            "last_edit_at advanced on the holder's save",
        )
        self.assertEqual(self._lock_count(), 1)

    # -- reject (held fresh by another) -------------------------------------

    def test_save_by_other_while_fresh_is_rejected_and_mutates_nothing(self):
        # `other` holds a FRESH lock; `me` (session user) tries to save -> reject.
        self._seed_lock(self.other, frappe.utils.now_datetime())
        before = self._lock_row()
        with self.assertRaises(frappe.ValidationError) as ctx:
            save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
                            committed_version=self.cv, rate=150.0)
        msg = str(ctx.exception)
        self.assertIn(_LOCK_HELD_MARKER, msg, "reject carries the stable lock marker")
        self.assertIn(
            pricing_lock._get_user_full_name(self.other), msg,
            "reject names the current holder",
        )
        # REJECT MUTATES NOTHING: no pricing row was written...
        self.assertEqual(
            frappe.get_all(_PRICING, filters={"boq": self.boq, "excel_row": 34, "col_letter": "E"}),
            [], "a rejected save creates no pricing record",
        )
        # ...and the lock holder + timestamp are untouched.
        after = self._lock_row()
        self.assertEqual(after["locked_by"], self.other, "holder unchanged after a reject")
        self.assertEqual(
            frappe.utils.get_datetime(after["last_edit_at"]),
            frappe.utils.get_datetime(before["last_edit_at"]),
            "the holder's last_edit_at is untouched by a rejected save",
        )

    # -- stale takeover -----------------------------------------------------

    def test_stale_lock_taken_over_by_new_user(self):
        # `other` holds a STALE lock (last edit > 5 min ago); `me` saves -> takeover.
        stale = frappe.utils.add_to_date(
            frappe.utils.now_datetime(), seconds=-(LOCK_STALE_SECONDS + 60)
        )
        self._seed_lock(self.other, stale)
        save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
                        committed_version=self.cv, rate=150.0)
        lock = self._lock_row()
        self.assertEqual(lock["locked_by"], self.me, "a stale lock flips to the new saver")
        self.assertGreater(
            frappe.utils.get_datetime(lock["last_edit_at"]),
            frappe.utils.get_datetime(stale),
            "last_edit_at refreshed on takeover",
        )
        self.assertEqual(self._lock_count(), 1, "takeover overwrites in place (no second row)")
        # The price WAS written (takeover proceeds).
        cur = frappe.get_all(_PRICING, filters={"boq": self.boq, "excel_row": 34,
                             "col_letter": "E", "is_current": 1}, pluck="rate")
        self.assertEqual(cur, [150.0])

    # -- THE ATOMICITY TEST (exactly one winner) ----------------------------

    def test_atomicity_concurrent_first_edit_exactly_one_winner(self):
        """Deterministically exercise the PK-collision path: a winner A already holds a
        FRESH lock; we force B's acquire to BELIEVE the sheet is free (patch the FIRST
        _read_lock to None) so B attempts the INSERT -- which COLLIDES on the deterministic
        primary key. The collision must RAISE (DuplicateEntryError), be caught, B re-reads
        the winner, and B is rejected. Proves exactly-one-winner: the duplicate insert does
        NOT create a second row, and the holder stays A."""
        now = frappe.utils.now_datetime()
        # A wins first (fresh holder).
        acquire_or_refresh(self.boq, self.sheet, self.cv, self.other, now)
        frappe.db.commit()
        name = _lock_identity(self.boq, self.sheet, self.cv)

        real_read = pricing_lock._read_lock
        calls = {"n": 0}

        def fake_read(boq, sheet_name, version):
            calls["n"] += 1
            if calls["n"] == 1:
                return None  # B's acquire sees "free" -> attempts the colliding insert
            return real_read(boq, sheet_name, version)  # re-read after collision -> finds A

        pricing_lock._read_lock = fake_read
        try:
            with self.assertRaises(frappe.ValidationError) as ctx:
                acquire_or_refresh(self.boq, self.sheet, self.cv, self.me, now)
        finally:
            pricing_lock._read_lock = real_read

        # The collision RAISED + was HANDLED (re-read A, rejected) -- not swallowed: if the
        # duplicate insert had been silently allowed, acquire would have returned (B holder),
        # not thrown the reject below.
        self.assertGreaterEqual(calls["n"], 2, "B re-read after the collision")
        self.assertIn(_LOCK_HELD_MARKER, str(ctx.exception), "B was rejected (collision handled)")
        # EXACTLY ONE row survived -- the duplicate insert collided, it did not duplicate.
        self.assertEqual(self._lock_count(), 1, "exactly one winner -- no duplicate row")
        # The holder is still A (the winner); B never overwrote.
        self.assertEqual(frappe.db.get_value(_LOCK_DT, name, "locked_by"), self.other)

    # -- lock_info shapes out of get_priced_rows (PURE READ) ----------------

    def test_lock_info_free(self):
        res = get_priced_rows(boq_name=self.boq, sheet_name=self.sheet)
        self.assertIsNone(res["lock_info"], "free sheet -> lock_info None")
        self.assertIs(res["editable"], True)

    def test_lock_info_mine(self):
        self._seed_lock(self.me, frappe.utils.now_datetime())
        res = get_priced_rows(boq_name=self.boq, sheet_name=self.sheet)
        li = res["lock_info"]
        self.assertIsNotNone(li)
        self.assertTrue(li["is_locked_by_me"], "session user holds it")
        self.assertFalse(li["is_stale"])
        self.assertEqual(li["locked_by_user"], self.me)
        self.assertIs(res["editable"], True, "editable when I hold it")

    def test_lock_info_other_fresh_blocks(self):
        self._seed_lock(self.other, frappe.utils.now_datetime())
        res = get_priced_rows(boq_name=self.boq, sheet_name=self.sheet)
        li = res["lock_info"]
        self.assertFalse(li["is_locked_by_me"], "another user holds it")
        self.assertEqual(li["locked_by_user"], self.other)
        self.assertEqual(li["locked_by_name"], pricing_lock._get_user_full_name(self.other))
        self.assertFalse(li["is_stale"])
        self.assertIs(res["editable"], False, "NOT editable when held fresh by another")
        # PURE READ: the read did not mutate the lock (still exactly one, holder unchanged).
        self.assertEqual(self._lock_count(), 1)
        self.assertEqual(self._lock_row()["locked_by"], self.other)

    def test_lock_info_other_stale_allows(self):
        stale = frappe.utils.add_to_date(
            frappe.utils.now_datetime(), seconds=-(LOCK_STALE_SECONDS + 60)
        )
        self._seed_lock(self.other, stale)
        res = get_priced_rows(boq_name=self.boq, sheet_name=self.sheet)
        li = res["lock_info"]
        self.assertFalse(li["is_locked_by_me"])
        self.assertTrue(li["is_stale"], "another user's lock is stale")
        self.assertIs(res["editable"], True, "a stale lock is acquirable -> editable")
        # get_priced_rows did NOT take over (still held by other; only save_cell_price acquires).
        self.assertEqual(self._lock_row()["locked_by"], self.other)


class TestLockPerSheetIsolation(FrappeTestCase):
    """Single-editor pricing lock -- PER-SHEET isolation (the MIRROR of slice A's same-sheet
    guarantee). Two users editing two DIFFERENT sheets of the SAME BoQ acquire two
    INDEPENDENT locks and never contend.

    TRUE BY CONSTRUCTION: the lock identity is name = sha1(boq \\x00 sheet_name \\x00 version),
    so a different sheet_name => a different primary key => a different lock row => no
    collision. These tests CERTIFY that deterministically (a substitute for a two-user live
    cert, which is not possible on a single local machine) AND guard against a regression
    that drops sheet_name from the identity. They drive BOTH layers: acquire_or_refresh (the
    lock core) AND save_cell_price (the real entry point the frontend calls). A same-sheet
    contrast test proves the lock DOES block -- so the different-sheet passes are meaningful.

    TWO committed sheets on ONE boq; sheet A carries a trailing space (#152). me = the
    session user (the real save_cell_price saver); other = a second real User holding the
    OTHER sheet via acquire_or_refresh -- mirroring the existing lock suite, which never
    frappe.set_user()s (so the suite is never left on a switched user). Driving me's save
    through the real save_cell_price endpoint still proves isolation at the layer the
    frontend calls; other-as-holder via acquire_or_refresh is permission-agnostic.
    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.test_project = _make_project()
        boq = frappe.new_doc("BOQs")
        boq.project = cls.test_project.name
        boq.boq_name = "Pricing Lock Isolation Test BoQ"
        boq.tax_treatment = "Pre-tax"
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        cls.boq = boq.name
        cls.cv = 1
        cls.sheet_a = "Iso Sheet A "  # VERBATIM trailing space (#152)
        cls.sheet_b = "Iso Sheet B"
        # TWO committed sheets on the SAME boq -- the helper keys on (boq, sheet_name), so two
        # distinct names build two distinct committed sheets + node sets on the one BOQs.
        cls.fixture_a = build_committed_sheet_fixture(cls.boq, cls.sheet_a, commit_version=cls.cv)
        cls.fixture_b = build_committed_sheet_fixture(cls.boq, cls.sheet_b, commit_version=cls.cv)
        # MANDATORY amount-formula gate: BOTH committed sheets carry amount cols F/I -> declare
        # covering formulas on EACH so save_cell_price reaches the lock logic on either sheet.
        _declare_fixture_amount_formulas(cls.boq, cls.sheet_a, cls.cv)
        _declare_fixture_amount_formulas(cls.boq, cls.sheet_b, cls.cv)
        cls.me = frappe.session.user
        cls.other = frappe.db.get_value(
            "User", {"name": ["not in", [cls.me, "Guest"]], "enabled": 1}, "name"
        )
        assert cls.other, "need a second real User to play the competing lock holder"

    @classmethod
    def tearDownClass(cls):
        # cleanup_committed_fixture deletes ALL nodes/sheets/pricing for the boq -> BOTH sheets.
        frappe.db.delete(_LOCK_DT, {"boq": cls.boq})
        cleanup_committed_fixture(cls.boq)
        _cleanup_project(cls.test_project.name)
        super().tearDownClass()

    def setUp(self):
        frappe.db.delete(_PRICING, {"boq": self.boq})
        frappe.db.delete(_LOCK_DT, {"boq": self.boq})
        frappe.db.commit()

    # -- helpers ------------------------------------------------------------

    def _lock_row(self, sheet_name):
        name = _lock_identity(self.boq, sheet_name, self.cv)
        return frappe.db.get_value(
            _LOCK_DT, name, ["locked_by", "last_edit_at"], as_dict=True
        )

    def _boq_lock_count(self):
        return frappe.db.count(_LOCK_DT, {"boq": self.boq})

    # -- Test 1: identity is per-(sheet, version) ---------------------------

    def test_lock_identity_is_per_sheet_and_per_version(self):
        """The construction proof: the deterministic lock name is keyed on sheet_name AND
        committed_version -- a one-liner guard against a regression dropping either."""
        ia = _lock_identity(self.boq, self.sheet_a, self.cv)
        ib = _lock_identity(self.boq, self.sheet_b, self.cv)
        ia_v2 = _lock_identity(self.boq, self.sheet_a, self.cv + 1)
        self.assertNotEqual(ia, ib, "different sheet_name -> different lock identity")
        self.assertNotEqual(ia, ia_v2, "different committed_version -> different lock identity")
        self.assertNotEqual(ib, ia_v2)

    # -- Test 2: two users, two sheets, both acquire, no contention ---------

    def test_two_users_two_sheets_both_acquire_no_contention(self):
        """acquire_or_refresh layer: me holds A, other holds B -- two independent rows, and
        neither acquire touched/blocked the other."""
        now = frappe.utils.now_datetime()
        # me acquires sheet A.
        acquire_or_refresh(self.boq, self.sheet_a, self.cv, self.me, now)
        frappe.db.commit()
        row_a = self._lock_row(self.sheet_a)
        self.assertIsNotNone(row_a, "sheet A lock acquired")
        self.assertEqual(row_a["locked_by"], self.me)
        a_edit_at = row_a["last_edit_at"]

        # other acquires sheet B -- must SUCCEED (no throw) despite me holding A.
        acquire_or_refresh(self.boq, self.sheet_b, self.cv, self.other, now)
        frappe.db.commit()
        row_b = self._lock_row(self.sheet_b)
        self.assertIsNotNone(row_b, "sheet B lock acquired independently (me holding A did not block)")
        self.assertEqual(row_b["locked_by"], self.other)

        # TWO distinct lock rows for this boq, with distinct identities.
        self.assertEqual(self._boq_lock_count(), 2, "two independent locks -- no collision")
        self.assertNotEqual(
            _lock_identity(self.boq, self.sheet_a, self.cv),
            _lock_identity(self.boq, self.sheet_b, self.cv),
        )

        # NO CONTENTION: other's sheet-B acquire did NOT touch sheet A (holder + timestamp).
        row_a_after = self._lock_row(self.sheet_a)
        self.assertEqual(row_a_after["locked_by"], self.me, "sheet A holder unchanged by B's acquire")
        self.assertEqual(
            frappe.utils.get_datetime(row_a_after["last_edit_at"]),
            frappe.utils.get_datetime(a_edit_at),
            "sheet A last_edit_at untouched by sheet B's acquire",
        )

    def test_read_lock_info_is_independent_per_sheet(self):
        """Cross-check via read_lock_info: each user sees their OWN sheet as theirs and the
        OTHER's sheet as not-theirs -- the locks are independent per sheet."""
        now = frappe.utils.now_datetime()
        acquire_or_refresh(self.boq, self.sheet_a, self.cv, self.me, now)
        acquire_or_refresh(self.boq, self.sheet_b, self.cv, self.other, now)
        frappe.db.commit()

        # other looking at sheet A: held by me, NOT by other.
        li_a = read_lock_info(self.boq, self.sheet_a, self.cv, self.other, now)
        self.assertIsNotNone(li_a)
        self.assertFalse(li_a["is_locked_by_me"], "sheet A is not other's")
        self.assertEqual(li_a["locked_by_user"], self.me)

        # me looking at sheet B: held by other, NOT by me.
        li_b = read_lock_info(self.boq, self.sheet_b, self.cv, self.me, now)
        self.assertIsNotNone(li_b)
        self.assertFalse(li_b["is_locked_by_me"], "sheet B is not mine")
        self.assertEqual(li_b["locked_by_user"], self.other)

        # Each user sees their OWN sheet as theirs.
        self.assertTrue(
            read_lock_info(self.boq, self.sheet_a, self.cv, self.me, now)["is_locked_by_me"],
            "me sees sheet A as mine",
        )
        self.assertTrue(
            read_lock_info(self.boq, self.sheet_b, self.cv, self.other, now)["is_locked_by_me"],
            "other sees sheet B as theirs",
        )

    # -- Test 3: same isolation at the save_cell_price entry point ----------

    def test_save_cell_price_isolation_across_sheets(self):
        """The real frontend path: me drives save_cell_price on sheet A while other holds
        sheet B -- the save succeeds (acquires A for me), and sheet B's lock is undisturbed."""
        now = frappe.utils.now_datetime()
        # other independently holds sheet B (fresh) via the lock core.
        acquire_or_refresh(self.boq, self.sheet_b, self.cv, self.other, now)
        frappe.db.commit()
        b_before = self._lock_row(self.sheet_b)

        # me drives the REAL endpoint on sheet A -- must SUCCEED despite other holding B.
        save_cell_price(boq_name=self.boq, sheet_name=self.sheet_a, excel_row=34,
                        col_letter="E", committed_version=self.cv, rate=150.0)

        # me's save acquired sheet A's lock for me...
        row_a = self._lock_row(self.sheet_a)
        self.assertIsNotNone(row_a, "save_cell_price acquired sheet A's lock")
        self.assertEqual(row_a["locked_by"], self.me, "holder is the saver (session user)")
        # ...and wrote a current pricing row on sheet A (the save was NOT blocked by B's lock).
        cur = frappe.get_all(
            _PRICING,
            filters={"boq": self.boq, "sheet_name": self.sheet_a, "excel_row": 34,
                     "col_letter": "E", "is_current": 1},
            pluck="rate",
        )
        self.assertEqual(cur, [150.0], "sheet A price written -- not blocked by sheet B's lock")

        # sheet B's lock is UNDISTURBED by me's save on sheet A (holder + timestamp).
        b_after = self._lock_row(self.sheet_b)
        self.assertEqual(b_after["locked_by"], self.other, "sheet B still held by other")
        self.assertEqual(
            frappe.utils.get_datetime(b_after["last_edit_at"]),
            frappe.utils.get_datetime(b_before["last_edit_at"]),
            "sheet B last_edit_at untouched by the sheet A save",
        )
        # TWO independent locks now exist (A=me via the endpoint, B=other).
        self.assertEqual(self._boq_lock_count(), 2)

    # -- Test 4: same-sheet STILL contends (the contrast guard) -------------

    def test_same_sheet_still_contends_contrast_guard(self):
        """CONTRAST: the SAME sheet (A) DOES still contend -- proves the different-sheet tests
        pass BECAUSE the sheets differ, not because the lock never blocks. Driven through the
        real save_cell_price endpoint: other holds A fresh -> me's save on A is rejected."""
        now = frappe.utils.now_datetime()
        acquire_or_refresh(self.boq, self.sheet_a, self.cv, self.other, now)  # other holds A fresh
        frappe.db.commit()
        with self.assertRaises(frappe.ValidationError) as ctx:
            save_cell_price(boq_name=self.boq, sheet_name=self.sheet_a, excel_row=34,
                            col_letter="E", committed_version=self.cv, rate=150.0)
        self.assertIn(
            _LOCK_HELD_MARKER, str(ctx.exception),
            "same-sheet save by another user is REJECTED -- the lock DOES block",
        )
        # The reject mutated nothing: sheet A is still held by other (unchanged).
        self.assertEqual(self._lock_row(self.sheet_a)["locked_by"], self.other)


class TestGetCommittedSheetGrid(FrappeTestCase):
    """The faithful committed-grid read for the read-only general-specs view.

    Seeds two committed grids DIRECTLY (no commit pipeline run): a CONFIGURED data sheet
    (non-empty column_role_map etc.) and an EMPTY-CONFIG general-specs sheet (empty
    column_role_map -- the SOW shape) -- proving the row return is NEVER gated on config.
    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.test_project = _make_project()
        cls.cv = 1

        boq = frappe.new_doc("BOQs")
        boq.project = cls.test_project.name
        boq.boq_name = "TEST faithful grid BoQ"
        boq.insert(ignore_permissions=True, ignore_links=True)
        cls.boq = boq.name

        # (1) CONFIGURED sheet -- a non-empty column-config snapshot + 2 grid rows.
        cls.cfg_sheet = "Electrical"
        cls._make_grid(
            cls.cfg_sheet,
            rows=[
                {"row_number": 5, "cells": {"A": "1.1", "B": "Cabling", "C": 12}},
                {"row_number": 6, "cells": {"A": "1.2", "B": "Trunking", "C": 8}},
            ],
        )
        cls._make_boq_sheet(
            cls.cfg_sheet,
            column_role_map={"A": {"role": "sl_no", "area": None},
                             "B": {"role": "description", "area": None}},
            column_headers={"A": "Sl.No", "B": "Description"},
            area_dimensions=["Phase 1"],
            header_row=2,
            header_row_count=1,
        )

        # (2) EMPTY-CONFIG general-specs sheet -- the SOW shape (empty maps), rows still present.
        cls.gen_sheet = "SOW"
        cls._make_grid(
            cls.gen_sheet,
            rows=[
                {"row_number": 1, "cells": {"A": "Scope of Work"}},
                {"row_number": 2, "cells": {"A": "1. General", "B": "All works as per drawings"}},
                {"row_number": 3, "cells": {"A": "2. Exclusions"}},
            ],
        )
        cls._make_boq_sheet(
            cls.gen_sheet,
            column_role_map={},
            column_headers={},
            area_dimensions=[],
            header_row=0,
            header_row_count=1,
        )
        frappe.db.commit()

    @classmethod
    def _make_grid(cls, sheet_name, rows):
        grid = frappe.new_doc("BoQ Committed Sheet Grid")
        grid.boq = cls.boq
        grid.source_sheet_name = sheet_name
        grid.commit_version = cls.cv
        grid.is_current = 1
        grid.committed_at = "2026-06-21 10:00:00"
        # Seed rows OUT of row_order to prove the order_by row_order asc.
        for order, r in reversed(list(enumerate(rows))):
            grid.append("rows", {
                "row_number": r["row_number"],
                "row_order": order,
                "cells": r["cells"],
            })
        grid.insert(ignore_permissions=True, ignore_links=True)
        return grid.name

    @classmethod
    def _make_boq_sheet(cls, sheet_name, column_role_map, column_headers,
                        area_dimensions, header_row, header_row_count):
        bs = frappe.new_doc("BoQ Sheet")
        bs.boq = cls.boq
        bs.sheet_name = sheet_name
        bs.sheet_order = 0
        bs.commit_version = cls.cv
        bs.is_current = 1
        # JSON fields stored as json.dumps strings (the commit pipeline's pattern).
        bs.column_role_map = json.dumps(column_role_map)
        bs.column_headers = json.dumps(column_headers)
        bs.area_dimensions = json.dumps(area_dimensions)
        bs.header_row = header_row
        bs.header_row_count = header_row_count
        bs.insert(ignore_permissions=True, ignore_links=True)
        return bs.name

    @classmethod
    def tearDownClass(cls):
        frappe.db.delete("BoQ Committed Sheet Grid", {"boq": cls.boq})
        frappe.db.delete("BoQ Sheet", {"boq": cls.boq})
        _cleanup_project(cls.test_project.name)
        super().tearDownClass()

    def test_configured_sheet_returns_rows_and_config(self):
        res = get_committed_sheet_grid(self.boq, self.cfg_sheet, self.cv)
        # Rows returned in row_order (seeded reversed) -> ascending by row_number here.
        self.assertEqual([r["row_number"] for r in res["rows"]], [5, 6])
        self.assertEqual(res["rows"][0]["cells"], {"A": "1.1", "B": "Cabling", "C": 12})
        # Config snapshot parsed back to Python objects.
        self.assertEqual(res["column_role_map"]["A"], {"role": "sl_no", "area": None})
        self.assertEqual(res["column_headers"], {"A": "Sl.No", "B": "Description"})
        self.assertEqual(res["area_dimensions"], ["Phase 1"])
        self.assertEqual(res["header_row"], 2)
        self.assertEqual(res["header_row_count"], 1)

    def test_empty_config_general_specs_returns_rows(self):
        """THE empty-config case: a general-specs sheet (SOW shape) with an empty
        column_role_map STILL returns its grid rows -- the return is never config-gated."""
        res = get_committed_sheet_grid(self.boq, self.gen_sheet, self.cv)
        self.assertEqual(len(res["rows"]), 3, "all 3 grid rows returned despite empty config")
        self.assertEqual([r["row_number"] for r in res["rows"]], [1, 2, 3])
        self.assertEqual(res["rows"][1]["cells"], {"A": "1. General", "B": "All works as per drawings"})
        # Config is empty -- and that is fine, not an error.
        self.assertEqual(res["column_role_map"], {})
        self.assertEqual(res["column_headers"], {})
        self.assertEqual(res["area_dimensions"], [])

    def test_missing_args_throw(self):
        with self.assertRaises(frappe.ValidationError):
            get_committed_sheet_grid(None, self.gen_sheet, self.cv)
        with self.assertRaises(frappe.ValidationError):
            get_committed_sheet_grid(self.boq, None, self.cv)
        with self.assertRaises(frappe.ValidationError):
            get_committed_sheet_grid(self.boq, self.gen_sheet, None)

    def test_unknown_boq_throws(self):
        with self.assertRaises(frappe.ValidationError):
            get_committed_sheet_grid("BOQ-DOES-NOT-EXIST-999", self.gen_sheet, self.cv)

    def test_unknown_sheet_or_version_throws(self):
        with self.assertRaises(frappe.ValidationError):
            get_committed_sheet_grid(self.boq, "Nonexistent Sheet", self.cv)
        with self.assertRaises(frappe.ValidationError):
            get_committed_sheet_grid(self.boq, self.gen_sheet, 999)


class TestPriceabilityGuard(FrappeTestCase):
    """Slice 3e: save_cell_price rejects a rate on a NON-priceable committed row (node_type
    'Other') by default, and ACCEPTS it only when allow_non_priceable is asserted. Reuses the
    shared committed fixture (Preamble row 6 + Line Items 34/35) and ADDS one 'Other' node
    (a note row) at source_row 50 -- seeded in THIS class's own setup (the shared builder is
    untouched). Also asserts node_type now rides the delivered committed/priced row."""

    OTHER_ROW = 50  # the Other node's Excel source_row_number

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.test_project = _make_project()
        boq = frappe.new_doc("BOQs")
        boq.project = cls.test_project.name
        boq.boq_name = "Priceability Guard Test BoQ"
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        cls.boq = boq.name
        cls.sheet = "Guard Fix "  # VERBATIM trailing space (#152)
        cls.cv = 1
        cls.fixture = build_committed_sheet_fixture(cls.boq, cls.sheet, commit_version=cls.cv)

        # ADD an 'Other' (non-priceable) node to the SAME committed sheet -- a note row.
        # boq auto-fills from sheet (P4-2); Other carries no qty/level/parent requirement.
        other = frappe.new_doc("BOQ Nodes")
        other.sheet = cls.fixture["bqsh"]
        other.node_type = "Other"
        other.row_class = "note"
        other.description = "NOTE: rates are exclusive of taxes"
        other.sort_order = 3
        other.source_row_number = cls.OTHER_ROW
        other.commit_version = cls.cv
        other.is_current = 1
        other.committed_at = frappe.utils.now()
        other.insert(ignore_permissions=True)
        cls.other_node = other.name
        frappe.db.commit()
        # MANDATORY amount-formula gate: declare covering formulas so the sheet is formula-COMPLETE
        # -- the formula gate fires BEFORE the priceability block, so without this the formula
        # throw would PRE-EMPT the priceability throw and these tests would assert the wrong
        # message. With the sheet complete, save_cell_price reaches the priceability check.
        _declare_fixture_amount_formulas(cls.boq, cls.sheet, cls.cv)

    @classmethod
    def tearDownClass(cls):
        cleanup_committed_fixture(cls.boq)
        _cleanup_project(cls.test_project.name)
        super().tearDownClass()

    def setUp(self):
        frappe.db.delete(_PRICING, {"boq": self.boq})
        frappe.db.delete(_LOCK_DT, {"boq": self.boq})
        frappe.db.commit()

    def _price_count(self, excel_row):
        return frappe.db.count(
            _PRICING, {"boq": self.boq, "sheet_name": self.sheet, "excel_row": excel_row}
        )

    def _lock_count(self):
        return frappe.db.count(_LOCK_DT, {"boq": self.boq})

    # -- the guard: non-priceable rejected without override --------------------

    def test_non_priceable_rejected_without_override(self):
        with self.assertRaises(frappe.ValidationError) as ctx:
            save_cell_price(
                boq_name=self.boq, sheet_name=self.sheet, excel_row=self.OTHER_ROW,
                col_letter="E", committed_version=self.cv, rate=99.0,
            )
        self.assertIn("not priceable", str(ctx.exception).lower())
        # A rejected write mutated NOTHING -- no price row AND the lock was never acquired
        # (the guard sits BEFORE acquire_or_refresh).
        self.assertEqual(self._price_count(self.OTHER_ROW), 0, "no price row written")
        self.assertEqual(self._lock_count(), 0, "the lock was not acquired on a rejected write")

    # -- the override: non-priceable accepted with allow_non_priceable ---------

    def test_non_priceable_accepted_with_override(self):
        res = save_cell_price(
            boq_name=self.boq, sheet_name=self.sheet, excel_row=self.OTHER_ROW,
            col_letter="E", committed_version=self.cv, rate=99.0,
            allow_non_priceable=True,
        )
        self.assertTrue(res["ok"])
        cur = frappe.get_all(
            _PRICING,
            filters={"boq": self.boq, "sheet_name": self.sheet, "excel_row": self.OTHER_ROW,
                     "col_letter": "E", "committed_version": self.cv, "is_current": 1},
            fields=["rate", "node", "is_filled"],
        )
        self.assertEqual(len(cur), 1, "the override accepted the write -> one current price row")
        self.assertEqual(cur[0]["rate"], 99.0)
        self.assertEqual(cur[0]["node"], self.other_node, "stores the resolved Other node pointer")

    def test_override_accepts_http_string_true(self):
        # HTTP coercion: the whitelisted endpoint receives a STRING. "true" must read truthy.
        res = save_cell_price(
            boq_name=self.boq, sheet_name=self.sheet, excel_row=self.OTHER_ROW,
            col_letter="E", committed_version=self.cv, rate=12.0,
            allow_non_priceable="true",
        )
        self.assertTrue(res["ok"])
        self.assertEqual(self._price_count(self.OTHER_ROW), 1)

    # -- priceable rows save normally either way (regression) ------------------

    def test_priceable_saves_without_override(self):
        res = save_cell_price(
            boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
            committed_version=self.cv, rate=150.0, area="Phase 1",
        )
        self.assertTrue(res["ok"])
        self.assertEqual(self._price_count(34), 1)

    def test_priceable_saves_with_override_too(self):
        # The override does not break a priceable save (it just isn't needed there).
        res = save_cell_price(
            boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
            committed_version=self.cv, rate=150.0, area="Phase 1", allow_non_priceable=True,
        )
        self.assertTrue(res["ok"])
        self.assertEqual(self._price_count(34), 1)

    # -- node_type rides the delivered row (the gate's data dependency) --------

    def test_node_type_on_committed_and_priced_rows(self):
        committed = get_committed_rows(boq_name=self.boq, sheet_name=self.sheet)
        by_row = {r["source_row_number"]: r for r in committed["rows"]}
        self.assertIn("node_type", by_row[6], "node_type is surfaced on the committed row")
        self.assertEqual(by_row[6]["node_type"], "Preamble")
        self.assertEqual(by_row[34]["node_type"], "Line Item")
        self.assertEqual(by_row[self.OTHER_ROW]["node_type"], "Other")
        # get_priced_rows passes the committed rows through -> node_type rides it too.
        priced = get_priced_rows(boq_name=self.boq, sheet_name=self.sheet)
        p_by_row = {r["source_row_number"]: r for r in priced["rows"]}
        self.assertEqual(p_by_row[self.OTHER_ROW]["node_type"], "Other")


class TestPreambleQtyBearingGuard(FrappeTestCase):
    """The ASYMMETRIC rate-edit gate (owner-locked): on save_cell_price,
      - a ZERO-QTY PREAMBLE is rejected (read-only) unless override;
      - a QTY-BEARING PREAMBLE (scalar qty OR any per-area qty non-zero) is accepted;
      - a LINE ITEM is ALWAYS accepted (a zero-qty Line Item is a valid rate-only line);
      - a non-priceable type ("Other") is rejected unless override (unchanged);
      - allow_non_priceable unlocks BOTH a zero-qty Preamble AND a non-priceable type.

    Reuses the shared committed fixture (Preamble row 6 is ZERO-QTY; Line Items 34/35 are
    qty-bearing) and ADDS, in THIS class's setup: a qty-bearing Preamble (scalar qty, row 7),
    a qty-bearing-via-area Preamble (scalar qty 0 + a non-zero BOQ Node Qty By Area child,
    row 8), a ZERO-QTY Line Item (row 51), and an 'Other' node (row 52). The shared builder is
    untouched. sheet_name carries a trailing space (#152)."""

    QTY_PREAMBLE_ROW = 7      # Preamble, scalar qty non-zero -> editable
    AREA_PREAMBLE_ROW = 8     # Preamble, scalar qty 0 but a non-zero per-area child -> editable
    ZERO_LINE_ITEM_ROW = 51   # Line Item, zero qty -> STILL editable (rate-only)
    OTHER_ROW = 52            # Other -> non-priceable, rejected unless override
    ZERO_PREAMBLE_ROW = 6     # the shared fixture's Preamble (zero-qty) -> read-only

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.test_project = _make_project()
        boq = frappe.new_doc("BOQs")
        boq.project = cls.test_project.name
        boq.boq_name = "Preamble Qty Gate Test BoQ"
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        cls.boq = boq.name
        cls.sheet = "Qty Gate "  # VERBATIM trailing space (#152)
        cls.cv = 1
        cls.fixture = build_committed_sheet_fixture(cls.boq, cls.sheet, commit_version=cls.cv)
        bqsh = cls.fixture["bqsh"]
        now = frappe.utils.now()

        def _node(node_type, row_class, source_row, sort_order, qty=None, area_qty=None, level=None):
            n = frappe.new_doc("BOQ Nodes")
            n.sheet = bqsh
            n.node_type = node_type
            n.row_class = row_class
            n.description = f"{node_type} @ {source_row}"
            n.sort_order = sort_order
            n.source_row_number = source_row
            if qty is not None:
                n.qty = qty
            if level is not None:
                n.level = level
            n.commit_version = cls.cv
            n.is_current = 1
            n.committed_at = now
            if area_qty is not None:
                for area_name, q in area_qty.items():
                    n.append("qty_by_area", {"area_name": area_name, "qty": q})
            n.insert(ignore_permissions=True)
            return n.name

        # Preamble WITH a scalar qty -> editable.
        cls.qty_preamble = _node("Preamble", "preamble", cls.QTY_PREAMBLE_ROW, 4, qty=10.0, level=1)
        # Preamble with scalar qty 0 but a NON-ZERO per-area child -> editable (the child read).
        cls.area_preamble = _node(
            "Preamble", "preamble", cls.AREA_PREAMBLE_ROW, 5, qty=0.0, level=1,
            area_qty={"Phase 1": 0.0, "Phase 2": 7.0},
        )
        # Line Item with ZERO qty (and a zero per-area child) -> STILL editable (rate-only).
        cls.zero_line_item = _node(
            "Line Item", "line_item", cls.ZERO_LINE_ITEM_ROW, 6, qty=0.0,
            area_qty={"Phase 1": 0.0, "Phase 2": 0.0},
        )
        # A non-priceable 'Other' node -> rejected unless override.
        cls.other_node = _node("Other", "note", cls.OTHER_ROW, 7)
        frappe.db.commit()
        # MANDATORY amount-formula gate: declare covering formulas so the sheet is formula-COMPLETE
        # -- the formula gate fires BEFORE the qty/priceability block, so without this it would
        # PRE-EMPT the qty-gate throws these tests assert. With the sheet complete, save_cell_price
        # reaches the asymmetric qty gate.
        _declare_fixture_amount_formulas(cls.boq, cls.sheet, cls.cv)

    @classmethod
    def tearDownClass(cls):
        cleanup_committed_fixture(cls.boq)
        _cleanup_project(cls.test_project.name)
        super().tearDownClass()

    def setUp(self):
        frappe.db.delete(_PRICING, {"boq": self.boq})
        frappe.db.delete(_LOCK_DT, {"boq": self.boq})
        frappe.db.commit()

    def _price_count(self, excel_row):
        return frappe.db.count(
            _PRICING, {"boq": self.boq, "sheet_name": self.sheet, "excel_row": excel_row}
        )

    def _lock_count(self):
        return frappe.db.count(_LOCK_DT, {"boq": self.boq})

    def _save(self, excel_row, **kw):
        return save_cell_price(
            boq_name=self.boq, sheet_name=self.sheet, excel_row=excel_row,
            col_letter="E", committed_version=self.cv, rate=99.0, **kw,
        )

    # -- Preamble: zero-qty rejected, qty-bearing accepted ---------------------

    def test_zero_qty_preamble_rejected_without_override(self):
        with self.assertRaises(frappe.ValidationError) as ctx:
            self._save(self.ZERO_PREAMBLE_ROW)
        self.assertIn("not priceable", str(ctx.exception).lower())
        # A rejected write mutated NOTHING (the guard sits BEFORE acquire_or_refresh).
        self.assertEqual(self._price_count(self.ZERO_PREAMBLE_ROW), 0, "no price row written")
        self.assertEqual(self._lock_count(), 0, "the lock was not acquired on a rejected write")

    def test_qty_bearing_preamble_accepted_scalar_qty(self):
        res = self._save(self.QTY_PREAMBLE_ROW)
        self.assertTrue(res["ok"])
        self.assertEqual(self._price_count(self.QTY_PREAMBLE_ROW), 1)

    def test_qty_bearing_preamble_accepted_via_area_child(self):
        # scalar qty is 0 but a per-area child qty is non-zero -> the child read makes it editable.
        res = self._save(self.AREA_PREAMBLE_ROW)
        self.assertTrue(res["ok"])
        self.assertEqual(self._price_count(self.AREA_PREAMBLE_ROW), 1)

    def test_zero_qty_preamble_accepted_with_override(self):
        res = self._save(self.ZERO_PREAMBLE_ROW, allow_non_priceable=True)
        self.assertTrue(res["ok"])
        self.assertEqual(self._price_count(self.ZERO_PREAMBLE_ROW), 1)

    # -- Line Item: always editable, even at zero qty (rate-only) --------------

    def test_zero_qty_line_item_accepted_without_override(self):
        res = self._save(self.ZERO_LINE_ITEM_ROW)
        self.assertTrue(res["ok"], "a zero-qty Line Item is a valid rate-only line -> editable")
        self.assertEqual(self._price_count(self.ZERO_LINE_ITEM_ROW), 1)

    def test_qty_bearing_line_item_accepted(self):
        # the shared fixture's Line Item 34 (qty-bearing) -- unchanged behaviour.
        res = self._save(34, area="Phase 1")
        self.assertTrue(res["ok"])
        self.assertEqual(self._price_count(34), 1)

    # -- non-priceable type: unchanged (rejected unless override) --------------

    def test_other_type_rejected_without_override(self):
        with self.assertRaises(frappe.ValidationError) as ctx:
            self._save(self.OTHER_ROW)
        self.assertIn("not priceable", str(ctx.exception).lower())
        self.assertEqual(self._price_count(self.OTHER_ROW), 0)

    def test_other_type_accepted_with_override(self):
        res = self._save(self.OTHER_ROW, allow_non_priceable=True)
        self.assertTrue(res["ok"])
        self.assertEqual(self._price_count(self.OTHER_ROW), 1)


class TestRowRemark(FrappeTestCase):
    """Slice 4a: the per-ROW remark layer (BoQ Cell Remark) -- save_row_remark + the
    get_priced_rows merge. Reuses the shared per-area committed fixture (line items at
    excel_row 34/35). sheet_name carries a trailing space (#152). Each test starts with a
    clean remark layer + lock (the committed fixture persists)."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.test_project = _make_project()
        boq = frappe.new_doc("BOQs")
        boq.project = cls.test_project.name
        boq.boq_name = "Row-Remark Test BoQ"
        boq.tax_treatment = "Pre-tax"
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        cls.boq = boq.name
        cls.sheet = "Remark Fix "  # VERBATIM trailing space (#152)
        cls.cv = 1
        cls.fixture = build_committed_sheet_fixture(cls.boq, cls.sheet, commit_version=cls.cv)
        cls.me = frappe.session.user
        cls.other = frappe.db.get_value(
            "User", {"name": ["not in", [cls.me, "Guest"]], "enabled": 1}, "name"
        )
        assert cls.other, "need a second real User to play the competing lock holder"

    @classmethod
    def tearDownClass(cls):
        frappe.db.delete(_LOCK_DT, {"boq": cls.boq})
        frappe.db.delete(_REMARK_DT, {"boq": cls.boq})
        cleanup_committed_fixture(cls.boq)
        _cleanup_project(cls.test_project.name)
        super().tearDownClass()

    def setUp(self):
        frappe.db.delete(_REMARK_DT, {"boq": self.boq})
        frappe.db.delete(_LOCK_DT, {"boq": self.boq})
        frappe.db.commit()

    # -- helpers ------------------------------------------------------------

    def _current(self, excel_row):
        return frappe.get_all(
            _REMARK_DT,
            filters={"boq": self.boq, "sheet_name": self.sheet, "excel_row": excel_row,
                     "committed_version": self.cv, "is_current": 1},
            fields=["name", "remark", "remark_version", "description"],
        )

    def _seed_lock(self, user):
        acquire_or_refresh(self.boq, self.sheet, self.cv, user, frappe.utils.now_datetime())
        frappe.db.commit()

    # -- POSITIVE -----------------------------------------------------------

    def test_save_creates_current_remark_v1(self):
        res = save_row_remark(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                              committed_version=self.cv, remark="check the quantity",
                              description="cable 1.1.2")
        self.assertTrue(res["ok"])
        self.assertEqual(res["remark_version"], 1)
        self.assertEqual(res["froze_prior"], 0, "first save freezes nothing")
        self.assertFalse(res["cleared"])
        cur = self._current(34)
        self.assertEqual(len(cur), 1, "exactly one current remark record")
        self.assertEqual(cur[0]["remark"], "check the quantity")
        self.assertEqual(cur[0]["description"], "cable 1.1.2", "guard carried, not branched on")

    def test_resave_freezes_prior_and_supersedes(self):
        save_row_remark(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                        committed_version=self.cv, remark="first")
        res2 = save_row_remark(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                               committed_version=self.cv, remark="second")
        self.assertEqual(res2["remark_version"], 2)
        self.assertEqual(res2["froze_prior"], 1, "the re-save froze the prior current")
        cur = self._current(34)
        self.assertEqual(len(cur), 1, "exactly ONE current after re-save (freeze-and-supersede)")
        self.assertEqual(cur[0]["remark"], "second")
        self.assertEqual(cur[0]["remark_version"], 2)

    def test_clear_blank_reads_as_no_remark(self):
        save_row_remark(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                        committed_version=self.cv, remark="to be cleared")
        res = save_row_remark(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                              committed_version=self.cv, remark="   ")
        self.assertTrue(res["cleared"], "blank/whitespace clears the remark")
        self.assertEqual(res["froze_prior"], 1)
        self.assertEqual(self._current(34), [], "no current remark after a clear")
        got = get_sheet_remarks(boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv)
        self.assertEqual(got["remarks"], [], "a cleared remark does not surface")

    def test_cap_250_ok_251_throws(self):
        ok = save_row_remark(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                             committed_version=self.cv, remark="x" * 250)
        self.assertTrue(ok["ok"], "exactly 250 chars is allowed")
        with self.assertRaises(frappe.ValidationError):
            save_row_remark(boq_name=self.boq, sheet_name=self.sheet, excel_row=35,
                            committed_version=self.cv, remark="y" * 251)
        self.assertEqual(self._current(35), [], "the over-cap remark wrote nothing")

    # -- NEGATIVE -----------------------------------------------------------

    def test_nonexistent_row_throws(self):
        with self.assertRaises(frappe.ValidationError):
            save_row_remark(boq_name=self.boq, sheet_name=self.sheet, excel_row=9999,
                            committed_version=self.cv, remark="orphan")
        self.assertEqual(self._current(9999), [], "no remark for a non-existent row")

    def test_lock_held_by_other_rejects_and_mutates_nothing(self):
        self._seed_lock(self.other)  # other holds a FRESH lock
        with self.assertRaises(frappe.ValidationError) as ctx:
            save_row_remark(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                            committed_version=self.cv, remark="blocked")
        self.assertIn(_LOCK_HELD_MARKER, str(ctx.exception), "reject carries the lock marker")
        self.assertEqual(self._current(34), [], "a lock-rejected remark wrote nothing")

    # -- MERGE into get_priced_rows -----------------------------------------

    def test_get_priced_rows_surfaces_remark(self):
        save_row_remark(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                        committed_version=self.cv, remark="see drawing rev B")
        res = get_priced_rows(boq_name=self.boq, sheet_name=self.sheet)
        by_row = {r["source_row_number"]: r for r in res["rows"]}
        self.assertEqual(by_row[34]["remark"], "see drawing rev B", "remark merged onto its row")
        self.assertIsNone(by_row[35]["remark"], "a row with no remark surfaces remark None")


class TestCellColor(FrappeTestCase):
    """Slice 4a: the per-CELL color layer (BoQ Cell Color) -- save_cell_color + the
    get_priced_rows merge. Reuses the shared committed fixture (line items 34/35) and ADDS
    one 'Other' (non-priceable) node at source_row 50 to prove color is allowed where a
    PRICE would be rejected. sheet_name carries a trailing space (#152)."""

    OTHER_ROW = 50

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.test_project = _make_project()
        boq = frappe.new_doc("BOQs")
        boq.project = cls.test_project.name
        boq.boq_name = "Cell-Color Test BoQ"
        boq.tax_treatment = "Pre-tax"
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        cls.boq = boq.name
        cls.sheet = "Color Fix "  # VERBATIM trailing space (#152)
        cls.cv = 1
        cls.fixture = build_committed_sheet_fixture(cls.boq, cls.sheet, commit_version=cls.cv)

        # An 'Other' (non-priceable) node -- a note row at source_row 50.
        other = frappe.new_doc("BOQ Nodes")
        other.sheet = cls.fixture["bqsh"]
        other.node_type = "Other"
        other.row_class = "note"
        other.description = "NOTE: rates exclusive of taxes"
        other.sort_order = 3
        other.source_row_number = cls.OTHER_ROW
        other.commit_version = cls.cv
        other.is_current = 1
        other.committed_at = frappe.utils.now()
        other.insert(ignore_permissions=True)
        cls.other_node = other.name
        frappe.db.commit()

        cls.me = frappe.session.user
        cls.other_user = frappe.db.get_value(
            "User", {"name": ["not in", [cls.me, "Guest"]], "enabled": 1}, "name"
        )
        assert cls.other_user, "need a second real User to play the competing lock holder"

    @classmethod
    def tearDownClass(cls):
        frappe.db.delete(_LOCK_DT, {"boq": cls.boq})
        frappe.db.delete(_COLOR_DT, {"boq": cls.boq})
        cleanup_committed_fixture(cls.boq)
        _cleanup_project(cls.test_project.name)
        super().tearDownClass()

    def setUp(self):
        frappe.db.delete(_COLOR_DT, {"boq": self.boq})
        frappe.db.delete(_PRICING, {"boq": self.boq})
        frappe.db.delete(_LOCK_DT, {"boq": self.boq})
        frappe.db.commit()

    # -- helpers ------------------------------------------------------------

    def _current(self, excel_row, col_letter):
        return frappe.get_all(
            _COLOR_DT,
            filters={"boq": self.boq, "sheet_name": self.sheet, "excel_row": excel_row,
                     "col_letter": col_letter, "committed_version": self.cv, "is_current": 1},
            fields=["name", "color", "color_version"],
        )

    def _seed_lock(self, user):
        acquire_or_refresh(self.boq, self.sheet, self.cv, user, frappe.utils.now_datetime())
        frappe.db.commit()

    # -- POSITIVE -----------------------------------------------------------

    def test_save_creates_current_color_v1(self):
        res = save_cell_color(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                              col_letter="E", committed_version=self.cv, color="red",
                              description="cable 1.1.2")
        self.assertTrue(res["ok"])
        self.assertEqual(res["color_version"], 1)
        self.assertFalse(res["cleared"])
        cur = self._current(34, "E")
        self.assertEqual(len(cur), 1)
        self.assertEqual(cur[0]["color"], "red")

    def test_resave_freezes_prior_and_supersedes(self):
        save_cell_color(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                        col_letter="E", committed_version=self.cv, color="red")
        res2 = save_cell_color(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                               col_letter="E", committed_version=self.cv, color="blue")
        self.assertEqual(res2["color_version"], 2)
        self.assertEqual(res2["froze_prior"], 1)
        cur = self._current(34, "E")
        self.assertEqual(len(cur), 1, "exactly ONE current after re-save")
        self.assertEqual(cur[0]["color"], "blue")

    def test_all_8_tokens_accepted(self):
        for token in ("red", "orange", "yellow", "green", "blue", "purple", "pink", "grey"):
            res = save_cell_color(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                                  col_letter="E", committed_version=self.cv, color=token)
            self.assertTrue(res["ok"], f"token {token} accepted")
        self.assertEqual(self._current(34, "E")[0]["color"], "grey", "last token is current")

    def test_invalid_color_throws(self):
        with self.assertRaises(frappe.ValidationError):
            save_cell_color(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                            col_letter="E", committed_version=self.cv, color="teal")
        self.assertEqual(self._current(34, "E"), [], "an invalid color wrote nothing")

    def test_color_allowed_on_non_priceable_cell(self):
        # THE contrast: a PRICE on the Other row is rejected without the override...
        with self.assertRaises(frappe.ValidationError):
            save_cell_price(boq_name=self.boq, sheet_name=self.sheet, excel_row=self.OTHER_ROW,
                            col_letter="E", committed_version=self.cv, rate=10.0)
        # ...but a COLOR on the SAME non-priceable cell is accepted (no priceability gate).
        res = save_cell_color(boq_name=self.boq, sheet_name=self.sheet, excel_row=self.OTHER_ROW,
                              col_letter="E", committed_version=self.cv, color="green")
        self.assertTrue(res["ok"])
        cur = self._current(self.OTHER_ROW, "E")
        self.assertEqual(len(cur), 1, "color saved on a non-priceable cell")
        self.assertEqual(cur[0]["color"], "green")

    def test_clear_blank_no_current(self):
        save_cell_color(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                        col_letter="E", committed_version=self.cv, color="red")
        res = save_cell_color(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                              col_letter="E", committed_version=self.cv, color="  ")
        self.assertTrue(res["cleared"])
        self.assertEqual(self._current(34, "E"), [], "no current color after a clear")
        got = get_sheet_colors(boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv)
        self.assertEqual(got["colors"], [])

    # -- NEGATIVE -----------------------------------------------------------

    def test_nonexistent_row_throws(self):
        with self.assertRaises(frappe.ValidationError):
            save_cell_color(boq_name=self.boq, sheet_name=self.sheet, excel_row=9999,
                            col_letter="E", committed_version=self.cv, color="red")

    def test_lock_held_by_other_rejects_and_mutates_nothing(self):
        self._seed_lock(self.other_user)
        with self.assertRaises(frappe.ValidationError) as ctx:
            save_cell_color(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                            col_letter="E", committed_version=self.cv, color="red")
        self.assertIn(_LOCK_HELD_MARKER, str(ctx.exception))
        self.assertEqual(self._current(34, "E"), [], "a lock-rejected color wrote nothing")

    # -- MERGE into get_priced_rows -----------------------------------------

    def test_get_priced_rows_surfaces_color_by_cell(self):
        save_cell_color(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                        col_letter="E", committed_version=self.cv, color="yellow")
        res = get_priced_rows(boq_name=self.boq, sheet_name=self.sheet)
        by_row = {r["source_row_number"]: r for r in res["rows"]}
        self.assertEqual(by_row[34]["color_by_cell"], {"E": "yellow"}, "color merged per cell")
        self.assertNotIn("color_by_cell", by_row[35], "a row with no color carries no color_by_cell")


def _build_scalar_amount_committed_sheet(boq_name: str, sheet_name: str, commit_version: int = 1) -> str:
    """Insert a minimal committed sheet whose AMOUNT column is SCALAR (area=None) -- the
    branch the per-area shared fixture (amount cols are *_by_area) does NOT exercise. Columns
    A=sl_no, B=description, C=qty_total, D=rate_combined (scalar rate), E=amount_total (scalar
    amount). 1 Preamble + 1 Line Item at Excel row 10. Returns the line-item node name.
    sheet_name VERBATIM (#152). Cleaned up by cleanup_committed_fixture (deletes all committed
    rows for the boq)."""
    now = frappe.utils.now()
    bs = frappe.new_doc("BoQ Sheet")
    bs.boq = boq_name
    bs.sheet_name = sheet_name  # VERBATIM (#152)
    bs.sheet_order = 3
    bs.treat_as = "data"
    bs.header_row = 1
    bs.header_row_count = 1
    bs.column_role_map = {
        "A": {"role": "sl_no", "area": None},
        "B": {"role": "description", "area": None},
        "C": {"role": "qty_total", "area": None},
        "D": {"role": "rate_combined", "area": None},   # SCALAR rate
        "E": {"role": "amount_total", "area": None},    # SCALAR amount (the formula target)
    }
    bs.column_headers = {}
    bs.area_dimensions = json.dumps([])
    bs.commit_version = commit_version
    bs.is_current = 1
    bs.committed_at = now
    bs.insert(ignore_permissions=True)

    pre = frappe.new_doc("BOQ Nodes")
    pre.sheet = bs.name
    pre.node_type = "Preamble"
    pre.row_class = "preamble"
    pre.level = 1
    pre.description = "SCALAR AMOUNT PREAMBLE"
    pre.code = "SA.0"
    pre.sort_order = 0
    pre.source_row_number = 8
    pre.commit_version = commit_version
    pre.is_current = 1
    pre.committed_at = now
    pre.insert(ignore_permissions=True)

    li = frappe.new_doc("BOQ Nodes")
    li.sheet = bs.name
    li.node_type = "Line Item"
    li.row_class = "line_item"
    li.description = "scalar amount item"
    li.code = "SA.1"
    li.parent_node = pre.name
    li.qty = 5.0
    li.unit = "Nos"
    li.combined_rate = 999.0
    li.total_amount = 4995.0
    li.source_row_number = 10
    li.sort_order = 1
    li.commit_version = commit_version
    li.is_current = 1
    li.committed_at = now
    li.insert(ignore_permissions=True)

    frappe.db.commit()
    return li.name


# A well-formed token tree: qty x rate (the area-WILDCARD default; area-bound operands carry
# value_key=None = "bind to the current area" -- the F1 default-template shape).
_WILDCARD_QTY_X_RATE = {
    "op": "*",
    "operands": [
        {"ref": {"value_field": "qty_by_area", "value_key": None, "rate_subkey": None}},
        {"ref": {"value_field": "rate_by_area", "value_key": None, "rate_subkey": "combined_rate"}},
    ],
}
# The PER-AREA OVERRIDE analog: operands name a concrete area (value_key set).
_PHASE1_QTY_X_RATE = {
    "op": "*",
    "operands": [
        {"ref": {"value_field": "qty_by_area", "value_key": "Phase 1", "rate_subkey": None}},
        {"ref": {"value_field": "rate_by_area", "value_key": "Phase 1", "rate_subkey": "combined_rate"}},
    ],
}
# A SCALAR formula: scalar qty_total x scalar rate_combined.
_SCALAR_QTY_X_RATE = {
    "op": "*",
    "operands": [
        {"ref": {"value_field": "qty_total", "value_key": None, "rate_subkey": None}},
        {"ref": {"value_field": "rate_combined", "value_key": None, "rate_subkey": None}},
    ],
}


class TestAmountFormula(FrappeTestCase):
    """Formula Builder F1: the BoQ Cell Amount Formula doctype + save_amount_formula /
    get_sheet_amount_formulas + the get_priced_rows column_formulas merge.

    Reuses the shared per-area committed fixture (build_committed_sheet_fixture: amount cols
    F [Phase 1, total] + I [Phase 2, install]; rate cols E/H combined) + a local SCALAR-amount
    committed sheet (amount col E = amount_total scalar). sheet_name carries a trailing space
    (#152). The committed tier is read-only here; each test starts with a clean formula layer
    + lock. `me` = session user; `other` = a second real User for the lock-contention test.
    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.test_project = _make_project()
        boq = frappe.new_doc("BOQs")
        boq.project = cls.test_project.name
        boq.boq_name = "Amount-Formula Test BoQ"
        boq.tax_treatment = "Pre-tax"
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        cls.boq = boq.name
        cls.cv = 1
        cls.sheet = "Formula Fix "  # VERBATIM trailing space (#152)
        cls.fixture = build_committed_sheet_fixture(cls.boq, cls.sheet, commit_version=cls.cv)
        cls.scalar_sheet = "Formula Scalar Fix "  # VERBATIM trailing space (#152)
        cls.scalar_node = _build_scalar_amount_committed_sheet(cls.boq, cls.scalar_sheet, cls.cv)
        cls.me = frappe.session.user
        cls.other = frappe.db.get_value(
            "User", {"name": ["not in", [cls.me, "Guest"]], "enabled": 1}, "name"
        )
        assert cls.other, "need a second real User to play the competing lock holder"

    @classmethod
    def tearDownClass(cls):
        frappe.db.delete(_FORMULA_DT, {"boq": cls.boq})
        frappe.db.delete(_LOCK_DT, {"boq": cls.boq})
        cleanup_committed_fixture(cls.boq)
        _cleanup_project(cls.test_project.name)
        super().tearDownClass()

    def setUp(self):
        frappe.db.delete(_FORMULA_DT, {"boq": self.boq})
        frappe.db.delete(_LOCK_DT, {"boq": self.boq})
        frappe.db.commit()

    # -- helpers ------------------------------------------------------------

    def _current(self, sheet, tvf, tvk, trs):
        """Current records for one column identity (None-safe identity match in Python --
        avoids NULL-filter subtleties)."""
        recs = frappe.get_all(
            _FORMULA_DT,
            filters={"boq": self.boq, "sheet_name": sheet, "committed_version": self.cv,
                     "is_current": 1},
            fields=["name", "target_value_field", "target_value_key", "target_rate_subkey",
                    "formula", "formula_version"],
        )
        return [r for r in recs
                if r["target_value_field"] == tvf and r["target_value_key"] == tvk
                and r["target_rate_subkey"] == trs]

    def _all_current(self, sheet):
        return frappe.get_all(
            _FORMULA_DT,
            filters={"boq": self.boq, "sheet_name": sheet, "committed_version": self.cv,
                     "is_current": 1},
            fields=["target_value_field", "target_value_key", "target_rate_subkey"],
        )

    def _count(self, sheet):
        return frappe.db.count(_FORMULA_DT, {"boq": self.boq, "sheet_name": sheet})

    def _seed_lock(self, user):
        acquire_or_refresh(self.boq, self.sheet, self.cv, user, frappe.utils.now_datetime())
        frappe.db.commit()

    # -- POSITIVE: default / override / scalar ------------------------------

    def test_save_default_wildcard_current_v1(self):
        # A per-area DEFAULT for the "total" amount column: target_value_key NULL (wildcard).
        res = save_amount_formula(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
            target_value_field="amount_by_area", target_value_key=None,
            target_rate_subkey="total", formula=_WILDCARD_QTY_X_RATE, target_col="F",
            description="amount total",
        )
        self.assertTrue(res["ok"])
        self.assertEqual(res["formula_version"], 1)
        self.assertEqual(res["froze_prior"], 0)
        self.assertFalse(res["cleared"])
        cur = self._current(self.sheet, "amount_by_area", None, "total")
        self.assertEqual(len(cur), 1, "exactly one current default record")
        # formula stored verbatim (parsed back to the same object).
        self.assertEqual(json.loads(cur[0]["formula"]), _WILDCARD_QTY_X_RATE)

    def test_default_and_override_coexist_as_distinct(self):
        # DEFAULT (key NULL) + OVERRIDE (concrete area) for the SAME logical column -> two
        # SEPARATE current records (distinct identities), both current.
        save_amount_formula(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
            target_value_field="amount_by_area", target_value_key=None,
            target_rate_subkey="total", formula=_WILDCARD_QTY_X_RATE,
        )
        save_amount_formula(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
            target_value_field="amount_by_area", target_value_key="Phase 1",
            target_rate_subkey="total", formula=_PHASE1_QTY_X_RATE,
        )
        default_cur = self._current(self.sheet, "amount_by_area", None, "total")
        override_cur = self._current(self.sheet, "amount_by_area", "Phase 1", "total")
        self.assertEqual(len(default_cur), 1, "the wildcard default is its own current record")
        self.assertEqual(len(override_cur), 1, "the per-area override is a DISTINCT current record")
        self.assertEqual(len(self._all_current(self.sheet)), 2,
                         "default + override coexist (distinct identities)")

    def test_save_scalar_amount_formula(self):
        res = save_amount_formula(
            boq_name=self.boq, sheet_name=self.scalar_sheet, committed_version=self.cv,
            target_value_field="amount_total", target_value_key=None,
            target_rate_subkey=None, formula=_SCALAR_QTY_X_RATE, target_col="E",
        )
        self.assertTrue(res["ok"])
        cur = self._current(self.scalar_sheet, "amount_total", None, None)
        self.assertEqual(len(cur), 1, "one current scalar formula")

    # -- freeze-and-supersede + clear ---------------------------------------

    def test_resave_freezes_prior_and_supersedes(self):
        save_amount_formula(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
            target_value_field="amount_by_area", target_value_key=None,
            target_rate_subkey="total", formula=_WILDCARD_QTY_X_RATE,
        )
        res2 = save_amount_formula(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
            target_value_field="amount_by_area", target_value_key=None,
            target_rate_subkey="total", formula=_PHASE1_QTY_X_RATE,  # a different (valid) tree
        )
        self.assertEqual(res2["formula_version"], 2)
        self.assertEqual(res2["froze_prior"], 1, "the re-save froze the prior current")
        cur = self._current(self.sheet, "amount_by_area", None, "total")
        self.assertEqual(len(cur), 1, "exactly ONE current after re-save (the invariant)")
        self.assertEqual(cur[0]["formula_version"], 2)
        # total versions for this identity = 2 (v1 frozen, v2 current).
        allv = frappe.get_all(
            _FORMULA_DT,
            filters={"boq": self.boq, "sheet_name": self.sheet, "committed_version": self.cv,
                     "target_value_field": "amount_by_area", "target_rate_subkey": "total",
                     "target_value_key": ["is", "not set"]},
            fields=["formula_version", "is_current"], order_by="formula_version asc",
        )
        self.assertEqual([v["formula_version"] for v in allv], [1, 2])
        self.assertEqual([v["is_current"] for v in allv], [0, 1], "v1 frozen, v2 current")

    def test_clear_blank_formula_removes_current(self):
        save_amount_formula(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
            target_value_field="amount_by_area", target_value_key=None,
            target_rate_subkey="total", formula=_WILDCARD_QTY_X_RATE,
        )
        res = save_amount_formula(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
            target_value_field="amount_by_area", target_value_key=None,
            target_rate_subkey="total", formula="",  # CLEAR
        )
        self.assertTrue(res["cleared"])
        self.assertEqual(res["froze_prior"], 1, "clear froze the prior current")
        self.assertEqual(self._current(self.sheet, "amount_by_area", None, "total"), [],
                         "no current formula after clear")

    def test_accepts_json_string_formula(self):
        # The wire may send the formula as a JSON STRING (not a dict) -- both accepted.
        res = save_amount_formula(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
            target_value_field="amount_by_area", target_value_key=None,
            target_rate_subkey="total", formula=json.dumps(_WILDCARD_QTY_X_RATE),
        )
        self.assertTrue(res["ok"])
        cur = self._current(self.sheet, "amount_by_area", None, "total")
        self.assertEqual(json.loads(cur[0]["formula"]), _WILDCARD_QTY_X_RATE)

    # -- NEGATIVE: amount-target gate ---------------------------------------

    def test_reject_non_amount_target_rate(self):
        # A RATE target (value_field rate_by_area) is rejected -> nothing written.
        with self.assertRaises(frappe.ValidationError):
            save_amount_formula(
                boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
                target_value_field="rate_by_area", target_value_key="Phase 1",
                target_rate_subkey="combined_rate", formula=_WILDCARD_QTY_X_RATE,
            )
        self.assertEqual(self._count(self.sheet), 0, "a non-amount target wrote nothing")

    def test_reject_non_amount_target_qty(self):
        with self.assertRaises(frappe.ValidationError):
            save_amount_formula(
                boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
                target_value_field="qty_by_area", target_value_key="Phase 1",
                target_rate_subkey=None, formula=_WILDCARD_QTY_X_RATE,
            )
        self.assertEqual(self._count(self.sheet), 0)

    def test_reject_amount_target_not_on_sheet(self):
        # An amount value_field that has no matching committed column (rate_subkey 'supply'
        # is not mapped on this sheet) -> throw, nothing written.
        with self.assertRaises(frappe.ValidationError):
            save_amount_formula(
                boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
                target_value_field="amount_by_area", target_value_key=None,
                target_rate_subkey="supply", formula=_WILDCARD_QTY_X_RATE,
            )
        self.assertEqual(self._count(self.sheet), 0)

    # -- NEGATIVE: structural validation ------------------------------------

    def test_reject_literal_node(self):
        with self.assertRaises(frappe.ValidationError):
            save_amount_formula(
                boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
                target_value_field="amount_by_area", target_value_key=None,
                target_rate_subkey="total",
                formula={"op": "+", "operands": [{"literal": 5},
                                                 {"ref": {"value_field": "qty_by_area",
                                                          "value_key": None, "rate_subkey": None}}]},
            )
        self.assertEqual(self._count(self.sheet), 0, "a literal node wrote nothing")

    def test_reject_bad_operator(self):
        with self.assertRaises(frappe.ValidationError):
            save_amount_formula(
                boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
                target_value_field="amount_by_area", target_value_key=None,
                target_rate_subkey="total",
                formula={"op": "-", "operands": [
                    {"ref": {"value_field": "qty_by_area", "value_key": None, "rate_subkey": None}}]},
            )
        self.assertEqual(self._count(self.sheet), 0)

    def test_reject_empty_operands(self):
        with self.assertRaises(frappe.ValidationError):
            save_amount_formula(
                boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
                target_value_field="amount_by_area", target_value_key=None,
                target_rate_subkey="total", formula={"op": "+", "operands": []},
            )
        self.assertEqual(self._count(self.sheet), 0)

    def test_reject_malformed_validates_before_lock(self):
        # A malformed formula must mutate nothing -- including NOT acquiring the lock (the
        # structural check is BEFORE acquire_or_refresh).
        with self.assertRaises(frappe.ValidationError):
            save_amount_formula(
                boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
                target_value_field="amount_by_area", target_value_key=None,
                target_rate_subkey="total", formula={"nonsense": True},
            )
        self.assertEqual(self._count(self.sheet), 0)
        self.assertEqual(frappe.db.count(_LOCK_DT, {"boq": self.boq}), 0,
                         "a malformed-formula reject never acquired the lock")

    # -- NEGATIVE: lock contention ------------------------------------------

    def test_lock_held_by_other_rejects_and_mutates_nothing(self):
        self._seed_lock(self.other)
        with self.assertRaises(frappe.ValidationError) as ctx:
            save_amount_formula(
                boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
                target_value_field="amount_by_area", target_value_key=None,
                target_rate_subkey="total", formula=_WILDCARD_QTY_X_RATE,
            )
        self.assertIn(_LOCK_HELD_MARKER, str(ctx.exception))
        self.assertEqual(self._count(self.sheet), 0, "a lock-rejected formula wrote nothing")

    # -- READ + MERGE -------------------------------------------------------

    def test_get_sheet_amount_formulas_returns_current_parsed(self):
        save_amount_formula(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
            target_value_field="amount_by_area", target_value_key=None,
            target_rate_subkey="total", formula=_WILDCARD_QTY_X_RATE,
        )
        res = get_sheet_amount_formulas(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv
        )
        self.assertEqual(len(res["formulas"]), 1)
        f = res["formulas"][0]
        self.assertEqual(f["formula"], _WILDCARD_QTY_X_RATE, "formula returned PARSED (object)")
        self.assertIsNone(f["target_value_key"], "default is target_value_key NULL")

    def test_get_sheet_amount_formulas_empty_when_none(self):
        res = get_sheet_amount_formulas(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv
        )
        self.assertEqual(res, {"formulas": []})

    def test_get_priced_rows_surfaces_column_formulas(self):
        # default + override both appear as DISTINCT per-column entries in column_formulas.
        save_amount_formula(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
            target_value_field="amount_by_area", target_value_key=None,
            target_rate_subkey="total", formula=_WILDCARD_QTY_X_RATE, target_col="F",
        )
        save_amount_formula(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
            target_value_field="amount_by_area", target_value_key="Phase 1",
            target_rate_subkey="total", formula=_PHASE1_QTY_X_RATE, target_col="F",
        )
        res = get_priced_rows(boq_name=self.boq, sheet_name=self.sheet)
        self.assertIn("column_formulas", res, "the envelope carries column_formulas")
        cf = res["column_formulas"]
        self.assertEqual(len(cf), 2, "default + override both surface as distinct entries")
        # shape: per-column (NOT per-row); keys present + formula parsed.
        keys = {(e["target_value_field"], e["target_value_key"], e["target_rate_subkey"]) for e in cf}
        self.assertEqual(
            keys,
            {("amount_by_area", None, "total"), ("amount_by_area", "Phase 1", "total")},
        )
        for e in cf:
            self.assertIn("formula", e)
            self.assertIn("target_col", e)
            self.assertIsInstance(e["formula"], dict, "formula is a parsed object")
        # And NOT stamped onto any row.
        for row in res["rows"]:
            self.assertNotIn("column_formulas", row)
            self.assertNotIn("formula", row)

    def test_get_priced_rows_column_formulas_empty_uncommitted(self):
        res = get_priced_rows(boq_name=self.boq, sheet_name="No Such Sheet ZZ ")
        self.assertEqual(res["column_formulas"], [], "uncommitted sheet -> empty column_formulas")

    # -- missing args -------------------------------------------------------

    def test_missing_args_throw(self):
        with self.assertRaises(frappe.ValidationError):
            save_amount_formula(
                boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
                target_value_field=None, formula=_WILDCARD_QTY_X_RATE,  # no target_value_field
            )
        with self.assertRaises(frappe.ValidationError):
            get_sheet_amount_formulas(boq_name=self.boq, sheet_name=self.sheet)  # no version


class TestCellDismissal(FrappeTestCase):
    """Slice 4b-ACKNOWLEDGE: the per-(row, flag_kind) DISMISSAL layer (BoQ Cell Dismissal) --
    save_cell_dismissal + get_sheet_dismissals + the get_priced_rows merge + the save_cell_price
    RE-ARM. A dismissal HIDES a review-strip entry (a computed flag or a remark) from the active
    view WITHOUT changing the underlying condition. Reuses the shared committed fixture (line
    items 34/35, preamble row 6) and ADDS one 'Other' (non-priceable) node at source_row 50 to
    drive the success-only re-arm proof via a priceability reject. sheet_name carries a trailing
    space (#152). Each test starts with a clean dismissal + pricing + lock layer."""

    OTHER_ROW = 50

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.test_project = _make_project()
        boq = frappe.new_doc("BOQs")
        boq.project = cls.test_project.name
        boq.boq_name = "Cell-Dismissal Test BoQ"
        boq.tax_treatment = "Pre-tax"
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        cls.boq = boq.name
        cls.sheet = "Dismiss Fix "  # VERBATIM trailing space (#152)
        cls.cv = 1
        cls.fixture = build_committed_sheet_fixture(cls.boq, cls.sheet, commit_version=cls.cv)

        # An 'Other' (non-priceable) node -- a note row at source_row 50. A rate save on it is
        # rejected by the priceability guard (no override), which is the T7 success-only proof.
        other = frappe.new_doc("BOQ Nodes")
        other.sheet = cls.fixture["bqsh"]
        other.node_type = "Other"
        other.row_class = "note"
        other.description = "NOTE: rates exclusive of taxes"
        other.sort_order = 3
        other.source_row_number = cls.OTHER_ROW
        other.commit_version = cls.cv
        other.is_current = 1
        other.committed_at = frappe.utils.now()
        other.insert(ignore_permissions=True)
        cls.other_node = other.name
        frappe.db.commit()

        cls.me = frappe.session.user
        cls.other = frappe.db.get_value(
            "User", {"name": ["not in", [cls.me, "Guest"]], "enabled": 1}, "name"
        )
        assert cls.other, "need a second real User to play the competing lock holder"
        # MANDATORY amount-formula gate: declare covering formulas so the SUCCESS-path rate save
        # (the re-arm proof on a Line Item) is not rejected by the gate. setUp clears the lock
        # this leaves; the formula records persist (cleared in tearDownClass).
        _declare_fixture_amount_formulas(cls.boq, cls.sheet, cls.cv)

    @classmethod
    def tearDownClass(cls):
        frappe.db.delete(_FORMULA_DT, {"boq": cls.boq})
        frappe.db.delete(_DISMISSAL_DT, {"boq": cls.boq})
        frappe.db.delete(_PRICING, {"boq": cls.boq})
        frappe.db.delete(_LOCK_DT, {"boq": cls.boq})
        cleanup_committed_fixture(cls.boq)
        _cleanup_project(cls.test_project.name)
        super().tearDownClass()

    def setUp(self):
        frappe.db.delete(_DISMISSAL_DT, {"boq": self.boq})
        frappe.db.delete(_PRICING, {"boq": self.boq})
        frappe.db.delete(_LOCK_DT, {"boq": self.boq})
        frappe.db.commit()

    # -- helpers ------------------------------------------------------------

    def _current(self, excel_row, flag_kind):
        return frappe.get_all(
            _DISMISSAL_DT,
            filters={"boq": self.boq, "sheet_name": self.sheet, "excel_row": excel_row,
                     "flag_kind": flag_kind, "committed_version": self.cv, "is_current": 1},
            fields=["name", "flag_kind", "dismissal_version", "description", "dismissed_by"],
        )

    def _all_versions(self, excel_row, flag_kind):
        return frappe.get_all(
            _DISMISSAL_DT,
            filters={"boq": self.boq, "sheet_name": self.sheet, "excel_row": excel_row,
                     "flag_kind": flag_kind, "committed_version": self.cv},
            fields=["dismissal_version", "is_current"],
            order_by="dismissal_version asc",
        )

    def _seed_lock(self, user):
        acquire_or_refresh(self.boq, self.sheet, self.cv, user, frappe.utils.now_datetime())
        frappe.db.commit()

    # -- T1: dismiss -> one current at (row,kind); the read surfaces it ------

    def test_dismiss_creates_current_v1(self):
        res = save_cell_dismissal(
            boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
            committed_version=self.cv, flag_kind="needs_rate", dismissed=True,
            description="cable 1.1.2",
        )
        self.assertTrue(res["ok"])
        self.assertEqual(res["dismissal_version"], 1)
        self.assertEqual(res["froze_prior"], 0, "first dismiss freezes nothing")
        self.assertTrue(res["dismissed"])
        cur = self._current(34, "needs_rate")
        self.assertEqual(len(cur), 1, "exactly one current dismissal record")
        self.assertEqual(cur[0]["description"], "cable 1.1.2", "guard carried, not branched on")
        self.assertEqual(cur[0]["dismissed_by"], self.me, "stamps the dismissing user")
        got = get_sheet_dismissals(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv
        )["dismissals"]
        self.assertEqual(len(got), 1)
        self.assertEqual(got[0]["excel_row"], 34)
        self.assertEqual(got[0]["flag_kind"], "needs_rate")

    # -- T2: re-dismiss same (row,kind) -> freeze v1, insert v2, one current --

    def test_redismiss_freezes_and_supersedes(self):
        save_cell_dismissal(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                            committed_version=self.cv, flag_kind="needs_rate", dismissed=True)
        res2 = save_cell_dismissal(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                                   committed_version=self.cv, flag_kind="needs_rate", dismissed=True)
        self.assertEqual(res2["dismissal_version"], 2)
        self.assertEqual(res2["froze_prior"], 1, "the re-dismiss froze the prior current")
        cur = self._current(34, "needs_rate")
        self.assertEqual(len(cur), 1, "exactly ONE current after re-dismiss (freeze-and-supersede)")
        self.assertEqual(cur[0]["dismissal_version"], 2)

    # -- T3: two kinds, same row -> two independent currents (the discriminator) --

    def test_two_kinds_same_row_independent(self):
        save_cell_dismissal(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                            committed_version=self.cv, flag_kind="needs_rate", dismissed=True)
        save_cell_dismissal(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                            committed_version=self.cv, flag_kind="qty_anomaly", dismissed=True)
        self.assertEqual(len(self._current(34, "needs_rate")), 1)
        self.assertEqual(len(self._current(34, "qty_anomaly")), 1)
        got = get_sheet_dismissals(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv
        )["dismissals"]
        self.assertEqual(len(got), 2, "two independent dismissals on the same row")

    # -- T4: un-dismiss (falsy) -> current frozen, none current, nothing inserted --

    def test_undismiss_freezes_inserts_nothing(self):
        save_cell_dismissal(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                            committed_version=self.cv, flag_kind="needs_rate", dismissed=True)
        res = save_cell_dismissal(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                                  committed_version=self.cv, flag_kind="needs_rate", dismissed=False)
        self.assertFalse(res["dismissed"])
        self.assertEqual(res["froze_prior"], 1)
        self.assertIsNone(res["name"], "un-dismiss inserts no new current")
        self.assertIsNone(res["dismissal_version"])
        self.assertEqual(self._current(34, "needs_rate"), [], "no current dismissal after un-dismiss")
        allv = self._all_versions(34, "needs_rate")
        self.assertEqual(len(allv), 1, "still only v1 -- the un-dismiss inserted nothing")
        self.assertEqual(allv[0]["is_current"], 0, "v1 frozen")
        got = get_sheet_dismissals(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv
        )["dismissals"]
        self.assertEqual(got, [], "an un-dismissed entry does not surface")

    # -- T5: RE-ARM -- a rate write on the row freezes its computed dismissal --

    def test_rate_edit_rearms_computed_dismissal(self):
        save_cell_dismissal(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                            committed_version=self.cv, flag_kind="needs_rate", dismissed=True)
        res = save_cell_price(
            boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
            committed_version=self.cv, rate=150.0, area="Phase 1",
        )
        self.assertEqual(res["rearmed_dismissals"], 1, "the rate write re-armed the needs_rate dismissal")
        self.assertEqual(self._current(34, "needs_rate"), [], "the dismissal is no longer current")

    # -- T6: RE-ARM EXCLUDES REMARK (the load-bearing exclusion) -------------

    def test_rate_edit_excludes_remark_dismissal(self):
        save_cell_dismissal(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                            committed_version=self.cv, flag_kind="remark", dismissed=True)
        save_cell_dismissal(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                            committed_version=self.cv, flag_kind="needs_rate", dismissed=True)
        res = save_cell_price(
            boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
            committed_version=self.cv, rate=150.0, area="Phase 1",
        )
        self.assertEqual(res["rearmed_dismissals"], 1, "ONLY the computed kind re-armed")
        self.assertEqual(self._current(34, "needs_rate"), [], "needs_rate frozen by the rate edit")
        self.assertEqual(len(self._current(34, "remark")), 1, "the remark dismissal SURVIVES a rate edit")

    # -- T7: RE-ARM only on SUCCESS -- a rejected rate write mutates nothing --

    def test_rejected_rate_save_leaves_dismissals(self):
        # Dismiss qty_anomaly on the non-priceable Other row, then attempt a rate save there
        # WITHOUT the override -> the priceability guard throws BEFORE the re-arm runs.
        save_cell_dismissal(boq_name=self.boq, sheet_name=self.sheet, excel_row=self.OTHER_ROW,
                            committed_version=self.cv, flag_kind="qty_anomaly", dismissed=True)
        with self.assertRaises(frappe.ValidationError) as ctx:
            save_cell_price(
                boq_name=self.boq, sheet_name=self.sheet, excel_row=self.OTHER_ROW,
                col_letter="E", committed_version=self.cv, rate=99.0,
            )
        self.assertIn("not priceable", str(ctx.exception).lower())
        self.assertEqual(len(self._current(self.OTHER_ROW, "qty_anomaly")), 1,
                         "a rejected rate save left the dismissal untouched (re-arm never ran)")

    # -- T8: NEG -- dismissal for a non-existent row throws, mutates nothing --

    def test_nonexistent_row_throws(self):
        with self.assertRaises(frappe.ValidationError):
            save_cell_dismissal(boq_name=self.boq, sheet_name=self.sheet, excel_row=9999,
                                committed_version=self.cv, flag_kind="needs_rate", dismissed=True)
        self.assertEqual(self._current(9999, "needs_rate"), [], "no dismissal for a non-existent row")

    # -- T9: NEG -- bad / missing flag_kind throws --------------------------

    def test_bad_flag_kind_throws(self):
        with self.assertRaises(frappe.ValidationError) as ctx:
            save_cell_dismissal(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                                committed_version=self.cv, flag_kind="bogus", dismissed=True)
        self.assertIn("flag_kind", str(ctx.exception).lower())
        self.assertEqual(self._current(34, "bogus"), [], "a bad-kind dismissal wrote nothing")

    def test_missing_flag_kind_throws(self):
        with self.assertRaises(frappe.ValidationError):
            save_cell_dismissal(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                                committed_version=self.cv, flag_kind=None, dismissed=True)

    # -- the lock guard: a fresh lock by another user rejects + mutates nothing --

    def test_lock_held_by_other_rejects_and_mutates_nothing(self):
        self._seed_lock(self.other)  # other holds a FRESH lock
        with self.assertRaises(frappe.ValidationError) as ctx:
            save_cell_dismissal(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                                committed_version=self.cv, flag_kind="needs_rate", dismissed=True)
        self.assertIn(_LOCK_HELD_MARKER, str(ctx.exception), "reject carries the lock marker")
        self.assertEqual(self._current(34, "needs_rate"), [], "a lock-rejected dismissal wrote nothing")

    # -- HTTP coercion: the wire sends `dismissed` as a string --------------

    def test_dismissed_http_string_coercion(self):
        # "true" reads truthy -> dismiss.
        save_cell_dismissal(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                            committed_version=self.cv, flag_kind="needs_rate", dismissed="true")
        self.assertEqual(len(self._current(34, "needs_rate")), 1, '"true" dismisses')
        # "false" reads falsy -> un-dismiss (freeze, insert nothing).
        save_cell_dismissal(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                            committed_version=self.cv, flag_kind="needs_rate", dismissed="false")
        self.assertEqual(self._current(34, "needs_rate"), [], '"false" un-dismisses')

    # -- T10: get_priced_rows surfaces the additive dismissals key ----------

    def test_get_priced_rows_surfaces_dismissals(self):
        save_cell_dismissal(boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                            committed_version=self.cv, flag_kind="needs_rate", dismissed=True)
        res = get_priced_rows(boq_name=self.boq, sheet_name=self.sheet)
        self.assertIn("dismissals", res, "the additive sheet-level key is present")
        self.assertEqual(
            res["dismissals"], [{"excel_row": 34, "flag_kind": "needs_rate"}],
            "the current dismissal is delivered as a flat sheet-level list",
        )

    def test_get_priced_rows_dismissals_empty_when_none(self):
        res = get_priced_rows(boq_name=self.boq, sheet_name=self.sheet)
        self.assertIn("dismissals", res, "the key is always present (backwards-compat shape)")
        self.assertEqual(res["dismissals"], [], "a no-dismissal sheet returns it empty")


class TestMandatoryFormulaGate(FrappeTestCase):
    """The MANDATORY amount-formula gate (Phase 5): save_cell_price rejects ANY rate write until
    every amount column on the sheet has a declared formula. The gate is ABSOLUTE -- the
    allow_non_priceable override does NOT bypass it (it loosens ONLY the type/qty axis). The
    predicate is _sheet_formulas_complete (per-COVERAGE: an area-wildcard default covers all its
    per-area columns; zero amount columns -> trivially complete). Declaration (save_amount_formula)
    stays usable while rates are locked.

    Uses the SHARED per-area committed fixture (amount cols F [Phase 1, total] + I [Phase 2,
    install]) -- INCOMPLETE by default (no formulas) -- plus a SCALAR-RATE committed sheet (zero
    amount columns -> trivially complete). sheet_name carries a trailing space (#152)."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.test_project = _make_project()
        boq = frappe.new_doc("BOQs")
        boq.project = cls.test_project.name
        boq.boq_name = "Formula-Gate Test BoQ"
        boq.tax_treatment = "Pre-tax"
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        cls.boq = boq.name
        cls.cv = 1
        cls.sheet = "Gate Fix "  # amount-bearing per-area fixture -> INCOMPLETE by default (#152)
        cls.fixture = build_committed_sheet_fixture(cls.boq, cls.sheet, commit_version=cls.cv)
        cls.noamt_sheet = "Gate NoAmt "  # scalar rate, ZERO amount cols -> trivially complete
        cls.noamt_node = _build_scalar_rate_committed_sheet(cls.boq, cls.noamt_sheet, cls.cv)

    @classmethod
    def tearDownClass(cls):
        frappe.db.delete(_FORMULA_DT, {"boq": cls.boq})
        frappe.db.delete(_PRICING, {"boq": cls.boq})
        frappe.db.delete(_LOCK_DT, {"boq": cls.boq})
        cleanup_committed_fixture(cls.boq)
        _cleanup_project(cls.test_project.name)
        super().tearDownClass()

    def setUp(self):
        # Each test starts INCOMPLETE: clean formula layer + pricing + lock (fixture persists).
        frappe.db.delete(_FORMULA_DT, {"boq": self.boq})
        frappe.db.delete(_PRICING, {"boq": self.boq})
        frappe.db.delete(_LOCK_DT, {"boq": self.boq})
        frappe.db.commit()

    # -- helpers ------------------------------------------------------------

    def _declare(self, rate_subkey):
        """Declare ONE wildcard-default amount formula for the per-area fixture sheet."""
        save_amount_formula(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
            target_value_field="amount_by_area", target_value_key=None,
            target_rate_subkey=rate_subkey, formula=_FIXTURE_FORMULA_LEAF,
        )

    def _price(self, **kw):
        return save_cell_price(
            boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
            committed_version=self.cv, rate=100.0, area="Phase 1", rate_kind="combined_rate",
            description="cable", **kw,
        )

    # -- _sheet_formulas_complete (the predicate) ---------------------------

    def test_predicate_zero_amount_cols_is_trivially_complete(self):
        # The scalar-rate sheet has NO amount columns -> trivially complete (nothing to declare).
        self.assertTrue(_sheet_formulas_complete(self.boq, self.noamt_sheet, self.cv))

    def test_predicate_false_when_uncovered_or_partial(self):
        # No formulas -> INCOMPLETE (two amount cols F[total] + I[install] uncovered).
        self.assertFalse(_sheet_formulas_complete(self.boq, self.sheet, self.cv))
        # Only ONE of the two distinct rate_subkeys covered -> still INCOMPLETE.
        self._declare("total")
        self.assertFalse(_sheet_formulas_complete(self.boq, self.sheet, self.cv))

    def test_predicate_true_when_all_covered(self):
        self._declare("total")
        self._declare("install")
        self.assertTrue(_sheet_formulas_complete(self.boq, self.sheet, self.cv))

    # -- save_cell_price gate (override can NOT bypass) ----------------------

    def test_save_cell_price_rejected_when_incomplete_even_with_override(self):
        # Row 34/col E is a qty-bearing Line Item (priceable) -> the ONLY possible reject here is
        # the formula gate. allow_non_priceable=True must NOT bypass it.
        with self.assertRaises(frappe.ValidationError) as ctx:
            self._price(allow_non_priceable=True)
        self.assertIn("declared formula", str(ctx.exception),
                      "the FORMULA gate rejected it (not priceability) -- override can't bypass")
        # Reject mutated NOTHING: no pricing row, no lock acquired (the gate is before both).
        self.assertEqual(frappe.db.count(_PRICING, {"boq": self.boq}), 0, "no price row written")
        self.assertEqual(frappe.db.count(_LOCK_DT, {"boq": self.boq}), 0, "no lock acquired")

    def test_save_cell_price_succeeds_once_complete(self):
        self._declare("total")
        self._declare("install")
        frappe.db.delete(_LOCK_DT, {"boq": self.boq})  # clear the lock declaration left
        frappe.db.commit()
        res = self._price()
        self.assertTrue(res["ok"])
        self.assertEqual(res["is_filled"], 1)
        self.assertEqual(frappe.db.count(_PRICING, {"boq": self.boq, "is_current": 1}), 1)

    # -- declaration under the gate (the usability seam) --------------------

    def test_declaration_works_while_sheet_is_rate_locked(self):
        # The sheet is INCOMPLETE (rates locked). Declaring a formula MUST still succeed --
        # save_amount_formula carries no rate-editability precondition.
        self.assertFalse(_sheet_formulas_complete(self.boq, self.sheet, self.cv))
        self._declare("total")  # must not raise
        recs = get_sheet_amount_formulas(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv
        )["formulas"]
        self.assertEqual(len(recs), 1, "the formula was declared despite the sheet being locked")


# A qty x rate(combined) amount formula -- references a RATE operand (rate_by_area wildcard,
# combined_rate), so its amount column DEPENDS on the per-area combined rate. Bound per-area, a
# wildcard rate leaf resolves to that area's combined rate (col E for Phase 1, col H for Phase 2).
_QTY_TIMES_RATE_FORMULA = json.dumps({
    "op": "*",
    "operands": [
        {"ref": {"value_field": "qty_by_area", "value_key": None, "rate_subkey": None}},
        {"ref": {"value_field": "rate_by_area", "value_key": None, "rate_subkey": "combined_rate"}},
    ],
})


class TestReconciliationChoice(FrappeTestCase):
    """Cluster B: the per-CELL formula-vs-document reconciliation choice (BoQ Cell Reconciliation
    Choice + save_cell_reconciliation_choice / get_sheet_reconciliation_choices + the surgical,
    column-aware re-arm in save_cell_price / save_amount_formula).

    Fixture columns (COMMITTED_FIXTURE_ROLE_MAP): E = rate Phase 1 combined; F = amount Phase 1
    total; H = rate Phase 2 combined; I = amount Phase 2 install. Line items at rows 34 + 35.
    Per-test formulas: total (col F) = qty x rate(combined) [references the Phase-1 combined rate
    -> col E]; install (col I) = qty-only [references NO rate]. So a rate save on E re-arms a
    choice on F at the SAME row only -- never col I (qty-only) nor F at a DIFFERENT row.
    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.test_project = _make_project()
        boq = frappe.new_doc("BOQs")
        boq.project = cls.test_project.name
        boq.boq_name = "Recon Choice Test BoQ"
        boq.tax_treatment = "Pre-tax"
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        cls.boq = boq.name
        cls.sheet = "Recon Fix "  # VERBATIM trailing space (#152)
        cls.cv = 1
        cls.fixture = build_committed_sheet_fixture(cls.boq, cls.sheet, commit_version=cls.cv)

    @classmethod
    def tearDownClass(cls):
        cleanup_committed_fixture(cls.boq)
        _cleanup_project(cls.test_project.name)
        super().tearDownClass()

    def setUp(self):
        # Clean every overlay each test, then re-declare BOTH covering formulas (so the mandatory
        # gate passes for the rate-save re-arm test) -- total references a rate, install does not.
        for dt in (_CHOICE_DT, _PRICING, _FORMULA_DT, _LOCK_DT):
            frappe.db.delete(dt, {"boq": self.boq})
        frappe.db.commit()
        save_amount_formula(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
            target_value_field="amount_by_area", target_value_key=None,
            target_rate_subkey="total", formula=_QTY_TIMES_RATE_FORMULA,
        )
        save_amount_formula(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
            target_value_field="amount_by_area", target_value_key=None,
            target_rate_subkey="install", formula=_FIXTURE_FORMULA_LEAF,
        )
        frappe.db.delete(_LOCK_DT, {"boq": self.boq})
        frappe.db.commit()

    # -- helpers ------------------------------------------------------------

    def _current_choice(self, excel_row, col_letter):
        return frappe.get_all(
            _CHOICE_DT,
            filters={"boq": self.boq, "sheet_name": self.sheet, "excel_row": excel_row,
                     "col_letter": col_letter, "committed_version": self.cv, "is_current": 1},
            fields=["name", "choice", "choice_version"],
        )

    def _all_choice_versions(self, excel_row, col_letter):
        return frappe.get_all(
            _CHOICE_DT,
            filters={"boq": self.boq, "sheet_name": self.sheet, "excel_row": excel_row,
                     "col_letter": col_letter, "committed_version": self.cv},
            fields=["choice_version", "is_current", "choice"],
            order_by="choice_version asc",
        )

    def _save_choice(self, excel_row, col_letter, choice):
        return save_cell_reconciliation_choice(
            boq_name=self.boq, sheet_name=self.sheet, excel_row=excel_row,
            col_letter=col_letter, committed_version=self.cv, choice=choice,
        )

    # -- POSITIVE: CRUD + freeze-and-supersede ------------------------------

    def test_save_creates_current_choice_v1(self):
        res = self._save_choice(34, "F", "keep_document")
        self.assertTrue(res["ok"])
        self.assertEqual(res["choice_version"], 1)
        self.assertEqual(res["froze_prior"], 0, "first choice freezes nothing")
        self.assertEqual(res["choice"], "keep_document")
        cur = self._current_choice(34, "F")
        self.assertEqual(len(cur), 1, "exactly one current choice")
        self.assertEqual(cur[0]["choice"], "keep_document")

    def test_resave_freezes_prior_and_supersedes(self):
        self._save_choice(34, "F", "keep_document")
        res2 = self._save_choice(34, "F", "take_formula")
        self.assertEqual(res2["choice_version"], 2)
        self.assertEqual(res2["froze_prior"], 1, "the prior current was frozen")
        cur = self._current_choice(34, "F")
        self.assertEqual(len(cur), 1, "still exactly ONE current after re-save (the invariant)")
        self.assertEqual(cur[0]["choice"], "take_formula")
        allv = self._all_choice_versions(34, "F")
        self.assertEqual(len(allv), 2, "both versions retained (frozen, never deleted)")
        self.assertEqual([v["is_current"] for v in allv], [0, 1])

    def test_keep_vs_take_are_distinct_per_cell(self):
        self._save_choice(34, "F", "keep_document")
        self._save_choice(34, "I", "take_formula")
        f = self._current_choice(34, "F")
        i = self._current_choice(34, "I")
        self.assertEqual(f[0]["choice"], "keep_document")
        self.assertEqual(i[0]["choice"], "take_formula")

    def test_clear_freezes_only_no_current(self):
        self._save_choice(34, "F", "keep_document")
        res = save_cell_reconciliation_choice(
            boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="F",
            committed_version=self.cv, choice="",  # blank -> CLEAR
        )
        self.assertEqual(res["froze_prior"], 1)
        self.assertEqual(res["is_current"], 0)
        self.assertIsNone(res["choice"])
        self.assertEqual(len(self._current_choice(34, "F")), 0, "cleared -> no current (unset)")
        # but the frozen historical record is retained (never deleted)
        self.assertEqual(len(self._all_choice_versions(34, "F")), 1)

    # -- READ ---------------------------------------------------------------

    def test_get_sheet_reconciliation_choices(self):
        self._save_choice(34, "F", "keep_document")
        self._save_choice(34, "I", "take_formula")
        out = get_sheet_reconciliation_choices(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
        )["choices"]
        by_col = {c["col_letter"]: c["choice"] for c in out}
        self.assertEqual(by_col, {"F": "keep_document", "I": "take_formula"})

    def test_get_priced_rows_includes_reconciliation_choices(self):
        self._save_choice(34, "F", "take_formula")
        res = get_priced_rows(boq_name=self.boq, sheet_name=self.sheet)
        self.assertIn("reconciliation_choices", res)
        self.assertEqual(
            res["reconciliation_choices"],
            [{"excel_row": 34, "col_letter": "F", "choice": "take_formula"}],
        )

    # -- NEGATIVE -----------------------------------------------------------

    def test_save_to_nonexistent_cell_throws(self):
        with self.assertRaises(frappe.ValidationError):
            self._save_choice(9999, "F", "keep_document")

    def test_save_to_nonexistent_sheet_throws(self):
        with self.assertRaises(frappe.ValidationError):
            save_cell_reconciliation_choice(
                boq_name=self.boq, sheet_name="No Such Sheet ", excel_row=34, col_letter="F",
                committed_version=self.cv, choice="keep_document",
            )

    def test_save_to_nonexistent_boq_throws(self):
        with self.assertRaises(frappe.ValidationError):
            save_cell_reconciliation_choice(
                boq_name="NOPE-DOES-NOT-EXIST", sheet_name=self.sheet, excel_row=34,
                col_letter="F", committed_version=self.cv, choice="keep_document",
            )

    def test_missing_col_letter_throws(self):
        with self.assertRaises(frappe.ValidationError):
            save_cell_reconciliation_choice(
                boq_name=self.boq, sheet_name=self.sheet, excel_row=34,
                committed_version=self.cv, choice="keep_document",
            )

    def test_missing_committed_version_throws(self):
        with self.assertRaises(frappe.ValidationError):
            save_cell_reconciliation_choice(
                boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="F",
                choice="keep_document",
            )

    def test_invalid_choice_token_throws(self):
        with self.assertRaises(frappe.ValidationError):
            self._save_choice(34, "F", "bogus_choice")

    # -- INVALIDATION (D3): surgical, per-cell, column-aware re-arm ----------

    def test_rate_save_rearms_only_the_referencing_column_choice(self):
        # Choices on F@34 (references E via qty x rate), I@34 (qty-only -> no rate), F@35 (refs E
        # but a DIFFERENT row).
        self._save_choice(34, "F", "keep_document")
        self._save_choice(34, "I", "keep_document")
        self._save_choice(35, "F", "keep_document")
        # Save a rate on E (rate Phase 1 combined) at row 34.
        res = save_cell_price(
            boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="E",
            committed_version=self.cv, rate=125.0, area="Phase 1", rate_kind="combined_rate",
        )
        self.assertEqual(res["rearmed_choices"], 1, "exactly the one referencing cell re-armed")
        self.assertEqual(len(self._current_choice(34, "F")), 0, "F@34 references E -> re-armed")
        self.assertEqual(len(self._current_choice(34, "I")), 1, "I@34 is qty-only -> NOT re-armed")
        self.assertEqual(len(self._current_choice(35, "F")), 1, "F@35 is a different row -> NOT re-armed")

    def test_rate_save_on_unreferenced_rate_rearms_nothing(self):
        # H = rate Phase 2 combined. F's formula binds to Phase 1, so it references E (Phase 1),
        # NOT H. A rate save on H@34 must re-arm NOTHING (F@34's choice survives).
        self._save_choice(34, "F", "keep_document")
        res = save_cell_price(
            boq_name=self.boq, sheet_name=self.sheet, excel_row=34, col_letter="H",
            committed_version=self.cv, rate=80.0, area="Phase 2", rate_kind="combined_rate",
        )
        self.assertEqual(res["rearmed_choices"], 0, "F binds Phase 1 -> H (Phase 2) feeds it nothing")
        self.assertEqual(len(self._current_choice(34, "F")), 1, "F@34 choice survives")

    def test_formula_remove_clears_column_choices_silently(self):
        self._save_choice(34, "F", "keep_document")
        self._save_choice(35, "F", "take_formula")
        res = save_amount_formula(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
            target_value_field="amount_by_area", target_value_key=None,
            target_rate_subkey="total", formula="",  # CLEAR
        )
        self.assertTrue(res["cleared"])
        self.assertEqual(res["rearmed_choices"], 2, "both F rows cleared on formula remove")
        self.assertEqual(len(self._current_choice(34, "F")), 0)
        self.assertEqual(len(self._current_choice(35, "F")), 0)

    def test_formula_save_rearms_column_choices(self):
        self._save_choice(34, "F", "keep_document")
        res = save_amount_formula(
            boq_name=self.boq, sheet_name=self.sheet, committed_version=self.cv,
            target_value_field="amount_by_area", target_value_key=None,
            target_rate_subkey="total", formula=_FIXTURE_FORMULA_LEAF,  # a NEW formula
        )
        self.assertFalse(res["cleared"])
        self.assertGreaterEqual(res["rearmed_choices"], 1)
        self.assertEqual(len(self._current_choice(34, "F")), 0, "the column's choice re-armed on save")


# ====================================================================================
# Slice 5a -- Excel write-back backend (export_writeback)
# ====================================================================================
import base64 as _b64
import os as _os
import tempfile as _tempfile

import openpyxl as _openpyxl

from nirmaan_stack.api.boq.wizard.export_writeback import (
    _COLOR_HEX,
    _apply_colors,
    _assert_fidelity,
    _col_is_empty,
    _fidelity_snapshot,
    _generate_priced_workbook,
    _resolve_sheet_plan,
    _rightmost_mapped_col_index,
    _stamp_rates,
    _write_remark_column,
    export_priced_workbook,
)

# The committed fixture's role map (mirrors test_review_screen.COMMITTED_FIXTURE_ROLE_MAP):
# A-I mapped, so the TRUE data edge is I (9) and the remark column is J (10).
_FIX_ROLE_MAP = {
    "A": {"role": "sl_no", "area": None},
    "B": {"role": "description", "area": None},
    "C": {"role": "unit", "area": None},
    "D": {"role": "qty", "area": "Phase 1"},
    "E": {"role": "rate_combined_by_area", "area": "Phase 1"},
    "F": {"role": "amount_total_by_area", "area": "Phase 1"},
    "G": {"role": "qty", "area": "Phase 2"},
    "H": {"role": "rate_combined_by_area", "area": "Phase 2"},
    "I": {"role": "amount_install_by_area", "area": "Phase 2"},
}


def _new_ws(title="S"):
    """A fresh single-sheet workbook for pure-helper tests."""
    wb = _openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    return wb, ws


def _save_tmp(wb):
    """Save a workbook to a throwaway temp path; caller unlinks."""
    fd, path = _tempfile.mkstemp(suffix=".xlsx")
    _os.close(fd)
    wb.save(path)
    return path


class TestExportWritebackPureHelpers(FrappeTestCase):
    """Pure worksheet/fidelity helpers -- hermetic, no DB, synthetic workbooks."""

    def test_stamp_rates_lands_value_preserves_paired_formula(self):
        wb, ws = _new_ws()
        ws["D36"] = 10
        ws["E36"] = None                  # a blank rate cell
        ws["F36"] = "=D36*E36"            # paired amount formula
        skipped = _stamp_rates(ws, [{"excel_row": 36, "col_letter": "E", "rate": 123.45}])
        self.assertEqual(ws["E36"].value, 123.45, "rate stamped into the blank cell")
        self.assertEqual(ws["F36"].value, "=D36*E36", "paired amount formula preserved")
        self.assertEqual(skipped, [], "a value cell is not skipped")

    def test_stamp_rates_skips_formula_cell(self):
        wb, ws = _new_ws()
        ws["E10"] = "=SUM(H10:I10)"       # the VRF combined-rate case (a formula rate cell)
        skipped = _stamp_rates(ws, [{"excel_row": 10, "col_letter": "E", "rate": 999.0}])
        self.assertEqual(ws["E10"].value, "=SUM(H10:I10)", "formula rate cell NOT overwritten")
        self.assertEqual(skipped, [{"excel_row": 10, "col_letter": "E"}], "formula cell reported skipped")

    def test_stamp_rates_distinguishes_formula_vs_value_vs_blank(self):
        wb, ws = _new_ws()
        ws["E1"] = "=A1+B1"   # formula -> skip
        ws["E2"] = 500        # static value -> overwrite
        ws["E3"] = None       # blank -> overwrite
        skipped = _stamp_rates(ws, [
            {"excel_row": 1, "col_letter": "E", "rate": 1.0},
            {"excel_row": 2, "col_letter": "E", "rate": 2.0},
            {"excel_row": 3, "col_letter": "E", "rate": 3.0},
        ])
        self.assertEqual(ws["E1"].value, "=A1+B1")
        self.assertEqual(ws["E2"].value, 2.0)
        self.assertEqual(ws["E3"].value, 3.0)
        self.assertEqual(skipped, [{"excel_row": 1, "col_letter": "E"}],
                         "only the formula cell is skipped")

    def test_apply_colors_fills_any_cell_incl_formula(self):
        wb, ws = _new_ws()
        ws["B5"] = "description text"       # a NON-rate cell
        ws["F36"] = "=D36*E36"             # a formula cell
        applied = _apply_colors(ws, [
            {"excel_row": 5, "col_letter": "B", "color": "red"},
            {"excel_row": 36, "col_letter": "F", "color": "green"},
        ])
        self.assertEqual(applied, 2)
        self.assertEqual(ws["B5"].fill.fill_type, "solid", "fill applied to non-rate cell")
        self.assertIn(_COLOR_HEX["red"], ws["B5"].fill.fgColor.rgb or "")
        self.assertEqual(ws["F36"].value, "=D36*E36", "formula intact under a color fill")
        self.assertIn(_COLOR_HEX["green"], ws["F36"].fill.fgColor.rgb or "")

    def test_apply_colors_skips_unknown_token(self):
        wb, ws = _new_ws()
        applied = _apply_colors(ws, [{"excel_row": 1, "col_letter": "A", "color": "chartreuse"}])
        self.assertEqual(applied, 0, "an unknown token is skipped, never invented")

    def test_rightmost_mapped_col_index(self):
        self.assertEqual(_rightmost_mapped_col_index(_FIX_ROLE_MAP), 9, "I is the edge")
        self.assertEqual(_rightmost_mapped_col_index({}), 0)
        self.assertEqual(_rightmost_mapped_col_index({"A": {}, "AB": {}}), 28, "AB beyond Z")

    def test_write_remark_column_at_true_edge(self):
        wb, ws = _new_ws()
        col = _write_remark_column(
            ws, [{"excel_row": 34, "remark": "check with Nitesh"}], _FIX_ROLE_MAP, 3
        )
        self.assertEqual(col, "J", "one past the true edge (I) -- NOT inflated max_column")
        self.assertEqual(ws["J3"].value, "Nirmaan Remarks", "header at the data header row")
        self.assertEqual(ws["J34"].value, "check with Nitesh")

    def test_write_remark_column_refuses_nonempty_target(self):
        wb, ws = _new_ws()
        ws["J10"] = "real data already here"   # J is the would-be remark column
        with self.assertRaises(frappe.exceptions.ValidationError):
            _write_remark_column(ws, [{"excel_row": 34, "remark": "x"}], _FIX_ROLE_MAP, 3)

    def test_col_is_empty(self):
        wb, ws = _new_ws()
        self.assertTrue(_col_is_empty(ws, 10))
        ws["J7"] = 0          # a real 0 is data, not empty
        self.assertFalse(_col_is_empty(ws, 10))

    def test_fidelity_snapshot_clean_round_trip_passes(self):
        wb, ws = _new_ws()
        ws["F1"] = "=A1+B1"
        ws["F2"] = "=A2+B2"
        ws.merge_cells("A1:B1")
        before = _fidelity_snapshot(wb)
        path = _save_tmp(wb)
        try:
            wb2 = _openpyxl.load_workbook(path, data_only=False)
            after = _fidelity_snapshot(wb2)
            wb2.close()
        finally:
            _os.unlink(path)
        self.assertEqual(before["formulas"], 2)
        self.assertEqual(before["merges"], 1)
        _assert_fidelity(before, after)  # must NOT raise on a clean round-trip

    def test_assert_fidelity_fails_on_divergence(self):
        base = {"formulas": 5, "merges": 2, "sheets": 3, "defined_names": 4585}
        # Each of the four invariants must be caught (non-vacuous).
        for k in ("formulas", "merges", "sheets", "defined_names"):
            bad = dict(base)
            bad[k] = base[k] - 1
            with self.assertRaises(frappe.exceptions.ValidationError, msg=f"{k} divergence must raise"):
                _assert_fidelity(base, bad)
        _assert_fidelity(base, dict(base))  # identical -> no raise


class TestExportWritebackEndToEnd(FrappeTestCase):
    """The worker against the committed fixture + a synthetic workbook injected (bypassing
    the S3 fetch, which is a one-line reused helper). Exercises version resolution, stamping,
    the fidelity guard, last_exported_at, and the skipped-formula report."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.test_project = _make_project()
        boq = frappe.new_doc("BOQs")
        boq.project = cls.test_project.name
        boq.boq_name = "Export WB Test BoQ"
        boq.tax_treatment = "Pre-tax"
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        cls.boq = boq.name
        cls.sheet = "Export Fix "          # VERBATIM trailing space (#152)
        cls.sheet_stripped = "Export Fix"   # openpyxl tab title (Excel trims trailing space)
        cls.cv = 1
        cls.fixture = build_committed_sheet_fixture(cls.boq, cls.sheet, commit_version=cls.cv)

    @classmethod
    def tearDownClass(cls):
        cleanup_committed_fixture(cls.boq)
        _cleanup_project(cls.test_project.name)
        super().tearDownClass()

    def setUp(self):
        frappe.db.delete(_PRICING, {"boq": self.boq})
        frappe.db.delete(_COLOR_DT, {"boq": self.boq})
        frappe.db.delete(_REMARK_DT, {"boq": self.boq})
        # reset last_exported_at on the committed sheet so each test starts unstamped
        frappe.db.set_value("BoQ Sheet", self.fixture["bqsh"], "last_exported_at", None,
                            update_modified=False)
        frappe.db.commit()

    # -- builders -----------------------------------------------------------
    def _add_price(self, excel_row, col_letter, rate):
        doc = frappe.new_doc(_PRICING)
        doc.boq = self.boq
        doc.sheet_name = self.sheet  # VERBATIM
        doc.excel_row = excel_row
        doc.col_letter = col_letter
        doc.committed_version = self.cv
        doc.rate = rate
        doc.is_filled = 1
        doc.pricing_version = 1
        doc.is_current = 1
        doc.priced_at = frappe.utils.now()
        doc.is_finalized = 0
        doc.insert(ignore_permissions=True)
        frappe.db.commit()

    def _add_color(self, excel_row, col_letter, color):
        doc = frappe.new_doc(_COLOR_DT)
        doc.boq = self.boq
        doc.sheet_name = self.sheet
        doc.excel_row = excel_row
        doc.col_letter = col_letter
        doc.committed_version = self.cv
        doc.color = color
        doc.color_version = 1
        doc.is_current = 1
        doc.colored_at = frappe.utils.now()
        doc.insert(ignore_permissions=True)
        frappe.db.commit()

    def _add_remark(self, excel_row, remark):
        doc = frappe.new_doc(_REMARK_DT)
        doc.boq = self.boq
        doc.sheet_name = self.sheet
        doc.excel_row = excel_row
        doc.committed_version = self.cv
        doc.remark = remark
        doc.remark_version = 1
        doc.is_current = 1
        doc.remarked_at = frappe.utils.now()
        doc.insert(ignore_permissions=True)
        frappe.db.commit()

    def _synthetic_path(self, formula_rate=False):
        """A workbook whose tab matches the fixture sheet (stripped title), with the role-map
        columns + a paired amount formula (fidelity anchor) and the priced rate cells."""
        wb = _openpyxl.Workbook()
        ws = wb.active
        ws.title = self.sheet_stripped
        for r in (34, 35):
            ws["D{}".format(r)] = 10
            ws["E{}".format(r)] = "=H{0}:I{0}".format(r) if formula_rate else None
            ws["F{}".format(r)] = "=D{0}*E{0}".format(r)   # amount formula -> the fidelity anchor
        return _save_tmp(wb)

    def _load_b64(self, content_base64):
        raw = _b64.b64decode(content_base64)
        fd, path = _tempfile.mkstemp(suffix=".xlsx")
        _os.write(fd, raw)
        _os.close(fd)
        wb = _openpyxl.load_workbook(path, data_only=False)
        _os.unlink(path)
        return wb

    # -- POSITIVE -----------------------------------------------------------
    def test_generate_stamps_rate_and_sets_last_exported_at(self):
        self._add_price(34, "E", 250.0)
        path = self._synthetic_path()
        try:
            res = _generate_priced_workbook(self.boq, [self.sheet], path)
        finally:
            _os.unlink(path)
        self.assertEqual(res["exported_sheets"], [self.sheet])
        self.assertEqual(res["skipped_formula_columns"], {})
        wb = self._load_b64(res["content_base64"])
        ws = wb[self.sheet_stripped]
        self.assertEqual(ws["E34"].value, 250.0, "rate stamped")
        self.assertEqual(ws["F34"].value, "=D34*E34", "paired amount formula preserved")
        wb.close()
        # last_exported_at stamped on the committed BoQ Sheet via set_value
        stamped = frappe.db.get_value("BoQ Sheet", self.fixture["bqsh"], "last_exported_at")
        self.assertIsNotNone(stamped, "last_exported_at stamped on the exported sheet")

    def test_generate_reports_skipped_formula_rate_column(self):
        self._add_price(34, "E", 999.0)
        path = self._synthetic_path(formula_rate=True)  # E34/E35 are formulas
        try:
            res = _generate_priced_workbook(self.boq, [self.sheet], path)
        finally:
            _os.unlink(path)
        self.assertEqual(res["skipped_formula_columns"], {self.sheet: ["E"]},
                         "the formula rate column is reported skipped")
        wb = self._load_b64(res["content_base64"])
        self.assertEqual(wb[self.sheet_stripped]["E34"].value, "=H34:I34",
                         "formula rate cell left untouched")
        wb.close()

    def test_generate_applies_remark_column(self):
        self._add_remark(34, "verify qty")
        path = self._synthetic_path()
        try:
            res = _generate_priced_workbook(self.boq, [self.sheet], path)
        finally:
            _os.unlink(path)
        self.assertEqual(res["remark_columns"], {self.sheet: "J"})
        wb = self._load_b64(res["content_base64"])
        ws = wb[self.sheet_stripped]
        self.assertEqual(ws["J3"].value, "Nirmaan Remarks")
        self.assertEqual(ws["J34"].value, "verify qty")
        wb.close()

    def test_generate_color_fill_present_in_output(self):
        self._add_color(34, "B", "blue")     # color on a non-rate (description) column
        path = self._synthetic_path()
        try:
            res = _generate_priced_workbook(self.boq, [self.sheet], path)
        finally:
            _os.unlink(path)
        wb = self._load_b64(res["content_base64"])
        self.assertEqual(wb[self.sheet_stripped]["B34"].fill.fill_type, "solid",
                         "color fill survived the save round-trip")
        wb.close()

    # -- NEGATIVE -----------------------------------------------------------
    def test_resolve_plan_throws_for_uncommitted_sheet(self):
        with self.assertRaises(frappe.exceptions.ValidationError):
            _resolve_sheet_plan(self.boq, "No Such Sheet")

    def test_endpoint_unknown_boq_throws(self):
        with self.assertRaises(frappe.exceptions.ValidationError):
            export_priced_workbook(boq_name="BOQ-NOPE-99999", sheet_names=[self.sheet])

    def test_endpoint_empty_sheet_names_throws(self):
        with self.assertRaises(frappe.exceptions.ValidationError):
            export_priced_workbook(boq_name=self.boq, sheet_names=[])

    def test_endpoint_missing_boq_throws(self):
        with self.assertRaises(frappe.exceptions.ValidationError):
            export_priced_workbook(boq_name=None, sheet_names=[self.sheet])
