# Copyright (c) 2026, Nirmaan (Stratos Infra Technologies Pvt. Ltd.) and contributors
# For license information, please see license.txt

"""Template-origin priced-workbook generator (ADR-0013 Amendment A2-D2 / rectification R4).

A template-cloned BoQ has NO source workbook (BOQs.source_file_url is None), so its priced
Excel cannot be produced by the copy-on-write stamp path in `export_writeback.py`. This module
BUILDS a fresh .xlsx FROM SCRATCH out of the committed tier:

  - the FAITHFUL GRID  (BoQ Committed Sheet Grid + rows -> cell {col_letter: value})
  - the RATES          (BoQ Cell Pricing, is_current + is_filled -> the rate cell)
  - the AMOUNTS         as LIVE Excel formulas translated from BoQ Cell Amount Formula
                        (the stored token-tree AST -> `=<qtyCell>*<rateCell>`)
  - a synthesized HEADER row (the master carries no header labels; roles -> labels)
  - a per-sheet GRAND-TOTAL row (=SUM per amount column)
  - a REGENERATED SUMMARY sheet (first) with LIVE cross-sheet formulas + a GST/GRAND-TOTAL
    block driven by BOQs.tax_treatment
  - a MAKE-LIST / general-specs sheet dumping preamble_text line-by-line

It NEVER touches S3, never reads node amounts (ZERO for template BoQs -- capture-only), and
fail-safes any amount cell whose formula operand cannot be resolved to a BLANK cell.

REJECT-MUTATES-NOTHING: the whole workbook is built + serialized to bytes BEFORE any
last_exported_at stamp; a raise anywhere stamps nothing and returns nothing.

Public API (called ONLY from export_writeback.export_priced_workbook's is_template branch):
  generate_template_priced_workbook(boq_name, sheet_names) -> dict
"""
from __future__ import annotations

import base64
import io
import json
from typing import Any

import frappe
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string

# Reuse the upload-path helpers + constants (they operate on any worksheet). Imported at
# module load: this module is imported LAZILY from export_writeback's endpoint, so
# export_writeback is fully loaded first -> no import cycle.
from nirmaan_stack.api.boq.wizard.export_writeback import (
    _XLSX_CONTENT_TYPE,
    _BOQ_SHEET,
    _apply_colors,
    _apply_priced_highlight,
    _safe_export_basename,
    _stamp_rates,
    _write_remark_column,
)

_GRID = "BoQ Committed Sheet Grid"
_GRID_ROW = "BoQ Committed Sheet Grid Row"
_PRICING = "BoQ Cell Pricing"
_COLOR = "BoQ Cell Color"
_REMARK = "BoQ Cell Remark"
_FORMULA = "BoQ Cell Amount Formula"

_SUMMARY_TITLE = "Summary"
_CURRENCY_FMT = "#,##0.00"

# The regenerated Summary drops the ORIGINAL 6th per-sqft column -> exactly 5 columns.
_SUMMARY_HEADERS = ("SL. No.", "Particulars", "Supply", "Install", "Total")

# ── role classification (mirrors review_screen / classifier -- kept LOCAL, not imported, so
#    this leaf module has no dependency back into the review layer) ────────────────────────
_ALL_AMOUNT_ROLES = frozenset({
    "amount_supply", "amount_install", "amount_total",
    "amount_supply_by_area", "amount_install_by_area", "amount_total_by_area",
})
_CURRENCY_ROLES = frozenset({
    "rate_supply", "rate_install", "rate_combined",
    "rate_supply_by_area", "rate_install_by_area", "rate_combined_by_area",
}) | _ALL_AMOUNT_ROLES

# Synthesized header labels (the master's column_headers are empty -> synthesize from role).
_HEADER_LABELS = {
    "sl_no": "Sl. No.",
    "description": "Description",
    "unit": "Unit",
    "make_model": "Make / Model",
    "qty_total": "Total Quantity",
    "rate_supply": "Rate (Supply)",
    "rate_install": "Rate (Install)",
    "rate_combined": "Rate",
    "amount_supply": "Amount (Supply)",
    "amount_install": "Amount (Install)",
    "amount_total": "Amount",
    "row_notes": "Notes",
}
# By-area rate/amount roles -> a base label the area name is appended to.
_BY_AREA_BASE_LABEL = {
    "rate_supply_by_area": "Rate (Supply)",
    "rate_install_by_area": "Rate (Install)",
    "rate_combined_by_area": "Rate",
    "amount_supply_by_area": "Amount (Supply)",
    "amount_install_by_area": "Amount (Install)",
    "amount_total_by_area": "Amount",
}

# Column widths by role (Excel character units).
_ROLE_WIDTH = {
    "sl_no": 8, "description": 48, "unit": 10, "make_model": 22,
    "qty_total": 14, "qty": 12,
    "rate_supply": 14, "rate_install": 14, "rate_combined": 14,
    "amount_supply": 16, "amount_install": 16, "amount_total": 16,
    "row_notes": 30,
}

_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="E8EDF3")
_BOLD = Font(bold=True)


# ── formula AST -> Excel translator (leaf ref -> column letter -> row-relative address) ────
# value_field -> the column_role_map role (scalar leaves). Inverse of review_screen's
# _SINGLETON_ROLE_TO_FIELD, kept LOCAL. Only value_fields a formula operand can carry matter
# (qty_total + the three scalar rates); the descriptor scalars are included for completeness.
_SCALAR_FIELD_TO_ROLE = {
    "qty_total": "qty_total",
    "rate_supply": "rate_supply",
    "rate_install": "rate_install",
    "rate_combined": "rate_combined",
    "amount_supply": "amount_supply",
    "amount_install": "amount_install",
    "amount_total": "amount_total",
    "sl_no_value": "sl_no",
    "description": "description",
    "unit": "unit",
    "make_model": "make_model",
    "row_notes": "row_notes",
}
# rate_by_area / amount_by_area rate_subkey -> role (inverse of classifier maps, kept LOCAL).
_RATE_SUBKEY_TO_ROLE = {
    "supply_rate": "rate_supply_by_area",
    "install_rate": "rate_install_by_area",
    "combined_rate": "rate_combined_by_area",
}
_AMOUNT_SUBKEY_TO_ROLE = {
    "supply": "amount_supply_by_area",
    "install": "amount_install_by_area",
    "total": "amount_total_by_area",
}
_OP_INFIX = {"*": "*", "+": "+"}


def _json(val: Any, default: Any) -> Any:
    """Coerce a JSON-column value (already-parsed object, raw JSON string, or None) to a
    Python object; fall back to `default` on None/blank/parse-failure."""
    if val is None or val == "":
        return default
    if isinstance(val, str):
        try:
            return json.loads(val)
        except (ValueError, TypeError):
            return default
    return val


def header_label_for(role: Any, area: Any) -> str:
    """Synthesize a display header label from a column's (role, area). A per-area qty column
    is labelled by its AREA NAME; a by-area rate/amount column appends the area to its base
    label; everything else uses the fixed role label (empty string for an unknown role)."""
    if not role:
        return ""
    if role == "qty":
        return str(area) if area else "Quantity"
    if role in _HEADER_LABELS:
        return _HEADER_LABELS[role]
    if role in _BY_AREA_BASE_LABEL:
        base = _BY_AREA_BASE_LABEL[role]
        return f"{base} - {area}" if area else base
    return ""


def _col_for_role_area(role_map: dict, role: str, area: Any) -> str | None:
    """The Excel column letter mapped to (role, area) in the column_role_map, or None. area
    is normalized (None/'' equal) so a scalar role matches an area-less column entry."""
    want_area = area or None
    for col, spec in (role_map or {}).items():
        if not isinstance(spec, dict):
            continue
        if spec.get("role") == role and (spec.get("area") or None) == want_area:
            return col
    return None


def resolve_ref_col(ref: dict, role_map: dict) -> str | None:
    """A formula leaf ref {value_field, value_key, rate_subkey} -> the Excel column letter it
    reads, resolved through column_role_map. Returns None when unresolvable (an area-bound
    wildcard with no value_key, or a value_field/subkey with no mapped column) -> the caller
    fail-safes the whole amount cell to BLANK."""
    if not isinstance(ref, dict):
        return None
    vf = ref.get("value_field")
    vk = ref.get("value_key")
    sk = ref.get("rate_subkey")
    if vf in _SCALAR_FIELD_TO_ROLE:
        return _col_for_role_area(role_map, _SCALAR_FIELD_TO_ROLE[vf], None)
    if vf == "qty_by_area":
        return _col_for_role_area(role_map, "qty", vk) if vk else None
    if vf == "rate_by_area":
        role = _RATE_SUBKEY_TO_ROLE.get(sk)
        return _col_for_role_area(role_map, role, vk) if (role and vk) else None
    if vf == "amount_by_area":
        role = _AMOUNT_SUBKEY_TO_ROLE.get(sk)
        return _col_for_role_area(role_map, role, vk) if (role and vk) else None
    return None


# Scalar amount target_value_field -> its column_role_map role (for target-column resolution).
_AMOUNT_FIELD_TO_ROLE = {
    "amount_total": "amount_total",
    "amount_supply": "amount_supply",
    "amount_install": "amount_install",
}


def resolve_target_col(formula_rec: dict, role_map: dict) -> str | None:
    """The Excel column an amount formula WRITES into. Prefers resolving the target through the
    authoritative column_role_map (target_value_field [+ target_value_key/target_rate_subkey for a
    by-area amount] -> role -> letter); falls back to the stored target_col GUARD (which the
    doctype notes can drift on re-commit). None only when neither resolves -> the formula is
    skipped (nothing to place)."""
    tv = formula_rec.get("target_value_field")
    role = _AMOUNT_FIELD_TO_ROLE.get(tv)
    if role:
        col = _col_for_role_area(role_map, role, None)
        if col:
            return col
    if tv == "amount_by_area":
        role = _AMOUNT_SUBKEY_TO_ROLE.get(formula_rec.get("target_rate_subkey"))
        vk = formula_rec.get("target_value_key")
        if role and vk:
            col = _col_for_role_area(role_map, role, vk)
            if col:
                return col
    return formula_rec.get("target_col") or None


def ast_to_excel(node: Any, excel_row: int, role_map: dict) -> str | None:
    """Translate a formula token-tree node to an Excel formula BODY (no leading '='),
    row-relative to excel_row. GENERAL: an operator node joins its operand bodies with the
    op's infix symbol (wrapped in parens); a leaf ref resolves to `<col><row>`. Returns None
    if ANY operand cannot be resolved OR the op is unsupported -> a BLANK amount cell."""
    if not isinstance(node, dict):
        return None
    if "ref" in node:
        col = resolve_ref_col(node.get("ref"), role_map)
        return f"{col}{excel_row}" if col else None
    if "op" in node:
        op = node.get("op")
        if op not in _OP_INFIX:
            return None
        operands = node.get("operands") or []
        if not operands:
            return None
        parts: list[str] = []
        for child in operands:
            body = ast_to_excel(child, excel_row, role_map)
            if body is None:
                return None
            parts.append(body)
        return "(" + _OP_INFIX[op].join(parts) + ")"
    return None


def _quote_sheet(title: str) -> str:
    """Excel single-quoted sheet reference token, internal single-quotes doubled. Always
    quoted (safe for names with spaces/special chars) so a cross-sheet ref never breaks."""
    return "'" + str(title).replace("'", "''") + "'"


# ── DB reads ───────────────────────────────────────────────────────────────────────
def _resolve_template_sheet(boq_name: str, sheet_name: str) -> Any:
    """The CURRENT committed BoQ Sheet for (boq, sheet_name) with the fields the from-scratch
    build needs. Throws if there is no current committed version (mirrors export_writeback's
    _resolve_sheet_plan contract). sheet_name matched VERBATIM (#152)."""
    row = frappe.db.get_value(
        _BOQ_SHEET,
        {"boq": boq_name, "sheet_name": sheet_name, "is_current": 1},
        ["name", "commit_version", "treat_as", "column_role_map", "column_headers",
         "header_row", "header_row_count", "sheet_order", "sheet_label", "area_dimensions"],
        as_dict=True,
    )
    if not row:
        frappe.throw(
            f"No current committed version for sheet '{sheet_name}' in BoQ '{boq_name}'. "
            "Commit the sheet before exporting.",
            title="No committed version",
        )
    row.column_role_map = _json(row.column_role_map, {}) or {}
    row.column_headers = _json(row.column_headers, {}) or {}
    row.area_dimensions = _json(row.area_dimensions, []) or []
    return row


def _read_grid_rows(boq_name: str, sheet_name: str) -> list[dict]:
    """The committed faithful grid rows [{row_number, cells}] in row_order. sheet_name
    matched VERBATIM (#152)."""
    grid_name = frappe.db.get_value(
        _GRID, {"boq": boq_name, "source_sheet_name": sheet_name, "is_current": 1}, "name"
    )
    if not grid_name:
        return []
    rows = frappe.db.get_all(
        _GRID_ROW,
        filters={"parent": grid_name, "parenttype": _GRID, "parentfield": "rows"},
        fields=["row_number", "row_order", "cells"],
        order_by="row_order asc",
    )
    return [{"row_number": r["row_number"], "cells": _json(r["cells"], {}) or {}} for r in rows]


def _preamble_text_for(boq_doc: Any, sheet_name: str) -> str:
    """The carried preamble_text for a committed general-specs sheet (BOQs.general_specs_sheets
    child, source_sheet_name VERBATIM #152), or ''."""
    for gs in (boq_doc.general_specs_sheets or []):
        if gs.source_sheet_name == sheet_name:
            return gs.preamble_text or ""
    return ""


# ── per-sheet builders ───────────────────────────────────────────────────────────────
def _build_preamble_sheet(wb: Workbook, sheet_name: str, boq_doc: Any) -> None:
    """A general-specs (Make List etc.) sheet: dump the carried preamble_text line-by-line
    into column A, verbatim. openpyxl assigns the actual title (Excel 31-char / illegal-char
    rules); the sheet is display-only (never referenced by the Summary)."""
    ws = wb.create_sheet(title=_excel_title(sheet_name))
    text = _preamble_text_for(boq_doc, sheet_name)
    for i, line in enumerate((text or "").splitlines(), start=1):
        ws.cell(row=i, column=1).value = line
    ws.column_dimensions["A"].width = 90


def _build_data_sheet(wb: Workbook, boq_name: str, sheet_name: str, plan: Any) -> dict:
    """Build ONE data worksheet from scratch and return its Summary descriptor
    {title, label, grand: {role: cell_addr}, remark_col}. Ordering follows the spec:
    grid -> header -> rates -> amount formulas -> total-qty (multi-area) -> grand-total ->
    styling -> colors -> priced highlight -> remark column."""
    ws = wb.create_sheet(title=_excel_title(sheet_name))
    title = ws.title  # the ACTUAL title openpyxl assigned (used by the Summary cross-ref)
    role_map = plan.column_role_map
    header_row = int(plan.header_row) if plan.header_row else 1
    cv = plan.commit_version

    # (a) faithful grid cells -- EXCEPT the rate/amount columns, which are owned by pricing
    #     (step c) and the amount formulas (step d). A template BoQ's committed grid carries
    #     placeholder 0s in those columns (_invert_rows_to_grid iterates the full role map), and
    #     writing them would print 0 / 0.00 on UNPRICED rows. Skip them here so unpriced rows stay
    #     blank; step (c) overlays real rates and step (d) writes amount formulas on priced rows.
    currency_cols = {c for c, s in role_map.items()
                     if isinstance(s, dict) and s.get("role") in _CURRENCY_ROLES}
    grid_rows = _read_grid_rows(boq_name, sheet_name)
    data_rows: list[int] = []
    for gr in grid_rows:
        rn = gr.get("row_number")
        if rn is None:
            continue
        r = int(rn)
        data_rows.append(r)
        for col, val in (gr.get("cells") or {}).items():
            if val is not None and col not in currency_cols:
                ws[f"{col}{r}"] = val

    # (b) synthesized header row (written AFTER the grid so it wins any collision).
    for col, spec in role_map.items():
        if not isinstance(spec, dict):
            continue
        label = header_label_for(spec.get("role"), spec.get("area"))
        if label:
            ws[f"{col}{header_row}"] = label

    # (c) rates overlay (BoQ Cell Pricing). Reuse _stamp_rates -> `written` drives the teal
    #     priced-cell highlight; no rate cell holds a formula so nothing is skipped.
    pricing = frappe.get_all(
        _PRICING,
        filters={"boq": boq_name, "sheet_name": sheet_name, "committed_version": cv,
                 "is_current": 1, "is_filled": 1},
        fields=["excel_row", "col_letter", "rate"],
        order_by="excel_row asc, col_letter asc",
    )
    _skipped, written = _stamp_rates(ws, pricing)

    # (d) AMOUNT cells as live Excel formulas (translate each current BoQ Cell Amount Formula).
    #     Written ONLY on PRICED rows (a stamped rate exists for that row). This MIRRORS the
    #     pricing editor's evaluateAmountCell -- an amount is BLANK unless the row is priced --
    #     and the upload path (which preserves blank amount cells on unpriced/heading/group-
    #     preamble rows). A pure group preamble is seeded with a qty but never priced, so gating
    #     on qty would still print a spurious =qtyCell*rateCell -> 0.00 on it; gating on the rate
    #     keeps those rows (and every unpriced row) blank, exactly like the on-screen grid.
    priced_rows = {int(p["excel_row"]) for p in pricing if p.get("excel_row") is not None}
    formulas = frappe.get_all(
        _FORMULA,
        filters={"boq": boq_name, "sheet_name": sheet_name, "committed_version": cv,
                 "is_current": 1},
        fields=["target_col", "target_value_field", "target_value_key", "target_rate_subkey",
                "formula"],
    )
    for f in formulas:
        target_col = resolve_target_col(f, role_map)
        ast = _json(f.get("formula"), None)
        if not target_col or ast is None:
            continue
        for r in data_rows:
            if r not in priced_rows:
                continue  # unpriced / heading / group-preamble row -> blank amount (no 0.00)
            body = ast_to_excel(ast, r, role_map)
            if body is not None:  # fail-safe: unresolved operand -> leave the cell BLANK
                ws[f"{target_col}{r}"] = "=" + body

    # (e) Total-Quantity: MULTI-area -> =SUM(firstArea:lastArea) per data row (area qty cols
    #     are contiguous, inserted immediately before Total); single-area keeps the committed
    #     static grid value (untouched).
    area_qty_cols = [
        c for c, s in role_map.items()
        if isinstance(s, dict) and s.get("role") == "qty" and s.get("area")
    ]
    total_col = _col_for_role_area(role_map, "qty_total", None)
    if area_qty_cols and total_col:
        ordered = sorted(area_qty_cols, key=column_index_from_string)
        first, last = ordered[0], ordered[-1]
        for r in data_rows:
            ws[f"{total_col}{r}"] = f"=SUM({first}{r}:{last}{r})"

    # (f) per-sheet GRAND-TOTAL row (=SUM per amount column). Record the amount grand-total
    #     addresses for the Summary cross-sheet formulas.
    amount_cols = {
        c: s for c, s in role_map.items()
        if isinstance(s, dict) and s.get("role") in _ALL_AMOUNT_ROLES
    }
    grand_row = (max(data_rows) if data_rows else header_row) + 1
    desc_col = _col_for_role_area(role_map, "description", None) or "A"
    ws[f"{desc_col}{grand_row}"] = "TOTAL"
    grand_addr: dict[str, str] = {}
    if data_rows:
        top, bot = min(data_rows), max(data_rows)
        for col, spec in amount_cols.items():
            ws[f"{col}{grand_row}"] = f"=SUM({col}{top}:{col}{bot})"
            grand_addr[spec.get("role")] = f"{col}{grand_row}"

    # (g) styling + annotation overlays.
    _style_data_sheet(ws, role_map, header_row, data_rows, grand_row, desc_col)
    colors = frappe.get_all(
        _COLOR,
        filters={"boq": boq_name, "sheet_name": sheet_name, "committed_version": cv,
                 "is_current": 1},
        fields=["excel_row", "col_letter", "color"],
    )
    _apply_colors(ws, colors)
    _apply_priced_highlight(ws, written)  # teal LAST -> wins over a user color on a rate cell

    remark_col = None
    remarks = frappe.get_all(
        _REMARK,
        filters={"boq": boq_name, "sheet_name": sheet_name, "committed_version": cv,
                 "is_current": 1},
        fields=["excel_row", "remark"],
        order_by="excel_row asc",
    )
    if remarks:
        remark_col = _write_remark_column(ws, remarks, role_map, header_row)

    return {
        "title": title,
        "label": (plan.sheet_label or "").strip() or (sheet_name or "").strip() or sheet_name,
        "grand": grand_addr,
        "remark_col": remark_col,
    }


def _style_data_sheet(ws, role_map: dict, header_row: int, data_rows: list[int],
                      grand_row: int, desc_col: str) -> None:
    """Light styling: bold+filled header, column widths, currency number_format on rate/amount
    cells (data rows + grand total), thin borders over the header+data+grand region. Sets only
    font/fill/width/number_format/border -- never a cell VALUE."""
    rows_to_border = [header_row, *data_rows, grand_row]
    for col, spec in role_map.items():
        if not isinstance(spec, dict):
            continue
        role = spec.get("role")
        idx = None
        try:
            idx = column_index_from_string(col)
        except (ValueError, TypeError):
            idx = None
        # header cell.
        h = ws[f"{col}{header_row}"]
        h.font = _BOLD
        h.fill = _HEADER_FILL
        h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # width.
        if idx is not None and role in _ROLE_WIDTH:
            ws.column_dimensions[col].width = _ROLE_WIDTH[role]
        # currency number_format on data + grand-total cells of currency columns.
        if role in _CURRENCY_ROLES:
            for r in [*data_rows, grand_row]:
                ws[f"{col}{r}"].number_format = _CURRENCY_FMT
        # thin borders over the region.
        for r in rows_to_border:
            ws[f"{col}{r}"].border = _BORDER
    # bold the grand-total label.
    ws[f"{desc_col}{grand_row}"].font = _BOLD


# ── Summary sheet ────────────────────────────────────────────────────────────────────
def fill_summary_sheet(ws, data_summaries: list[dict], tax_treatment: str) -> None:
    """Populate the REGENERATED Summary worksheet (5 columns; the original 6th per-sqft column
    is dropped). One row per exported DATA sheet with LIVE cross-sheet formulas to that sheet's
    amount grand-total cells, then the tax block:
      - "Total Amount Excluding Taxes" = SUM of the discipline Total cells;
      - Pre-tax  -> "GST @ 18%" (=subtotal*0.18) + "GRAND TOTAL" (=subtotal*1.18);
      - Post-tax -> "GRAND TOTAL" (=subtotal), NO GST line.
    PURE over openpyxl (no DB) so it is unit-testable directly."""
    for i, label in enumerate(_SUMMARY_HEADERS, start=1):
        c = ws.cell(row=1, column=i)
        c.value = label
        c.font = _BOLD
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    r = 2
    first_data_row = r
    for i, s in enumerate(data_summaries, start=1):
        qt = _quote_sheet(s["title"])
        grand = s.get("grand") or {}
        supply = grand.get("amount_supply")
        install = grand.get("amount_install")
        total = grand.get("amount_total")
        ws.cell(row=r, column=1).value = i
        ws.cell(row=r, column=2).value = s.get("label")
        if supply:
            ws.cell(row=r, column=3).value = f"={qt}!{supply}"
        if install:
            ws.cell(row=r, column=4).value = f"={qt}!{install}"
        if total:
            ws.cell(row=r, column=5).value = f"={qt}!{total}"
        else:
            # No scalar amount_total column -> Total = this row's own Supply + Install.
            ws.cell(row=r, column=5).value = f"=C{r}+D{r}"
        for col in (3, 4, 5):
            ws.cell(row=r, column=col).number_format = _CURRENCY_FMT
        r += 1

    last_data_row = r - 1

    subtotal_row = r
    ws.cell(row=subtotal_row, column=2).value = "Total Amount Excluding Taxes"
    ws.cell(row=subtotal_row, column=2).font = _BOLD
    stc = ws.cell(row=subtotal_row, column=5)
    stc.value = (
        f"=SUM(E{first_data_row}:E{last_data_row})" if data_summaries else 0
    )
    stc.number_format = _CURRENCY_FMT
    stc.font = _BOLD
    r += 1

    if (tax_treatment or "Pre-tax") == "Pre-tax":
        gst_row = r
        ws.cell(row=gst_row, column=2).value = "GST @ 18%"
        gc = ws.cell(row=gst_row, column=5)
        gc.value = f"=E{subtotal_row}*0.18"
        gc.number_format = _CURRENCY_FMT
        r += 1
        grand_row = r
        ws.cell(row=grand_row, column=2).value = "GRAND TOTAL"
        gtc = ws.cell(row=grand_row, column=5)
        gtc.value = f"=E{subtotal_row}*1.18"
        gtc.number_format = _CURRENCY_FMT
    else:  # Post-tax -> no GST line.
        grand_row = r
        ws.cell(row=grand_row, column=2).value = "GRAND TOTAL"
        gtc = ws.cell(row=grand_row, column=5)
        gtc.value = f"=E{subtotal_row}"
        gtc.number_format = _CURRENCY_FMT

    ws.cell(row=grand_row, column=2).font = _BOLD
    ws.cell(row=grand_row, column=5).font = _BOLD
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 42
    for col in ("C", "D", "E"):
        ws.column_dimensions[col].width = 18


# ── Excel title safety ────────────────────────────────────────────────────────────────
_ILLEGAL_TITLE_CHARS = str.maketrans({c: " " for c in r"\/*?:[]"})


def _excel_title(sheet_name: str) -> str:
    """A worksheet title honoring Excel's rules (<=31 chars, no \\ / * ? : [ ]). The committed
    sheet_name originated from Excel so it is normally already legal; this only guards a
    pathological name. The Summary uses ws.title (the ACTUAL assigned title) for its
    cross-sheet refs, so any transform here is transparent to the reference."""
    t = (sheet_name or "Sheet").translate(_ILLEGAL_TITLE_CHARS)[:31]
    return t or "Sheet"


# ── public entry ──────────────────────────────────────────────────────────────────────
def generate_template_priced_workbook(boq_name: str, sheet_names: list[str]) -> dict:
    """Build a priced .xlsx FROM SCRATCH for a template-origin BoQ's ticked sheets and return
    the SAME payload shape as the upload path. Data sheets are built in sheet_order; the
    regenerated Summary is inserted FIRST. last_exported_at is stamped ONLY after the bytes are
    produced (reject-mutates-nothing)."""
    plans = {sn: _resolve_template_sheet(boq_name, sn) for sn in sheet_names}
    boq_doc = frappe.get_doc("BOQs", boq_name)
    tax_treatment = boq_doc.tax_treatment or "Pre-tax"
    display_name = boq_doc.boq_name

    # Deterministic order: committed sheet_order, then verbatim name as a stable tiebreak.
    ordered = sorted(sheet_names, key=lambda sn: ((plans[sn].sheet_order or 0), sn))

    wb = Workbook()
    default_ws = wb.active  # openpyxl's initial "Sheet" -- removed after the real sheets exist

    data_summaries: list[dict] = []
    exported: list[str] = []
    remark_cols: dict[str, str] = {}

    for sn in ordered:
        plan = plans[sn]
        if plan.treat_as == "master_preamble":
            _build_preamble_sheet(wb, sn, boq_doc)  # Make List / general-specs dump
            exported.append(sn)
            continue
        summary = _build_data_sheet(wb, boq_name, sn, plan)
        data_summaries.append(summary)
        if summary.get("remark_col"):
            remark_cols[sn] = summary["remark_col"]
        exported.append(sn)

    # Regenerated Summary FIRST (built after the data sheets so their grand-total addresses
    # are known for the cross-sheet formulas).
    summary_ws = wb.create_sheet(title=_SUMMARY_TITLE, index=0)
    fill_summary_sheet(summary_ws, data_summaries, tax_treatment)

    wb.remove(default_ws)

    # Serialize to bytes IN MEMORY (no tempfile, no S3). A raise above never reaches here, so
    # the stamps below run only on a fully-built workbook.
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    data = buf.getvalue()

    # Stamp last_exported_at per exported sheet via set_value (NOT doc.save -- BoQ Sheet's
    # list-valued area_dimensions JSON throws on a full save; mirror export_writeback).
    now = frappe.utils.now()
    for sn in exported:
        frappe.db.set_value(_BOQ_SHEET, plans[sn].name, "last_exported_at", now,
                            update_modified=False)
    frappe.db.commit()

    ts = now[:19].replace("-", "").replace(":", "").replace(" ", "_")
    filename = f"{_safe_export_basename(display_name, boq_name)}_priced_{ts}.xlsx"
    return {
        "filename": filename,
        "content_type": _XLSX_CONTENT_TYPE,
        "content_base64": base64.b64encode(data).decode("ascii"),
        "exported_sheets": exported,
        "skipped_formula_columns": {},   # template amounts are WRITTEN as formulas -- none skipped
        "remark_columns": remark_cols,
        "last_exported_at": now,
    }
