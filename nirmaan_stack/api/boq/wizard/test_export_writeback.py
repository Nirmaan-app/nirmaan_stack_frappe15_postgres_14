# Copyright (c) 2026, Nirmaan (Stratos Infra Technologies Pvt. Ltd.) and Contributors
# See license.txt

"""Tests for the priced-workbook export -- specifically the TEMPLATE-ORIGIN from-scratch
branch (ADR-0013 A2-D2 / R4).

A template-cloned BoQ has NO source workbook, so export_priced_workbook GENERATES the priced
.xlsx from the committed tier (grid + pricing + amount formulas) instead of stamping a copy of
the S3 original. Coverage:

  PURE helpers (no DB): header synthesis, the formula-AST -> Excel translator, ref resolution,
  and the regenerated Summary / GST builder (Pre-tax vs Post-tax).

  DB end-to-end: seed a committed template BoQ (single-area data sheet + multi-area data sheet
  + a Make-List general-specs sheet), call export_priced_workbook, decode the base64, load the
  produced workbook, and assert the sheet set, grid cells, overlaid rates, amount cells being
  live Excel formulas, the multi-area Total = =SUM, the per-sheet grand-total row, and the
  Summary's live cross-sheet refs + GST line.

  GUARD: the upload path branch is unchanged (an origin='upload' BoQ with no source_file_url
  still throws the original "has no source_file_url set").

sheet_name carries a trailing space (#152) throughout.
"""
import base64
import io
import json

import frappe
from frappe.tests.utils import FrappeTestCase
from openpyxl import Workbook, load_workbook

from nirmaan_stack.api.boq.wizard import export_template_workbook as etw
from nirmaan_stack.api.boq.wizard.export_writeback import export_priced_workbook
from nirmaan_stack.api.boq.wizard.test_review_screen import _make_project, _cleanup_project


# The canonical stored amount formula: a simple qty_total * scalar-rate multiply.
def _multiply_ast(rate_field="rate_combined"):
    return {
        "op": "*",
        "operands": [
            {"ref": {"value_field": "qty_total", "value_key": None, "rate_subkey": None}},
            {"ref": {"value_field": rate_field, "value_key": None, "rate_subkey": None}},
        ],
    }


# ---------------------------------------------------------------------------
# PURE helper tests (no DB)
# ---------------------------------------------------------------------------
class TestTemplateExportPureHelpers(FrappeTestCase):

    _RM = {
        "A": {"role": "sl_no", "area": None},
        "B": {"role": "description", "area": None},
        "C": {"role": "unit", "area": None},
        "D": {"role": "qty_total", "area": None},
        "E": {"role": "rate_combined", "area": None},
        "F": {"role": "amount_total", "area": None},
    }
    _RM_AREA = {
        **_RM,
        "G": {"role": "qty", "area": "Zone A"},
        "H": {"role": "rate_supply_by_area", "area": "Zone A"},
    }

    def test_header_label_synthesis(self):
        self.assertEqual(etw.header_label_for("sl_no", None), "Sl. No.")
        self.assertEqual(etw.header_label_for("description", None), "Description")
        self.assertEqual(etw.header_label_for("unit", None), "Unit")
        self.assertEqual(etw.header_label_for("qty_total", None), "Total Quantity")
        self.assertEqual(etw.header_label_for("rate_supply", None), "Rate (Supply)")
        self.assertEqual(etw.header_label_for("rate_install", None), "Rate (Install)")
        self.assertEqual(etw.header_label_for("rate_combined", None), "Rate")
        self.assertEqual(etw.header_label_for("amount_supply", None), "Amount (Supply)")
        self.assertEqual(etw.header_label_for("amount_install", None), "Amount (Install)")
        self.assertEqual(etw.header_label_for("amount_total", None), "Amount")
        self.assertEqual(etw.header_label_for("row_notes", None), "Notes")
        # per-area qty column -> the AREA name.
        self.assertEqual(etw.header_label_for("qty", "Block B"), "Block B")
        # by-area rate/amount -> base label + area.
        self.assertEqual(etw.header_label_for("rate_supply_by_area", "Z1"), "Rate (Supply) - Z1")
        # unknown / non-display role -> empty.
        self.assertEqual(etw.header_label_for("ignore", None), "")
        self.assertEqual(etw.header_label_for(None, None), "")

    def test_resolve_ref_col(self):
        self.assertEqual(
            etw.resolve_ref_col({"value_field": "qty_total", "value_key": None}, self._RM), "D")
        self.assertEqual(
            etw.resolve_ref_col({"value_field": "rate_combined", "value_key": None}, self._RM), "E")
        # a scalar field with no matching column -> None.
        self.assertIsNone(
            etw.resolve_ref_col({"value_field": "rate_supply", "value_key": None}, self._RM))
        # area-bound refs resolve via (role, area).
        self.assertEqual(
            etw.resolve_ref_col(
                {"value_field": "qty_by_area", "value_key": "Zone A"}, self._RM_AREA), "G")
        self.assertEqual(
            etw.resolve_ref_col(
                {"value_field": "rate_by_area", "value_key": "Zone A", "rate_subkey": "supply_rate"},
                self._RM_AREA), "H")
        # an area-bound WILDCARD (no value_key) is unresolvable -> None (fail-safe).
        self.assertIsNone(
            etw.resolve_ref_col(
                {"value_field": "rate_by_area", "value_key": None, "rate_subkey": "supply_rate"},
                self._RM_AREA))

    def test_ast_to_excel_multiply(self):
        ast = _multiply_ast()
        self.assertEqual(etw.ast_to_excel(ast, 5, self._RM), "(D5*E5)")
        self.assertEqual(etw.ast_to_excel(ast, 42, self._RM), "(D42*E42)")

    def test_ast_to_excel_single_leaf(self):
        leaf = {"ref": {"value_field": "qty_total", "value_key": None}}
        self.assertEqual(etw.ast_to_excel(leaf, 3, self._RM), "D3")

    def test_ast_to_excel_failsafe_on_unresolvable_operand(self):
        # rate_supply has no column in _RM -> the WHOLE formula fails safe to None -> BLANK cell.
        ast = _multiply_ast(rate_field="rate_supply")
        self.assertIsNone(etw.ast_to_excel(ast, 5, self._RM))

    def test_ast_to_excel_unsupported_op_and_bad_shape(self):
        self.assertIsNone(etw.ast_to_excel({"op": "/", "operands": []}, 5, self._RM))
        self.assertIsNone(etw.ast_to_excel({"op": "*", "operands": []}, 5, self._RM))
        self.assertIsNone(etw.ast_to_excel("not a dict", 5, self._RM))
        self.assertIsNone(etw.ast_to_excel({}, 5, self._RM))

    def test_ast_to_excel_nested_add_inside_multiply(self):
        ast = {
            "op": "*",
            "operands": [
                {"ref": {"value_field": "qty_total", "value_key": None}},
                {"op": "+", "operands": [
                    {"ref": {"value_field": "rate_combined", "value_key": None}},
                    {"ref": {"value_field": "rate_combined", "value_key": None}},
                ]},
            ],
        }
        self.assertEqual(etw.ast_to_excel(ast, 7, self._RM), "(D7*(E7+E7))")

    def test_quote_sheet(self):
        self.assertEqual(etw._quote_sheet("My Sheet"), "'My Sheet'")
        self.assertEqual(etw._quote_sheet("Sheet "), "'Sheet '")           # trailing space kept
        self.assertEqual(etw._quote_sheet("O'Brien"), "'O''Brien'")        # internal quote doubled

    def test_excel_title_sanitizes_and_bounds(self):
        self.assertEqual(etw._excel_title("Normal"), "Normal")
        self.assertEqual(etw._excel_title("A/B*C?[x]"), "A B C  x ")       # illegal chars -> spaces
        self.assertEqual(len(etw._excel_title("x" * 40)), 31)              # bounded to 31
        self.assertEqual(etw._excel_title(""), "Sheet")

    def test_fill_summary_pretax(self):
        wb = Workbook()
        ws = wb.active
        data = [
            {"title": "Elec ", "label": "Electrical",
             "grand": {"amount_supply": "D20", "amount_install": "E20", "amount_total": "F20"}},
            {"title": "Plumb", "label": "Plumbing", "grand": {"amount_total": "F15"}},
        ]
        etw.fill_summary_sheet(ws, data, "Pre-tax")
        self.assertEqual(ws["A1"].value, "SL. No.")
        self.assertEqual(ws["E1"].value, "Total")
        self.assertEqual(ws["B2"].value, "Electrical")
        self.assertEqual(ws["C2"].value, "='Elec '!D20")   # trailing-space title quoted
        self.assertEqual(ws["D2"].value, "='Elec '!E20")
        self.assertEqual(ws["E2"].value, "='Elec '!F20")
        self.assertEqual(ws["E3"].value, "='Plumb'!F15")
        # subtotal row 4, GST 5, GRAND 6.
        self.assertEqual(ws["B4"].value, "Total Amount Excluding Taxes")
        self.assertEqual(ws["E4"].value, "=SUM(E2:E3)")
        self.assertEqual(ws["B5"].value, "GST @ 18%")
        self.assertEqual(ws["E5"].value, "=E4*0.18")
        self.assertEqual(ws["B6"].value, "GRAND TOTAL")
        self.assertEqual(ws["E6"].value, "=E4*1.18")

    def test_fill_summary_posttax_has_no_gst(self):
        wb = Workbook()
        ws = wb.active
        data = [{"title": "Elec", "label": "Electrical", "grand": {"amount_total": "F20"}}]
        etw.fill_summary_sheet(ws, data, "Post-tax")
        self.assertEqual(ws["B3"].value, "Total Amount Excluding Taxes")
        self.assertEqual(ws["B4"].value, "GRAND TOTAL")
        self.assertEqual(ws["E4"].value, "=E3")            # grand == subtotal, no GST multiplier
        labels = [ws.cell(row=r, column=2).value for r in range(1, 8)]
        self.assertNotIn("GST @ 18%", labels)

    def test_fill_summary_total_falls_back_to_supply_plus_install(self):
        # A sheet with only split amount columns (no scalar amount_total) -> Total = C+D.
        wb = Workbook()
        ws = wb.active
        etw.fill_summary_sheet(
            ws, [{"title": "S", "label": "Split",
                  "grand": {"amount_supply": "D9", "amount_install": "E9"}}], "Post-tax")
        self.assertEqual(ws["E2"].value, "=C2+D2")

    # -- blank-row compaction: row_has_content --------------------------------
    _CURRENCY = {"E", "F"}   # rate_combined + amount_total in _RM

    def test_row_has_content_true_for_a_real_data_cell(self):
        self.assertTrue(etw.row_has_content(
            {"B": "Wire", "D": 100}, self._CURRENCY, False, False))

    def test_row_has_content_ignores_currency_placeholder_columns(self):
        # An UNPRICED row whose only grid values sit in the rate/amount columns renders BLANK
        # (step (a) skips those columns), so it must not count as content.
        self.assertFalse(etw.row_has_content({"E": 0, "F": 0}, self._CURRENCY, False, False))

    def test_row_has_content_false_for_empty_and_whitespace(self):
        self.assertFalse(etw.row_has_content({}, self._CURRENCY, False, False))
        self.assertFalse(etw.row_has_content(None, self._CURRENCY, False, False))
        self.assertFalse(etw.row_has_content({"B": "   ", "C": None}, self._CURRENCY, False, False))

    def test_row_has_content_true_when_priced_or_remarked(self):
        self.assertTrue(etw.row_has_content({}, self._CURRENCY, True, False))    # stamped rate
        self.assertTrue(etw.row_has_content({}, self._CURRENCY, False, True))    # remark text

    def test_row_has_content_counts_unmapped_columns(self):
        # Step (a) writes ANY non-currency cell, including a column outside the role map.
        self.assertTrue(etw.row_has_content({"Z": "stray"}, self._CURRENCY, False, False))

    # -- blank-row compaction: build_compact_row_map --------------------------
    def test_row_map_is_identity_when_every_row_has_content(self):
        # The pre-compaction shape: contiguous content rows are untouched.
        self.assertEqual(etw.build_compact_row_map([2, 3, 4], 2, 4, 2), {2: 2, 3: 3, 4: 4})

    def test_row_map_collapses_a_blank_run_to_exactly_one(self):
        # 3,4,5 blank -> ONE blank output row (3); 6 lands at 4.
        self.assertEqual(etw.build_compact_row_map([2, 6], 2, 6, 2), {2: 2, 6: 4})

    def test_row_map_preserves_a_lone_blank(self):
        # A single authored spacer is a section break -- it survives.
        self.assertEqual(etw.build_compact_row_map([2, 4], 2, 4, 2), {2: 2, 4: 4})

    def test_row_map_treats_holes_and_blanks_as_one_run(self):
        # Deselection holes (absent rows) and blank grid rows collapse together.
        self.assertEqual(etw.build_compact_row_map([2, 9], 2, 9, 2), {2: 2, 9: 4})

    def test_row_map_collapses_leading_and_trailing_runs_too(self):
        # ONE uniform rule, no head/tail special case: a leading gap becomes one blank row.
        self.assertEqual(etw.build_compact_row_map([5], 2, 8, 2), {5: 3})

    def test_row_map_empty_when_no_content(self):
        self.assertEqual(etw.build_compact_row_map([], 2, 9, 2), {})

    # -- blank-row compaction: _remap_excel_rows ------------------------------
    def test_remap_excel_rows_re_addresses_and_drops_compacted_rows(self):
        recs = [{"excel_row": 2, "col_letter": "E", "rate": 25},
                {"excel_row": 9, "col_letter": "E", "rate": 200},
                {"excel_row": 7, "col_letter": "E", "rate": 99}]   # 7 was compacted away
        out = etw._remap_excel_rows(recs, {2: 2, 9: 4})
        self.assertEqual(
            out,
            [{"excel_row": 2, "col_letter": "E", "rate": 25},
             {"excel_row": 4, "col_letter": "E", "rate": 200}],
        )

    def test_remap_excel_rows_handles_empty_input(self):
        self.assertEqual(etw._remap_excel_rows([], {2: 2}), [])
        self.assertEqual(etw._remap_excel_rows(None, {2: 2}), [])


# ---------------------------------------------------------------------------
# DB-backed seeding helper
# ---------------------------------------------------------------------------
def _seed_template_boq(project_name):
    """Create a committed TEMPLATE BoQ with: a single-area data sheet, a multi-area data sheet,
    and a Make-List general-specs sheet -- each with its committed grid (+ pricing + amount
    formula for the data sheets). Returns the BOQs docname."""
    boq = frappe.new_doc("BOQs")
    boq.project = project_name
    boq.boq_name = "Template Export Test BoQ"
    boq.tax_treatment = "Pre-tax"
    boq.origin = "template"
    boq.area_dimensions = json.dumps(["Zone A", "Zone B"])
    boq.append("general_specs_sheets", {
        "source_sheet_name": "Make List ",
        "preamble_text": "Line 1\nLine 2\nLine 3",
    })
    boq.insert(ignore_permissions=True)
    frappe.db.commit()
    boq_name = boq.name
    now = frappe.utils.now()

    # ---- single-area data sheet "Electrical " ----
    elec_map = {
        "A": {"role": "sl_no", "area": None},
        "B": {"role": "description", "area": None},
        "C": {"role": "unit", "area": None},
        "D": {"role": "qty_total", "area": None},
        "E": {"role": "rate_combined", "area": None},
        "F": {"role": "amount_total", "area": None},
    }
    _seed_committed_sheet(
        boq_name, "Electrical ", "data", 1, elec_map, [], sheet_order=1,
        sheet_label="Electrical", now=now,
        grid=[
            {"row_number": 2, "row_order": 0,
             "cells": {"A": "1", "B": "Wire", "C": "Rmt", "D": 100}},
            {"row_number": 3, "row_order": 1,
             "cells": {"A": "2", "B": "Switch", "C": "Nos", "D": 50}},
        ],
        pricing=[{"excel_row": 2, "col_letter": "E", "rate": 25},
                 {"excel_row": 3, "col_letter": "E", "rate": 200}],
        formula={"target_col": "F", "target_value_field": "amount_total",
                 "formula": _multiply_ast()},
    )

    # ---- multi-area data sheet "HVAC " (Zone A=D, Zone B=E, Total=F) ----
    hvac_map = {
        "A": {"role": "sl_no", "area": None},
        "B": {"role": "description", "area": None},
        "C": {"role": "unit", "area": None},
        "D": {"role": "qty", "area": "Zone A"},
        "E": {"role": "qty", "area": "Zone B"},
        "F": {"role": "qty_total", "area": None},
        "G": {"role": "rate_combined", "area": None},
        "H": {"role": "amount_total", "area": None},
    }
    _seed_committed_sheet(
        boq_name, "HVAC ", "data", 1, hvac_map, ["Zone A", "Zone B"], sheet_order=2,
        sheet_label="HVAC", now=now,
        grid=[
            {"row_number": 2, "row_order": 0,
             "cells": {"A": "1", "B": "AC unit", "C": "Nos", "D": 3, "E": 2, "F": 5}},
            {"row_number": 3, "row_order": 1,
             "cells": {"A": "2", "B": "Duct", "C": "Rmt", "D": 10, "E": 20, "F": 30}},
        ],
        pricing=[{"excel_row": 2, "col_letter": "G", "rate": 1000},
                 {"excel_row": 3, "col_letter": "G", "rate": 50}],
        formula={"target_col": "H", "target_value_field": "amount_total",
                 "formula": _multiply_ast()},
    )

    # ---- Make-List general-specs sheet (grid-only, no pricing/formula) ----
    _seed_committed_sheet(
        boq_name, "Make List ", "master_preamble", 1, {}, [], sheet_order=3,
        sheet_label="Make List", now=now,
        grid=[{"row_number": 1, "row_order": 0, "cells": {"A": "Line 1\nLine 2\nLine 3"}}],
        pricing=[], formula=None, sheet_disposition="grid_only",
    )
    frappe.db.commit()
    return boq_name


def _seed_committed_sheet(boq_name, sheet_name, treat_as, cv, role_map, area_dims,
                          sheet_order, sheet_label, now, grid, pricing, formula,
                          sheet_disposition="grid_and_nodes"):
    sheet = frappe.new_doc("BoQ Sheet")
    sheet.boq = boq_name
    sheet.sheet_name = sheet_name
    sheet.sheet_order = sheet_order
    sheet.sheet_label = sheet_label
    sheet.treat_as = treat_as
    sheet.header_row = 1
    sheet.header_row_count = 1
    sheet.column_role_map = json.dumps(role_map)
    sheet.column_headers = json.dumps({})
    sheet.area_dimensions = json.dumps(area_dims)
    sheet.commit_version = cv
    sheet.is_current = 1
    sheet.committed_at = now
    sheet.insert(ignore_permissions=True)

    g = frappe.new_doc("BoQ Committed Sheet Grid")
    g.boq = boq_name
    g.source_sheet_name = sheet_name
    g.sheet_disposition = sheet_disposition
    g.commit_version = cv
    g.is_current = 1
    g.committed_at = now
    for row in grid:
        g.append("rows", row)
    g.insert(ignore_permissions=True)

    for p in pricing:
        pr = frappe.new_doc("BoQ Cell Pricing")
        pr.boq = boq_name
        pr.sheet_name = sheet_name
        pr.excel_row = p["excel_row"]
        pr.col_letter = p["col_letter"]
        pr.committed_version = cv
        pr.rate = p["rate"]
        pr.is_filled = 1
        pr.pricing_version = 1
        pr.is_current = 1
        pr.priced_at = now
        pr.insert(ignore_permissions=True)

    if formula:
        af = frappe.new_doc("BoQ Cell Amount Formula")
        af.boq = boq_name
        af.sheet_name = sheet_name
        af.committed_version = cv
        af.target_value_field = formula["target_value_field"]
        af.target_col = formula["target_col"]
        af.formula = json.dumps(formula["formula"])
        af.formula_version = 1
        af.is_current = 1
        af.defined_at = now
        af.insert(ignore_permissions=True)


class TestTemplateExportEndToEnd(FrappeTestCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.project = _make_project()
        cls.boq = _seed_template_boq(cls.project.name)
        result = export_priced_workbook(
            boq_name=cls.boq,
            sheet_names=json.dumps(["Electrical ", "HVAC ", "Make List "]),
        )
        cls.result = result
        cls.wb = load_workbook(
            io.BytesIO(base64.b64decode(result["content_base64"])), data_only=False
        )

    @classmethod
    def tearDownClass(cls):
        for dt, flt in (
            ("BoQ Cell Amount Formula", {"boq": cls.boq}),
            ("BoQ Cell Pricing", {"boq": cls.boq}),
            ("BoQ Committed Sheet Grid", {"boq": cls.boq}),
            ("BoQ Sheet", {"boq": cls.boq}),
        ):
            frappe.db.delete(dt, flt)
        frappe.db.delete("BOQs", {"name": cls.boq})
        frappe.db.commit()
        _cleanup_project(cls.project.name)
        super().tearDownClass()

    # -- workbook shape --------------------------------------------------------
    def test_sheet_set_and_order(self):
        names = self.wb.sheetnames
        self.assertEqual(names[0], "Summary")               # Summary is FIRST
        self.assertIn("Electrical ", names)                 # trailing-space title (#152)
        self.assertIn("HVAC ", names)
        self.assertIn("Make List ", names)
        # data sheets in sheet_order after Summary.
        self.assertLess(names.index("Electrical "), names.index("HVAC "))

    def test_return_payload_shape(self):
        self.assertEqual(
            set(self.result["exported_sheets"]), {"Electrical ", "HVAC ", "Make List "})
        self.assertEqual(self.result["skipped_formula_columns"], {})   # formulas WRITTEN, none skipped
        self.assertTrue(self.result["filename"].endswith(".xlsx"))
        self.assertEqual(self.result["content_type"], etw._XLSX_CONTENT_TYPE)

    # -- single-area data sheet ------------------------------------------------
    def test_electrical_grid_header_rates_and_amounts(self):
        ws = self.wb["Electrical "]
        # synthesized header row.
        self.assertEqual(ws["A1"].value, "Sl. No.")
        self.assertEqual(ws["B1"].value, "Description")
        self.assertEqual(ws["D1"].value, "Total Quantity")
        self.assertEqual(ws["E1"].value, "Rate")
        self.assertEqual(ws["F1"].value, "Amount")
        # grid cells + static single-area qty (untouched).
        self.assertEqual(ws["B2"].value, "Wire")
        self.assertEqual(ws["D2"].value, 100)
        # overlaid rate.
        self.assertEqual(ws["E2"].value, 25)
        # AMOUNT is a live Excel formula qty*rate.
        self.assertEqual(ws["F2"].value, "=(D2*E2)")
        self.assertEqual(ws["F3"].value, "=(D3*E3)")

    def test_electrical_grand_total_row(self):
        ws = self.wb["Electrical "]
        # data rows 2..3 -> grand-total row 4.
        self.assertEqual(ws["B4"].value, "TOTAL")
        self.assertEqual(ws["F4"].value, "=SUM(F2:F3)")

    # -- multi-area data sheet -------------------------------------------------
    def test_hvac_total_quantity_is_sum_of_areas(self):
        ws = self.wb["HVAC "]
        # area qty columns D (Zone A) + E (Zone B); Total qty F = live SUM.
        self.assertEqual(ws["D1"].value, "Zone A")
        self.assertEqual(ws["E1"].value, "Zone B")
        self.assertEqual(ws["F2"].value, "=SUM(D2:E2)")
        self.assertEqual(ws["F3"].value, "=SUM(D3:E3)")
        # per-area qty grid cells preserved.
        self.assertEqual(ws["D2"].value, 3)
        self.assertEqual(ws["E2"].value, 2)

    def test_hvac_amount_formula_over_total_and_rate(self):
        ws = self.wb["HVAC "]
        self.assertEqual(ws["G2"].value, 1000)             # overlaid rate
        self.assertEqual(ws["H2"].value, "=(F2*G2)")       # amount = Total-qty * rate
        self.assertEqual(ws["H4"].value, "=SUM(H2:H3)")    # grand total

    # -- Make List -------------------------------------------------------------
    def test_make_list_dumps_preamble_lines(self):
        ws = self.wb["Make List "]
        self.assertEqual(ws["A1"].value, "Line 1")
        self.assertEqual(ws["A2"].value, "Line 2")
        self.assertEqual(ws["A3"].value, "Line 3")

    # -- Summary ---------------------------------------------------------------
    def test_summary_cross_sheet_formulas_and_gst(self):
        ws = self.wb["Summary"]
        self.assertEqual(ws["B2"].value, "Electrical")
        self.assertEqual(ws["B3"].value, "HVAC")
        # Electrical grand-total amount cell is F4; quoted trailing-space title.
        self.assertEqual(ws["E2"].value, "='Electrical '!F4")
        self.assertEqual(ws["E3"].value, "='HVAC '!H4")
        # tax block (Pre-tax): subtotal, GST, grand total.
        self.assertEqual(ws["B4"].value, "Total Amount Excluding Taxes")
        self.assertEqual(ws["E4"].value, "=SUM(E2:E3)")
        self.assertEqual(ws["B5"].value, "GST @ 18%")
        self.assertEqual(ws["E5"].value, "=E4*0.18")
        self.assertEqual(ws["B6"].value, "GRAND TOTAL")
        self.assertEqual(ws["E6"].value, "=E4*1.18")

    # -- last_exported_at stamped ----------------------------------------------
    def test_last_exported_at_stamped_on_each_sheet(self):
        for sn in ("Electrical ", "HVAC ", "Make List "):
            val = frappe.db.get_value(
                "BoQ Sheet", {"boq": self.boq, "sheet_name": sn, "is_current": 1},
                "last_exported_at")
            self.assertIsNotNone(val)


class TestTemplateExportBlankRowCompaction(FrappeTestCase):
    """END-TO-END blank-row compaction on a PRUNED template sheet.

    The committed grid keeps each row at its ORIGINAL template Excel row while the commit drops
    the deselected ones, so a pruned sheet exports with a hole at every removed row on top of the
    master's own spacer rows. The export compacts those: any run of consecutive blank source rows
    becomes exactly ONE blank output row, a lone spacer survives, and EVERY absolute-row writer
    (grid, rates, amount formulas, colours, remarks, grand total) follows the same map.

    Source layout (header_row 1)          ->  Output
      2  Wire      content                ->  2
      3  4  5      blank grid rows        ->  3   (one blank)
      6  Switch    content                ->  4
      7  8         ABSENT (deselected)    ->  5   (merged with 9 -- one blank)
      9            currency placeholders  ->  (blank: an unpriced rate/amount-only row)
      10 Panel     content + rate + color ->  6
      11           lone blank spacer      ->  7   (preserved)
      12           remark only            ->  8
                                              9   TOTAL
    """

    _ROLE_MAP = {
        "A": {"role": "sl_no", "area": None},
        "B": {"role": "description", "area": None},
        "C": {"role": "unit", "area": None},
        "D": {"role": "qty_total", "area": None},
        "E": {"role": "rate_combined", "area": None},
        "F": {"role": "amount_total", "area": None},
    }

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.project = _make_project()
        boq = frappe.new_doc("BOQs")
        boq.project = cls.project.name
        boq.boq_name = "Compaction Test BoQ"
        boq.tax_treatment = "Pre-tax"
        boq.origin = "template"
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        cls.boq = boq.name
        now = frappe.utils.now()

        _seed_committed_sheet(
            cls.boq, "Compact ", "data", 1, cls._ROLE_MAP, [], sheet_order=1,
            sheet_label="Compact", now=now,
            grid=[
                {"row_number": 2, "row_order": 0,
                 "cells": {"A": "1", "B": "Wire", "C": "Rmt", "D": 100}},
                {"row_number": 3, "row_order": 1, "cells": {}},
                {"row_number": 4, "row_order": 2, "cells": {"B": "   "}},   # whitespace == blank
                {"row_number": 5, "row_order": 3, "cells": {}},
                {"row_number": 6, "row_order": 4,
                 "cells": {"A": "2", "B": "Switch", "C": "Nos", "D": 50}},
                # rows 7 + 8 absent -- deselected at commit (is_excluded=1).
                {"row_number": 9, "row_order": 5, "cells": {"E": 0, "F": 0}},  # unpriced placeholders
                {"row_number": 10, "row_order": 6,
                 "cells": {"A": "3", "B": "Panel", "C": "Nos", "D": 5}},
                {"row_number": 11, "row_order": 7, "cells": {}},             # lone spacer
                {"row_number": 12, "row_order": 8, "cells": {}},             # remark-only row
            ],
            pricing=[{"excel_row": 2, "col_letter": "E", "rate": 25},
                     {"excel_row": 10, "col_letter": "E", "rate": 200}],
            formula={"target_col": "F", "target_value_field": "amount_total",
                     "formula": _multiply_ast()},
        )

        rk = frappe.new_doc("BoQ Cell Remark")
        rk.boq, rk.sheet_name, rk.excel_row = cls.boq, "Compact ", 12
        rk.committed_version, rk.remark_version, rk.is_current = 1, 1, 1
        rk.remark = "Check with client"
        rk.insert(ignore_permissions=True)

        cl = frappe.new_doc("BoQ Cell Color")
        cl.boq, cl.sheet_name, cl.excel_row = cls.boq, "Compact ", 10
        cl.col_letter, cl.color = "B", "yellow"
        cl.committed_version, cl.color_version, cl.is_current = 1, 1, 1
        cl.insert(ignore_permissions=True)
        frappe.db.commit()

        cls.result = export_priced_workbook(
            boq_name=cls.boq, sheet_names=json.dumps(["Compact "]))
        cls.wb = load_workbook(
            io.BytesIO(base64.b64decode(cls.result["content_base64"])), data_only=False)
        cls.ws = cls.wb["Compact "]
        # Snapshot the extent AS PRODUCED: openpyxl MATERIALISES a cell on indexed access, so a
        # later test reading an out-of-range address would otherwise grow max_row under us.
        cls.max_row = cls.ws.max_row

    @classmethod
    def tearDownClass(cls):
        for dt in ("BoQ Cell Color", "BoQ Cell Remark", "BoQ Cell Amount Formula",
                   "BoQ Cell Pricing", "BoQ Committed Sheet Grid", "BoQ Sheet"):
            frappe.db.delete(dt, {"boq": cls.boq})
        frappe.db.delete("BOQs", {"name": cls.boq})
        frappe.db.commit()
        _cleanup_project(cls.project.name)
        super().tearDownClass()

    def test_content_rows_are_compacted_onto_consecutive_output_rows(self):
        self.assertEqual(self.ws["B2"].value, "Wire")     # src 2  -> 2
        self.assertEqual(self.ws["B4"].value, "Switch")   # src 6  -> 4
        self.assertEqual(self.ws["B6"].value, "Panel")    # src 10 -> 6

    def test_blank_runs_collapse_to_one_and_a_lone_spacer_survives(self):
        # src 3/4/5 -> one blank at 3; src 7/8/9 -> one blank at 5; src 11 (lone) -> blank at 7.
        for r in (3, 5, 7):
            self.assertTrue(
                all(self.ws.cell(row=r, column=c).value in (None, "") for c in range(1, 7)),
                f"output row {r} should be blank",
            )

    def test_rates_follow_the_row_map(self):
        # THE corruption guard: the src-10 rate must land on output row 6, never on row 10.
        # iter_rows (not indexed access) so the assertion cannot materialise an out-of-range cell.
        rate_rows = {
            c.row: c.value
            for (c,) in self.ws.iter_rows(min_col=5, max_col=5, min_row=2)
            if c.value is not None
        }
        self.assertEqual(rate_rows, {2: 25, 6: 200})

    def test_amount_formulas_reference_the_compacted_rows(self):
        self.assertEqual(self.ws["F2"].value, "=(D2*E2)")
        self.assertEqual(self.ws["F6"].value, "=(D6*E6)")
        self.assertIsNone(self.ws["F4"].value)   # src 6 is unpriced -> blank amount

    def test_grand_total_spans_the_compacted_range(self):
        self.assertEqual(self.ws["B9"].value, "TOTAL")
        self.assertEqual(self.ws["F9"].value, "=SUM(F2:F8)")

    def test_remark_and_colour_follow_the_row_map(self):
        remark_col = self.result["remark_columns"]["Compact "]
        self.assertEqual(self.ws[f"{remark_col}8"].value, "Check with client")  # src 12 -> 8
        self.assertEqual(self.ws["B6"].fill.fgColor.rgb[-6:].upper(), "FFEB9C")  # src 10 -> 6

    def test_sheet_does_not_span_the_original_row_numbers(self):
        # 12 source rows -> 9 output rows (8 data/blank + TOTAL). Nothing past the grand total.
        self.assertEqual(self.max_row, 9)


class TestUploadPathBranchUnchanged(FrappeTestCase):
    """The is_template branch is a NO-OP for origin != 'template': an upload-origin BoQ with no
    source_file_url still hits the ORIGINAL guard throw (byte-identical upload path)."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.project = _make_project()
        boq = frappe.new_doc("BOQs")
        boq.project = cls.project.name
        boq.boq_name = "Upload Guard BoQ"
        boq.tax_treatment = "Pre-tax"
        # origin defaults to "upload"; source_file_url left unset.
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        cls.boq = boq.name

    @classmethod
    def tearDownClass(cls):
        frappe.db.delete("BOQs", {"name": cls.boq})
        frappe.db.commit()
        _cleanup_project(cls.project.name)
        super().tearDownClass()

    def test_upload_origin_without_source_file_throws(self):
        with self.assertRaises(frappe.ValidationError):
            export_priced_workbook(boq_name=self.boq, sheet_names=json.dumps(["Any Sheet"]))


# ---------------------------------------------------------------------------
# STANDING GUARD (owner-locked, slice BCS-S1): BCS never reaches the export
# ---------------------------------------------------------------------------
class TestBcsCostRatesNeverReachTheExport(FrappeTestCase):
    """BCS is our INTERNAL cost and margin. The priced workbook is handed to the CLIENT,
    so a BCS value appearing in it would leak what the job costs us.

    The exclusion holds by CONSTRUCTION: the export reads "BoQ Cell Pricing" and names
    three fields explicitly (excel_row, col_letter, rate), while BCS lives in its own
    doctype (BoQ Row BCS Rate) with no col_letter at all. This test PINS that
    construction. It is expected to pass the day it is written -- its value is the day it
    fails, which is the day someone folds BCS onto BoQ Cell Pricing, adds a BCS stamping
    pass, or widens the export's field list.

    NOT VACUOUS: the test first asserts the BCS rows genuinely exist for the exported
    sheets and version, so a pass can never mean "there was nothing to leak"."""

    # Deliberately absurd sentinels that cannot collide with the fixture's own numbers
    # (rates 25 / 200 / 1000 / 50; quantities 3 / 5 / 10 / 20 / 30 / 50 / 100).
    #
    # ⚠️ _COMBINED ADDED AT BCS-S3a, AND ITS ABSENCE MATTERED. BCS-S2b widened storage to a
    # THIRD field, combined_rate, for a sheet that quotes one undifferentiated figure -- and it
    # is the ONLY field such a sheet uses. This class kept seeding just the two halves, so on
    # the whole axis S2b opened it was passing because there was nothing there to leak, which is
    # exactly the vacuity the class docstring says it exists to avoid. A combined-rate sheet's
    # cost could have reached the CLIENT workbook with the standing guard green.
    _SUPPLY = 987654.21
    _INSTALL = 123456.78
    _COMBINED = 456789.33

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.project = _make_project()
        cls.boq = _seed_template_boq(cls.project.name)
        cls.sheets = ["Electrical ", "HVAC ", "Make List "]

        # Cost EVERY priced row of both data sheets, at the CURRENT committed version.
        for sheet, rows in (("Electrical ", (2, 3)), ("HVAC ", (2, 3))):
            for excel_row in rows:
                doc = frappe.new_doc("BoQ Row BCS Rate")
                doc.boq = cls.boq
                doc.sheet_name = sheet          # VERBATIM (#152)
                doc.excel_row = excel_row
                doc.committed_version = 1
                doc.supply_rate = cls._SUPPLY
                doc.install_rate = cls._INSTALL
                doc.combined_rate = cls._COMBINED
                doc.is_filled = 1
                doc.bcs_version = 1
                doc.is_current = 1
                doc.bcs_rated_at = frappe.utils.now()
                doc.insert(ignore_permissions=True)
        frappe.db.commit()

        result = export_priced_workbook(
            boq_name=cls.boq, sheet_names=json.dumps(cls.sheets)
        )
        cls.result = result
        cls.wb = load_workbook(
            io.BytesIO(base64.b64decode(result["content_base64"])), data_only=False
        )

    @classmethod
    def tearDownClass(cls):
        for dt, flt in (
            ("BoQ Row BCS Rate", {"boq": cls.boq}),
            ("BoQ Cell Amount Formula", {"boq": cls.boq}),
            ("BoQ Cell Pricing", {"boq": cls.boq}),
            ("BoQ Committed Sheet Grid", {"boq": cls.boq}),
            ("BoQ Sheet", {"boq": cls.boq}),
        ):
            frappe.db.delete(dt, flt)
        frappe.db.delete("BOQs", {"name": cls.boq})
        frappe.db.commit()
        _cleanup_project(cls.project.name)
        super().tearDownClass()

    def _all_cell_values(self):
        """Every cell value in every worksheet of the produced workbook, as strings."""
        out = []
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value not in (None, ""):
                        out.append(str(cell.value))
        return out

    def test_the_bcs_rows_really_exist_for_the_exported_version(self):
        """Anti-vacuity guard: without this, an empty BCS table would make every
        assertion below pass while proving nothing at all."""
        rows = frappe.get_all(
            "BoQ Row BCS Rate",
            filters={"boq": self.boq, "committed_version": 1, "is_current": 1},
            fields=["sheet_name", "excel_row", "supply_rate", "install_rate", "combined_rate"],
        )
        self.assertEqual(len(rows), 4, "4 costed rows must exist during the export")
        self.assertTrue(all(r.supply_rate == self._SUPPLY for r in rows))
        self.assertTrue(all(r.install_rate == self._INSTALL for r in rows))
        # BCS-S3a: the third field must genuinely be stored, or the leak assertions below
        # would pass on it for the same "nothing to leak" reason they were written to exclude.
        self.assertTrue(all(r.combined_rate == self._COMBINED for r in rows))

    def test_no_bcs_rate_value_appears_anywhere_in_the_exported_workbook(self):
        values = self._all_cell_values()
        self.assertTrue(values, "the export produced a workbook with content")
        for sentinel in (self._SUPPLY, self._INSTALL, self._COMBINED):
            for rendered in (str(sentinel), str(int(sentinel))):
                hits = [v for v in values if rendered in v]
                self.assertEqual(
                    hits, [],
                    f"BCS cost {sentinel} leaked into the CLIENT workbook: {hits}",
                )

    def test_no_bcs_derived_total_or_margin_leaks_either(self):
        """Total Amount and % Profit are computed, never stored -- so they must not appear
        in the export by any route. Checks the per-unit sum and, for each fixture row, the
        quantity-multiplied total that a BCS Total Amount column would carry.

        BOTH RENDER FORMS, as its sibling above already does: an integer-rendered total
        ('1111111' rather than '1111111.0') would otherwise slip straight through.

        BCS-S3a adds the COMBINED-sheet shape. ⚠️ It is checked as its OWN per-unit cost, NOT
        added to the halves: `bcs.py:16` states combined_rate is not a total of the other two
        and must never be summed with them, so a `_SUPPLY + _INSTALL + _COMBINED` candidate
        would be a figure the product never computes -- a sentinel for an arithmetic that does
        not exist, which is a test that can only ever pass."""
        values = self._all_cell_values()
        halves = self._SUPPLY + self._INSTALL
        candidates = {halves, self._COMBINED}
        for qty in (100, 50, 5, 30):        # the fixture's committed quantities
            candidates.add(halves * qty)
            candidates.add(self._COMBINED * qty)
        for candidate in candidates:
            for rendered in (str(candidate), str(int(candidate))):
                self.assertEqual(
                    [v for v in values if rendered in v], [],
                    f"a BCS-derived figure ({candidate}, rendered {rendered!r}) leaked "
                    f"into the CLIENT workbook",
                )

    def test_the_export_still_carries_the_client_facing_rates(self):
        """The complement: excluding BCS must not have excluded ordinary pricing. The
        client's own rates are still stamped, so the guard above is about BCS alone."""
        ws = self.wb["Electrical "]
        self.assertEqual(ws["E2"].value, 25)
        self.assertEqual(ws["E3"].value, 200)

    def test_export_writeback_module_never_names_the_bcs_doctype(self):
        """Belt-and-braces on the CONSTRUCTION itself: the export module must not mention
        BCS at all. This is what makes the exclusion structural rather than incidental.

        MATCHED CASE-INSENSITIVELY: the token list mixes cases ('bcs' against the doctype
        'BoQ Row BCS Rate'), so a case-sensitive search was inconsistent with itself -- a
        module-level `BCS_DOCTYPE = ...` or a `SUPPLY_RATE` constant would have passed."""
        import inspect

        from nirmaan_stack.api.boq.wizard import export_writeback

        src = inspect.getsource(export_writeback).lower()
        # BCS-S3a: `combined_rate` joins the token list for the same reason the other two are
        # on it -- the S2b field was never added, so the ONE field a combined-rate sheet uses
        # was the one name this guard did not watch for.
        for token in ("BoQ Row BCS Rate", "supply_rate", "install_rate", "combined_rate", "bcs"):
            self.assertNotIn(
                token.lower(), src,
                f"export_writeback.py must never reference {token!r} (in any casing) -- "
                f"BCS is internal cost and the priced workbook is client-facing",
            )
