# Copyright (c) 2026, Nirmaan (Stratos Infra Technologies Pvt. Ltd.) and contributors
# For license information, please see license.txt

"""BoQ Excel write-back -- the priced-workbook generator (Phase 5 Slice 5a).

Given a committed BoQ + a ticked sheet subset, this module produces a COPY of the
original .xlsx with the user's pricing + annotations stamped IN, and returns its bytes
for download. It NEVER touches the original on S3 (copy-on-write) and NEVER writes
amounts, formulas, or structure -- the tendering document is client-owned (governing
principle 0): we write RATES only (and the user's own color/remark annotations).

What it stamps, per ticked sheet, onto a fresh copy:
  - RATES (BoQ Cell Pricing, is_current + is_filled): the rate value INTO its
    (excel_row, col_letter) cell -- but PER-CELL FORMULA SKIP: if the original cell is a
    formula (cell.data_type == 'f'), it is LEFT UNTOUCHED and reported as a skipped
    formula-rate column (e.g. a VRF combined-rate =SUM(supply,install) cell). The paired
    amount formula recomputes when the client opens the workbook.
  - COLORS (BoQ Cell Color): an openpyxl PatternFill at the tagged (excel_row, col_letter)
    -- any column, incl. non-rate (a description cell). A fill never alters the cell value
    or formula, so it is NOT subject to the formula-skip rule.
  - PRICED-CELL VERIFICATION HIGHLIGHT (system, always-on): a muted-teal PatternFill on every
    rate cell the write-back ACTUALLY stamped -- a Nirmaan-facing "which cells did we write"
    aid. Skipped formula rate cells get NO highlight (so teal doubles as the live signal of
    the rates-only + formula-skip rule). Applied AFTER the user-color pass, so on a stamped
    rate cell that also carries a user color tag the system teal WINS (the one place a system
    fill beats a user fill -- the aid must be exhaustive over written cells).
  - REMARKS (BoQ Cell Remark, per-row): a NEW TRAILING COLUMN one past the TRUE data edge
    (the rightmost MAPPED column from the committed column_role_map -- NOT openpyxl
    max_column, which is inflated by empty styled cells). A hard empty-column safety check
    refuses to write if that column carries real data.

After stamping + saving the copy, a POST-SAVE FIDELITY ASSERTION re-opens the saved file
and verifies the amount-formula count, merged-range count, worksheet count, and
defined-name count are unchanged vs the pre-stamp copy -- a mismatch FAILS the export
(reject-mutates-nothing; nothing is handed to a client and last_exported_at is not stamped).

Grid-only general-specs sheets (treat_as == "master_preamble") carry no rates/nodes; they
pass through UNTOUCHED but still count as exported (their last_exported_at is stamped).

Return shape (base64-in-JSON, NOT frappe.local.response.filecontent): the file-only
download idiom cannot carry the skipped-formula report alongside the bytes, so one JSON
response carries BOTH -- 5b decodes content_base64 -> Blob -> browser download and surfaces
skipped_formula_columns.

Public API:
  export_priced_workbook(boq_name, sheet_names) -> dict   [whitelisted POST]
"""
from __future__ import annotations

import base64
import json
import os
import shutil
from typing import Any

import frappe
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

from nirmaan_stack.api.boq.wizard.sheet_preview import _fetch_boq_file_to_tempfile

_PRICING = "BoQ Cell Pricing"
_COLOR = "BoQ Cell Color"
_REMARK = "BoQ Cell Remark"
_BOQ_SHEET = "BoQ Sheet"

_XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# The header label written atop the appended remark column.
_REMARK_HEADER = "Nirmaan Remarks"

# Token -> Excel fill hex (ARGB without the alpha; openpyxl PatternFill accepts RRGGBB).
# DECIDED HERE (Slice 5a -- the one place a color hex is chosen): light, distinct tints so
# the cell's own text stays readable under the fill. MUST stay in sync with the 8 tokens in
# boq_cell_color.json (red/orange/yellow/green/blue/purple/pink/grey).
_COLOR_HEX: dict[str, str] = {
    "red": "FFC7CE",     # light red
    "orange": "FFD9A0",  # light orange
    "yellow": "FFEB9C",  # light yellow
    "green": "C6EFCE",   # light green
    "blue": "BDD7EE",    # light blue
    "purple": "E1D5F7",  # light purple
    "pink": "FBD4E4",    # light pink
    "grey": "D9D9D9",    # light grey
}

# System "priced-cell verification highlight" fill -- a Nirmaan-facing aid marking every rate
# cell the write-back ACTUALLY stamped (teal => "we wrote a rate here"; no-teal on a rate cell
# => "we left the client's formula untouched"). Muted teal, DELIBERATELY a SEPARATE constant
# from the 8 user tokens in _COLOR_HEX (and distinct from user green C6EFCE / user blue BDD7EE)
# so a SYSTEM mark can never read as a USER color tag. On a stamped rate cell that also carries
# a user color tag, this system fill WINS (applied AFTER _apply_colors) -- the verification aid
# must be exhaustive over written cells; that is the ONE place a system fill beats a user fill.
_PRICED_HIGHLIGHT_HEX = "B7E4D8"


# ── arg coercion ─────────────────────────────────────────────────────────────────
def _coerce_names(sheet_names: Any) -> list[str]:
    """Normalize sheet_names (JSON string from HTTP, or a Python list) to a non-empty list.
    Mirrors commit_pipeline._coerce_subset."""
    if sheet_names is None:
        frappe.throw("sheet_names is required.", title="Missing field: sheet_names")
    if isinstance(sheet_names, str):
        try:
            sheet_names = json.loads(sheet_names)
        except (ValueError, TypeError):
            frappe.throw("sheet_names must be a JSON list of sheet names.",
                         title="Invalid sheet_names")
    if not isinstance(sheet_names, (list, tuple)):
        frappe.throw("sheet_names must be a list of sheet names.", title="Invalid sheet_names")
    names = [s for s in sheet_names]
    if not names:
        frappe.throw("sheet_names must name at least one sheet.", title="Empty sheet_names")
    return names


# ── pure worksheet helpers (unit-tested directly against synthetic workbooks) ─────
def _find_ws(wb, sheet_name: str):
    """Resolve a worksheet by VERBATIM (#152) title, falling back to a stripped compare
    (trailing-space sheet names exist). Throws if the sheet is not in the workbook."""
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    for ws in wb.worksheets:
        if ws.title.strip() == sheet_name.strip():
            return ws
    frappe.throw(
        f"Sheet '{sheet_name}' not found in the workbook. Available: {wb.sheetnames}",
        title="Sheet not found",
    )


def _rightmost_mapped_col_index(column_role_map: dict) -> int:
    """The TRUE data edge: the rightmost MAPPED Excel column index (1-based) from the
    committed column_role_map. 0 when the map is empty. This is the deliberate basis for
    the remark column (NOT openpyxl max_column, which recon found inflated by empty styled
    cells out to AC/Z/AS)."""
    if not column_role_map:
        return 0
    idxs = []
    for col in column_role_map.keys():
        try:
            idxs.append(column_index_from_string(str(col)))
        except (ValueError, TypeError):
            continue
    return max(idxs) if idxs else 0


def _col_is_empty(ws, col_idx: int) -> bool:
    """True iff every cell in column col_idx is empty across the sheet (no value, no
    formula). A formula cell has its formula string as .value, so it is caught too."""
    for r in range(1, (ws.max_row or 0) + 1):
        if ws.cell(row=r, column=col_idx).value not in (None, ""):
            return False
    return True


def _stamp_rates(ws, pricing_rows: list[dict]) -> tuple[list[dict], list[tuple]]:
    """Stamp each filled rate into its (col_letter, excel_row) cell -- RATES ONLY.
    PER-CELL FORMULA SKIP: a cell whose data_type == 'f' is LEFT UNTOUCHED and recorded as
    skipped (decided against the REAL file, never inferred from the role name).

    Returns (skipped, written): the list of skipped {excel_row, col_letter} entries AND the
    list of (col_letter, excel_row) addresses ACTUALLY written. The priced-cell highlight is
    driven by `written` (the truly-stamped cells), NOT the input pricing list, so a skipped
    formula cell is EXCLUDED from the highlight -- the teal mark is the live signal of the
    rates-only + formula-skip safety rule."""
    skipped: list[dict] = []
    written: list[tuple] = []
    for p in pricing_rows:
        col = p["col_letter"]
        row = int(p["excel_row"])
        cell = ws[f"{col}{row}"]
        if cell.data_type == "f":  # a formula -> never overwrite (e.g. combined =SUM(supply,install))
            skipped.append({"excel_row": row, "col_letter": col})
            continue
        cell.value = p["rate"]
        written.append((col, row))
    return skipped, written


def _apply_priced_highlight(ws, written_cells: list[tuple]) -> int:
    """Apply the system priced-cell verification fill (muted teal) to every rate cell the
    write-back ACTUALLY stamped. MUST run AFTER _apply_colors so on a collision with a user
    color tag the system teal lands last and WINS (RULE 2). A fill sets ONLY .fill -- the
    rate value just stamped into the cell is untouched. Only stamped RATE cells get this; a
    user color on a NON-stamped cell is left as the user set it. Returns the count filled."""
    for col, row in written_cells:
        ws[f"{col}{row}"].fill = PatternFill(fill_type="solid", fgColor=_PRICED_HIGHLIGHT_HEX)
    return len(written_cells)


def _apply_colors(ws, color_rows: list[dict]) -> int:
    """Apply a solid PatternFill at each tagged (col_letter, excel_row) -- ANY column,
    incl. non-rate. A fill sets ONLY the cell's fill; .value / formula are never touched
    (so a colored formula cell keeps its formula). Returns the count applied."""
    applied = 0
    for c in color_rows:
        hexv = _COLOR_HEX.get(c["color"])
        if not hexv:
            continue  # unknown token -> skip defensively (never invent a fill)
        cell = ws[f'{c["col_letter"]}{int(c["excel_row"])}']
        cell.fill = PatternFill(fill_type="solid", fgColor=hexv)  # value/formula untouched
        applied += 1
    return applied


def _write_remark_column(ws, remark_rows: list[dict], column_role_map: dict, header_row) -> str:
    """Append the remark column one past the TRUE data edge and write the header + each
    remark at its excel_row. HARD SAFETY: refuse (throw) if the target column is not
    genuinely empty across the sheet -- never overwrite real data. Returns the column letter."""
    rightmost = _rightmost_mapped_col_index(column_role_map)
    if rightmost <= 0:
        frappe.throw(
            f"Cannot place the remark column on sheet '{ws.title}': no mapped columns in "
            "the committed column_role_map (true data edge is undefined).",
            title="Remark column undefined",
        )
    target_idx = rightmost + 1
    target_letter = get_column_letter(target_idx)
    if not _col_is_empty(ws, target_idx):
        frappe.throw(
            f"Remark target column {target_letter} on sheet '{ws.title}' is not empty -- "
            "refusing to overwrite existing data.",
            title="Remark column not empty",
        )
    hrow = header_row if (header_row and int(header_row) >= 1) else 1
    ws.cell(row=int(hrow), column=target_idx).value = _REMARK_HEADER
    for r in remark_rows:
        text = r.get("remark")
        if text in (None, ""):
            continue
        ws.cell(row=int(r["excel_row"]), column=target_idx).value = text
    return target_letter


# ── fidelity guard ────────────────────────────────────────────────────────────────
def _fidelity_snapshot(wb) -> dict:
    """Count the four invariants the write-back must NOT disturb: amount/any formula cells,
    merged ranges, worksheets, and defined names."""
    formulas = 0
    merges = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    formulas += 1
        merges += len(list(ws.merged_cells.ranges))
    return {
        "formulas": formulas,
        "merges": merges,
        "sheets": len(wb.sheetnames),
        "defined_names": len(list(wb.defined_names)),
    }


def _assert_fidelity(before: dict, after: dict) -> None:
    """Throw if any of the four invariants diverged between the pre-stamp and saved copy."""
    diffs = [
        f"{k}: {before[k]} -> {after[k]}"
        for k in ("formulas", "merges", "sheets", "defined_names")
        if before.get(k) != after.get(k)
    ]
    if diffs:
        frappe.throw(
            "Export fidelity check FAILED -- the round-trip would alter the workbook: "
            + "; ".join(diffs)
            + ". No file was produced.",
            title="Export fidelity failure",
        )


# ── per-sheet committed-version + config resolution ───────────────────────────────
def _resolve_sheet_plan(boq_name: str, sheet_name: str) -> dict:
    """Resolve the CURRENT committed BoQ Sheet for (boq, sheet_name): its name,
    commit_version, treat_as, column_role_map, header_row. The client NEVER passes a
    version -- it is read here. Throws if there is no current committed version."""
    row = frappe.db.get_value(
        _BOQ_SHEET,
        {"boq": boq_name, "sheet_name": sheet_name, "is_current": 1},
        ["name", "commit_version", "treat_as", "column_role_map", "header_row"],
        as_dict=True,
    )
    if not row:
        frappe.throw(
            f"No current committed version for sheet '{sheet_name}' in BoQ '{boq_name}'. "
            "Commit the sheet before exporting.",
            title="No committed version",
        )
    crm = row.column_role_map
    if isinstance(crm, str):
        try:
            crm = json.loads(crm)
        except (ValueError, TypeError):
            crm = {}
    row.column_role_map = crm or {}
    return row


# ── the worker (testable: takes an already-fetched workbook copy path) ─────────────
def _generate_priced_workbook(boq_name: str, sheet_names: list[str], src_path: str) -> dict:
    """Stamp the committed pricing/colors/remarks for each sheet onto the workbook at
    src_path (a throwaway COPY), run the fidelity guard, stamp last_exported_at per exported
    sheet, and return the download payload. src_path is loaded data_only=False so formulas
    survive (data_only=True would load cached VALUES and DESTROY formulas on save -- the trap).

    Separated from the whitelisted endpoint so tests can inject a synthetic workbook and
    bypass the S3 fetch."""
    plans = {sn: _resolve_sheet_plan(boq_name, sn) for sn in sheet_names}

    # MUST be data_only=False -- preserve formulas as formula strings (see the trap above).
    wb = openpyxl.load_workbook(src_path, data_only=False)
    before = _fidelity_snapshot(wb)

    exported: list[str] = []
    skipped_by_sheet: dict[str, list[str]] = {}
    remark_cols: dict[str, str] = {}

    for sn in sheet_names:
        plan = plans[sn]
        ws = _find_ws(wb, sn)

        # Grid-only general-specs sheets carry no rates/nodes -> pass through untouched,
        # but still count as exported (owner decision: ticked => last_exported_at stamped).
        if plan.treat_as == "master_preamble":
            exported.append(sn)
            continue

        cv = plan.commit_version
        pricing = frappe.get_all(
            _PRICING,
            filters={"boq": boq_name, "sheet_name": sn, "committed_version": cv,
                     "is_current": 1, "is_filled": 1},
            fields=["excel_row", "col_letter", "rate"],
            order_by="excel_row asc, col_letter asc",
        )
        colors = frappe.get_all(
            _COLOR,
            filters={"boq": boq_name, "sheet_name": sn, "committed_version": cv, "is_current": 1},
            fields=["excel_row", "col_letter", "color"],
        )
        remarks = frappe.get_all(
            _REMARK,
            filters={"boq": boq_name, "sheet_name": sn, "committed_version": cv, "is_current": 1},
            fields=["excel_row", "remark"],
            order_by="excel_row asc",
        )

        skipped, written = _stamp_rates(ws, pricing)
        _apply_colors(ws, colors)
        # System priced-cell highlight LAST so it WINS over any user color on a stamped rate
        # cell (RULE 2); driven by `written`, so skipped formula cells get NO teal (RULE 1).
        _apply_priced_highlight(ws, written)
        if remarks:
            remark_cols[sn] = _write_remark_column(ws, remarks, plan.column_role_map, plan.header_row)

        if skipped:
            skipped_by_sheet[sn] = sorted({s["col_letter"] for s in skipped})
        exported.append(sn)

    # Save the stamped copy, re-open it, and assert fidelity BEFORE returning / stamping
    # last_exported_at (reject-mutates-nothing on a fidelity failure).
    saved = src_path + ".priced.xlsx"
    try:
        wb.save(saved)
        wb.close()
        wb2 = openpyxl.load_workbook(saved, data_only=False)
        after = _fidelity_snapshot(wb2)
        wb2.close()
        _assert_fidelity(before, after)
        with open(saved, "rb") as f:
            data = f.read()
    finally:
        try:
            os.unlink(saved)
        except OSError:
            pass

    # Stamp last_exported_at per exported sheet via set_value (NOT doc.save -- BoQ Sheet's
    # list-valued area_dimensions JSON throws on a full save; mirror the commit pipeline).
    now = frappe.utils.now()
    for sn in exported:
        frappe.db.set_value(_BOQ_SHEET, plans[sn].name, "last_exported_at", now,
                            update_modified=False)
    frappe.db.commit()

    ts = frappe.utils.now()[:19].replace("-", "").replace(":", "").replace(" ", "_")
    filename = f"{boq_name}_priced_{ts}.xlsx"
    return {
        "filename": filename,
        "content_type": _XLSX_CONTENT_TYPE,
        "content_base64": base64.b64encode(data).decode("ascii"),
        "exported_sheets": exported,
        "skipped_formula_columns": skipped_by_sheet,
        "remark_columns": remark_cols,
        "last_exported_at": now,
    }


@frappe.whitelist(methods=["POST"])
def export_priced_workbook(boq_name: str = None, sheet_names: Any = None) -> dict:
    """Generate a priced .xlsx for a ticked subset of a committed BoQ's sheets and return
    its bytes (base64) + the skipped-formula report. The CURRENT committed version per sheet
    is resolved server-side (the client never passes a version).

    COPY-ON-WRITE: the original is fetched from S3 to a tempfile and shutil-copied to a
    second tempfile that is the ONLY thing stamped + saved; the original temp is never
    written and nothing is ever uploaded back to S3 (the source remains the immutable
    forever-reference).

    Returns {filename, content_type, content_base64, exported_sheets,
    skipped_formula_columns, remark_columns, last_exported_at}. 5b decodes content_base64 ->
    Blob -> browser download and surfaces skipped_formula_columns.
    URL: /api/method/nirmaan_stack.api.boq.wizard.export_writeback.export_priced_workbook
    """
    if not boq_name:
        frappe.throw("boq_name is required.", title="Missing field: boq_name")
    if not frappe.db.exists("BOQs", boq_name):
        frappe.throw(f"BOQs '{boq_name}' not found.", title="Not found")
    names = _coerce_names(sheet_names)

    source_file_url = frappe.db.get_value("BOQs", boq_name, "source_file_url")
    if not source_file_url:
        frappe.throw(f"BOQs '{boq_name}' has no source_file_url set.", title="Missing source file")

    fetched = None
    copy = None
    try:
        fetched = _fetch_boq_file_to_tempfile(source_file_url)
        copy = fetched + ".work.xlsx"
        shutil.copy(fetched, copy)  # COPY-ON-WRITE -- stamp the copy, never the original temp/S3
        return _generate_priced_workbook(boq_name, names, copy)
    finally:
        for p in (fetched, copy):
            if p:
                try:
                    os.unlink(p)
                except OSError:
                    pass
