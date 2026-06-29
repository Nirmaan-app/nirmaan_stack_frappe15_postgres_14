"""
BoQ commit pipeline -- the commit shell + grid write-body + NODE TREE (Phase 5
Slices 3a + 3b).

This writes REAL parsed BoQ data into the committed schema.

THE TWO-LAYER MODEL
  Every committable sheet commits its COMPLETE original as a FAITHFUL GRID (all 6
  row classifications, in original position) into `BoQ Committed Sheet Grid` (3a).
  A FINALIZED sheet ALSO commits a NODE TREE (3b; X): its review rows are read,
  resolved (resolve_effective), mapped, and written as BOQ Nodes + BOQ Node Qty By
  Area children, parent-wired, with provenance carried. X commits EVERY classified row
  except spacer (preamble/line_item -> Preamble/Line Item; note/subtotal_marker/
  header_repeat -> Other; row_class carries the full taxonomy), so the node tree is a
  complete semantic mirror. General-specs sheets are grid-only.

DISPOSITION (from the Slice-2 commit gate, mapped onto the grid's discriminator):
  gate "general_specs" -> sheet_disposition "grid_only"
  gate "finalized"     -> sheet_disposition "grid_and_nodes" (gets the node tree)

NODE WRITE (3b) -- two-pass + deferred list-JSON fields
  Pass 1 inserts each node PARENT-LESS, with NO list-valued JSON field set
  (attached_notes / edit_log left null). Pass 2 sets parent_node + doc.save() in
  ancestor-depth order so paths cascade correctly AND the controller's parent-chain
  validation re-runs (a free integrity check) -- safe because the list fields are
  still null at save (Frappe's get_valid_dict rejects a list-valued JSON field, the
  same wall BoQ Sheet.area_dimensions hit in 3a). Pass 3 writes the list-JSON
  fields (attached_notes, edit_log) via frappe.db.set_value(json.dumps(...)), a
  targeted column write that bypasses full-doc serialization. (Dict-valued JSON
  like append_notes_raw is set in pass 1 -- only LISTs trip get_valid_dict.)

CAPTURE-ONLY (Slice 2.5)
  Reviewed money values (rate_*/amount_*) are carried VERBATIM; nothing is
  recomputed. combined_rate consistency is a WARNING, not a block (3b). Float->
  Currency coercion is left to Frappe.

THREE-TIER VERSIONING
  One SHARED commit_version per (boq, sheet) covers the grid, the BoQ Sheet, and
  all nodes for that commit. On re-commit each tier FREEZES its prior current row
  (is_current 1 -> 0 via frappe.db.set_value -- NEVER doc.save(): the BoQ Sheet's
  list-valued area_dimensions would re-serialize and hit get_valid_dict) and lands
  a new is_current=1 row under the shared version. v1 nodes stay attached to frozen
  sheet v1 -- grid v1 + sheet v1 + nodes v1 remain a coherent frozen snapshot. The
  BoQ Sheet is now VERSIONED (freeze-and-supersede), NOT deleted (3a's raw-delete
  retired).

SINGLE-OPEN FILE PATH
  The workbook is fetched + opened ONCE; sheets are looped, each grid built by the
  SHARED sheet_preview._extract_grid_rows inside the single open workbook.

VERBATIM SHEET IDENTITY (#152)
  source_sheet_name / sheet_name matched byte-for-byte on write AND freeze lookup
  -- never trimmed (trailing-space sheet names exist; PostgreSQL `=` on varchar is
  byte-exact).

TRANSACTION BOUNDARY
  One trailing frappe.db.commit() PER SHEET (no savepoints). A sheet's grid +
  BoQ Sheet + node tree land together. PER-SHEET FAILURE ISOLATION (Slice 5): the
  loop wraps each sheet in try/except; a sheet that raises is rolled back
  (frappe.db.rollback() -- MANDATORY, since catching the exception suppresses
  Frappe's request-level rollback and the freeze-before-write would otherwise be
  flushed by the next sheet's commit and orphan it) and recorded in the returned
  failed[]; the loop continues. Earlier committed sheets are durable. Mixed
  committed/failed is a valid outcome (no all-or-nothing wrapper).

Public API:
  commit_boq(boq_name, sheet_subset) -> dict   [whitelisted, POST]
"""
from __future__ import annotations

import json
import os
from typing import Any

import frappe
import openpyxl

from nirmaan_stack.api.boq.wizard.commit_gate import compute_committable_sheets
from nirmaan_stack.api.boq.wizard.review_screen import resolve_effective
from nirmaan_stack.api.boq.wizard.sheet_preview import (
    _extract_grid_rows,
    _fetch_boq_file_to_tempfile,
)
# The level helpers + the shared cls->node_type+level derivation + the effective-
# classification constants are OWNED by commit_validation (so the real commit and the
# pre-commit preflight can never diverge). _build_node_pass1 calls
# derive_node_type_and_level; _commit_node_tree calls _compute_levelless_preamble_levels
# (now returning (assigned, level_warnings) -- the real commit IGNORES the warnings;
# they surface only in the preflight).
from nirmaan_stack.api.boq.wizard.commit_validation import (
    _PREAMBLE_CLS,
    _LINE_ITEM_CLS,
    _GRID_ONLY_CLASSIFICATIONS,
    _compute_levelless_preamble_levels,
    derive_node_type_and_level,
    RESOLVE_EFFECTIVE_COMMIT_INPUT_FIELDS,
)

_GRID_DOCTYPE = "BoQ Committed Sheet Grid"
_SHEET_DOCTYPE = "BoQ Sheet"
_NODE_DOCTYPE = "BOQ Nodes"

# Gate disposition -> grid discriminator (the only mapping; do not re-derive).
_DISPOSITION_TO_SHEET_DISPOSITION = {
    "general_specs": "grid_only",
    "finalized": "grid_and_nodes",
}
# Gate disposition -> committed BoQ Sheet.treat_as snapshot value.
_DISPOSITION_TO_TREAT_AS = {
    "general_specs": "master_preamble",
    "finalized": "data",
}

# X: commit ALL classified rows as semantic nodes EXCEPT spacer, so the committed node
# tree is a complete semantic mirror of the reviewed BoQ. node_type is the PRICEABILITY
# axis -- preamble -> Preamble, line_item -> Line Item, every OTHER committed class
# (note / subtotal_marker / header_repeat) -> Other; the full effective classification is
# carried verbatim on row_class (the taxonomy axis, set in pass 1). SPACER stays GRID-ONLY
# (pure layout, no semantic content -- the faithful grid preserves its position). A future
# pricing walk stays exactly `node_type in (Preamble, Line Item)`, unchanged.
# The effective-classification constants (_PREAMBLE_CLS / _LINE_ITEM_CLS /
# _GRID_ONLY_CLASSIFICATIONS, plus _SPACER_CLS / _PRICEABLE_CLASSIFICATIONS) are OWNED by
# commit_validation (single source of truth); the few this module still references are
# imported above.

# BoQ Review Row fields the node body reads (the get_review_rows source-of-truth set,
# narrowed to what the mapping needs).
_REVIEW_ROW_FIELDS = [
    "name", "row_index", "source_row_number",
    # SHARED with the preflight (commit_validation) so resolve_effective gets identical
    # inputs in both paths -- human > parser, NO ai_* (parity invariant; see the constant).
    *RESOLVE_EFFECTIVE_COMMIT_INPUT_FIELDS,
    "sl_no_value", "description", "unit", "make_model",
    "is_rate_only", "is_synthetic",
    "qty_total", "qty_by_area",
    "rate_supply", "rate_install", "rate_combined", "rate_by_area",
    "amount_total", "amount_supply", "amount_install", "amount_by_area",
    "row_notes", "append_notes_raw", "attached_notes",
    "edit_log", "edited_by", "edited_at", "remarks",
]


def _coerce_subset(sheet_subset: Any) -> list[str]:
    """Normalize sheet_subset (JSON string from HTTP, or a Python list) to a list."""
    if sheet_subset is None:
        frappe.throw("sheet_subset is required.", title="Missing field: sheet_subset")
    if isinstance(sheet_subset, str):
        try:
            sheet_subset = json.loads(sheet_subset)
        except (ValueError, TypeError):
            frappe.throw("sheet_subset must be a JSON list of sheet names.",
                         title="Invalid sheet_subset")
    if not isinstance(sheet_subset, (list, tuple)):
        frappe.throw("sheet_subset must be a list of sheet names.",
                     title="Invalid sheet_subset")
    subset = [s for s in sheet_subset]
    if not subset:
        frappe.throw("sheet_subset must name at least one sheet.",
                     title="Empty sheet_subset")
    return subset


def _commit_failure_reason(e: Exception) -> str:
    """A best-effort, renderable reason for a per-sheet commit failure, with a safe
    fallback. Prefers the exception's own message (a frappe.throw carries its text on
    the raised ValidationError), lightly stripped of HTML; NEVER returns empty (the
    fallback covers a message-less exception, e.g. a bare RuntimeError). Deliberately
    simple -- no error classification."""
    raw = ""
    try:
        raw = str(e) or ""
    except Exception:
        raw = ""
    try:
        raw = frappe.utils.strip_html_tags(raw)
    except Exception:
        pass
    raw = raw.strip()
    return raw or "Commit failed for this sheet -- see server logs."


def _record_commit_failure(boq_name: str, sheet_name: str, reason: str) -> None:
    """Persist a per-sheet commit-failure stamp ADDITIVELY on the BoQ Sheet Draft
    (Slice F1 -- the durable analog of Slice 1a's parse_failure_*). Writes ONLY the two
    failure fields (commit_failure_reason / commit_failure_at) -- NEVER wizard_status.

    sheet_name matched VERBATIM (#152 -- trailing-space names exist; PostgreSQL `=` on
    varchar is byte-exact). A missing child row is silently skipped. NO commit here: the
    caller owns the commit. In the commit_boq per-sheet except this is called AFTER the
    mandatory frappe.db.rollback() (so the stamp is not swept away by it) and is followed
    by the caller's own explicit frappe.db.commit() -- required because commit_boq has no
    trailing commit, so a last-sheet failure would otherwise never be flushed. The draft
    is a DIFFERENT doctype from the rolled-back grid/sheet/node tiers, so this set_value is
    the only pending write and cannot resurrect a rolled-back tier write (T2 orphan-
    prevention preserved).
    """
    child_name = frappe.db.get_value(
        "BoQ Sheet Draft",
        {"parent": boq_name, "parenttype": "BOQs", "sheet_name": sheet_name},
        "name",
    )
    if not child_name:
        return
    frappe.db.set_value("BoQ Sheet Draft", child_name, {
        "commit_failure_reason": reason,
        "commit_failure_at": frappe.utils.now(),
    }, update_modified=False)


def _clear_commit_failure(boq_name: str, sheet_name: str) -> None:
    """Clear any prior per-sheet commit-failure stamp on a SUCCESSFUL commit (Slice F1 --
    mirrors Slice 1a's clear-on-success). Sets both failure fields to None so a re-commit
    that now succeeds drops the stale notice.

    sheet_name matched VERBATIM (#152). Missing child row silently skipped. NO commit:
    this is folded into _commit_one_sheet's trailing per-sheet frappe.db.commit(), so the
    clear lands ATOMICALLY with the successful grid/sheet/node write.
    """
    child_name = frappe.db.get_value(
        "BoQ Sheet Draft",
        {"parent": boq_name, "parenttype": "BOQs", "sheet_name": sheet_name},
        "name",
    )
    if not child_name:
        return
    frappe.db.set_value("BoQ Sheet Draft", child_name, {
        "commit_failure_reason": None,
        "commit_failure_at": None,
    }, update_modified=False)


def _coerce_config(sheet_config: Any) -> dict:
    """The draft sheet_config blob may arrive as a dict or a JSON string."""
    if not sheet_config:
        return {}
    if isinstance(sheet_config, str):
        try:
            return json.loads(sheet_config) or {}
        except (ValueError, TypeError):
            return {}
    if isinstance(sheet_config, dict):
        return sheet_config
    return {}


def _current_names(doctype: str, boq: str, name_field: str, name_value: str) -> list[str]:
    """The single centralized 'read the current row(s)' accessor across all three
    committed tiers (Slice 3b). Returns the names of is_current=1 rows for a tier,
    parameterized by the tier's identity field:
      grid      -> ("BoQ Committed Sheet Grid", boq, "source_sheet_name", sheet_name)
      BoQ Sheet -> ("BoQ Sheet",                boq, "sheet_name",        sheet_name)
      nodes     -> ("BOQ Nodes",                boq, "sheet",             boq_sheet_name)
    name_value is matched VERBATIM (#152). Normally returns 0 or 1 name.
    """
    return frappe.get_all(
        doctype,
        filters={"boq": boq, name_field: name_value, "is_current": 1},
        pluck="name",
    )


def _next_commit_version(boq: str, sheet_name: str) -> int:
    """The SHARED commit version for this (boq, sheet) commit -- one number for the
    grid, the BoQ Sheet, and all nodes. Anchored on the grid (always written for
    both dispositions): max prior grid version + 1 (first commit = 1)."""
    agg = frappe.get_all(
        _GRID_DOCTYPE,
        filters={"boq": boq, "source_sheet_name": sheet_name},
        fields=["max(commit_version) as mv"],
    )
    return ((agg[0].mv if agg else None) or 0) + 1


@frappe.whitelist(methods=["POST"])
def commit_boq(boq_name: str = None, sheet_subset: Any = None) -> dict:
    """Commit a subset of a BoQ's sheets into the committed schema (Phase 5 Slice 3a/3b).

    For each requested sheet (under ONE shared commit_version): build + persist its
    faithful grid (both dispositions), version its committed BoQ Sheet with the
    column-config snapshot (freeze-and-supersede), and -- for a FINALIZED sheet --
    write the node tree (BOQ Nodes + per-area children, parent-wired, provenance).

    SERVER-SIDE GATE RE-CHECK: every sheet in sheet_subset MUST be in the live
    commit gate's eligible set, or the whole call is rejected (throw) before any
    write or file fetch. The caller's subset is never trusted. (Eligibility is a
    precondition -- a whole-call throw; it is NOT a per-sheet runtime failure.)

    PER-SHEET FAILURE ISOLATION (Slice 5): once past the gate, each sheet commits in
    its own transaction. A sheet that raises mid-write is rolled back (MANDATORY -- so
    its freeze-before-write does not orphan it once the exception is caught) and
    recorded in failed[]; the loop continues. Earlier successful sheets stay durable.
    MIXED STATE (some committed, some failed) is a valid resting outcome -- there is
    NO all-or-nothing wrapper.

    Args:
      boq_name:     an existing BOQs document name.
      sheet_subset: list (or JSON-string list) of sheet names to commit. Required.

    Returns:
      {"boq_name": str,
       "committed": [{"sheet_name", "disposition", "sheet_disposition",
                      "grid_name", "boq_sheet_name", "commit_version", "row_count",
                      "froze_prior", "froze_prior_sheet", "node_count",
                      "froze_nodes"}, ...],
       "failed": [{"sheet_name", "reason"}, ...]}   # empty list when all succeed
    """
    if not boq_name:
        frappe.throw("boq_name is required.", title="Missing field: boq_name")
    if not frappe.db.exists("BOQs", boq_name):
        frappe.throw(f"BOQs '{boq_name}' not found.", title="Not found")

    subset = _coerce_subset(sheet_subset)

    boq_doc = frappe.get_doc("BOQs", boq_name)

    # Live gate (SEPARATE from parse-eligibility; same read pattern the gate uses).
    general_specs_sheet_names: set[str] = {
        row.source_sheet_name
        for row in (boq_doc.general_specs_sheets or [])
        if row.source_sheet_name
    }
    committable = compute_committable_sheets(
        boq_doc.sheet_drafts, general_specs_sheet_names
    )
    disposition_by_sheet = {c["sheet_name"]: c["disposition"] for c in committable}

    # Gate re-check -- reject any requested sheet that is not commit-eligible.
    ineligible = [s for s in subset if s not in disposition_by_sheet]
    if ineligible:
        frappe.throw(
            "These sheets are not commit-eligible and cannot be committed: "
            f"{ineligible}. Eligible sheets: {sorted(disposition_by_sheet)}.",
            title="Not commit-eligible",
        )

    # Draft lookup (verbatim sheet_name) for the column-config snapshot.
    drafts_by_name = {d.sheet_name: d for d in boq_doc.sheet_drafts}

    source_file_url = frappe.db.get_value("BOQs", boq_name, "source_file_url")
    if not source_file_url:
        frappe.throw(
            f"BOQs '{boq_name}' has no source_file_url set.",
            title="Missing source file",
        )

    tempfile_path = None
    wb = None
    committed: list[dict] = []
    failed: list[dict] = []
    try:
        # SINGLE fetch + SINGLE open; loop the sheets inside.
        tempfile_path = _fetch_boq_file_to_tempfile(source_file_url)
        wb = openpyxl.load_workbook(tempfile_path, data_only=True, read_only=True)

        for sheet_name in subset:
            # PER-SHEET FAILURE ISOLATION (Slice 5): a sheet that raises mid-write is
            # rolled back + recorded in failed[]; the loop continues. Earlier sheets
            # stay committed (each _commit_one_sheet commits before the next begins),
            # so MIXED STATE (some committed, some failed) is a valid resting outcome.
            try:
                if sheet_name not in wb.sheetnames:
                    # A genuinely-absent sheet is a PER-SHEET failure (consistent with
                    # the new contract), NOT a whole-call abort -- the throw is inside
                    # the per-sheet try so it lands in failed[]. (An ineligible sheet is
                    # still rejected upfront by the gate re-check, before this loop.)
                    frappe.throw(
                        f"Sheet '{sheet_name}' not found in the workbook. "
                        f"Available sheets: {wb.sheetnames}",
                        title="Sheet not found",
                    )
                ws = wb[sheet_name]
                grid_rows = _extract_grid_rows(ws)  # shared transform, single open wb

                disposition = disposition_by_sheet[sheet_name]
                draft = drafts_by_name.get(sheet_name)
                result = _commit_one_sheet(
                    boq_name, sheet_name, disposition, grid_rows, draft
                )
                committed.append(result)
            except Exception as e:
                # MANDATORY rollback. Catching the exception SUPPRESSES Frappe's
                # request-level rollback, so we MUST roll back here ourselves -- BEFORE
                # recording + continuing. Otherwise this sheet's freeze-before-write
                # (the prior version's is_current set to 0) + its partial new writes stay
                # pending, and the NEXT sheet's frappe.db.commit() would FLUSH them,
                # orphaning the sheet (prior frozen, new write incomplete -> NO
                # is_current=1 version). Already-committed earlier sheets are durable and
                # untouched by this rollback (commit made them permanent).
                frappe.db.rollback()
                # PERSIST the failure durably (Slice F1) -- AFTER the rollback above (so the
                # stamp is not swept away by it) and with its OWN explicit commit (commit_boq
                # has NO trailing commit; if this is the LAST sheet in the subset, no later
                # _commit_one_sheet commit would ever flush the stamp -> it would be lost).
                # The draft is a DIFFERENT doctype from the rolled-back tiers, so after the
                # rollback this set_value is the only pending write -- it cannot re-flush a
                # rolled-back tier write, so the orphan-prevention above is preserved.
                reason = _commit_failure_reason(e)
                _record_commit_failure(boq_name, sheet_name, reason)
                frappe.db.commit()
                failed.append({
                    "sheet_name": sheet_name,
                    "reason": reason,
                })
                frappe.log_error(
                    message=f"commit_boq: sheet {sheet_name!r} failed\n\n{frappe.get_traceback()}",
                    title="BoQ commit per-sheet failure",
                )
                continue
    finally:
        if wb is not None:
            wb.close()
        if tempfile_path is not None:
            try:
                os.unlink(tempfile_path)
            except OSError:
                pass

    return {"boq_name": boq_name, "committed": committed, "failed": failed}


def _commit_one_sheet(
    boq_name: str,
    sheet_name: str,
    disposition: str,
    grid_rows: list[dict],
    draft: Any,
) -> dict:
    """Persist ONE sheet's three tiers (grid + BoQ Sheet [+ node tree]), then commit.

    Per-sheet transaction boundary: a single trailing frappe.db.commit() so the
    grid, the versioned BoQ Sheet, and (for a finalized sheet) the node tree land
    atomically for this sheet under ONE shared commit_version.

    grid_rows is the output of sheet_preview._extract_grid_rows: a list of
    {"row_number", "cells": {col_letter: value}} in source order.
    """
    sheet_disposition = _DISPOSITION_TO_SHEET_DISPOSITION[disposition]
    # ONE shared commit_version for grid + BoQ Sheet + nodes this commit.
    commit_version = _next_commit_version(boq_name, sheet_name)
    committed_at = frappe.utils.now()

    grid_doc, froze_grid = _write_grid(
        boq_name, sheet_name, sheet_disposition, grid_rows, commit_version, committed_at
    )
    boq_sheet_name, prior_sheet_names = _write_committed_boq_sheet(
        boq_name, sheet_name, disposition, draft, commit_version, committed_at
    )

    node_result = {"node_count": 0, "froze_nodes": 0}
    if disposition == "finalized":
        node_result = _commit_node_tree(
            boq_name, sheet_name, boq_sheet_name, prior_sheet_names,
            commit_version, committed_at,
        )

    # OUTPUT-FIDELITY RECONCILIATION (Slice 2) -- grid tier. Verify the persisted faithful
    # grid equals the extracted grid_rows BEFORE the commit. (The node tier reconciles
    # inline at the tail of _commit_node_tree.) A divergence raises -> per-sheet failed[].
    _reconcile_grid(sheet_name, grid_doc.name, grid_rows)

    # CLEAR any prior commit-failure stamp on success (Slice F1) -- folded into the
    # per-sheet commit below so the clear is ATOMIC with this successful commit (a
    # re-commit that now succeeds drops its stale failure notice).
    _clear_commit_failure(boq_name, sheet_name)

    frappe.db.commit()

    return {
        "sheet_name": sheet_name,
        "disposition": disposition,
        "sheet_disposition": sheet_disposition,
        "grid_name": grid_doc.name,
        "boq_sheet_name": boq_sheet_name,
        "commit_version": commit_version,
        "row_count": len(grid_rows),
        "froze_prior": froze_grid,                  # grid froze (3a back-compat key)
        "froze_prior_sheet": bool(prior_sheet_names),
        "node_count": node_result["node_count"],
        "froze_nodes": node_result["froze_nodes"],
    }


def _write_grid(
    boq_name: str,
    sheet_name: str,
    sheet_disposition: str,
    grid_rows: list[dict],
    commit_version: int,
    committed_at: str,
) -> tuple[Any, bool]:
    """Write the new faithful grid (is_current=1) under the SHARED commit_version,
    freezing the prior current via set_value. Verbatim (#152) (boq, source_sheet_name)
    match on the freeze lookup. Returns (grid_doc, froze_prior)."""
    prior_current = _current_names(_GRID_DOCTYPE, boq_name, "source_sheet_name", sheet_name)
    for name in prior_current:
        frappe.db.set_value(_GRID_DOCTYPE, name, "is_current", 0)

    grid = frappe.new_doc(_GRID_DOCTYPE)
    grid.boq = boq_name
    grid.source_sheet_name = sheet_name  # verbatim, no strip (#152)
    grid.sheet_disposition = sheet_disposition
    grid.commit_version = commit_version
    grid.is_current = 1
    grid.committed_at = committed_at
    for order, row in enumerate(grid_rows):
        grid.append("rows", {
            "row_number": row["row_number"],
            "row_order": order,
            "cells": row["cells"],  # dict-JSON child field accepts a Python dict
        })
    grid.insert(ignore_permissions=True)

    return grid, bool(prior_current)


def _write_committed_boq_sheet(
    boq_name: str,
    sheet_name: str,
    disposition: str,
    draft: Any,
    commit_version: int,
    committed_at: str,
) -> tuple[str, list[str]]:
    """Version the committed BoQ Sheet: FREEZE the prior current, INSERT a fresh
    current (is_current=1) under the SHARED commit_version. Returns (new_name,
    prior_current_names).

    FREEZE-AND-SUPERSEDE (Slice 3b) -- 3a's raw-delete is RETIRED. The prior sheet
    is frozen via frappe.db.set_value("BoQ Sheet", name, "is_current", 0), a TARGETED
    column write. We must NOT doc.save() the prior sheet: BoQ Sheet.area_dimensions
    is a list-valued JSON field that Frappe re-parses to a Python list on load, and
    doc.save()'s get_valid_dict would reject it ("cannot be a list") -- the same wall
    the 3a delete_doc hit. set_value bypasses full-doc serialization. The prior sheet
    + its nodes thus survive as a coherent frozen version (no delete, no orphan).

    The column-config snapshot (column_role_map + column_headers + area_dimensions,
    plus header_row / header_row_count / treat_as) is pinned VERBATIM from the
    draft's sheet_config at commit time.
    """
    prior_current = _current_names(_SHEET_DOCTYPE, boq_name, "sheet_name", sheet_name)
    for name in prior_current:
        frappe.db.set_value(_SHEET_DOCTYPE, name, "is_current", 0)

    cfg = _coerce_config(getattr(draft, "sheet_config", None) if draft else None)

    bs = frappe.new_doc(_SHEET_DOCTYPE)
    bs.boq = boq_name
    bs.sheet_name = sheet_name  # verbatim, no strip (#152)
    bs.sheet_order = (getattr(draft, "sheet_order", None) if draft else None) or 0
    bs.sheet_label = getattr(draft, "sheet_label", None) if draft else None
    bs.treat_as = _DISPOSITION_TO_TREAT_AS[disposition]
    bs.header_row = cfg.get("header_row")
    bs.header_row_count = cfg.get("header_row_count") or 1

    # JSON snapshot fields. area_dimensions is a LIST -- json.dumps it (a JSON field
    # rejects a raw Python list on insert); dumps the dict ones too for uniformity.
    bs.column_role_map = json.dumps(cfg.get("column_role_map") or {})
    bs.column_headers = json.dumps(cfg.get("column_headers") or {})
    bs.area_dimensions = json.dumps(cfg.get("area_dimensions") or [])

    bs.commit_version = commit_version
    bs.is_current = 1
    bs.committed_at = committed_at

    # Work-header assignments carried from the draft.
    if draft is not None:
        for wp in (getattr(draft, "work_packages", None) or []):
            wh = getattr(wp, "work_header", None)
            if wh:
                bs.append("work_packages", {"work_header": wh})

    bs.insert(ignore_permissions=True)
    return bs.name, prior_current


# ---------------------------------------------------------------------------
# Node tree (Slice 3b)
# ---------------------------------------------------------------------------

_JSON_FIELDS_TO_PARSE = (
    "qty_by_area", "rate_by_area", "amount_by_area",
    "attached_notes", "append_notes_raw", "edit_log",
)


def _parse_json_fields(d: dict) -> dict:
    """db.get_all returns JSON fields as strings; parse the ones the node body needs."""
    for f in _JSON_FIELDS_TO_PARSE:
        v = d.get(f)
        if isinstance(v, str) and v:
            try:
                d[f] = json.loads(v)
            except (ValueError, TypeError):
                pass
    return d


def _commit_node_tree(
    boq_name: str,
    sheet_name: str,
    boq_sheet_name: str,
    prior_sheet_names: list[str],
    commit_version: int,
    committed_at: str,
) -> dict:
    """Build + persist the committed node tree for a FINALIZED sheet (Slice 3b).

    Reads the sheet's BoQ Review Rows, resolves the effective tree, and writes EVERY
    classified row EXCEPT spacer as a BOQ Node (X): priceable rows keep node_type
    Preamble / Line Item, non-priceable committed rows (note / subtotal_marker /
    header_repeat) become node_type Other; row_class carries the full taxonomy. Per-area
    BOQ Node Qty By Area children are exploded for priceable rows (non-priceable rows
    have none). Three passes (see module docstring). Carries reviewed values VERBATIM
    (capture-only). Returns {"node_count", "froze_nodes"}.
    """
    # 0. Freeze the prior commit's current nodes (attached to the now-frozen prior
    #    BoQ Sheet[s]). set_value only -- never touch parent_node; v1 nodes stay a
    #    coherent frozen snapshot under frozen sheet v1.
    froze_nodes = 0
    for ps in prior_sheet_names:
        for n in _current_names(_NODE_DOCTYPE, boq_name, "sheet", ps):
            frappe.db.set_value(_NODE_DOCTYPE, n, "is_current", 0)
            froze_nodes += 1

    # 1. Read review rows (verbatim sheet_name, #152) + resolve effective values.
    raw_rows = frappe.db.get_all(
        "BoQ Review Row",
        filters={"boq": boq_name, "sheet_name": sheet_name},
        fields=_REVIEW_ROW_FIELDS,
        order_by="row_index asc",
    )
    node_rows: list[tuple[dict, dict]] = []   # (row_dict, eff) for node rows only
    eff_parent_by_idx: dict[int, Any] = {}
    for r in raw_rows:
        d = _parse_json_fields(dict(r))
        eff = resolve_effective(d)
        if eff["effective_classification"] in _GRID_ONLY_CLASSIFICATIONS:
            continue  # spacer -> grid-only (pure layout, captured in the faithful grid)
        node_rows.append((d, eff))
        eff_parent_by_idx[d["row_index"]] = eff["effective_parent_index"]

    # Level-less preambles: precompute their assigned level from the whole sheet's
    # effective tree (children / shallowest-defined-level), Phase-5 guard fix. The real
    # commit is SILENT -- the returned level_warnings (old #22 msgprint) are DISCARDED
    # here; they surface only in the pre-commit preflight (validate_node_plan).
    assigned_levels, _level_warnings = _compute_levelless_preamble_levels(
        sheet_name, node_rows, eff_parent_by_idx
    )

    # 2. PASS 1 -- insert every node PARENT-LESS, with NO list-valued JSON field set
    #    (attached_notes / edit_log deferred to pass 3 so pass-2 doc.save() is safe).
    docs_by_idx: dict[int, Any] = {}
    name_by_idx: dict[int, str] = {}
    for d, eff in node_rows:
        node = _build_node_pass1(
            boq_sheet_name, d, eff, commit_version, committed_at, assigned_levels
        )
        node.insert(ignore_permissions=True)
        docs_by_idx[d["row_index"]] = node
        name_by_idx[d["row_index"]] = node.name

    # 3. PASS 2 -- wire parents via doc.save() in ancestor-depth order, so each
    #    parent's path is final before its children read it AND the controller's
    #    parent-chain validation re-runs (a free integrity check). Roots (or a row
    #    whose effective parent is not itself a node) skip -- path=name from pass-1
    #    after_insert. doc.save() is list-JSON-safe because attached_notes/edit_log
    #    are still null on the held doc.
    depth_by_idx = _node_depths(eff_parent_by_idx, name_by_idx)
    for idx in sorted(name_by_idx, key=lambda i: depth_by_idx[i]):
        eff_parent = eff_parent_by_idx.get(idx)
        if eff_parent is None or eff_parent not in name_by_idx:
            continue
        node = docs_by_idx[idx]
        node.parent_node = name_by_idx[eff_parent]
        node.save(ignore_permissions=True)

    # 4. PASS 3 -- write the DEFERRED list-valued JSON fields via set_value(json.dumps).
    #    Targeted column writes bypass get_valid_dict's "cannot be a list". (Dict-valued
    #    append_notes_raw was set in pass 1 -- only LISTs trip the wall.)
    for d, _eff in node_rows:
        name = name_by_idx[d["row_index"]]
        updates: dict[str, str] = {}
        att = d.get("attached_notes")
        if att:
            updates["attached_notes"] = json.dumps(att)
        elog = d.get("edit_log")
        if elog:
            updates["edit_log"] = json.dumps(elog)
        if updates:
            frappe.db.set_value(_NODE_DOCTYPE, name, updates, update_modified=False)

    # 5. OUTPUT-FIDELITY RECONCILIATION (Slice 2): verify every just-written node faithfully
    #    equals what the commit PRODUCED, BEFORE the per-sheet commit. Reuses the in-hand
    #    maps (no re-query; no re-run of resolve_effective / the level functions). A
    #    divergence raises -> commit_boq's per-sheet isolation rolls back + records failed[].
    _reconcile_node_tree(
        sheet_name, boq_sheet_name, commit_version,
        node_rows, eff_parent_by_idx, name_by_idx, docs_by_idx,
    )

    return {"node_count": len(node_rows), "froze_nodes": froze_nodes}


def _build_node_pass1(
    boq_sheet_name: str,
    d: dict,
    eff: dict,
    commit_version: int,
    committed_at: str,
    assigned_levels: dict,
) -> Any:
    """Build a parent-less BOQ Nodes doc from a resolved review row (pass 1).

    node_type/level mapping, the P4-3 verbatim field mapping (word-order reversal,
    Float->Currency left to Frappe), flags + human-layer-as-provenance + provenance
    ids + versioning, and the per-area child explosion. Does NOT set the list-valued
    JSON fields (attached_notes / edit_log) -- those land in pass 3.

    assigned_levels: {row_index: level} for level-less preambles, precomputed by
    _compute_levelless_preamble_levels (needs the whole sheet's children/min-level).
    """
    node = frappe.new_doc(_NODE_DOCTYPE)
    node.sheet = boq_sheet_name  # boq auto-fills from the sheet (P4-2 sync-guard)

    cls = eff["effective_classification"]
    node.row_class = cls  # X: full taxonomy axis (carried verbatim for every node)
    # node_type + level come from the SHARED derivation (commit_validation), so the real
    # commit and the pre-commit preflight can never diverge. Preamble -> ("Preamble",
    # real-or-assigned level); Line Item -> ("Line Item", None: level must NOT be set, the
    # controller throws if it is); else (note/subtotal_marker/header_repeat) -> ("Other",
    # None). level is set ONLY when non-None, so a Line Item / Other node stays level-less.
    node_type, level = derive_node_type_and_level(d, eff, assigned_levels)
    node.node_type = node_type
    if level is not None:
        node.level = level

    # Identity / content (P4-3: sl_no_value->code, row_index->sort_order).
    node.code = d.get("sl_no_value")
    node.sort_order = d.get("row_index")
    node.source_row_number = d.get("source_row_number")
    # description is reqd by the controller; abort-guard an empty one with a placeholder.
    node.description = d.get("description") or d.get("sl_no_value") or "(untitled)"
    node.unit = d.get("unit")
    node.make_model = d.get("make_model")

    # qty stays Float. A Line Item requires qty (0 for rate-only items).
    qty = d.get("qty_total")
    if node.node_type == "Line Item" and qty is None:
        qty = 0
    node.qty = qty

    # Money: word-order reversal, Float -> Currency (Frappe coerces; no manual round).
    node.supply_rate = d.get("rate_supply")
    node.install_rate = d.get("rate_install")
    node.combined_rate = d.get("rate_combined")
    node.supply_amount = d.get("amount_supply")
    node.install_amount = d.get("amount_install")
    node.total_amount = d.get("amount_total")

    # Flags carried VERBATIM.
    node.is_rate_only = 1 if d.get("is_rate_only") else 0
    node.is_synthetic = 1 if d.get("is_synthetic") else 0

    # HUMAN LAYER -- carried as PROVENANCE ONLY. The committed tier NEVER branches on
    # these; the LIVE wiring is parent_node + node_type + level (the resolved effective
    # values from resolve_effective). human_* are written, never read by commit logic.
    node.human_classification = d.get("human_classification")
    hp = d.get("human_parent")
    node.human_parent = hp if hp is not None else -1  # -1 sentinel (agreement #54)
    node.human_is_root = 1 if d.get("human_is_root") else 0

    # Notes (Option A -- verbatim, NO aggregation / up-roll / cross-flow).
    node.notes = d.get("row_notes")
    apn = d.get("append_notes_raw")
    if apn:
        node.append_notes_raw = apn  # dict-valued JSON -- safe at insert + pass-2 save

    # Provenance.
    node.review_row_name = d.get("name")              # human-readable tie (may dangle)
    node.commit_provenance_id = frappe.generate_hash()  # durable key

    # Versioning (shared with grid + sheet for this commit).
    node.commit_version = commit_version
    node.is_current = 1
    node.committed_at = committed_at

    # Per-area children (dual-shape; NEVER downgrade granularity).
    for child in _explode_area_children(
        d.get("qty_by_area"), d.get("rate_by_area"), d.get("amount_by_area")
    ):
        node.append("qty_by_area", child)

    return node


def _explode_area_children(qty_ba: Any, rate_ba: Any, amount_ba: Any) -> list[dict]:
    """Explode the review row's per-area JSON into BOQ Node Qty By Area child dicts.

    DUAL-SHAPE, NEVER DOWNGRADE (owner-locked): per-area supply/install granularity
    is a real business requirement and MUST survive the commit.
      qty_by_area    -- always flat {area: float}            -> child.qty
      rate_by_area   -- NESTED {area: {supply_rate, install_rate, combined_rate}}
                        -> each preserved on the matching child column;
                        FLAT bare scalar {area: float} -> combined_rate ONLY
                        (deliberate default; supply/install left empty, not invented)
      amount_by_area -- NESTED {area: {supply, install, total}}
                        -> supply_amount / install_amount / total_amount each preserved;
                        FLAT bare scalar {area: float} -> total_amount ONLY (deliberate)
    area_name is the dict key, VERBATIM (#152). Area order = first-seen across
    qty -> rate -> amount.
    """
    qty_ba = qty_ba if isinstance(qty_ba, dict) else {}
    rate_ba = rate_ba if isinstance(rate_ba, dict) else {}
    amount_ba = amount_ba if isinstance(amount_ba, dict) else {}

    areas: list[str] = []
    seen: set[str] = set()
    for src in (qty_ba, rate_ba, amount_ba):
        for a in src:
            if a not in seen:
                seen.add(a)
                areas.append(a)

    children: list[dict] = []
    for area in areas:
        child: dict[str, Any] = {"area_name": area}

        # The child's qty is reqd=1 (BOQ Node Qty By Area schema). An area present only
        # in rate/amount (not qty_by_area) is a real zero-qty area in the breakdown ->
        # default qty 0.0 (satisfies the constraint without inventing a quantity).
        qv = qty_ba.get(area)
        child["qty"] = qv if isinstance(qv, (int, float)) else 0.0

        rv = rate_ba.get(area)
        if isinstance(rv, dict):  # NESTED -- preserve each kind present
            if rv.get("supply_rate") is not None:
                child["supply_rate"] = rv["supply_rate"]
            if rv.get("install_rate") is not None:
                child["install_rate"] = rv["install_rate"]
            if rv.get("combined_rate") is not None:
                child["combined_rate"] = rv["combined_rate"]
        elif isinstance(rv, (int, float)):  # FLAT scalar -> combined only
            child["combined_rate"] = rv

        av = amount_ba.get(area)
        if isinstance(av, dict):  # NESTED -- preserve each kind present
            if av.get("supply") is not None:
                child["supply_amount"] = av["supply"]
            if av.get("install") is not None:
                child["install_amount"] = av["install"]
            if av.get("total") is not None:
                child["total_amount"] = av["total"]
        elif isinstance(av, (int, float)):  # FLAT scalar -> total only
            child["total_amount"] = av

        children.append(child)

    return children


def _node_depths(eff_parent_by_idx: dict[int, Any], name_by_idx: dict[int, str]) -> dict[int, int]:
    """Depth of each node row in the effective tree (root = 0), for parent-wire
    ordering. Cycle-safe + missing-parent-safe: an idx whose effective parent is None,
    not itself a node, or part of a cycle is treated as depth 0."""
    depth: dict[int, int] = {}

    def _d(idx: int, stack: frozenset) -> int:
        if idx in depth:
            return depth[idx]
        p = eff_parent_by_idx.get(idx)
        if p is None or p not in name_by_idx or p in stack:
            depth[idx] = 0
            return 0
        depth[idx] = _d(p, stack | {idx}) + 1
        return depth[idx]

    for idx in name_by_idx:
        _d(idx, frozenset())
    return depth


# ---------------------------------------------------------------------------
# Output-fidelity reconciliation (Slice 2)
# ---------------------------------------------------------------------------
# Runs on EVERY per-sheet commit, at the tail of the write path, BEFORE the per-sheet
# frappe.db.commit(). It verifies ONE thing: that the value each transform PRODUCED was
# COPIED to the DB faithfully -- OUTPUT-FIDELITY, not logic-correctness. It does NOT judge
# whether a produced value is *correct* (valid parent, no cycles, sane rates, declared
# area, level-monotonicity = VALIDITY = tendering's job = OUT OF SCOPE).
#
#   COPIED fields  -> compare stored == transform(review-row source), applying the rename
#                     + coercion + the deterministic transform; numerics rounded to the
#                     field's own precision (so float->Currency coercion never false-flags).
#   DERIVED fields (EXACTLY TWO) -> compare stored == the CAPTURED producer output:
#     parent_node = name_by_idx.get(eff_parent_by_idx.get(idx))  -- the captured eff-parent
#       map + insertion record (the exact lookup pass-2 used); NEVER re-runs resolve_effective.
#     level       = docs_by_idx[idx].level  -- the in-memory value the producer built;
#       NEVER re-runs the level functions. (Thin tripwire: nothing mutates level today, so
#       it cannot fail under current code -- kept for near-zero cost regression value.)
#
# A divergence frappe.throws a specific message; commit_boq's per-sheet try/except then
# rolls back + records {sheet_name, reason} in failed[] (Slice-5 isolation, unchanged).
# Source values are REUSED from node_rows / grid_rows already in hand (no re-query drift);
# stored values are read fresh in-transaction (Frappe sees the just-written, uncommitted rows).

_NODE_MONEY_FIELDS = (
    "supply_rate", "install_rate", "combined_rate",
    "supply_amount", "install_amount", "total_amount",
)
# Node money field -> its review-row source (P4-3 word-order reversal).
_NODE_MONEY_SOURCE = {
    "supply_rate": "rate_supply", "install_rate": "rate_install",
    "combined_rate": "rate_combined",
    "supply_amount": "amount_supply", "install_amount": "amount_install",
    "total_amount": "amount_total",
}
_CHILD_MONEY_FIELDS = (
    "supply_rate", "install_rate", "combined_rate",
    "supply_amount", "install_amount", "total_amount",
)


def _text_eq(a: Any, b: Any) -> bool:
    """Text compare treating None and '' as equally 'absent' (Frappe may store either)."""
    return (a or None) == (b or None)


def _json_norm(v: Any) -> Any:
    """Normalize a JSON value for comparison: parse a JSON string, and treat
    None / '' / {} / [] as the single 'empty' value None."""
    if isinstance(v, str):
        try:
            v = frappe.parse_json(v)
        except Exception:
            pass
    if v is None or v == "" or v == {} or v == []:
        return None
    return v


def _json_eq(a: Any, b: Any) -> bool:
    return _json_norm(a) == _json_norm(b)


def _node_type_for(cls: str) -> str:
    """Produced node_type for an effective classification (mirrors _build_node_pass1)."""
    if cls == _PREAMBLE_CLS:
        return "Preamble"
    if cls == _LINE_ITEM_CLS:
        return "Line Item"
    return "Other"


def _reconcile_node_tree(
    sheet_name: str,
    boq_sheet_name: str,
    commit_version: int,
    node_rows: list,
    eff_parent_by_idx: dict,
    name_by_idx: dict,
    docs_by_idx: dict,
) -> None:
    """OUTPUT-FIDELITY check for the committed node tree (+ per-area children).

    Raises frappe.ValidationError on the first divergent node. Reuses the in-hand maps
    (no re-query of the review rows; no re-run of resolve_effective / the level functions).
    """
    flt = frappe.utils.flt
    money_prec = {f: frappe.get_precision(_NODE_DOCTYPE, f) for f in _NODE_MONEY_FIELDS}
    qty_prec = frappe.get_precision(_NODE_DOCTYPE, "qty")
    child_money_prec = {
        f: frappe.get_precision("BOQ Node Qty By Area", f) for f in _CHILD_MONEY_FIELDS
    }
    child_qty_prec = frappe.get_precision("BOQ Node Qty By Area", "qty")

    # COUNT: every produced node row must have a current stored node (catches a finalized
    # sheet that should have nodes but persisted none).
    stored_count = frappe.db.count(
        _NODE_DOCTYPE,
        {"sheet": boq_sheet_name, "is_current": 1, "commit_version": commit_version},
    )
    if stored_count != len(node_rows):
        frappe.throw(
            f"Commit reconciliation: sheet {sheet_name!r} produced {len(node_rows)} "
            f"node(s) but {stored_count} were committed.",
            title="Commit reconciliation failed",
        )

    read_fields = [
        "code", "sort_order", "source_row_number", "description", "unit", "make_model",
        "node_type", "row_class", "qty", "parent_node", "level",
        "is_rate_only", "is_synthetic",
        "human_classification", "human_parent", "human_is_root",
        "notes", "append_notes_raw", "attached_notes", "edit_log",
        *_NODE_MONEY_FIELDS,
    ]

    for d, eff in node_rows:
        idx = d["row_index"]
        name = name_by_idx[idx]
        stored = frappe.db.get_value(_NODE_DOCTYPE, name, read_fields, as_dict=True)
        cls = eff["effective_classification"]
        node_type = _node_type_for(cls)
        mism: list[str] = []

        def _t(field, expected):
            if not _text_eq(stored.get(field), expected):
                mism.append(f"{field}: produced {expected!r} != stored {stored.get(field)!r}")

        def _exact(field, expected):
            if stored.get(field) != expected:
                mism.append(f"{field}: produced {expected!r} != stored {stored.get(field)!r}")

        def _num(field, expected, prec):
            if flt(expected or 0, prec) != flt(stored.get(field) or 0, prec):
                mism.append(f"{field}: produced {expected!r} != stored {stored.get(field)!r}")

        def _jsn(field, expected):
            if not _json_eq(stored.get(field), expected):
                mism.append(f"{field}: produced {expected!r} != stored {stored.get(field)!r}")

        # --- COPIED scalars (transform(source)) ---
        _t("code", d.get("sl_no_value"))
        _exact("sort_order", d.get("row_index"))
        _exact("source_row_number", d.get("source_row_number"))
        _t("description", d.get("description") or d.get("sl_no_value") or "(untitled)")
        _t("unit", d.get("unit"))
        _t("make_model", d.get("make_model"))
        _exact("node_type", node_type)
        _t("row_class", cls)
        _exact("is_rate_only", 1 if d.get("is_rate_only") else 0)
        _exact("is_synthetic", 1 if d.get("is_synthetic") else 0)
        _exact("human_is_root", 1 if d.get("human_is_root") else 0)
        _t("human_classification", d.get("human_classification"))
        hp = d.get("human_parent")
        _exact("human_parent", hp if hp is not None else -1)  # -1 sentinel (N-2)
        _t("notes", d.get("row_notes"))  # row_notes -> notes rename

        # qty: Line-Item None -> 0; others carry source as-is (N-4). Float precision.
        qty_expected = d.get("qty_total")
        if node_type == "Line Item" and qty_expected is None:
            qty_expected = 0
        _num("qty", qty_expected, qty_prec)

        # money: word-order reversal, Currency precision.
        for nf in _NODE_MONEY_FIELDS:
            _num(nf, d.get(_NODE_MONEY_SOURCE[nf]), money_prec[nf])

        # JSON carried fields (empty {}/[]/None all normalize to absent).
        _jsn("append_notes_raw", d.get("append_notes_raw"))
        _jsn("attached_notes", d.get("attached_notes"))
        _jsn("edit_log", d.get("edit_log"))

        # --- DERIVED: parent_node = captured eff-parent map (NOT re-run resolve_effective) ---
        expected_parent = name_by_idx.get(eff_parent_by_idx.get(idx))
        _exact("parent_node", expected_parent)

        # --- DERIVED: level = captured in-memory producer doc (thin tripwire) ---
        produced_level = docs_by_idx[idx].level
        if (produced_level or 0) != (stored.get("level") or 0):
            mism.append(
                f"level: produced {produced_level!r} != stored {stored.get('level')!r}"
            )

        # --- per-area children (deterministic explosion -> verify persisted faithfully) ---
        expected_children = _explode_area_children(
            d.get("qty_by_area"), d.get("rate_by_area"), d.get("amount_by_area")
        )
        stored_children = frappe.db.get_all(
            "BOQ Node Qty By Area",
            filters={"parent": name},
            fields=["area_name", "qty", *_CHILD_MONEY_FIELDS],
        )
        if len(expected_children) != len(stored_children):
            mism.append(
                f"qty_by_area: produced {len(expected_children)} child(ren) != "
                f"stored {len(stored_children)}"
            )
        else:
            stored_by_area = {c["area_name"]: c for c in stored_children}
            for ec in expected_children:
                sc = stored_by_area.get(ec["area_name"])
                if sc is None:
                    mism.append(
                        f"qty_by_area: produced area {ec['area_name']!r} missing in stored"
                    )
                    continue
                if flt(ec.get("qty") or 0, child_qty_prec) != flt(sc.get("qty") or 0, child_qty_prec):
                    mism.append(
                        f"qty_by_area[{ec['area_name']}].qty: produced {ec.get('qty')!r} "
                        f"!= stored {sc.get('qty')!r}"
                    )
                # FLAT-stays-flat: an absent kind in produced normalizes to 0 == stored 0.
                for cf in _CHILD_MONEY_FIELDS:
                    if flt(ec.get(cf) or 0, child_money_prec[cf]) != flt(sc.get(cf) or 0, child_money_prec[cf]):
                        mism.append(
                            f"qty_by_area[{ec['area_name']}].{cf}: produced {ec.get(cf)!r} "
                            f"!= stored {sc.get(cf)!r}"
                        )

        if mism:
            frappe.throw(
                f"Commit reconciliation: sheet {sheet_name!r} row_index {idx} (node {name}) "
                f"diverged from the produced values -- " + "; ".join(mism[:8]),
                title="Commit reconciliation failed",
            )


def _reconcile_grid(sheet_name: str, grid_name: str, grid_rows: list) -> None:
    """OUTPUT-FIDELITY check for the committed faithful grid: the persisted grid rows
    must equal the extracted grid_rows (row_number + cells) in order. Raises on divergence."""
    stored = frappe.db.get_all(
        "BoQ Committed Sheet Grid Row",
        filters={"parent": grid_name},
        fields=["row_number", "row_order", "cells"],
        order_by="row_order asc",
    )
    if len(stored) != len(grid_rows):
        frappe.throw(
            f"Commit reconciliation: sheet {sheet_name!r} grid produced {len(grid_rows)} "
            f"row(s) but {len(stored)} were committed.",
            title="Commit reconciliation failed",
        )
    stored_by_order = {s["row_order"]: s for s in stored}
    for order, gr in enumerate(grid_rows):
        s = stored_by_order.get(order)
        if s is None:
            frappe.throw(
                f"Commit reconciliation: sheet {sheet_name!r} grid row_order {order} missing.",
                title="Commit reconciliation failed",
            )
        if s["row_number"] != gr["row_number"]:
            frappe.throw(
                f"Commit reconciliation: sheet {sheet_name!r} grid row_order {order} "
                f"row_number produced {gr['row_number']} != stored {s['row_number']}.",
                title="Commit reconciliation failed",
            )
        if _json_norm(s["cells"]) != _json_norm(gr["cells"]):
            frappe.throw(
                f"Commit reconciliation: sheet {sheet_name!r} grid row {gr['row_number']} "
                f"cells diverged from the extracted grid.",
                title="Commit reconciliation failed",
            )
