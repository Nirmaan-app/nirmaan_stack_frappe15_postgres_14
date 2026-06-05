"""
Tests for parse_run.py (Slice 1 + Slice 2).

Four test groups:
  TestAssembleMappingConfig  -- DB tests: assemble_mapping_config correctness (incl. FIX 1)
  TestFlattenFaithfulness    -- pure-Python: flatten_resolved_row / flatten_parsed_boq
  TestBoQReviewRowRoundTrip  -- DB tests: insert + read-back via Frappe ORM
  TestRunParseWorker         -- DB tests: _run_parse_worker lifecycle (Slice 2)
"""
import json
import os
import tempfile
import unittest
from pathlib import Path

import frappe
from frappe.tests.utils import FrappeTestCase

from nirmaan_stack.services.boq_parser.config import (
    ColumnRole,
    MappingConfig,
    MasterBoqMetadata,
    SheetConfig,
)
from nirmaan_stack.services.boq_parser.orchestrator import parse_boq
from nirmaan_stack.services.boq_parser.tests.fixtures.generate_synthetic import generate_all

from nirmaan_stack.api.boq.wizard.parse_run import (
    _LIST_JSON_FIELDS,
    _publish_parse_event,
    _run_parse_worker,
    assemble_mapping_config,
    flatten_parsed_boq,
    flatten_resolved_row,
    run_parse,
)
from nirmaan_stack.api.boq.wizard.update_sheet_draft import (
    set_sheet_config,
    set_sheet_status,
)

_FIXTURES = Path(__file__).parent.parent.parent.parent / "services" / "boq_parser" / "tests" / "fixtures"


def _p(name: str) -> str:
    return str(_FIXTURES / name)


# ---------------------------------------------------------------------------
# Fixture configs (mirrors test_orchestrator.py)
# ---------------------------------------------------------------------------

def _simple_config(boq_name: str = "test_boq", project: str = "test") -> MappingConfig:
    return MappingConfig(
        project=project,
        master_boq=MasterBoqMetadata(boq_name=boq_name),
        sheets=[SheetConfig(
            sheet_name="Sheet1",
            header_row=1,
            column_role_map={
                "A": ColumnRole(role="sl_no"),
                "B": ColumnRole(role="description"),
                "C": ColumnRole(role="unit"),
                "D": ColumnRole(role="qty"),
                "E": ColumnRole(role="rate_supply"),
                "F": ColumnRole(role="amount_supply"),
            },
        )],
    )


def _multi_area_config(boq_name: str = "test_boq", project: str = "test") -> MappingConfig:
    return MappingConfig(
        project=project,
        master_boq=MasterBoqMetadata(boq_name=boq_name),
        sheets=[SheetConfig(
            sheet_name="Multi Area",
            header_row=1,
            area_dimensions=["Floor 1", "Floor 2"],
            column_role_map={
                "A": ColumnRole(role="sl_no"),
                "B": ColumnRole(role="description"),
                "C": ColumnRole(role="unit"),
                "D": ColumnRole(role="qty", area="Floor 1"),
                "E": ColumnRole(role="qty", area="Floor 2"),
                "F": ColumnRole(role="qty_total"),
                "G": ColumnRole(role="rate_supply"),
                "H": ColumnRole(role="amount_by_area", area="Floor 1"),
                "I": ColumnRole(role="amount_by_area", area="Floor 2"),
                "J": ColumnRole(role="amount_total"),
            },
        )],
    )


# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

def _make_project():
    proj = frappe.new_doc("Projects")
    proj.project_name = f"TEST_PARSE_RUN_{frappe.generate_hash(length=6)}"
    proj.project_start_date = frappe.utils.now()[:19]
    proj.project_end_date = frappe.utils.add_to_date(frappe.utils.now()[:19], years=1)[:19]
    proj.project_scopes = {"scopes": []}
    proj.insert(ignore_permissions=True)
    frappe.db.commit()
    return proj


def _cleanup_project(project_name: str):
    for boq in frappe.get_all("BOQs", filters={"project": project_name}, fields=["name"]):
        frappe.db.delete("BoQ Review Row", {"boq": boq.name})
        frappe.delete_doc("BOQs", boq.name, force=True, ignore_permissions=True)
    frappe.delete_doc("Projects", project_name, force=True, ignore_permissions=True)
    frappe.db.commit()


# ---------------------------------------------------------------------------
# Group 1: assemble_mapping_config
# ---------------------------------------------------------------------------

class TestAssembleMappingConfig(FrappeTestCase):
    """
    Verifies that assemble_mapping_config correctly routes each wizard_status
    to the right SheetConfig treatment and collects not_eligible entries.
    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.test_project = _make_project()

    @classmethod
    def tearDownClass(cls):
        _cleanup_project(cls.test_project.name)
        super().tearDownClass()

    def _make_boq_with_sheets(self):
        sc = SheetConfig(
            sheet_name="Sheet1",
            header_row=1,
            column_role_map={"A": ColumnRole(role="sl_no"), "B": ColumnRole(role="description")},
        )
        blob = json.dumps(sc.model_dump())

        boq = frappe.new_doc("BOQs")
        boq.project = self.__class__.test_project.name
        boq.boq_name = f"Assemble Test {frappe.generate_hash(length=4)}"
        boq.tax_treatment = "Pre-tax"
        boq.notes = "test notes"
        boq.append("general_specs_sheets", {"source_sheet_name": "Sheet5", "preamble_text": ""})
        boq.append("sheet_drafts", {
            "sheet_name": "Sheet1", "sheet_order": 1,
            "wizard_status": "Reviewed", "sheet_config": blob,
        })
        boq.append("sheet_drafts", {
            "sheet_name": "Sheet2", "sheet_order": 2,
            "wizard_status": "Hidden",
        })
        boq.append("sheet_drafts", {
            "sheet_name": "Sheet3", "sheet_order": 3,
            "wizard_status": "Skip",
        })
        boq.append("sheet_drafts", {
            "sheet_name": "Sheet4", "sheet_order": 4,
            "wizard_status": "Pending",
        })
        boq.append("sheet_drafts", {
            "sheet_name": "Sheet5", "sheet_order": 5,
            "wizard_status": "Reviewed",
        })
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        return boq

    def tearDown(self):
        for boq in frappe.get_all(
            "BOQs",
            filters={"project": self.__class__.test_project.name},
            fields=["name"],
        ):
            frappe.delete_doc("BOQs", boq.name, force=True, ignore_permissions=True)
        frappe.db.commit()

    def test_reviewed_sheet_included_as_data(self):
        """A Reviewed sheet with a valid blob is included as a data SheetConfig."""
        boq = self._make_boq_with_sheets()
        config, not_eligible = assemble_mapping_config(boq.name)
        reviewed = next((s for s in config.sheets if s.sheet_name == "Sheet1"), None)
        self.assertIsNotNone(reviewed)
        self.assertFalse(reviewed.skip)
        self.assertEqual(reviewed.treat_as, "data")

    def test_hidden_sheet_included_with_skip_true(self):
        """A Hidden sheet gets skip=True."""
        boq = self._make_boq_with_sheets()
        config, _ = assemble_mapping_config(boq.name)
        hidden = next((s for s in config.sheets if s.sheet_name == "Sheet2"), None)
        self.assertIsNotNone(hidden)
        self.assertTrue(hidden.skip)

    def test_skip_sheet_included_with_skip_true(self):
        """A Skip sheet gets skip=True."""
        boq = self._make_boq_with_sheets()
        config, _ = assemble_mapping_config(boq.name)
        skipped = next((s for s in config.sheets if s.sheet_name == "Sheet3"), None)
        self.assertIsNotNone(skipped)
        self.assertTrue(skipped.skip)

    def test_pending_sheet_excluded_and_in_not_eligible(self):
        """A Pending sheet is excluded from the config and collected in not_eligible."""
        boq = self._make_boq_with_sheets()
        config, not_eligible = assemble_mapping_config(boq.name)
        self.assertIn("Sheet4", not_eligible)
        sheet_names = [s.sheet_name for s in config.sheets]
        self.assertNotIn("Sheet4", sheet_names)

    def test_general_specs_sheet_treated_as_master_preamble(self):
        """The sheet pointed to by BOQs.general_specs_sheet gets treat_as=master_preamble."""
        boq = self._make_boq_with_sheets()
        config, _ = assemble_mapping_config(boq.name)
        gs = next((s for s in config.sheets if s.sheet_name == "Sheet5"), None)
        self.assertIsNotNone(gs)
        self.assertEqual(gs.treat_as, "master_preamble")

    def test_mapping_config_validates_successfully(self):
        """The assembled MappingConfig passes Pydantic validation without error."""
        boq = self._make_boq_with_sheets()
        config, not_eligible = assemble_mapping_config(boq.name)
        MappingConfig.model_validate(config.model_dump())
        self.assertEqual(config.master_boq.boq_name, boq.boq_name)
        self.assertEqual(config.master_boq.tax_treatment, "Pre-tax")
        self.assertEqual(config.master_boq.notes, "test notes")

    def test_config_has_correct_sheet_count(self):
        """Exactly 4 SheetConfigs: 1 data + 2 skip + 1 master_preamble."""
        boq = self._make_boq_with_sheets()
        config, _ = assemble_mapping_config(boq.name)
        self.assertEqual(len(config.sheets), 4)

    def test_parsed_status_included_like_reviewed(self):
        """A 'Parsed' sheet (next lifecycle state after Reviewed) is treated as data."""
        sc = SheetConfig(
            sheet_name="SheetParsed",
            header_row=1,
            column_role_map={"A": ColumnRole(role="sl_no")},
        )
        blob = json.dumps(sc.model_dump())

        boq = frappe.new_doc("BOQs")
        boq.project = self.__class__.test_project.name
        boq.boq_name = f"Parsed Test {frappe.generate_hash(length=4)}"
        boq.tax_treatment = "Pre-tax"
        boq.append("sheet_drafts", {
            "sheet_name": "SheetParsed", "sheet_order": 1,
            "wizard_status": "Parsed", "sheet_config": blob,
        })
        boq.insert(ignore_permissions=True)
        frappe.db.commit()

        config, not_eligible = assemble_mapping_config(boq.name)
        parsed_sc = next((s for s in config.sheets if s.sheet_name == "SheetParsed"), None)
        self.assertIsNotNone(parsed_sc)
        self.assertFalse(parsed_sc.skip)
        self.assertEqual(parsed_sc.treat_as, "data")
        self.assertNotIn("SheetParsed", not_eligible)

    def test_reviewed_without_config_blob_goes_to_not_eligible(self):
        """Reviewed but no sheet_config blob -> not_eligible, not a hard error."""
        sc = SheetConfig(
            sheet_name="OtherSheet",
            header_row=1,
            column_role_map={"A": ColumnRole(role="sl_no")},
        )
        boq = frappe.new_doc("BOQs")
        boq.project = self.__class__.test_project.name
        boq.boq_name = f"No Blob Test {frappe.generate_hash(length=4)}"
        boq.tax_treatment = "Pre-tax"
        boq.append("sheet_drafts", {
            "sheet_name": "OtherSheet", "sheet_order": 1,
            "wizard_status": "Reviewed", "sheet_config": json.dumps(sc.model_dump()),
        })
        boq.append("sheet_drafts", {
            "sheet_name": "NoBlob", "sheet_order": 2,
            "wizard_status": "Reviewed",
        })
        boq.insert(ignore_permissions=True)
        frappe.db.commit()

        config, not_eligible = assemble_mapping_config(boq.name)
        self.assertIn("NoBlob", not_eligible)
        sheet_names = [s.sheet_name for s in config.sheets]
        self.assertNotIn("NoBlob", sheet_names)

    def test_skip_sheet_designated_as_general_specs_routes_to_master_preamble(self):
        """
        Rule-order fix: a sheet with wizard_status='Skip' that is in the
        general_specs_sheets child table must route to treat_as='master_preamble',
        NOT skip=True.  Mirrors real-data case: BOQ-26-00145 sheet 'SOW'
        (wizard_status='Skip', designated via child table).
        """
        boq = frappe.new_doc("BOQs")
        boq.project = self.__class__.test_project.name
        boq.boq_name = f"Skip+Pointer Test {frappe.generate_hash(length=4)}"
        boq.tax_treatment = "Pre-tax"
        boq.append("general_specs_sheets", {"source_sheet_name": "SOW", "preamble_text": ""})
        # SOW is stored as Skip (as it would be after hub Skip designation before
        # the user later designates it as general-specs via the child table)
        boq.append("sheet_drafts", {
            "sheet_name": "SOW", "sheet_order": 1,
            "wizard_status": "Skip",
        })
        # Include one data sheet so assemble_mapping_config does not raise
        sc_blob = json.dumps({
            "area_dimensions": [], "column_role_map": {"A": {"role": "sl_no", "area": None}},
            "header_row": 1, "header_row_count": 1,
            "skip_top_rows_after_header": [], "top_header_rows_override": None,
        })
        boq.append("sheet_drafts", {
            "sheet_name": "Data Sheet", "sheet_order": 2,
            "wizard_status": "Reviewed", "sheet_config": sc_blob,
        })
        boq.insert(ignore_permissions=True)
        frappe.db.commit()

        config, not_eligible = assemble_mapping_config(boq.name)

        sow_sc = next((s for s in config.sheets if s.sheet_name == "SOW"), None)
        self.assertIsNotNone(sow_sc, "SOW not found in config.sheets")
        self.assertFalse(
            sow_sc.skip,
            "SOW routed as skip=True; pointer should outrank wizard_status='Skip'",
        )
        self.assertEqual(
            sow_sc.treat_as, "master_preamble",
            f"SOW treat_as={sow_sc.treat_as!r}; expected 'master_preamble'",
        )
        self.assertNotIn("SOW", not_eligible)

    def test_unknown_boq_raises_validation_error(self):
        with self.assertRaises(frappe.ValidationError):
            assemble_mapping_config("BOQ-DOES-NOT-EXIST-99999")

    def test_fix1_production_blob_without_sheet_name_is_eligible(self):
        """
        FIX 1: A realistic 6-key production blob (no 'sheet_name' key) must be ELIGIBLE.

        Production wizard blobs saved by set_sheet_config have exactly these keys:
          area_dimensions, column_role_map, header_row, header_row_count,
          skip_top_rows_after_header, top_header_rows_override
        They NEVER contain 'sheet_name', 'skip', or 'treat_as' (verified live on
        BOQ-26-00150 and BOQ-26-00145).  Without the injection added in FIX 1,
        SheetConfig.model_validate raises a ValidationError (sheet_name is required with
        no default) and the sheet falls into not_eligible silently.
        """
        production_blob = json.dumps({
            "area_dimensions": [],
            "column_role_map": {
                "A": {"role": "sl_no", "area": None},
                "B": {"role": "description", "area": None},
                "C": {"role": "unit", "area": None},
                "D": {"role": "qty", "area": None},
                "E": {"role": "rate_supply", "area": None},
                "F": {"role": "amount_supply", "area": None},
            },
            "header_row": 2,
            "header_row_count": 1,
            "skip_top_rows_after_header": [],
            "top_header_rows_override": None,
        })
        boq = frappe.new_doc("BOQs")
        boq.project = self.__class__.test_project.name
        boq.boq_name = f"Fix1 Test {frappe.generate_hash(length=4)}"
        boq.tax_treatment = "Pre-tax"
        boq.append("sheet_drafts", {
            "sheet_name": "Electrical BOQ",
            "sheet_order": 1,
            "wizard_status": "Reviewed",
            "sheet_config": production_blob,
        })
        boq.insert(ignore_permissions=True)
        frappe.db.commit()

        config, not_eligible = assemble_mapping_config(boq.name)

        self.assertNotIn(
            "Electrical BOQ", not_eligible,
            "FIX 1 regression: production-shape blob (no sheet_name) landed in not_eligible",
        )
        sc = next((s for s in config.sheets if s.sheet_name == "Electrical BOQ"), None)
        self.assertIsNotNone(sc, "Electrical BOQ not in config.sheets")
        self.assertEqual(sc.sheet_name, "Electrical BOQ")
        self.assertFalse(sc.skip)
        self.assertEqual(sc.treat_as, "data")
        self.assertEqual(sc.header_row, 2)

    def test_two_general_specs_sheets_both_route_to_master_preamble(self):
        """Two entries in general_specs_sheets route both sheets to treat_as='master_preamble'."""
        sc = SheetConfig(
            sheet_name="DataSheet",
            header_row=1,
            column_role_map={"A": ColumnRole(role="sl_no")},
        )
        blob = json.dumps(sc.model_dump())
        boq = frappe.new_doc("BOQs")
        boq.project = self.__class__.test_project.name
        boq.boq_name = f"Two GS Test {frappe.generate_hash(length=4)}"
        boq.tax_treatment = "Pre-tax"
        boq.append("general_specs_sheets", {"source_sheet_name": "SOW", "preamble_text": ""})
        boq.append("general_specs_sheets", {"source_sheet_name": "General Notes", "preamble_text": ""})
        boq.append("sheet_drafts", {"sheet_name": "SOW", "sheet_order": 1, "wizard_status": "Skip"})
        boq.append("sheet_drafts", {"sheet_name": "General Notes", "sheet_order": 2, "wizard_status": "Skip"})
        boq.append("sheet_drafts", {
            "sheet_name": "DataSheet", "sheet_order": 3,
            "wizard_status": "Reviewed", "sheet_config": blob,
        })
        boq.insert(ignore_permissions=True)
        frappe.db.commit()

        # Verify two distinct source_sheet_names exist in child table
        gss_rows = frappe.db.get_all(
            "BoQ General Specs Sheet",
            filters={"parent": boq.name, "parenttype": "BOQs"},
            fields=["source_sheet_name"],
        )
        self.assertEqual(len(gss_rows), 2)
        self.assertEqual({r.source_sheet_name for r in gss_rows}, {"SOW", "General Notes"})

        # Both sheets must route to master_preamble
        config, _ = assemble_mapping_config(boq.name)
        sow_sc = next((s for s in config.sheets if s.sheet_name == "SOW"), None)
        self.assertIsNotNone(sow_sc)
        self.assertEqual(sow_sc.treat_as, "master_preamble")
        gn_sc = next((s for s in config.sheets if s.sheet_name == "General Notes"), None)
        self.assertIsNotNone(gn_sc)
        self.assertEqual(gn_sc.treat_as, "master_preamble")


# ---------------------------------------------------------------------------
# Group 2: flatten faithfulness (pure Python -- no DB needed)
# ---------------------------------------------------------------------------

class TestFlattenFaithfulness(unittest.TestCase):
    """
    Verifies that flatten_resolved_row and flatten_parsed_boq faithfully map
    all parser output fields to the expected BoQ Review Row field dict.
    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        generate_all()

    # ---------------------------------------------------------------- #
    # Case (i): single-area -- synthetic_simple.xlsx                   #
    # ---------------------------------------------------------------- #

    def test_single_area_row_count_matches_resolved_rows(self):
        """flatten_parsed_boq produces one dict per resolved_row across all sheets."""
        parsed = parse_boq(_p("synthetic_simple.xlsx"), _simple_config())
        flat = flatten_parsed_boq(parsed, "FAKE-BOQ-001")
        total_resolved = sum(len(ps.resolved_rows) for ps in parsed.sheets)
        self.assertEqual(len(flat), total_resolved)

    def test_single_area_classification_preserved(self):
        """classification field holds the RowClassification .value string."""
        parsed = parse_boq(_p("synthetic_simple.xlsx"), _simple_config())
        flat = flatten_parsed_boq(parsed, "FAKE-BOQ-001")
        classifications = {d["classification"] for d in flat}
        self.assertTrue(classifications.issubset({
            "preamble", "line_item", "note", "subtotal_marker", "spacer", "header_repeat"
        }))
        self.assertIn("line_item", classifications)

    def test_single_area_source_row_number_is_positive_int(self):
        """source_row_number is a positive int for every flattened row."""
        parsed = parse_boq(_p("synthetic_simple.xlsx"), _simple_config())
        flat = flatten_parsed_boq(parsed, "FAKE-BOQ-001")
        for d in flat:
            self.assertIsInstance(d["source_row_number"], int)
            self.assertGreater(d["source_row_number"], 0)

    def test_single_area_row_index_sequential(self):
        """row_index is 0-based and sequential within each sheet's resolved_rows."""
        parsed = parse_boq(_p("synthetic_simple.xlsx"), _simple_config())
        flat = flatten_parsed_boq(parsed, "FAKE-BOQ-001")
        indices = [d["row_index"] for d in flat]
        self.assertEqual(indices, list(range(len(indices))))

    def test_single_area_sheet_name_injected(self):
        """All flattened dicts carry the correct sheet_name from ParsedSheet."""
        parsed = parse_boq(_p("synthetic_simple.xlsx"), _simple_config())
        flat = flatten_parsed_boq(parsed, "FAKE-BOQ-001")
        for d in flat:
            self.assertEqual(d["sheet_name"], "Sheet1")

    def test_single_area_boq_name_injected(self):
        """flatten_parsed_boq injects the boq_name into every row dict."""
        parsed = parse_boq(_p("synthetic_simple.xlsx"), _simple_config())
        flat = flatten_parsed_boq(parsed, "FAKE-BOQ-001")
        for d in flat:
            self.assertEqual(d["boq"], "FAKE-BOQ-001")

    def test_single_area_parent_index_and_path_coherent(self):
        """For rows with a real parent (parent_index >= 0), path contains a slash."""
        parsed = parse_boq(_p("synthetic_simple.xlsx"), _simple_config())
        flat = flatten_parsed_boq(parsed, "FAKE-BOQ-001")
        for d in flat:
            # -1 is the "no parent" sentinel (root rows); skip them.
            if d["parent_index"] not in (None, -1):
                self.assertIsInstance(d["parent_index"], int)
                self.assertIn("/", d["path"])

    def test_single_area_description_for_known_items(self):
        """The three known line items are in the flattened output."""
        parsed = parse_boq(_p("synthetic_simple.xlsx"), _simple_config())
        flat = flatten_parsed_boq(parsed, "FAKE-BOQ-001")
        descriptions = [d["description"] for d in flat if d["classification"] == "line_item"]
        self.assertIn("First item", descriptions)
        self.assertIn("Second item", descriptions)
        self.assertIn("Bold item", descriptions)

    def test_single_area_scalar_rate_supply_preserved(self):
        """rate_supply for 'First item' (E2=100) survives the flatten."""
        parsed = parse_boq(_p("synthetic_simple.xlsx"), _simple_config())
        flat = flatten_parsed_boq(parsed, "FAKE-BOQ-001")
        first = next(d for d in flat if d.get("description") == "First item")
        self.assertEqual(first["rate_supply"], 100.0)

    # ---------------------------------------------------------------- #
    # Case (ii): multi-area -- synthetic_multi_area.xlsx               #
    # ---------------------------------------------------------------- #

    def test_multi_area_qty_by_area_is_dict(self):
        """qty_by_area for a multi-area line item is a Python dict."""
        parsed = parse_boq(_p("synthetic_multi_area.xlsx"), _multi_area_config())
        flat = flatten_parsed_boq(parsed, "FAKE-BOQ-002")
        line_items = [d for d in flat if d["classification"] == "line_item"]
        self.assertGreater(len(line_items), 0)
        for li in line_items:
            self.assertIsInstance(li["qty_by_area"], dict)

    def test_multi_area_qty_by_area_keys_are_area_names(self):
        """qty_by_area keys match the declared area names."""
        parsed = parse_boq(_p("synthetic_multi_area.xlsx"), _multi_area_config())
        flat = flatten_parsed_boq(parsed, "FAKE-BOQ-002")
        painting = next(d for d in flat if d.get("description") == "Painting works")
        self.assertIn("Floor 1", painting["qty_by_area"])
        self.assertIn("Floor 2", painting["qty_by_area"])
        self.assertAlmostEqual(painting["qty_by_area"]["Floor 1"], 5.0)
        self.assertAlmostEqual(painting["qty_by_area"]["Floor 2"], 3.0)

    def test_multi_area_amount_by_area_preserved(self):
        """amount_by_area for 'Painting works' carries the per-area amounts."""
        parsed = parse_boq(_p("synthetic_multi_area.xlsx"), _multi_area_config())
        flat = flatten_parsed_boq(parsed, "FAKE-BOQ-002")
        painting = next(d for d in flat if d.get("description") == "Painting works")
        self.assertAlmostEqual(painting["amount_by_area"]["Floor 1"], 500.0)
        self.assertAlmostEqual(painting["amount_by_area"]["Floor 2"], 300.0)

    def test_multi_area_tiling_has_validation_warning(self):
        """
        'Tiling works' (Excel row 3) has qty sum 8 vs total 10 (diff=2 > +/-1).
        The validation_warning must survive into the flattened dict.
        """
        parsed = parse_boq(_p("synthetic_multi_area.xlsx"), _multi_area_config())
        flat = flatten_parsed_boq(parsed, "FAKE-BOQ-002")
        tiling = next(d for d in flat if d.get("description") == "Tiling works")
        self.assertIsInstance(tiling["validation_warnings"], list)
        self.assertGreater(
            len(tiling["validation_warnings"]), 0,
            "Expected qty mismatch warning on 'Tiling works' but got empty list",
        )
        self.assertTrue(
            any("qty" in w.lower() for w in tiling["validation_warnings"]),
            f"No qty-related warning found: {tiling['validation_warnings']}",
        )

    def test_multi_area_clean_row_has_no_validation_warnings(self):
        """'Painting works' (row 2) is clean -- validation_warnings should be empty."""
        parsed = parse_boq(_p("synthetic_multi_area.xlsx"), _multi_area_config())
        flat = flatten_parsed_boq(parsed, "FAKE-BOQ-002")
        painting = next(d for d in flat if d.get("description") == "Painting works")
        self.assertEqual(painting["validation_warnings"], [])

    # ---------------------------------------------------------------- #
    # Case (iii): programmatic needs_classification_review flag        #
    # ---------------------------------------------------------------- #

    def test_flagged_row_review_flag_survives_flatten(self):
        """
        No fixture produces needs_classification_review=True (recon-confirmed --
        the flag requires a PREAMBLE with children AND a price signal, which the
        synthetic fixtures do not produce). We construct a ResolvedRow
        programmatically to verify the flag + reason survive the flatten.
        """
        from nirmaan_stack.services.boq_parser.classifier import ClassifiedRow, RowClassification
        from nirmaan_stack.services.boq_parser.hierarchy import ResolvedRow
        from nirmaan_stack.services.boq_parser.reader import CellInfo, RawRow

        dummy_cell = CellInfo(
            value=None, formula=None, is_formula=False,
            is_merged_origin=False, merged_range=None,
            font_bold=False, fill_color_rgb=None, indent=0,
        )
        raw = RawRow(row_number=5, cells={"A": dummy_cell})
        cr = ClassifiedRow(raw_row=raw, classification=RowClassification.PREAMBLE)
        rr = ResolvedRow(
            classified_row=cr,
            parent_index=None,
            level=1,
            path="2",
            needs_classification_review=True,
            review_reason="priced_preamble_with_children",
        )

        d = flatten_resolved_row(rr, "SheetX", 2)

        self.assertTrue(d["needs_classification_review"])
        self.assertEqual(d["review_reason"], "priced_preamble_with_children")
        self.assertEqual(d["classification"], "preamble")
        self.assertEqual(d["source_row_number"], 5)
        self.assertEqual(d["row_index"], 2)
        self.assertEqual(d["path"], "2")

    def test_is_synthetic_false_by_default(self):
        """Parser never sets is_synthetic=True; the flatten maps False correctly."""
        from nirmaan_stack.services.boq_parser.classifier import ClassifiedRow, RowClassification
        from nirmaan_stack.services.boq_parser.hierarchy import ResolvedRow
        from nirmaan_stack.services.boq_parser.reader import RawRow

        raw = RawRow(row_number=1, cells={})
        cr = ClassifiedRow(raw_row=raw, classification=RowClassification.SPACER)
        rr = ResolvedRow(classified_row=cr)

        d = flatten_resolved_row(rr, "S", 0)
        self.assertFalse(d["is_synthetic"])

    def test_human_parent_is_minus1_sentinel_for_root_row(self):
        """
        flatten_resolved_row must write human_parent=-1 for a root row (parent_index=None).

        Regression for agreement #54 (the missing half): Frappe coerces unset Int fields
        to 0 on insert, and 0 is a valid row index. resolve_effective treats
        human_parent >= 0 as a real override, so human_parent=0 falsely reports every
        row as parented to row 0, collapsing the review tree.
        """
        from nirmaan_stack.services.boq_parser.classifier import ClassifiedRow, RowClassification
        from nirmaan_stack.services.boq_parser.hierarchy import ResolvedRow
        from nirmaan_stack.services.boq_parser.reader import RawRow

        raw = RawRow(row_number=1, cells={})
        cr = ClassifiedRow(raw_row=raw, classification=RowClassification.PREAMBLE)
        rr = ResolvedRow(classified_row=cr, parent_index=None, level=0, path="")

        d = flatten_resolved_row(rr, "Sheet1", 0)
        self.assertEqual(
            d["human_parent"], -1,
            "flatten_resolved_row must write human_parent=-1 for a root row (agreement #54)",
        )

    def test_human_parent_is_minus1_sentinel_for_child_row(self):
        """
        flatten_resolved_row must write human_parent=-1 regardless of structural parent.

        A freshly-parsed child row (parent_index=3) has no human override; human_parent
        must be the -1 sentinel, not 0. Confirms the fix applies to all rows, not just roots.
        """
        from nirmaan_stack.services.boq_parser.classifier import ClassifiedRow, RowClassification
        from nirmaan_stack.services.boq_parser.hierarchy import ResolvedRow
        from nirmaan_stack.services.boq_parser.reader import RawRow

        raw = RawRow(row_number=5, cells={})
        cr = ClassifiedRow(raw_row=raw, classification=RowClassification.LINE_ITEM)
        rr = ResolvedRow(classified_row=cr, parent_index=3, level=2, path="2/3")

        d = flatten_resolved_row(rr, "Sheet1", 4)
        self.assertEqual(
            d["human_parent"], -1,
            "flatten_resolved_row must write human_parent=-1 for child rows (no human override at parse time)",
        )
        # structural parent_index must still be correctly 3
        self.assertEqual(d["parent_index"], 3)

    def test_json_fields_are_python_objects_not_strings(self):
        """
        JSON fields in the flattened dict must be Python objects (list/dict),
        NOT pre-serialized JSON strings. Frappe's JSON fieldtype handles serialization.
        """
        parsed = parse_boq(_p("synthetic_multi_area.xlsx"), _multi_area_config())
        flat = flatten_parsed_boq(parsed, "FAKE-BOQ-003")
        for d in flat:
            for field_name in (
                "attached_notes", "validation_warnings", "classifier_warnings",
                "preamble_candidate_signals",
            ):
                self.assertIsInstance(
                    d[field_name], list,
                    f"{field_name} must be a Python list, got {type(d[field_name])}",
                )
            for field_name in ("qty_by_area", "amount_by_area", "rate_by_area", "append_notes_raw"):
                self.assertIsInstance(
                    d[field_name], dict,
                    f"{field_name} must be a Python dict, got {type(d[field_name])}",
                )


# ---------------------------------------------------------------------------
# Group 3: optional DB round-trip insert + read-back
# ---------------------------------------------------------------------------

class TestBoQReviewRowRoundTrip(FrappeTestCase):
    """
    Inserts a handful of flattened dicts as BoQ Review Row records and asserts
    Frappe can save and read them back with the expected field values.
    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        generate_all()
        cls.test_project = _make_project()
        boq = frappe.new_doc("BOQs")
        boq.project = cls.test_project.name
        boq.boq_name = "Round Trip Test BoQ"
        boq.tax_treatment = "Pre-tax"
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        cls.boq_name = boq.name

    @classmethod
    def tearDownClass(cls):
        frappe.db.delete("BoQ Review Row", {"boq": cls.boq_name})
        frappe.db.commit()
        _cleanup_project(cls.test_project.name)
        super().tearDownClass()

    def setUp(self):
        frappe.db.delete("BoQ Review Row", {"boq": self.boq_name})
        frappe.db.commit()

    def _insert_rows(self, flat_rows: list[dict]) -> list[str]:
        names = []
        for orig_dict in flat_rows:
            row_dict = dict(orig_dict)  # shallow copy -- don't mutate caller's dict
            for key in _LIST_JSON_FIELDS:  # module-level constant from parse_run
                if isinstance(row_dict.get(key), list):
                    row_dict[key] = json.dumps(row_dict[key])
            doc = frappe.new_doc("BoQ Review Row")
            doc.update(row_dict)
            doc.insert(ignore_permissions=True)
            names.append(doc.name)
        frappe.db.commit()
        return names

    def test_insert_simple_rows_and_read_back_classification(self):
        """Inserting flattened rows from synthetic_simple.xlsx and reading back works."""
        parsed = parse_boq(_p("synthetic_simple.xlsx"), _simple_config())
        flat = flatten_parsed_boq(parsed, self.boq_name)
        names = self._insert_rows(flat[:3])
        self.assertEqual(len(names), 3)

        for name in names:
            doc = frappe.get_doc("BoQ Review Row", name)
            self.assertEqual(doc.boq, self.boq_name)
            self.assertEqual(doc.sheet_name, "Sheet1")
            self.assertIn(doc.classification, [
                "preamble", "line_item", "note",
                "subtotal_marker", "spacer", "header_repeat",
            ])
            self.assertIsNotNone(doc.source_row_number)

    def test_insert_multi_area_row_and_read_back_json_fields(self):
        """A multi-area line item row can be inserted and its JSON fields read back."""
        parsed = parse_boq(_p("synthetic_multi_area.xlsx"), _multi_area_config())
        flat = flatten_parsed_boq(parsed, self.boq_name)
        painting_dict = next(d for d in flat if d.get("description") == "Painting works")
        names = self._insert_rows([painting_dict])

        doc = frappe.get_doc("BoQ Review Row", names[0])
        self.assertEqual(doc.description, "Painting works")
        self.assertEqual(doc.classification, "line_item")

        qty_by_area = doc.qty_by_area
        if isinstance(qty_by_area, str):
            qty_by_area = json.loads(qty_by_area)
        self.assertIn("Floor 1", qty_by_area)
        self.assertAlmostEqual(qty_by_area["Floor 1"], 5.0)

    def test_insert_row_with_validation_warning_and_read_back(self):
        """The validation_warnings list on 'Tiling works' survives DB round-trip."""
        parsed = parse_boq(_p("synthetic_multi_area.xlsx"), _multi_area_config())
        flat = flatten_parsed_boq(parsed, self.boq_name)
        tiling_dict = next(d for d in flat if d.get("description") == "Tiling works")
        names = self._insert_rows([tiling_dict])

        doc = frappe.get_doc("BoQ Review Row", names[0])
        warnings = doc.validation_warnings
        if isinstance(warnings, str):
            warnings = json.loads(warnings)
        self.assertIsInstance(warnings, list)
        self.assertGreater(len(warnings), 0)

    def test_insert_flagged_row_and_read_back_review_fields(self):
        """needs_classification_review=True + review_reason survive DB round-trip."""
        from nirmaan_stack.services.boq_parser.classifier import ClassifiedRow, RowClassification
        from nirmaan_stack.services.boq_parser.hierarchy import ResolvedRow
        from nirmaan_stack.services.boq_parser.reader import RawRow

        raw = RawRow(row_number=7, cells={})
        cr = ClassifiedRow(raw_row=raw, classification=RowClassification.PREAMBLE)
        rr = ResolvedRow(
            classified_row=cr,
            needs_classification_review=True,
            review_reason="priced_preamble_with_children",
            level=1,
            path="3",
        )
        d = flatten_resolved_row(rr, "TestSheet", 3)
        d["boq"] = self.boq_name
        names = self._insert_rows([d])

        doc = frappe.get_doc("BoQ Review Row", names[0])
        self.assertEqual(doc.needs_classification_review, 1)
        self.assertEqual(doc.review_reason, "priced_preamble_with_children")

    def test_human_parent_sentinel_root_resolves_effective_parent_none(self):
        """
        Regression for agreement #54 (DB round-trip): after insert, a root row's
        human_parent must be stored as -1 (not 0). resolve_effective must translate
        -1 -> None for both parent_index and human_parent, giving
        effective_parent_index = None (root), not 0 (false override).

        This was the live-cert failure on BOQ-26-00145: human_parent defaulted to 0
        (Frappe Int coercion on unset field), resolve_effective treated 0 as
        'parent is row 0', flattening the entire review tree.
        """
        from nirmaan_stack.services.boq_parser.classifier import ClassifiedRow, RowClassification
        from nirmaan_stack.services.boq_parser.hierarchy import ResolvedRow
        from nirmaan_stack.services.boq_parser.reader import RawRow
        from nirmaan_stack.api.boq.wizard.review_screen import resolve_effective

        raw = RawRow(row_number=2, cells={})
        cr = ClassifiedRow(raw_row=raw, classification=RowClassification.PREAMBLE)
        rr = ResolvedRow(classified_row=cr, parent_index=None, level=0, path="")
        d = flatten_resolved_row(rr, "TestSheet", 0)
        d["boq"] = self.boq_name
        names = self._insert_rows([d])

        doc = frappe.get_doc("BoQ Review Row", names[0])
        self.assertEqual(
            doc.human_parent, -1,
            "human_parent stored as 0 after insert (Frappe Int coercion); should be -1",
        )
        eff = resolve_effective(doc)
        self.assertIsNone(
            eff["effective_parent_index"],
            f"Root row effective_parent_index should be None, got {eff['effective_parent_index']!r}",
        )


# ---------------------------------------------------------------------------
# Group 4: _run_parse_worker lifecycle (Slice 2)
# ---------------------------------------------------------------------------

# Minimal production-shape blob (6 keys, NO sheet_name) matching what the wizard saves.
# FIX 1 in assemble_mapping_config injects sheet_name before model_validate.
_PROD_BLOB_TMPL = {
    "area_dimensions": [],
    "column_role_map": {
        "A": {"role": "sl_no", "area": None},
        "B": {"role": "description", "area": None},
        "C": {"role": "unit", "area": None},
        "D": {"role": "qty", "area": None},
        "E": {"role": "rate_supply", "area": None},
        "F": {"role": "amount_supply", "area": None},
    },
    "header_row": 1,
    "header_row_count": 1,
    "skip_top_rows_after_header": [],
    "top_header_rows_override": None,
}


class TestRunParseWorker(FrappeTestCase):
    """
    Tests for _run_parse_worker: status lifecycle, row insert/replace, subset parse,
    general-specs handling, pending reporting, forced failure path.

    Uses a tiny 3-sheet workbook (SheetA, SheetB, SOW) built in setUpClass.
    source_file_url is set to the local tempfile path -- triggers the local-fetch branch
    in _fetch_boq_file_to_tempfile (no S3 dependency in tests).
    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        generate_all()
        cls.test_project = _make_project()

        # Build a small 3-sheet workbook: SheetA + SheetB (data) + SOW (general-specs)
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "SheetA"
        ws1.append(["SL", "Description", "Unit", "Qty", "Rate", "Amount"])
        ws1.append([1, "Alpha Item", "nos", 2, 100.0, 200.0])
        ws2 = wb.create_sheet("SheetB")
        ws2.append(["SL", "Description", "Unit", "Qty", "Rate", "Amount"])
        ws2.append([1, "Beta Item", "nos", 3, 50.0, 150.0])
        ws_sow = wb.create_sheet("SOW")
        ws_sow.append(["General Specifications"])
        ws_sow.append(["1. All materials to meet IS standards."])

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        tmp.close()
        cls._worker_wb_path = tmp.name

    @classmethod
    def tearDownClass(cls):
        try:
            os.unlink(cls._worker_wb_path)
        except OSError:
            pass
        _cleanup_project(cls.test_project.name)
        super().tearDownClass()

    def tearDown(self):
        for boq in frappe.get_all(
            "BOQs", filters={"project": self.__class__.test_project.name}, fields=["name"]
        ):
            frappe.db.delete("BoQ Review Row", {"boq": boq.name})
            frappe.delete_doc("BOQs", boq.name, force=True, ignore_permissions=True)
        frappe.db.commit()

    def _make_boq(
        self,
        include_b: bool = False,
        include_sow: bool = False,
        sheet_a_status: str = "Reviewed",
        sheet_b_status: str = "Reviewed",
    ):
        """Create a BOQs with SheetA + optional SheetB / SOW. Blob uses production 6-key shape."""
        blob = json.dumps(_PROD_BLOB_TMPL)
        boq = frappe.new_doc("BOQs")
        boq.project = self.__class__.test_project.name
        boq.boq_name = f"Worker Test {frappe.generate_hash(length=6)}"
        boq.tax_treatment = "Pre-tax"
        boq.source_file_url = self.__class__._worker_wb_path
        if include_sow:
            boq.append("general_specs_sheets", {"source_sheet_name": "SOW", "preamble_text": ""})
        boq.append("sheet_drafts", {
            "sheet_name": "SheetA", "sheet_order": 1,
            "wizard_status": sheet_a_status, "sheet_config": blob,
        })
        if include_b:
            boq.append("sheet_drafts", {
                "sheet_name": "SheetB", "sheet_order": 2,
                "wizard_status": sheet_b_status, "sheet_config": blob,
            })
        if include_sow:
            boq.append("sheet_drafts", {
                "sheet_name": "SOW", "sheet_order": 3,
                "wizard_status": "Reviewed",
                # no sheet_config: Rule 2 (master_preamble) fires before Rule 3 (blob check)
            })
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        return boq

    def _run(self, boq_name, sheet_names=None):
        """Call _run_parse_worker synchronously (bypasses enqueue for test speed)."""
        _run_parse_worker(boq_name, sheet_names=sheet_names, user="Administrator")

    # -- status lifecycle ------------------------------------------------

    def test_reviewed_becomes_parsed_on_success(self):
        """Reviewed sheet's wizard_status transitions to 'Parsed' after a successful parse."""
        boq = self._make_boq()
        self._run(boq.name)
        boq.reload()
        self.assertEqual(
            next(d for d in boq.sheet_drafts if d.sheet_name == "SheetA").wizard_status,
            "Parsed",
        )

    def test_rows_inserted_on_success(self):
        """At least one BoQ Review Row is inserted for a successfully parsed data sheet."""
        boq = self._make_boq()
        self._run(boq.name)
        count = frappe.db.count("BoQ Review Row", {"boq": boq.name, "sheet_name": "SheetA"})
        self.assertGreater(count, 0)

    def test_parsed_at_stamped_on_success(self):
        """BOQs.parsed_at is set after at least one sheet parses successfully."""
        boq = self._make_boq()
        self._run(boq.name)
        parsed_at = frappe.db.get_value("BOQs", boq.name, "parsed_at")
        self.assertIsNotNone(parsed_at, "BOQs.parsed_at was not set after successful parse")

    def test_failure_path_sets_parse_failed_status(self):
        """When parse_boq raises, all eligible sheets transition to 'Parse failed'."""
        from unittest.mock import patch
        boq = self._make_boq()
        with patch(
            "nirmaan_stack.api.boq.wizard.parse_run.parse_boq",
            side_effect=RuntimeError("forced failure for test"),
        ):
            self._run(boq.name)  # must NOT raise; worker handles it gracefully

        boq.reload()
        self.assertEqual(
            next(d for d in boq.sheet_drafts if d.sheet_name == "SheetA").wizard_status,
            "Parse failed",
        )

    # -- re-parse / replace ----------------------------------------------

    def test_re_parse_does_not_duplicate_rows(self):
        """Parsing the same sheet twice produces the same row count, not doubled."""
        boq = self._make_boq()
        self._run(boq.name)
        count_1 = frappe.db.count("BoQ Review Row", {"boq": boq.name, "sheet_name": "SheetA"})

        # Second parse: sheet is now "Parsed" (treated same as Reviewed by assemble_mapping_config)
        self._run(boq.name)
        count_2 = frappe.db.count("BoQ Review Row", {"boq": boq.name, "sheet_name": "SheetA"})

        self.assertEqual(count_1, count_2, "Re-parse duplicated rows instead of replacing them")
        self.assertGreater(count_2, 0)

    # -- subset parse ----------------------------------------------------

    def test_subset_parse_only_affects_specified_sheet(self):
        """
        run_parse with sheet_names=['SheetA'] inserts/replaces SheetA rows only;
        a pre-existing sentinel row for SheetB is left untouched.
        """
        boq = self._make_boq(include_b=True)

        # Insert a sentinel row for SheetB
        sentinel = frappe.new_doc("BoQ Review Row")
        sentinel.boq = boq.name
        sentinel.sheet_name = "SheetB"
        sentinel.source_row_number = 999
        sentinel.row_index = 0
        sentinel.classification = "spacer"
        for field in _LIST_JSON_FIELDS:
            setattr(sentinel, field, json.dumps([]))
        sentinel.insert(ignore_permissions=True)
        frappe.db.commit()
        sentinel_name = sentinel.name

        self._run(boq.name, sheet_names=["SheetA"])

        # SheetA: rows present
        self.assertGreater(
            frappe.db.count("BoQ Review Row", {"boq": boq.name, "sheet_name": "SheetA"}), 0
        )
        # SheetB sentinel: untouched
        self.assertTrue(
            frappe.db.exists("BoQ Review Row", sentinel_name),
            "Subset parse deleted SheetB sentinel (should only touch SheetA)",
        )

    # -- pending / not-eligible ------------------------------------------

    def test_pending_sheet_is_not_parsed_and_produces_no_rows(self):
        """A Pending sheet remains Pending, produces no rows, and does not prevent other parses."""
        boq = self._make_boq(include_b=True, sheet_b_status="Pending")
        self._run(boq.name)

        # SheetA parsed, SheetB skipped
        self.assertGreater(
            frappe.db.count("BoQ Review Row", {"boq": boq.name, "sheet_name": "SheetA"}), 0
        )
        self.assertEqual(
            frappe.db.count("BoQ Review Row", {"boq": boq.name, "sheet_name": "SheetB"}), 0
        )
        boq.reload()
        self.assertEqual(
            next(d for d in boq.sheet_drafts if d.sheet_name == "SheetB").wizard_status,
            "Pending",
            "Pending sheet wizard_status was modified by the worker",
        )

    # -- general-specs / master-preamble ---------------------------------

    def test_general_specs_sheet_produces_no_rows(self):
        """master_preamble sheet (general_specs_sheet='SOW') emits no BoQ Review Rows."""
        boq = self._make_boq(include_sow=True)
        self._run(boq.name)

        self.assertEqual(
            frappe.db.count("BoQ Review Row", {"boq": boq.name, "sheet_name": "SOW"}),
            0,
            "SOW sheet produced BoQ Review Rows (must not -- it is master_preamble)",
        )
        # SheetA (data sheet) still parsed
        self.assertGreater(
            frappe.db.count("BoQ Review Row", {"boq": boq.name, "sheet_name": "SheetA"}), 0
        )

    def test_master_preamble_written_for_general_specs_sheet(self):
        """Preamble text stored in general_specs_sheets child row; SOW produces no data rows."""
        boq = self._make_boq(include_sow=True)
        self._run(boq.name)

        rows = frappe.db.get_all(
            "BoQ General Specs Sheet",
            filters={"parent": boq.name, "parenttype": "BOQs", "source_sheet_name": "SOW"},
            fields=["preamble_text"],
        )
        self.assertEqual(len(rows), 1, "Expected exactly 1 BoQ General Specs Sheet row for SOW")
        preamble_text = rows[0].preamble_text
        self.assertTrue(preamble_text, "preamble_text is empty or None")
        self.assertIn(
            "IS standards", preamble_text,
            f"Expected SOW text not found in preamble_text: {preamble_text!r}",
        )
        # general-specs sheet must never produce BoQ Review Rows
        self.assertEqual(
            frappe.db.count("BoQ Review Row", {"boq": boq.name, "sheet_name": "SOW"}),
            0,
            "SOW sheet produced BoQ Review Rows (must not -- it is master_preamble)",
        )

    def test_master_preamble_written_when_general_specs_sheet_has_skip_status(self):
        """
        Rule-order fix (real-data regression): SOW in general_specs_sheets child table
        but wizard_status='Skip' -- worker must still extract and store preamble.

        Differs from test_master_preamble_written_for_general_specs_sheet: that test
        creates SOW with wizard_status='Reviewed'. This uses wizard_status='Skip', which
        is the real-data shape from BOQ-26-00145 and is what the rule-order bug muted.
        """
        blob = json.dumps(_PROD_BLOB_TMPL)
        boq = frappe.new_doc("BOQs")
        boq.project = self.__class__.test_project.name
        boq.boq_name = f"Skip+Pointer Worker {frappe.generate_hash(length=6)}"
        boq.tax_treatment = "Pre-tax"
        boq.source_file_url = self.__class__._worker_wb_path
        boq.append("general_specs_sheets", {"source_sheet_name": "SOW", "preamble_text": ""})
        boq.append("sheet_drafts", {
            "sheet_name": "SheetA", "sheet_order": 1,
            "wizard_status": "Reviewed", "sheet_config": blob,
        })
        # SOW stored as "Skip" -- mirrors real data; designation in child table does not
        # change wizard_status per M2.16 (status is derived, not stored).
        boq.append("sheet_drafts", {
            "sheet_name": "SOW", "sheet_order": 3,
            "wizard_status": "Skip",
        })
        boq.insert(ignore_permissions=True)
        frappe.db.commit()

        self._run(boq.name)

        rows = frappe.db.get_all(
            "BoQ General Specs Sheet",
            filters={"parent": boq.name, "parenttype": "BOQs", "source_sheet_name": "SOW"},
            fields=["preamble_text"],
        )
        self.assertEqual(len(rows), 1, "No BoQ General Specs Sheet row for SOW after parse")
        preamble_text = rows[0].preamble_text
        self.assertTrue(preamble_text, "preamble_text is empty or None")
        self.assertIn(
            "IS standards", preamble_text,
            f"Expected SOW text not found in preamble_text: {preamble_text!r}",
        )
        # SOW must produce no BoQ Review Rows regardless of its wizard_status
        self.assertEqual(
            frappe.db.count("BoQ Review Row", {"boq": boq.name, "sheet_name": "SOW"}),
            0,
            "SOW sheet produced BoQ Review Rows (must not -- it is master_preamble)",
        )

    def test_no_general_specs_sheets_is_safe(self):
        """Empty general_specs_sheets child table is treated as 'none' -- no crash."""
        boq = self._make_boq()
        # _make_boq without include_sow creates no general_specs_sheets rows
        self._run(boq.name)
        self.assertGreater(
            frappe.db.count("BoQ Review Row", {"boq": boq.name, "sheet_name": "SheetA"}), 0
        )

    def test_general_specs_sheet_none_is_safe(self):
        """No general_specs_sheets designations (never set) -- no crash, SheetA rows inserted."""
        boq = self._make_boq()
        self._run(boq.name)
        self.assertGreater(
            frappe.db.count("BoQ Review Row", {"boq": boq.name, "sheet_name": "SheetA"}), 0
        )

    def test_re_parse_replaces_general_specs_preamble_not_duplicates(self):
        """Re-parsing a general-specs sheet replaces the child row's preamble_text, not duplicates."""
        boq = self._make_boq(include_sow=True)
        self._run(boq.name)
        rows_after_1 = frappe.db.get_all(
            "BoQ General Specs Sheet",
            filters={"parent": boq.name, "parenttype": "BOQs", "source_sheet_name": "SOW"},
            fields=["preamble_text"],
        )
        self.assertEqual(len(rows_after_1), 1, "Expected exactly 1 child row after first parse")

        # Second parse -- sheet is now 'Parsed' status (treated same as Reviewed)
        self._run(boq.name)
        rows_after_2 = frappe.db.get_all(
            "BoQ General Specs Sheet",
            filters={"parent": boq.name, "parenttype": "BOQs", "source_sheet_name": "SOW"},
            fields=["preamble_text"],
        )
        self.assertEqual(len(rows_after_2), 1, "Re-parse duplicated the child row instead of replacing")
        self.assertIn("IS standards", rows_after_2[0].preamble_text or "")

    def test_general_specs_sheet_not_marked_parsed_after_worker(self):
        """General-specs sheets (in child table) are never marked 'Parsed' by the worker."""
        boq = self._make_boq(include_sow=True)
        self._run(boq.name)

        boq.reload()
        sow_draft = next(d for d in boq.sheet_drafts if d.sheet_name == "SOW")
        self.assertNotEqual(
            sow_draft.wizard_status, "Parsed",
            "SOW (general-specs sheet) was wrongly marked 'Parsed' by the worker",
        )

    # -- skip / hidden ---------------------------------------------------

    def test_skip_and_hidden_sheets_produce_no_rows(self):
        """Skip and Hidden sheets pass through as skip=True in config -- produce no rows."""
        blob = json.dumps(_PROD_BLOB_TMPL)
        boq = frappe.new_doc("BOQs")
        boq.project = self.__class__.test_project.name
        boq.boq_name = f"SkipHidden Test {frappe.generate_hash(length=6)}"
        boq.tax_treatment = "Pre-tax"
        boq.source_file_url = self.__class__._worker_wb_path
        boq.append("sheet_drafts", {"sheet_name": "SheetA", "sheet_order": 1, "wizard_status": "Reviewed", "sheet_config": blob})
        boq.append("sheet_drafts", {"sheet_name": "SheetB", "sheet_order": 2, "wizard_status": "Skip"})
        boq.insert(ignore_permissions=True)
        frappe.db.commit()

        self._run(boq.name)

        self.assertEqual(
            frappe.db.count("BoQ Review Row", {"boq": boq.name, "sheet_name": "SheetB"}),
            0,
            "Skip sheet produced rows",
        )

    # -- FIX 2: list-JSON serialization ---------------------------------

    def test_fix2_list_json_fields_round_trip_via_worker(self):
        """
        FIX 2: the worker pre-serializes _LIST_JSON_FIELDS before doc.insert().
        After round-trip via Frappe ORM the fields read back as lists (not double-encoded).
        Dict-type JSON fields (qty_by_area etc.) read back as dicts.
        """
        boq = self._make_boq()
        self._run(boq.name)

        rows = frappe.get_all(
            "BoQ Review Row", filters={"boq": boq.name, "sheet_name": "SheetA"}, fields=["name"]
        )
        self.assertGreater(len(rows), 0)

        doc = frappe.get_doc("BoQ Review Row", rows[0]["name"])

        for field in _LIST_JSON_FIELDS:
            raw = getattr(doc, field)
            parsed = json.loads(raw) if isinstance(raw, str) else raw
            self.assertIsInstance(
                parsed, list,
                f"FIX 2 regression: {field} should be a list after round-trip, got {type(parsed)}",
            )
            # Confirm no double-encoding (a double-encoded list reads as a string, not a list)
            if isinstance(raw, str):
                self.assertNotIsInstance(
                    json.loads(raw), str,
                    f"{field} is double-encoded (round-trip yields a string, not a list)",
                )

        for field in ("qty_by_area", "amount_by_area", "rate_by_area", "append_notes_raw"):
            raw = getattr(doc, field)
            if raw and raw != "null":
                parsed = json.loads(raw) if isinstance(raw, str) else raw
                self.assertIsInstance(
                    parsed, dict,
                    f"{field} should be a dict after round-trip, got {type(parsed)}",
                )


# ---------------------------------------------------------------------------
# Group 5: parse-history fields (has_prior_parse + last_parsed_at) -- Slice 2b-backend-2
# ---------------------------------------------------------------------------

class TestParseHistoryFields(FrappeTestCase):
    """
    Tests for has_prior_parse + last_parsed_at on BoQ Sheet Draft.

    Design contract (LOCKED):
      - Worker stamps both fields when it writes wizard_status="Parsed" for a data sheet.
      - General-specs sheets are never marked Parsed, so they never receive stamps.
      - Nothing clears these fields: set_sheet_config dirty-drop and set_sheet_status
        both leave them untouched.  History is history.
    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        generate_all()
        cls.test_project = _make_project()

        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "DataSheet"
        ws1.append(["SL", "Description", "Unit", "Qty", "Rate", "Amount"])
        ws1.append([1, "History Item", "nos", 1, 100.0, 100.0])
        ws_sow = wb.create_sheet("SOW")
        ws_sow.append(["General Specifications"])
        ws_sow.append(["1. All work per IS standards."])

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        tmp.close()
        cls._wb_path = tmp.name

    @classmethod
    def tearDownClass(cls):
        try:
            os.unlink(cls._wb_path)
        except OSError:
            pass
        _cleanup_project(cls.test_project.name)
        super().tearDownClass()

    def tearDown(self):
        for boq in frappe.get_all(
            "BOQs", filters={"project": self.__class__.test_project.name}, fields=["name"]
        ):
            frappe.db.delete("BoQ Review Row", {"boq": boq.name})
            frappe.delete_doc("BOQs", boq.name, force=True, ignore_permissions=True)
        frappe.db.commit()

    def _make_boq(self, include_sow: bool = False):
        blob = json.dumps(_PROD_BLOB_TMPL)
        boq = frappe.new_doc("BOQs")
        boq.project = self.__class__.test_project.name
        boq.boq_name = f"History Test {frappe.generate_hash(length=6)}"
        boq.tax_treatment = "Pre-tax"
        boq.source_file_url = self.__class__._wb_path
        if include_sow:
            boq.append("general_specs_sheets", {"source_sheet_name": "SOW", "preamble_text": ""})
        boq.append("sheet_drafts", {
            "sheet_name": "DataSheet", "sheet_order": 1,
            "wizard_status": "Reviewed", "sheet_config": blob,
        })
        if include_sow:
            boq.append("sheet_drafts", {
                "sheet_name": "SOW", "sheet_order": 2,
                "wizard_status": "Reviewed",
            })
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        return boq

    def _run(self, boq_name, sheet_names=None):
        _run_parse_worker(boq_name, sheet_names=sheet_names, user="Administrator")

    def _get_history(self, boq_name: str, sheet_name: str) -> dict:
        """Return has_prior_parse + last_parsed_at for the given sheet draft."""
        return frappe.db.get_value(
            "BoQ Sheet Draft",
            {"parent": boq_name, "parenttype": "BOQs", "sheet_name": sheet_name},
            ["has_prior_parse", "last_parsed_at"],
            as_dict=True,
        )

    def test_data_sheet_has_prior_parse_and_last_parsed_at_set_after_parse(self):
        """Worker parses a data sheet -> has_prior_parse=1 and last_parsed_at is non-null."""
        boq = self._make_boq()
        self._run(boq.name)

        hist = self._get_history(boq.name, "DataSheet")
        self.assertEqual(hist.has_prior_parse, 1,
            "has_prior_parse was not set to 1 after successful parse")
        self.assertIsNotNone(hist.last_parsed_at,
            "last_parsed_at was not stamped after successful parse")

    def test_general_specs_sheet_never_stamped(self):
        """General-specs sheet is never marked Parsed -> has_prior_parse stays 0 / last_parsed_at null."""
        boq = self._make_boq(include_sow=True)
        self._run(boq.name)

        hist = self._get_history(boq.name, "SOW")
        self.assertFalse(
            hist.has_prior_parse,
            "has_prior_parse was set on a general-specs sheet (must never be Parsed)",
        )
        self.assertIsNone(
            hist.last_parsed_at,
            "last_parsed_at was set on a general-specs sheet (must never be Parsed)",
        )
        # DataSheet must still be stamped
        data_hist = self._get_history(boq.name, "DataSheet")
        self.assertEqual(data_hist.has_prior_parse, 1)

    def test_parse_history_survives_config_change_dirty_drop(self):
        """
        Dirty-signal contract (Check 5 support):
        After parse, a config change via set_sheet_config drops Parsed->Reviewed.
        has_prior_parse must STILL be 1 and last_parsed_at STILL be set.
        The combination wizard_status=Reviewed + has_prior_parse=1 is the dirty signal.
        """
        boq = self._make_boq()
        self._run(boq.name)

        hist_before = self._get_history(boq.name, "DataSheet")
        self.assertEqual(hist_before.has_prior_parse, 1)
        self.assertIsNotNone(hist_before.last_parsed_at)

        # Config change -> drops Parsed to Reviewed
        set_sheet_config(
            boq_name=boq.name,
            sheet_name="DataSheet",
            sheet_config={"header_row": 99, "area_dimensions": [], "column_role_map": {}},
        )

        hist_after = self._get_history(boq.name, "DataSheet")
        # Status dropped
        status = frappe.db.get_value(
            "BoQ Sheet Draft",
            {"parent": boq.name, "sheet_name": "DataSheet"},
            "wizard_status",
        )
        self.assertEqual(status, "Reviewed", "Expected dirty-drop to Reviewed")
        # History fields MUST NOT be cleared
        self.assertEqual(hist_after.has_prior_parse, 1,
            "has_prior_parse was cleared by set_sheet_config -- history must survive dirty-drop")
        self.assertEqual(hist_after.last_parsed_at, hist_before.last_parsed_at,
            "last_parsed_at was changed by set_sheet_config -- history must survive dirty-drop")

    def test_parse_history_survives_set_pending_via_set_sheet_status(self):
        """
        History is history: set_sheet_status("Pending") must leave has_prior_parse=1
        and last_parsed_at intact.  A Pending sheet that was parsed before legitimately
        carries its parse history.
        """
        boq = self._make_boq()
        self._run(boq.name)

        hist_before = self._get_history(boq.name, "DataSheet")
        self.assertEqual(hist_before.has_prior_parse, 1)
        ts_before = hist_before.last_parsed_at
        self.assertIsNotNone(ts_before)

        set_sheet_status(boq_name=boq.name, sheet_name="DataSheet", status="Pending")

        hist_after = self._get_history(boq.name, "DataSheet")
        self.assertEqual(hist_after.has_prior_parse, 1,
            "has_prior_parse cleared by set_sheet_status -- history must survive status changes")
        self.assertEqual(hist_after.last_parsed_at, ts_before,
            "last_parsed_at changed by set_sheet_status -- history must survive status changes")

    def test_reparse_restamps_last_parsed_at(self):
        """
        Re-parse of an already-parsed sheet must re-stamp last_parsed_at to the
        newer time; has_prior_parse stays 1.
        """
        boq = self._make_boq()
        self._run(boq.name)

        # Overwrite last_parsed_at with a known sentinel past time
        child_name = frappe.db.get_value(
            "BoQ Sheet Draft",
            {"parent": boq.name, "parenttype": "BOQs", "sheet_name": "DataSheet"},
            "name",
        )
        sentinel_ts = "2020-01-01 00:00:00"
        frappe.db.set_value("BoQ Sheet Draft", child_name, "last_parsed_at", sentinel_ts)
        frappe.db.commit()

        # Re-parse
        self._run(boq.name)

        hist = self._get_history(boq.name, "DataSheet")
        self.assertEqual(hist.has_prior_parse, 1)
        self.assertIsNotNone(hist.last_parsed_at)
        # Normalize: Frappe returns Datetime as datetime.datetime; compare as 19-char strings.
        hist_ts_str = str(hist.last_parsed_at)[:19] if hist.last_parsed_at else ""
        self.assertNotEqual(
            hist_ts_str, sentinel_ts,
            "last_parsed_at was not re-stamped on re-parse (still the sentinel past time)",
        )


# ---------------------------------------------------------------------------
# Group 6: parse_in_progress transient marker (Bucket-2 Slice 1)
# ---------------------------------------------------------------------------

class TestParseInProgressMarker(FrappeTestCase):
    """
    Tests for the transient parse_in_progress Check field on BOQs.

    Contracts:
      1. run_parse sets parse_in_progress=1 after a successful enqueue.
      2. _publish_parse_event clears parse_in_progress=0 on the success path.
      3. _publish_parse_event clears on error paths (internal, missing_file)
         -- proving the stuck-flag failure mode is closed.
      4. The clear in _publish_parse_event survives a prior frappe.db.rollback()
         because it runs as its own independent set_value + commit after the rollback.

    Uses a minimal BOQs (no workbook) -- marker logic is independent of parse content.
    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.test_project = _make_project()
        blob = json.dumps(_PROD_BLOB_TMPL)
        boq = frappe.new_doc("BOQs")
        boq.project = cls.test_project.name
        boq.boq_name = "Marker Test BoQ"
        boq.tax_treatment = "Pre-tax"
        boq.append("sheet_drafts", {
            "sheet_name": "SheetA", "sheet_order": 1,
            "wizard_status": "Reviewed", "sheet_config": blob,
        })
        boq.insert(ignore_permissions=True)
        frappe.db.commit()
        cls.boq_name = boq.name

    @classmethod
    def tearDownClass(cls):
        frappe.db.delete("BoQ Review Row", {"boq": cls.boq_name})
        _cleanup_project(cls.test_project.name)
        super().tearDownClass()

    def tearDown(self):
        # Reset to 0 after each test so marker state does not leak between tests.
        frappe.db.set_value("BOQs", self.boq_name, "parse_in_progress", 0)
        frappe.db.commit()

    def _get_marker(self):
        return int(frappe.db.get_value("BOQs", self.boq_name, "parse_in_progress") or 0)

    def _set_marker(self, value):
        frappe.db.set_value("BOQs", self.boq_name, "parse_in_progress", value)
        frappe.db.commit()

    # -- SET on enqueue --------------------------------------------------

    def test_run_parse_sets_parse_in_progress_after_enqueue(self):
        """run_parse sets parse_in_progress=1 on the BOQs row after successful enqueue."""
        from unittest.mock import MagicMock, patch
        self._set_marker(0)

        mock_job = MagicMock()
        mock_job.id = "mock-job-id"
        with patch.object(frappe, "enqueue", return_value=mock_job):
            run_parse(boq_name=self.boq_name)

        self.assertEqual(self._get_marker(), 1,
            "parse_in_progress was not set to 1 after successful enqueue")

    # -- CLEAR at the publish choke-point --------------------------------

    def test_publish_clears_marker_on_success(self):
        """_publish_parse_event clears parse_in_progress=0 on the success completion path."""
        from unittest.mock import patch
        self._set_marker(1)

        with patch.object(frappe, "publish_realtime"):
            _publish_parse_event(
                self.boq_name, "success", user=None,
                parsed_sheets=["SheetA"], not_parsed_sheets=[], failed_sheets=[],
            )

        self.assertEqual(self._get_marker(), 0,
            "parse_in_progress not cleared on success publish path")

    def test_publish_clears_marker_on_internal_error(self):
        """_publish_parse_event clears parse_in_progress=0 on the error/internal path.
        This is the stuck-flag regression: a failed parse must not leave the flag at 1."""
        from unittest.mock import patch
        self._set_marker(1)

        with patch.object(frappe, "publish_realtime"):
            _publish_parse_event(self.boq_name, "error", user=None, error_code="internal")

        self.assertEqual(self._get_marker(), 0,
            "parse_in_progress not cleared on error/internal path (stuck-flag risk)")

    def test_publish_clears_marker_on_missing_file_error(self):
        """_publish_parse_event clears parse_in_progress=0 on the missing_file early-return path."""
        from unittest.mock import patch
        self._set_marker(1)

        with patch.object(frappe, "publish_realtime"):
            _publish_parse_event(self.boq_name, "error", user=None, error_code="missing_file")

        self.assertEqual(self._get_marker(), 0,
            "parse_in_progress not cleared on error/missing_file path")

    def test_clear_survives_prior_rollback(self):
        """
        The clear in _publish_parse_event persists even after a prior frappe.db.rollback().

        The 'internal' error path calls rollback() then _publish_parse_event.
        After rollback() completes, _publish_parse_event starts a fresh transaction
        (set_value + commit) that is independent of the rolled-back transaction.
        parse_in_progress must be 0 after the publish, not stuck at 1.
        """
        from unittest.mock import patch
        self._set_marker(1)
        # _set_marker committed the 1; rollback here undoes nothing (empty transaction).
        # This mirrors the real error path where rollback() precedes _publish_parse_event.
        frappe.db.rollback()

        with patch.object(frappe, "publish_realtime"):
            _publish_parse_event(self.boq_name, "error", user=None, error_code="internal")

        self.assertEqual(self._get_marker(), 0,
            "parse_in_progress not cleared after rollback+publish (stuck-flag on error path)")
