# Copyright (c) 2026, Nirmaan (Stratos Infra Technologies Pvt. Ltd.) and contributors
# For license information, please see license.txt

"""The INTERNAL priced workbook -- the client export PLUS the BCS cost block (BCS-EXP-2).

`export_writeback.export_priced_workbook` produces the file we hand a CLIENT: their own
workbook with our rates stamped in. This module produces the file we keep for OURSELVES: the
same thing, plus what the work costs us, appended as new columns at the sheet's right-hand
edge exactly the way the `Nirmaan Remarks` column is appended.

⚠️ WHY THIS IS A SEPARATE MODULE AND A SEPARATE ENDPOINT, AND WHY `export_writeback.py` HAS
ZERO DIFF FROM THIS SLICE.

  1. A STANDING TEST FORBIDS IT. `test_export_writeback.TestBcsCostRatesNeverReachTheExport
     .test_export_writeback_module_never_names_the_bcs_doctype` greps that module's SOURCE,
     case-insensitively, for `BoQ Row BCS Rate` / `supply_rate` / `install_rate` /
     `combined_rate` / `bcs`. That guard is what makes "internal cost cannot reach a client
     workbook" STRUCTURAL rather than a matter of care. An `include_bcs=True` flag on the
     existing endpoint would break it, and rightly.
  2. THE OWNER ASKED FOR IT EXPLICITLY (2026-08-19): the new export must live separately and
     must not affect the existing one. It does not -- not one line of `export_writeback.py`
     changed, so its 49 tests are unchanged evidence about unchanged code.

WHAT IS SHARED AND WHAT IS NOT. Every STAMPING helper is IMPORTED from `export_writeback` and
has exactly one definition: finding the worksheet, the rate stamp with its per-cell formula
skip, the colour pass, the priced-cell teal, the remark column, the fidelity snapshot, the
committed-version resolver, the filename sanitiser. What is duplicated is the ~40-line
ORCHESTRATION loop, and only because the alternative -- parameterising the client path with
hooks -- would have changed the file the guard above protects. If a stamping RULE ever changes
it changes in one place and both exports inherit it; if the ORDER of the passes changes, this
module must be updated too, which is the price of the separation and is deliberate.

WHAT IT WRITES, per ticked sheet, after everything the client export writes:

  * the COST columns -- `BCS Cost (Supply)` + `BCS Cost (Installation)`, or the single
    `BCS Cost` on a combined-rate sheet. WHICH of those a sheet gets is
    `services.boq_bcs.sources.live_rate_kinds`, the owner's halves-beat-combined ruling.
    Values are read STRAIGHT off `BoQ Row BCS Rate`; nothing is computed.
  * `BCS Total Amount` -- an EXCEL FORMULA over the cost cells just written and the sheet's
    own quantity cells, NOT a number. See the ruling below.

⚠️ THERE IS NO `% Margin` COLUMN, BY OWNER RULING (2026-08-19, planning Q8). The margin's
denominator is "the figure the amount cell is SHOWING", which on an unpriced-source workbook
is the FORMULA value rather than the document's own 0 -- a reconciliation this module has no
way to reproduce in Excel. Rather than ship a column that is blank on exactly the sheets we
upload most, the export omits it and the download dialog tells the user to add it. Do NOT
"finish the job" by dividing by the sheet's amount column without re-opening that ruling.

⚠️ THE TOTAL IS A FORMULA, NOT A NUMBER, AND THAT IS THE WHOLE DESIGN (planning Q1 = option
B). No server code has ever computed a BCS number: the arithmetic lives in `bcsColumns.ts`
and depends on the pricing screen's live drafts and reconciliation choices. Porting it would
have created a second implementation of an owner-locked rule. Emitting Excel instead means
Excel computes, the delivered file shows its working, and a pricer can change a cost and watch
the total move. The cost of that choice is recorded honestly: a cell we write is what EXCEL
says, and where Excel and the screen could disagree the guard below keeps the SHAPE right
(a blank stays a blank) even when it cannot keep the number identical.

Public API:
  export_priced_workbook_with_bcs(boq_name, sheet_names) -> dict   [whitelisted POST]
"""
from __future__ import annotations

import base64
import os
import shutil
from typing import Any

import frappe
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# ── api -> api. Every stamping rule has ONE definition and it is over there. ──────
from nirmaan_stack.api.boq.wizard.export_writeback import (
    _XLSX_CONTENT_TYPE,
    _apply_colors,
    _apply_priced_highlight,
    _assert_fidelity,
    _coerce_names,
    _col_is_empty,
    _fidelity_snapshot,
    _find_ws,
    _MAX_EXCEL_COLS,
    _resolve_sheet_plan,
    _rightmost_mapped_col_index,
    _safe_export_basename,
    _stamp_rates,
    _write_remark_column,
)
# The formula-to-Excel vocabulary, shared so the operator set and the sheet-column resolution
# cannot drift from the template export's.
from nirmaan_stack.api.boq.wizard.export_template_workbook import _OP_INFIX, resolve_ref_col
from nirmaan_stack.api.boq.wizard.pricing import _committed_descriptors
from nirmaan_stack.api.boq.wizard.sheet_preview import _fetch_boq_file_to_tempfile
# api -> service, the one legal direction. These two are the BCS column rules, mirrored from
# `bcsColumns.ts` and pinned against it by `parity_cases.json` (slice BCS-EXP-1).
from nirmaan_stack.services.boq_bcs.sources import (
    derive_amount_columns,
    derive_qty_columns,
    live_rate_kinds,
)

_PRICING = "BoQ Cell Pricing"
_COLOR = "BoQ Cell Color"
_REMARK = "BoQ Cell Remark"
_BOQ_SHEET = "BoQ Sheet"
_BCS_RATE = "BoQ Row BCS Rate"
_FORMULA = "BoQ Cell Amount Formula"

# The rate kind -> the stored field it reads, and the column header it is written under.
# ⚠️ THE HEADERS ARE OWNER-RULED AND MUST READ EXACTLY AS THE GRID'S COLUMNS READ
# (`bcsColumns.BCS_RATE_LABEL`). The `BCS ` prefix marks which side of the sheet a figure
# belongs to -- BCS is what it costs US, everything else is what we charge the CLIENT -- and
# that matters most here, in a file where the two sit side by side. Do NOT shorten
# "Installation": the owner chose the longer word for the header a pricer reads.
_KIND_FIELD = {"supply": "supply_rate", "install": "install_rate", "combined": "combined_rate"}
_KIND_HEADER = {
    "supply": "BCS Cost (Supply)",
    "install": "BCS Cost (Installation)",
    "combined": "BCS Cost",
}
_TOTAL_HEADER = "BCS Total Amount"
_MARGIN_HEADER = "% Margin"

# The BCS block's "this cell carries a figure" fill (owner ruling 2026-08-19, after the live
# check). It is the counterpart of `export_writeback._apply_priced_highlight`'s teal on a
# stamped rate cell: the same idea -- mark what actually got written -- in the light blue the
# owner asked for.
#
# ⚠️ THIS DELIBERATELY REUSES THE USER PALETTE'S `blue` (`_COLOR_HEX["blue"]`), which the rate
# highlight deliberately AVOIDS doing. The rate highlight lives on a SHEET column, where a user
# colour tag can also land, so a shared hex there would make a system mark read as a user tag.
# The BCS columns are ones this module APPENDS: a user tag is addressed by (col_letter,
# excel_row) against the committed grid and can never resolve to a column that did not exist
# when the grid was committed. The collision is structurally unreachable here, which is what
# makes matching the palette's own "light blue" the least surprising choice rather than a
# near-miss shade nobody can name.
_BCS_FILLED_HEX = "BDD7EE"


def _fill_bcs_cells(ws, cells: list[tuple]) -> int:
    """Light-blue every BCS cell that actually carries a figure. Returns the count filled.

    ⚠️ ONLY CELLS WE WROTE, and the distinction is the whole point of the mark: a costed row
    gets a fill, an UNCOSTED row's blank stays visibly unfilled. Filling the whole column
    height would say "every row is costed", which is exactly the claim the COUNT guard exists
    to avoid making. The HEADER is not filled either -- it is a label, not a figure, and the
    rate highlight it mirrors marks no header.

    A fill sets ONLY `.fill`; the number or formula already in the cell is untouched.
    """
    for col, row in cells:
        ws[f"{col}{row}"].fill = PatternFill(fill_type="solid", fgColor=_BCS_FILLED_HEX)
    return len(cells)

# The formula operands that resolve to the cost columns THIS module writes. Mirrors
# `pricing._BCS_OPERAND_FIELDS` / `bcsColumns.BCS_OPERAND_FIELD`. `bcs_qty` is handled
# separately -- it has no column of its own and resolves to the sheet's quantity columns.
_COST_OPERAND_TO_KIND = {
    "bcs_supply": "supply",
    "bcs_install": "install",
    "bcs_combined": "combined",
}
_QTY_OPERAND = "bcs_qty"
_BCS_TOTAL_TARGET = "bcs_total"

# The two % Margin formula TARGETS. Mirrors `pricing._MARGIN_COST_TARGET` /
# `pricing._BOQ_TOTAL_TARGET` -- the numerator and the denominator of the ratio, each of
# which a sheet MAY declare a formula for and usually does not.
#
# ⚠️ THERE IS NO `bcs_margin` TARGET AND THERE MUST NEVER BE ONE. The RATIO itself is not
# editable: it needs the numeric literals 1 and 100, which the formula system rejects by
# design, and -- the real reason -- it carries the sign guard below. Making the shape
# editable would hand that guard back to the user with nothing enforcing it. `pricing.py`
# says the same thing at the same place; the two comments must stay in agreement.
_MARGIN_COST_TARGET = "bcs_margin_cost"
_BOQ_TOTAL_TARGET = "boq_total"

# The internal-file marker. Owner ruling (planning Q6): the FILENAME carries it and nothing
# else does -- no banner row, no red header cell.
_INTERNAL_SUFFIX = "priced_bcs_internal"


# ── access ────────────────────────────────────────────────────────────────────────
def _require_bcs_export_access() -> str:
    """Admins + estimation, which is EXACTLY the Pricing Module's read set, so the decision is
    REUSED rather than re-minted (planning Q7 = option (a)): `PRICING_ACCESS_SET` holds the
    DB-verified strings and `_require_pricing_access` holds the three-way check (Administrator,
    or role_profile_name, or any role, intersecting that set).

    ⚠️ THE VOICE IS RE-THROWN ON PURPOSE. That gate's own message says "You do not have access
    to the Pricing Module", which is true of the module it belongs to and confusing on a BoQ
    hub. Re-voicing the refusal while REUSING the decision keeps one definition of who may see
    a margin and still tells the user what they actually tried to do.

    ⚠️ AND THIS IS THE REAL BOUNDARY, not the hidden menu item. The frontend mirrors it for UX
    only."""
    from nirmaan_stack.api.pricing.workbook import _require_pricing_access

    try:
        return _require_pricing_access()
    except frappe.PermissionError:
        frappe.throw(
            "This download includes internal cost and margin data. It is available to "
            "administrators and the estimation team only.",
            frappe.PermissionError,
            title="Not permitted",
        )


# ── the cost block ────────────────────────────────────────────────────────────────
def _next_empty_col(ws, start_idx: int) -> int:
    """The first genuinely-empty column at or after `start_idx`, scanning RIGHT past anything
    that carries real data. The same rule -- and the same `_col_is_empty` -- the remark column
    has always used, so the block can never overwrite a stray annotation sitting past the
    mapped edge and can never dead-end the export either."""
    idx = start_idx
    while not _col_is_empty(ws, idx):
        idx += 1
        if idx > _MAX_EXCEL_COLS:
            frappe.throw(
                f"No empty column found on sheet '{ws.title}' to place the cost columns "
                f"(scanned past the {_MAX_EXCEL_COLS}-column limit).",
                title="Cost columns: no space",
            )
    return idx


def _cost_rows(boq_name: str, sheet_name: str, committed_version) -> dict:
    """The sheet's CURRENT cost rows, keyed by excel_row.

    ⚠️ PRESENCE IS THE WHOLE TEST, and it is what makes a blank stay blank. `save_row_bcs_rates`
    is a WHOLE-ROW SNAPSHOT WRITE that coerces every rate it is not given to 0.0, so a stored
    row always carries three numbers -- possibly all zero. The browser reads exactly the same
    way (`mergeBcsRowValues`: a saved record yields a string for every field, an absent one
    yields null on all three, and `bcsUnitCost` returns null only in the second case). So a row
    WITH a record shows a number on screen, including 0, and a row WITHOUT one shows a blank.
    This module writes a Total formula on precisely the first set."""
    rows = frappe.get_all(
        _BCS_RATE,
        filters={"boq": boq_name, "sheet_name": sheet_name,
                 "committed_version": committed_version, "is_current": 1},
        fields=["excel_row", "supply_rate", "install_rate", "combined_rate"],
        order_by="excel_row asc",
    )
    return {int(r["excel_row"]): r for r in rows}


def _formula_tree(boq_name: str, sheet_name: str, committed_version, target: str):
    """The sheet's declared formula tree for one BCS target, or None for that target's
    built-in rule. Mirrors `bcsColumns.pickBcsTotalFormula` / `pickBoqTotalFormula` /
    `pickMarginCostFormula` -- a BCS target carries no area and no rate kind, so the match is
    on the target token alone, and ONE reader therefore serves all three.

    LIVE on real data: 4 of the 5 costed sheets on the bench declare a `bcs_total` (all four
    spelling out `(bcs_supply + bcs_install) * qty_total`, which is the built-in rule written
    explicitly against the sheet's own quantity column). The two margin targets are declared
    far more rarely, which is exactly why their DEFAULTS carry the weight here."""
    rec = frappe.db.get_value(
        _FORMULA,
        {"boq": boq_name, "sheet_name": sheet_name,
         "committed_version": committed_version, "is_current": 1,
         "target_value_field": target},
        "formula",
    )
    if not rec:
        return None
    if isinstance(rec, str):
        import json

        try:
            return json.loads(rec)
        except (ValueError, TypeError):
            return None
    return rec


def _cell(col: str, row: int) -> str:
    return f"{col}{row}"


def _ref_to_excel(ref: dict, row: int, cost_cols: dict, qty_cols: list, role_map: dict,
                  total_col: str | None = None):
    """One formula leaf -> `(excel body, sheet-cell refs used)`, or `(None, [])` when it cannot
    be resolved -- in which case the caller fail-safes the whole cell to BLANK, exactly as
    `ast_to_excel` does.

    THREE VOCABULARIES MEET HERE, which is why this cannot simply BE `ast_to_excel`:
      * a COST operand resolves to a column this module just wrote;
      * `bcs_qty` has no column at all and resolves to the SUM of the sheet's quantity cells;
      * anything else is one of the sheet's own columns and goes to the SHARED
        `resolve_ref_col`, so a sheet-column ref means the same thing in both exports.

    The second return value is the list of SHEET cells the body reads. The caller guards on
    them: a cost cell we wrote ourselves is numeric whenever it exists, but a quantity cell in
    the client's workbook may be blank, and Excel silently reads a blank as 0."""
    if not isinstance(ref, dict):
        return None, []
    field = ref.get("value_field")
    kind = _COST_OPERAND_TO_KIND.get(field)
    if kind:
        col = cost_cols.get(kind)
        return (_cell(col, row), []) if col else (None, [])
    if field == _BCS_TOTAL_TARGET:
        # ⚠️ `bcs_total` IS BOTH A TARGET AND AN OPERAND, and that is deliberate: choosing
        # "BCS Total Amount" in a margin numerator must mean "whatever that column currently
        # computes", never a frozen copy of the rule it had when the margin was configured.
        # Here that reads across perfectly -- the operand becomes a REFERENCE to the Total
        # column, so Excel re-evaluates it exactly as the screen does.
        #
        # It resolves only AFTER the Total column exists, which is why the margin is written
        # last and why `total_col` is None on every call made before then. A numerator naming
        # it on a sheet that got no Total column is unresolvable -> the whole margin cell
        # fail-safes to blank, which is the honest answer: there is no such column to divide.
        #
        # NOTE the empty cell list. Unlike a quantity cell in the client's workbook, the
        # Total is a formula THIS module wrote, and it is already guarded -- it yields "" on
        # a row the screen leaves blank. A `""` divided into is not a silent zero; it makes
        # the margin `#VALUE!`, which is why the caller guards the DENOMINATOR instead.
        return (_cell(total_col, row), []) if total_col else (None, [])
    if field == _QTY_OPERAND:
        if not qty_cols:
            return None, []
        cells = [_cell(c, row) for c in qty_cols]
        return ("(" + "+".join(cells) + ")", cells)
    col = resolve_ref_col(ref, role_map)
    if not col:
        return None, []
    return _cell(col, row), [_cell(col, row)]


def _tree_to_excel(node: Any, row: int, cost_cols: dict, qty_cols: list, role_map: dict,
                   total_col: str | None = None):
    """Translate a BCS Total formula tree to an Excel body -> `(body, sheet cells read)`.
    Mirrors `ast_to_excel`'s shape (every operator node wrapped in its own parentheses, so an
    n-ary `(A-B-C)` reads in Excel exactly as `foldOperands` folds it) and fails SAFE the same
    way: any unresolvable operand or unsupported operator blanks the whole cell rather than
    emitting something plausible."""
    if not isinstance(node, dict):
        return None, []
    if "ref" in node:
        return _ref_to_excel(node.get("ref"), row, cost_cols, qty_cols, role_map, total_col)
    if "op" in node:
        op = node.get("op")
        if op not in _OP_INFIX:
            return None, []
        operands = node.get("operands") or []
        if not operands:
            return None, []
        parts, cells = [], []
        for child in operands:
            body, used = _tree_to_excel(child, row, cost_cols, qty_cols, role_map, total_col)
            if body is None:
                return None, []
            parts.append(body)
            cells.extend(used)
        return "(" + _OP_INFIX[op].join(parts) + ")", cells
    return None, []


def _builtin_total(row: int, kinds: list, cost_cols: dict, qty_cols: list):
    """The built-in rule -- `(the cost boxes summed) x (the quantity columns summed)` -- as
    Excel. Byte-equivalent in meaning to `bcsTotalAmount(qty, bcsUnitCost(...))`, which is what
    every sheet with no declared formula shows on screen."""
    cost = [_cell(cost_cols[k], row) for k in kinds if k in cost_cols]
    if not cost or not qty_cols:
        return None, []
    qty_cells = [_cell(c, row) for c in qty_cols]
    body = "(" + "+".join(cost) + ")*(" + "+".join(qty_cells) + ")"
    return body, qty_cells


def _amount_body(row: int, amount_cols: list, role_map: dict):
    """The % Margin DENOMINATOR when the sheet declares no `boq_total` formula -> the sum of
    its amount columns. Mirrors `bcsColumns.bcsRowAmount` over `derive_amount_columns`.

    Every cell is returned as a guarded cell, because these belong to the CLIENT's workbook
    and may legitimately be blank -- and Excel reads a blank as 0, which would then trip the
    sign guard and blank the margin. That is the right outcome, but it must be reached
    honestly: `COUNT = 0` says "this row has no amount", not "its amount is zero"."""
    cells = [_cell(c, row) for c in amount_cols]
    if not cells:
        return None, []
    body = cells[0] if len(cells) == 1 else "(" + "+".join(cells) + ")"
    return body, cells


def _margin_body(row: int, cost_body: str, amount_body: str) -> str:
    """★ `% Margin = (amount - cost) / amount x 100`, as Excel.

    ⚠️ THE DIRECTION IS OWNER-SETTLED AND WAS ONCE RELAYED BACKWARDS (corrected at BCS-S2d).
    Dividing by the AMOUNT means a sheet whose amount columns cover only the supply half
    reads LOWER, and goes sharply negative once the amount falls below the cost. That visible
    collapse IS the safety. Do NOT re-derive it as cost-over-amount or as a mark-up on cost;
    both read HIGHER on exactly the sheets that need a warning.

    Written as `(amount - cost) / amount` rather than the owner's `(1 - cost/amount)` because
    the two are the same expression rearranged (pinned by test on the browser side) and this
    form needs no literal `1`. The `* 100` is unavoidable and is the reason the ratio can
    never be an editable formula -- see `_MARGIN_COST_TARGET`."""
    return f"(({amount_body}-{cost_body})/{amount_body})*100"


def _guarded(body: str, sheet_cells: list, require_all: bool,
             extra_tests: list | None = None) -> str:
    """Wrap a Total body so a row the screen leaves BLANK is blank here too.

    ⚠️ WITHOUT THIS, EXCEL TURNS AN ABSENCE INTO A CLAIM. A quantity cell that is empty reads
    as 0 in Excel, so the Total would render a confident `0` -- "this row costs nothing" --
    where the screen renders an empty cell with the reason `no quantity`. `0` is a claim and an
    absence is not, and on a cost sheet that difference is the whole point of
    `BcsComputedCell`'s blank-with-a-reason.

    `COUNT` counts NUMERIC cells only, so a genuine 0 quantity still computes (COUNT is 1) and
    still reads 0 -- which is exactly what the screen does with it. The two shapes below mirror
    the two screen paths and are not interchangeable:
      * the BUILT-IN rule sums whatever resolves and blanks only when NOTHING does
        (`bcsRowQuantity`'s `any` flag)  -> blank iff COUNT = 0;
      * a DECLARED formula blanks when ANY operand is missing
        (`evaluateBcsTotalFormula` returns on the first unresolved ref) -> blank iff COUNT < n.

    `extra_tests` carries conditions that are NOT about a cell being absent -- today just the
    % Margin sign guard, which refuses a denominator that is zero or negative. They are ORed
    with the count test rather than wrapped in a second IF, so ONE builder emits every guard
    in this module and there is no second place for a guard to be written differently. Any
    number of tests collapses to one `IF`, and a single test skips the `OR` so the ordinary
    Total keeps the exact string it emitted before this parameter existed.
    """
    tests = []
    if sheet_cells:
        args = ",".join(sheet_cells)
        tests.append(
            f"COUNT({args})<{len(sheet_cells)}" if require_all else f"COUNT({args})=0"
        )
    tests.extend(extra_tests or [])
    if not tests:
        return "=" + body
    test = tests[0] if len(tests) == 1 else "OR(" + ",".join(tests) + ")"
    return f'=IF({test},"",{body})'


def _write_bcs_block(ws, plan, boq_name: str, sheet_name: str) -> dict:
    """Append this sheet's cost columns and its Total column. Returns a report:
    `{cost_columns, total_column, rows, formulas, reason}`. `reason` is set (and the rest
    empty) whenever the block was deliberately skipped -- the caller surfaces it, because a
    silently absent cost block on an internal cost file is worse than a visible refusal."""
    empty = {"cost_columns": {}, "total_column": None, "rows": 0, "formulas": 0}
    cv = plan.commit_version
    if not frappe.db.get_value(_BOQ_SHEET, plan.name, "bcs_enabled"):
        return {**empty, "reason": "cost tracking is switched off for this sheet"}

    descriptors = _committed_descriptors(boq_name, sheet_name, cv)
    kinds = live_rate_kinds(descriptors)
    if not kinds:
        return {**empty, "reason": "this sheet maps no rate column, so it has no cost columns"}

    cost_rows = _cost_rows(boq_name, sheet_name, cv)
    if not cost_rows:
        return {**empty, "reason": "no costs have been entered on this sheet yet"}

    # WHERE the block starts: the TRUE data edge (the rightmost MAPPED column), never
    # openpyxl's max_column, which recon found inflated by empty styled cells.
    role_map = plan.column_role_map or {}
    cursor = _rightmost_mapped_col_index(role_map)
    if cursor <= 0:
        return {**empty, "reason": "this sheet has no mapped columns, so the data edge is undefined"}
    cursor += 1

    hrow = int(plan.header_row) if (plan.header_row and int(plan.header_row) >= 1) else 1

    cost_cols: dict = {}
    filled: list = []          # every BCS cell that ends up carrying a figure
    for kind in kinds:
        idx = _next_empty_col(ws, cursor)
        letter = get_column_letter(idx)
        ws.cell(row=hrow, column=idx).value = _KIND_HEADER[kind]
        field = _KIND_FIELD[kind]
        for excel_row, rec in cost_rows.items():
            value = rec.get(field)
            if value is None:
                continue
            ws.cell(row=int(excel_row), column=idx).value = value
            filled.append((letter, int(excel_row)))
        cost_cols[kind] = letter
        cursor = idx + 1

    # The Total column. Its bodies are built FIRST: if not one row yields a formula there is
    # nothing to put in the column, and an empty column with a header is noise on a file whose
    # whole purpose is the two numbers side by side.
    tree = _formula_tree(boq_name, sheet_name, cv, _BCS_TOTAL_TARGET)
    qty_cols = [c.get("col") for c in derive_qty_columns(
        _sheet_qty_source(plan.name), descriptors) if c.get("col")]

    bodies: dict = {}
    for excel_row in cost_rows:
        if tree:
            body, cells = _tree_to_excel(tree, int(excel_row), cost_cols, qty_cols, role_map)
            require_all = True
        else:
            body, cells = _builtin_total(int(excel_row), kinds, cost_cols, qty_cols)
            require_all = False
        if body is None:
            continue
        bodies[int(excel_row)] = _guarded(body, cells, require_all)

    total_letter = None
    if bodies:
        idx = _next_empty_col(ws, cursor)
        total_letter = get_column_letter(idx)
        ws.cell(row=hrow, column=idx).value = _TOTAL_HEADER
        for excel_row, formula in bodies.items():
            ws.cell(row=excel_row, column=idx).value = formula
            filled.append((total_letter, int(excel_row)))
        cursor = idx + 1

    margin = _write_margin_column(
        ws, plan, boq_name, sheet_name, cv, descriptors, role_map, hrow, cursor,
        cost_cols, qty_cols, total_letter, set(cost_rows),
    )
    filled.extend(margin["cells"])

    # ONE fill pass over the whole block, last, so a cell is marked exactly once however many
    # columns it took to get here.
    _fill_bcs_cells(ws, filled)

    return {
        "cost_columns": cost_cols,
        "total_column": total_letter,
        "margin_column": margin["column"],
        "margin_skipped": margin["reason"],
        "rows": len(cost_rows),
        "formulas": len(bodies) + margin["formulas"],
        "reason": None if bodies else "this sheet's quantity columns could not be resolved, "
                                      "so no Total column was written",
    }


def _write_margin_column(ws, plan, boq_name: str, sheet_name: str, cv, descriptors: list,
                         role_map: dict, hrow: int, cursor: int, cost_cols: dict,
                         qty_cols: list, total_col: str | None, rows: set) -> dict:
    """★ THE % MARGIN COLUMN -- `(amount - cost) / amount x 100`, as a live Excel formula.

    Returns `{column, formulas, reason, cells}`; `reason` is set (and `column` None) whenever the
    column was deliberately skipped. It is its OWN reason, never folded into the block's,
    because a sheet can legitimately get costs and a Total and still have no margin -- there
    is nothing wrong with such a sheet, and saying "no Total" about it would be false.

    ⚠️ WHY THIS IS A FORMULA AND NOT A NUMBER, for the same reason the Total is: no server
    code has ever computed a BCS figure, and computing one here would be a second
    implementation of an owner-locked rule that lives in `bcsColumns.bcsMarginPercent`. As a
    formula it also stays LIVE -- edit a cost in the workbook and the margin follows, which
    is the whole reason for shipping the column rather than telling a user to add it. A
    hand-added column would compute the same ratio and carry NONE of the guards below.

    THE GUARDS, and each is load-bearing:
      * `<= 0` on the denominator. A ZERO amount has no margin to measure against; a NEGATIVE
        one FLIPS THE INEQUALITY, so an amount of -100 against a cost of 50 computes +150% --
        a loss displayed as a profit. That is the one failure mode this column treats as
        worse than a blank, because it is confidently wrong rather than visibly absent, and
        it is the single strongest argument for shipping the column instead of leaving the
        ratio to be typed by hand.
      * the COUNT guard on whatever sheet cells the denominator reads, inherited from
        `_guarded` -- Excel reads a blank cell as 0, so without it an unmapped amount would
        reach the sign guard as a zero and blank the margin for the RIGHT answer by the WRONG
        route. It would also be indistinguishable from a genuine zero.
      * NOT-FINITE needs no guard and gets none. Excel cannot reach it once the denominator
        is known non-zero, and a dead branch in a guard chain reads as a fourth rule.

    ⚠️ THE BLANK CANNOT CARRY ITS REASON HERE, and that is the one property lost in the move
    to Excel. On screen every blank BCS cell explains itself in its tooltip. A workbook has no
    tooltip -- so the guard is left legible IN THE CELL (`IF(OR(COUNT(G5)=0,G5<=0),"",...)`),
    where a reader who clicks a blank margin can see exactly which test refused it. That is
    the Excel-native form of the same promise, not an abandonment of it.
    """
    empty = {"column": None, "formulas": 0, "cells": []}

    # THE DENOMINATOR -- what we charge. A declared `boq_total` formula wins; otherwise the
    # sheet's own amount columns, resolved by the ported `derive_amount_columns`.
    amount_tree = _formula_tree(boq_name, sheet_name, cv, _BOQ_TOTAL_TARGET)
    amount_cols = [c.get("col") for c in derive_amount_columns(
        _sheet_amount_source(plan.name), descriptors) if c.get("col")]
    if not amount_tree and not amount_cols:
        return {**empty, "reason": "this sheet maps no amount column, so there is nothing to "
                                   "measure a margin against"}

    # THE NUMERATOR -- what it costs us. A declared `bcs_margin_cost` formula wins; otherwise
    # the BCS Total Amount column this module just wrote. With no Total column and no declared
    # numerator there is no cost figure to divide, so there is no margin -- not a zero.
    cost_tree = _formula_tree(boq_name, sheet_name, cv, _MARGIN_COST_TARGET)
    if not cost_tree and not total_col:
        return {**empty, "reason": "this sheet has no BCS Total Amount column, so its margin "
                                   "has no cost figure to measure"}

    bodies: dict = {}
    for excel_row in sorted(rows):
        row = int(excel_row)
        if amount_tree:
            amount, amount_cells = _tree_to_excel(
                amount_tree, row, cost_cols, qty_cols, role_map, total_col)
            require_all = True
        else:
            amount, amount_cells = _amount_body(row, amount_cols, role_map)
            require_all = False
        if amount is None:
            continue
        if cost_tree:
            cost, cost_cells = _tree_to_excel(
                cost_tree, row, cost_cols, qty_cols, role_map, total_col)
        else:
            cost, cost_cells = _cell(total_col, row), []
        if cost is None:
            continue
        # The sign guard names the denominator EXACTLY as the body divides by it, so the two
        # can never test different things.
        bodies[row] = _guarded(
            _margin_body(row, cost, amount),
            amount_cells + cost_cells,
            require_all,
            extra_tests=[f"{amount}<=0"],
        )

    if not bodies:
        return {**empty, "reason": "this sheet's amount columns could not be resolved on any "
                                   "costed row, so no % Margin column was written"}

    idx = _next_empty_col(ws, cursor)
    letter = get_column_letter(idx)
    ws.cell(row=hrow, column=idx).value = _MARGIN_HEADER
    cells = []
    for excel_row, formula in bodies.items():
        ws.cell(row=excel_row, column=idx).value = formula
        cells.append((letter, int(excel_row)))
    return {"column": letter, "formulas": len(bodies), "reason": None, "cells": cells}


def _sheet_amount_source(sheet_docname: str):
    """The sheet's stored BCS amount confirmation, or None. Same disposition as
    `_sheet_qty_source`: BCS-S12 removed the picker, so a sheet enabled since carries none and
    `derive_amount_columns` falls back to the sheet's own shape."""
    raw = frappe.db.get_value(_BOQ_SHEET, sheet_docname, "bcs_amount_source")
    if not raw:
        return None
    if isinstance(raw, str):
        import json

        try:
            return json.loads(raw)
        except (ValueError, TypeError):
            return None
    return raw


def _sheet_qty_source(sheet_docname: str):
    """The sheet's stored BCS quantity confirmation, or None. Pre-BCS-S12 sheets carry one and
    it WINS; sheets enabled since carry none, because S12 removed the picker and nothing has
    written the field since -- which is exactly why `derive_qty_columns` has a fallback."""
    raw = frappe.db.get_value(_BOQ_SHEET, sheet_docname, "bcs_qty_source")
    if not raw:
        return None
    if isinstance(raw, str):
        import json

        try:
            return json.loads(raw)
        except (ValueError, TypeError):
            return None
    return raw


# ── the worker ────────────────────────────────────────────────────────────────────
def _generate_internal_workbook(
    boq_name: str, sheet_names: list, src_path: str, display_name: str = None
) -> dict:
    """Stamp the client-facing layers AND the cost block onto the workbook at src_path (a
    throwaway COPY), assert fidelity, and return the download payload.

    ⚠️ THE PASS ORDER MIRRORS THE CLIENT EXPORT'S AND THEN ADDS ONE STEP. Rates, then user
    colours, then the system teal LAST so it wins on a stamped rate cell, then the REMARK
    COLUMN, then the COST BLOCK.

    ⚠️ REMARKS BEFORE THE BLOCK -- REVERSED at the live check (owner ruling 2026-08-19; this
    read the other way round under planning Q9). `Nirmaan Remarks` keeps the position it holds
    in the CLIENT export, at the right-hand edge of the client's own data, and everything
    internal sits beyond it. Both placers scan rightward past any occupied column, so the call
    order alone decides the layout -- there is no offset arithmetic and neither function knows
    the other exists.

    ⚠️ IT DOES NOT STAMP `last_exported_at`, BY OWNER RULING (planning Q2). That field means
    "when the CLIENT last got this sheet" and drives the amber changed-since-export chip.
    Stamping it from an internal download would make the chip claim the client holds something
    they have never been sent. This export therefore writes NOTHING to the database at all,
    which is also why it needs no commit.

    MUST be loaded data_only=False -- data_only=True loads cached VALUES and DESTROYS every
    formula on save, which on this export would take the client's own amount formulas with it.
    """
    import openpyxl

    plans = {sn: _resolve_sheet_plan(boq_name, sn) for sn in sheet_names}
    wb = openpyxl.load_workbook(src_path, data_only=False)
    before = _fidelity_snapshot(wb)

    exported: list = []
    skipped_by_sheet: dict = {}
    remark_cols: dict = {}
    cost_blocks: dict = {}
    cost_skipped: dict = {}
    formulas_written = 0

    for sn in sheet_names:
        plan = plans[sn]
        ws = _find_ws(wb, sn)

        # Grid-only general-specs sheets carry no rates and no nodes, so they carry no costs
        # either. They pass through untouched, exactly as in the client export.
        if plan.treat_as == "master_preamble":
            exported.append(sn)
            cost_skipped[sn] = "this is a general-specs sheet, which carries no priced rows"
            continue

        cv = plan.commit_version
        pricing = frappe.get_all(
            _PRICING,
            filters={"boq": boq_name, "sheet_name": sn, "committed_version": cv,
                     "is_current": 1, "is_filled": 1},
            fields=["excel_row", "col_letter", "rate"],
            order_by="excel_row asc, col_letter asc",
        )
        colors = frappe.get_all(
            _COLOR,
            filters={"boq": boq_name, "sheet_name": sn, "committed_version": cv, "is_current": 1},
            fields=["excel_row", "col_letter", "color"],
        )
        remarks = frappe.get_all(
            _REMARK,
            filters={"boq": boq_name, "sheet_name": sn, "committed_version": cv, "is_current": 1},
            fields=["excel_row", "remark"],
            order_by="excel_row asc",
        )

        skipped, written = _stamp_rates(ws, pricing)
        _apply_colors(ws, colors)
        _apply_priced_highlight(ws, written)

        # ⚠️ REMARKS FIRST, THEN THE BCS BLOCK (owner ruling 2026-08-19, after the live check
        # -- this REVERSES the original "block first, remarks last" ordering). `Nirmaan
        # Remarks` is a column people already know from the CLIENT export, so it keeps its
        # familiar position at the client data's right-hand edge and the internal cost block
        # sits beyond it -- everything internal to the right of everything shared.
        #
        # The ORDER OF THESE TWO CALLS IS THE WHOLE MECHANISM. Both placers scan rightward
        # from the true data edge past any occupied column, so whichever writes first claims
        # the nearer column and the second steps past it. There is no offset arithmetic
        # anywhere and neither function knows the other exists.
        if remarks:
            remark_cols[sn] = _write_remark_column(ws, remarks, plan.column_role_map,
                                                   plan.header_row)

        block = _write_bcs_block(ws, plan, boq_name, sn)
        formulas_written += block["formulas"]
        if block["cost_columns"]:
            cost_blocks[sn] = {
                "cost_columns": block["cost_columns"],
                "total_column": block["total_column"],
                # The margin rides the BLOCK, not `cost_skipped`: this sheet DID get a cost
                # block, so calling it a skipped block would be false. Its absence is reported
                # here beside the column it would have been, which is also what lets the hub
                # say "costs and a total, but no margin, because ..." in one line.
                "margin_column": block["margin_column"],
                "margin_skipped": block["margin_skipped"],
                "rows": block["rows"],
            }
        if block["reason"]:
            cost_skipped[sn] = block["reason"]

        if skipped:
            skipped_by_sheet[sn] = sorted({s["col_letter"] for s in skipped})
        exported.append(sn)

    # ⚠️ THE FIDELITY GUARD IS THE CLIENT EXPORT'S, WITH ONE TERM ADJUSTED (planning Q10). That
    # guard asserts the saved copy carries the SAME number of formulas, merges, worksheets and
    # defined names as the workbook we loaded -- it exists to catch openpyxl silently mangling
    # the client's own formulas on a round trip. This export adds formulas ON PURPOSE, so the
    # expectation is `before + exactly the number we wrote`, and every other term stays
    # untouched. Asserting nothing about formulas instead would have thrown away the guard's
    # main job to make room for its one legitimate exception.
    expected = {**before, "formulas": before["formulas"] + formulas_written}

    saved = src_path + ".bcs.xlsx"
    try:
        wb.save(saved)
        wb.close()
        wb2 = openpyxl.load_workbook(saved, data_only=False)
        after = _fidelity_snapshot(wb2)
        wb2.close()
        _assert_fidelity(expected, after)
        with open(saved, "rb") as f:
            data = f.read()
    finally:
        try:
            os.unlink(saved)
        except OSError:
            pass

    ts = frappe.utils.now()[:19].replace("-", "").replace(":", "").replace(" ", "_")
    filename = f"{_safe_export_basename(display_name, boq_name)}_{_INTERNAL_SUFFIX}_{ts}.xlsx"
    return {
        "filename": filename,
        "content_type": _XLSX_CONTENT_TYPE,
        "content_base64": base64.b64encode(data).decode("ascii"),
        "exported_sheets": exported,
        "skipped_formula_columns": skipped_by_sheet,
        "remark_columns": remark_cols,
        "cost_blocks": cost_blocks,
        "cost_skipped": cost_skipped,
    }


@frappe.whitelist(methods=["POST"])
def export_priced_workbook_with_bcs(boq_name: str = None, sheet_names: Any = None) -> dict:
    """Generate the INTERNAL priced .xlsx -- the client export plus the BCS cost block -- for a
    ticked subset of a committed BoQ's sheets, and return its bytes (base64).

    ADMIN + ESTIMATION ONLY, gated FIRST, before any lookup or read. COPY-ON-WRITE: the
    original is fetched from S3 to a tempfile and copied to a second tempfile that is the ONLY
    thing stamped and saved; nothing is ever uploaded back.

    ⚠️ TEMPLATE-ORIGIN BoQs ARE REFUSED, BY NAME (planning Q5). A BoQ cloned from a master
    template has no source workbook at all -- its priced export is BUILT FROM SCRATCH by
    `export_template_workbook`, a different generator with its own layout, its own row
    compaction and its own formula writer. Supporting the cost block there is a second
    implementation of this whole module, so v1 says so out loud rather than failing obscurely
    on a missing `source_file_url`.

    Returns {filename, content_type, content_base64, exported_sheets, skipped_formula_columns,
    remark_columns, cost_blocks, cost_skipped}. NOTHING is written to the database -- in
    particular `last_exported_at` is left alone (see `_generate_internal_workbook`).
    URL: /api/method/nirmaan_stack.api.boq.wizard.export_bcs_writeback.export_priced_workbook_with_bcs
    """
    _require_bcs_export_access()
    if not boq_name:
        frappe.throw("boq_name is required.", title="Missing field: boq_name")
    if not frappe.db.exists("BOQs", boq_name):
        frappe.throw(f"BOQs '{boq_name}' not found.", title="Not found")
    names = _coerce_names(sheet_names)

    if (frappe.db.get_value("BOQs", boq_name, "origin") or "upload") == "template":
        frappe.throw(
            "This BoQ was created from a master template, so it has no original workbook to "
            "add cost columns to. Use 'Download priced tender' for a template BoQ.",
            title="Not available for template BoQs",
        )

    source_file_url, display_name = frappe.db.get_value(
        "BOQs", boq_name, ["source_file_url", "boq_name"]
    )
    if not source_file_url:
        frappe.throw(f"BOQs '{boq_name}' has no source_file_url set.", title="Missing source file")

    fetched = None
    copy = None
    try:
        fetched = _fetch_boq_file_to_tempfile(source_file_url)
        copy = fetched + ".work.xlsx"
        shutil.copy(fetched, copy)  # COPY-ON-WRITE -- never the original temp, never S3
        return _generate_internal_workbook(boq_name, names, copy, display_name)
    finally:
        for p in (fetched, copy):
            if p:
                try:
                    os.unlink(p)
                except OSError:
                    pass
