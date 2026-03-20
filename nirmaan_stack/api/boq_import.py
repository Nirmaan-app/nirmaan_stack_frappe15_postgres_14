import frappe
from frappe import _
import json
import os
import csv
import re
import tempfile
import requests


HEADER_KEYWORDS = [
    'description', 'item', 'particular', 'particulars',
    'quantity', 'qty', 'unit', 'uom',
    'rate', 'amount', 'total', 'supply', 'installation',
    'sl', 'sno', 's.no', 'sr', 'no'
]

FIELD_KEYWORDS = {
    'description': ['description', 'item', 'particular', 'particulars', 'item description', 'name of item', 'name of work'],
    'unit': ['unit', 'uom', 'units'],
    'quantity': ['quantity', 'qty', 'nos', 'number'],
    'supply_rate': ['supply rate', 'supply', 'material rate', 'material cost'],
    'installation_rate': ['installation rate', 'installation', 'labour rate', 'labor rate', 'labour cost', 'labor cost'],
}


def _get_file_path(file_url):
    """Resolve Frappe file URL to absolute path. Downloads S3 files to a temp file."""
    if file_url.startswith('/private/files/') or file_url.startswith('/files/'):
        site_path = frappe.get_site_path()
        return os.path.join(site_path, file_url.lstrip('/'))

    # S3 attachment: download to temp file
    if 'frappe_s3_attachment.controller.generate_file' in file_url:
        from nirmaan_stack.api.frappe_s3_attachment import get_s3_temp_url
        download_url = get_s3_temp_url(file_url)
        if download_url and download_url != file_url:
            resp = requests.get(download_url, timeout=60)
            resp.raise_for_status()
            # Determine extension from file_name param or URL
            import urllib.parse
            parsed = urllib.parse.urlparse(file_url)
            params = urllib.parse.parse_qs(parsed.query)
            file_name = params.get('file_name', ['upload.xlsx'])[0]
            _, ext = os.path.splitext(file_name)
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext or '.xlsx')
            tmp.write(resp.content)
            tmp.close()
            return tmp.name

    return file_url


def _read_all_rows(filepath, sheet_index=None):
    """Read all rows from an Excel or CSV file into a 2D list of strings."""
    _, ext = os.path.splitext(filepath)

    if ext.lower() == '.csv':
        all_rows = []
        with open(filepath, 'r', newline='', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for row in reader:
                all_rows.append([str(c).strip() if c else '' for c in row])
        return all_rows

    # Excel (.xlsx)
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if sheet_index is not None and 0 <= sheet_index < len(wb.sheetnames):
        ws = wb.worksheets[sheet_index]
    else:
        ws = wb.active
    _resolve_merged_cells(ws)

    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([str(c).strip() if c is not None else '' for c in row])
    return all_rows


def _resolve_merged_cells(ws):
    """Propagate top-left value to all cells in merged ranges."""
    for merged_range in list(ws.merged_cells.ranges):
        min_row, min_col = merged_range.min_row, merged_range.min_col
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merged_range))
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                ws.cell(row=row, column=col).value = top_left_value


def _detect_header_row(rows, scan_limit=10):
    """Score each row to find the most likely header row."""
    best_score = -1
    best_row = 0

    limit = min(scan_limit, len(rows))
    for i in range(limit):
        row = rows[i]
        score = 0
        non_empty = [c for c in row if c is not None and str(c).strip()]

        if len(non_empty) < 3:
            score -= 5

        score += len(non_empty)

        # +2 if all non-empty cells are non-numeric strings
        all_strings = all(
            isinstance(c, str) or (c is not None and not _is_numeric(c))
            for c in non_empty
        )
        if all_strings and len(non_empty) > 0:
            score += 2

        # +3 per keyword match
        for cell in non_empty:
            lower = str(cell).lower().strip()
            for kw in HEADER_KEYWORDS:
                if kw in lower:
                    score += 3

        if score > best_score:
            best_score = score
            best_row = i

    return best_row


def _is_numeric(value):
    """Check if a value is numeric."""
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False


def _suggest_mapping(columns):
    """Auto-suggest column mapping based on keyword matching."""
    mapping = {}
    mapped_targets = set()

    for col_name in columns:
        if not col_name:
            continue
        lower = str(col_name).lower().strip()

        for target_field, keywords in FIELD_KEYWORDS.items():
            if target_field in mapped_targets:
                continue
            if any(kw in lower for kw in keywords):
                mapping[col_name] = target_field
                mapped_targets.add(target_field)
                break

    return mapping


def _to_float(value, default=0.0):
    """Safely convert a value to float."""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


# ─── Preamble Detection ──────────────────────────────────────────────────────

PREAMBLE_KEYWORDS = [
    'design', 'manufacture', 'supply', 'installation', 'standards', 'note',
    'specification', 'scope', 'general', 'clause', 'requirement', 'comply',
    'applicable', 'as per', 'in accordance', 'shall be', 'including',
]

_NUMBERING_RE = re.compile(r'^\d+(\.\d+)*\s')


def _is_preamble_row(description, quantity, supply_rate, installation_rate):
    """
    Classify a row as preamble (section heading / descriptive paragraph) vs item.

    A row is a PREAMBLE if ALL numeric fields (quantity, supply_rate, installation_rate)
    are zero or empty. If any numeric field is non-zero, it's an ITEM.
    """
    # If any numeric value is present, it's an item — not a preamble
    if quantity > 0 or supply_rate > 0 or installation_rate > 0:
        return False

    if not description or not description.strip():
        return False

    # All numeric fields are zero and description is non-empty → preamble
    return True


def _extract_row_fields(row, col_index_map):
    """Extract description, unit, quantity, supply_rate, installation_rate from a row."""
    desc_idx = col_index_map.get('description')
    unit_idx = col_index_map.get('unit')
    qty_idx = col_index_map.get('quantity')
    supply_idx = col_index_map.get('supply_rate')
    install_idx = col_index_map.get('installation_rate')

    description = row[desc_idx] if desc_idx is not None and desc_idx < len(row) else ''
    unit = row[unit_idx] if unit_idx is not None and unit_idx < len(row) else ''
    quantity = _to_float(row[qty_idx]) if qty_idx is not None and qty_idx < len(row) else 0.0
    supply_rate = _to_float(row[supply_idx]) if supply_idx is not None and supply_idx < len(row) else 0.0
    installation_rate = _to_float(row[install_idx]) if install_idx is not None and install_idx < len(row) else 0.0

    return description, unit, quantity, supply_rate, installation_rate


def _classify_rows(data_rows, col_index_map):
    """
    Classify rows and aggregate consecutive preamble rows into full blocks.

    Returns:
        row_classifications
        preambles
        items
        total_amount
    """

    preambles = []
    seen_preamble_texts = {}

    items = []
    row_classifications = []
    total_amount = 0.0

    current_preamble_id = None
    current_block_lines = []
    block_start_row = None

    for row_idx, row in enumerate(data_rows):
        description, unit, quantity, supply_rate, installation_rate = _extract_row_fields(row, col_index_map)

        desc_clean = (description or "").strip()

        # Empty row
        if not desc_clean:
            row_classifications.append({
                'row_index': row_idx,
                'type': 'empty'
            })
            continue

        is_preamble = _is_preamble_row(desc_clean, quantity, supply_rate, installation_rate)

        # ─────────────────────────────────────────────
        # PREAMBLE ROW → accumulate block
        # ─────────────────────────────────────────────
        if is_preamble:
            if block_start_row is None:
                block_start_row = row_idx

            current_block_lines.append(desc_clean)

            row_classifications.append({
                'row_index': row_idx,
                'type': 'preamble'
            })

            continue

        # ─────────────────────────────────────────────
        # ITEM ROW → first finalize any open block
        # ─────────────────────────────────────────────
        if current_block_lines:
            full_text = "\n".join(current_block_lines).strip()

            if full_text not in seen_preamble_texts:
                pid = len(preambles)
                preambles.append({
                    'id': pid,
                    'text': full_text
                })
                seen_preamble_texts[full_text] = pid

            current_preamble_id = seen_preamble_texts[full_text]

            current_block_lines = []
            block_start_row = None

        # Now process item
        total_rate = supply_rate + installation_rate
        amount = quantity * total_rate
        total_amount += amount

        items.append({
            'description': description,
            'unit': unit,
            'quantity': quantity,
            'supply_rate': supply_rate,
            'installation_rate': installation_rate,
            'total_rate': total_rate,
            'amount': amount,
            'preamble_id': current_preamble_id,
            'source_row_index': row_idx,
        })

        row_classifications.append({
            'row_index': row_idx,
            'type': 'item'
        })

    # ─────────────────────────────────────────────
    # Handle file ending with preamble block
    # ─────────────────────────────────────────────
    if current_block_lines:
        full_text = "\n".join(current_block_lines).strip()

        if full_text not in seen_preamble_texts:
            pid = len(preambles)
            preambles.append({
                'id': pid,
                'text': full_text
            })
            seen_preamble_texts[full_text] = pid

    return row_classifications, preambles, items, total_amount
@frappe.whitelist()
def parse_excel_preview(file_url, field_column_map=None, data_start_row=None,
                       sheet_index=None):
    """
    Parse an Excel/CSV file and return preview data for the BOQ import wizard.
    Returns detected header row, columns, preview rows, and suggested mapping.
    and preamble classifications when field_column_map is provided.
    """
    filepath = _get_file_path(file_url)
    if not os.path.exists(filepath):
        frappe.throw(_("File not found: {0}").format(file_url))

    sheet_idx = int(sheet_index) if sheet_index is not None else None
    all_rows = _read_all_rows(filepath, sheet_index=sheet_idx)

    if not all_rows:
        frappe.throw(_("The file appears to be empty"))

    # Detect header row (1-based for user display, 0-based internally)
    detected_header = _detect_header_row(all_rows)
    columns = all_rows[detected_header] if detected_header < len(all_rows) else []

    # Data rows after header
    data_rows = [
        row for row in all_rows[detected_header + 1:]
        if any(c.strip() for c in row)
    ]

    # Auto-suggest mapping
    suggested_mapping = _suggest_mapping(columns)

    result = {
        'detected_header_row': detected_header,
        'columns': columns,
        'preview_rows': data_rows[:100],
        'raw_rows': all_rows[:15],
        'total_rows': len(data_rows),
        'suggested_mapping': suggested_mapping,
    }

    # If field_column_map is provided, classify rows for preamble detection
    if field_column_map is not None:
        if isinstance(field_column_map, str):
            field_column_map = json.loads(field_column_map)
        start_row = int(data_start_row) if data_start_row is not None else 0
        col_index_map = {k: int(v) for k, v in field_column_map.items()}
        classify_data_rows = [
            row for row in all_rows[start_row:]
            if any(c.strip() for c in row)
        ]
        row_classifications, preambles, classified_items, _ = _classify_rows(
            classify_data_rows, col_index_map
        )
        result['row_classifications'] = row_classifications
        result['preambles'] = preambles
        result['items_count'] = len(classified_items)
        result['preambles_count'] = len(preambles)

    return result


@frappe.whitelist()
def import_boq_data(file_url, project, work_package, zone=None,
                    header_row=None, column_mapping=None,
                    field_column_map=None, data_start_row=None,
                    sheet_index=None):
    """
    Import BOQ data from an Excel file into a new BOQ document.

    Supports two modes:
    1. New format: field_column_map (targetField→colIndex) + data_start_row
    2. Legacy format: header_row + column_mapping (sourceColName→targetField)

    Preamble rows (section headings / descriptive paragraphs) are automatically
    detected and separated from actual items. Items are linked to their parent
    preamble via preamble_id.
    """
    filepath = _get_file_path(file_url)
    if not os.path.exists(filepath):
        frappe.throw(_("File not found: {0}").format(file_url))

    sheet_idx = int(sheet_index) if sheet_index is not None else None
    all_rows = _read_all_rows(filepath, sheet_index=sheet_idx)

    # Determine col_index_map and data_rows based on format
    if field_column_map is not None:
        # New format: field_column_map = {targetField: colIndex}
        if isinstance(field_column_map, str):
            field_column_map = json.loads(field_column_map)
        data_start_row = int(data_start_row) if data_start_row is not None else 0

        col_index_map = {k: int(v) for k, v in field_column_map.items()}
        data_rows = [
            row for row in all_rows[data_start_row:]
            if any(c.strip() for c in row)
        ]
        stored_header_row = data_start_row
        stored_mapping = json.dumps(field_column_map)
    else:
        # Legacy format: header_row + column_mapping
        header_row = int(header_row)
        if isinstance(column_mapping, str):
            column_mapping = json.loads(column_mapping)

        if header_row >= len(all_rows):
            frappe.throw(_("Header row index out of range"))

        columns = all_rows[header_row]
        data_rows = [
            row for row in all_rows[header_row + 1:]
            if any(c.strip() for c in row)
        ]

        col_index_map = {}
        for source_col, target_field in column_mapping.items():
            if not target_field:
                continue
            try:
                idx = columns.index(source_col)
                col_index_map[target_field] = idx
            except ValueError:
                continue

        stored_header_row = header_row + 1
        stored_mapping = json.dumps(column_mapping)

    # Classify rows — separate preambles from items
    _, preambles, items, total_amount = _classify_rows(data_rows, col_index_map)

    # Fix source_row_index to be absolute (relative to original sheet)
    base_offset = data_start_row if data_start_row else 0
    for item in items:
        item['source_row_index'] = base_offset + item['source_row_index'] + 1

    if not items:
        frappe.throw(_("No valid data rows found after applying the column mapping"))

    boq = frappe.get_doc({
        'doctype': 'BOQ',
        'project': project,
        'work_package': work_package,
        'zone': zone or '',
        'source_file': file_url,
        'status': 'Imported',
        'header_row': stored_header_row,
        'column_mapping': stored_mapping,
        'preambles': json.dumps(preambles) if preambles else None,
        'total_items': len(items),
        'total_amount': total_amount,
        'items': items,
    })

    boq.insert()
    frappe.db.commit()

    return {
        'status': 'success',
        'boq_name': boq.name,
        'items_count': len(items),
        'total_amount': total_amount,
        'preambles_count': len(preambles),
    }
