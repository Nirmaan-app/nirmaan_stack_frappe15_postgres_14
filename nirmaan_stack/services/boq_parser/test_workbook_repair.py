# Copyright (c) 2026, Nirmaan (Stratos Infra Technologies Pvt. Ltd.) and contributors
# See license.txt

"""Tests for `workbook_repair` -- the openpyxl-hostile-workbook repair pass.

Damaged fixtures are BUILT AT TEST TIME by injecting the exact out-of-spec XML found
in the two real customer BoQs into an existing synthetic fixture. Nothing binary is
committed, and the undamaged original stays on hand as the oracle: a repaired workbook
must read back IDENTICALLY to it, which is the only way to prove the repair changed
presentation and not data.
"""

import shutil
import tempfile
import unittest
import zipfile
from pathlib import Path

import openpyxl

from nirmaan_stack.services.boq_parser.reader import BoqReader
from nirmaan_stack.services.boq_parser.tests.fixtures.generate_synthetic import (
    generate_all,
)
from nirmaan_stack.services.boq_parser.workbook_repair import (
    needs_repair,
    repair_in_place,
)

_FIXTURES = Path(__file__).parent / "tests" / "fixtures"

# The literal defects measured on the real files (2026-08-20).
_BAD_FAMILY = '<family val="38"/>'
# NOTE the localSheetId -- a GLOBAL built-in is never parsed by openpyxl and so is
# deliberately NOT a defect. See `test_global_builtin_defined_names_are_left_alone`.
_BAD_DEFINED_NAME = (
    '<definedName name="_xlnm.Print_Titles" localSheetId="0">#N/A</definedName>'
)


def _rewrite_part(src: Path, dst: Path, part: str, mutate) -> None:
    """Copy `src` to `dst`, passing one zip member's text through `mutate`."""
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(
        dst, "w", zipfile.ZIP_DEFLATED
    ) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename == part:
                data = mutate(data.decode("utf-8")).encode("utf-8")
            zout.writestr(info, data)


class TestWorkbookRepair(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        generate_all()
        cls.clean = _FIXTURES / "synthetic_simple.xlsx"

    def setUp(self):
        self._dir = Path(tempfile.mkdtemp())

    def tearDown(self):
        shutil.rmtree(self._dir, ignore_errors=True)

    # -- helpers -------------------------------------------------------- #

    def _damaged_styles(self) -> Path:
        """A copy carrying the Fidelity defect: <family val="38"/>."""
        out = self._dir / "bad_styles.xlsx"
        _rewrite_part(
            self.clean,
            out,
            "xl/styles.xml",
            lambda x: x.replace("<fonts", _BAD_FAMILY + "<fonts", 1)
            if _BAD_FAMILY not in x
            else x,
        )
        return out

    def _damaged_defined_names(self) -> Path:
        """A copy carrying the Coimbatore defect: an _xlnm.* name valued #N/A."""
        out = self._dir / "bad_names.xlsx"
        _rewrite_part(
            self.clean,
            out,
            "xl/workbook.xml",
            lambda x: x.replace(
                "<sheets>", f"<definedNames>{_BAD_DEFINED_NAME}</definedNames><sheets>", 1
            ),
        )
        return out

    def _values(self, path: Path):
        wb = openpyxl.load_workbook(path, data_only=True)
        try:
            return {
                name: [list(r) for r in wb[name].iter_rows(values_only=True)]
                for name in wb.sheetnames
            }
        finally:
            wb.close()

    # -- a healthy workbook is untouched -------------------------------- #

    def test_healthy_workbook_needs_no_repair(self):
        self.assertFalse(needs_repair(str(self.clean)))

    def test_healthy_workbook_repair_is_a_noop(self):
        """repair_in_place must not fire, and must not rewrite the bytes."""
        target = self._dir / "healthy.xlsx"
        shutil.copy2(self.clean, target)
        before = target.read_bytes()

        self.assertEqual(repair_in_place(str(target)), [])
        self.assertEqual(target.read_bytes(), before)

    def _opens(self, path: Path) -> bool:
        try:
            openpyxl.load_workbook(path, data_only=True, read_only=True).close()
            return True
        except Exception:
            return False

    def test_detection_precision_against_the_real_fixture_corpus(self):
        """A fixture may be FLAGGED only if openpyxl genuinely refuses it.

        The corpus holds real customer BoQs, so this is the guard that keeps the rule
        honest against files nobody wrote for a test. It caught a real false positive:
        `Kohler-BOQ- 06-04-26.xlsx` carries `_xlnm.Print_Titles` = #N/A but GLOBALLY,
        which openpyxl never parses -- so it opens fine and must not be rewritten.
        """
        for path in sorted(_FIXTURES.glob("*.xlsx")):
            if not needs_repair(str(path)):
                continue
            with self.subTest(fixture=path.name):
                self.assertFalse(
                    self._opens(path),
                    f"{path.name} was flagged for repair but opens fine -- "
                    "the detection rule is too eager",
                )

    def test_flagged_fixtures_open_after_repair(self):
        """The other half: anything we flag, we must actually fix.

        Covers `R0_CIVIL INTERIOR & MEP_TABLESPACE...xlsx`, a real fixture that has
        been unreadable in this repo's own corpus.
        """
        flagged = [p for p in sorted(_FIXTURES.glob("*.xlsx")) if needs_repair(str(p))]
        for path in flagged:
            with self.subTest(fixture=path.name):
                target = self._dir / path.name
                shutil.copy2(path, target)
                self.assertNotEqual(repair_in_place(str(target)), [])
                self.assertTrue(self._opens(target), f"{path.name} still unreadable")

    def test_global_builtin_defined_names_are_left_alone(self):
        """A GLOBAL `_xlnm.*` = #N/A never reaches openpyxl's print-title parser."""
        out = self._dir / "global_builtin.xlsx"
        _rewrite_part(
            self.clean,
            out,
            "xl/workbook.xml",
            lambda x: x.replace(
                "<sheets>",
                '<definedNames><definedName name="_xlnm.Print_Titles">#N/A'
                "</definedName></definedNames><sheets>",
                1,
            ),
        )
        self.assertFalse(needs_repair(str(out)))
        self.assertEqual(repair_in_place(str(out)), [])

    # -- the font-family defect (Fidelity) ------------------------------ #

    def test_font_family_defect_is_detected_and_repaired(self):
        damaged = self._damaged_styles()
        self.assertTrue(needs_repair(str(damaged)))

        fired = repair_in_place(str(damaged))
        self.assertIn("font_family_clamped", fired)
        self.assertFalse(needs_repair(str(damaged)))

        styles = zipfile.ZipFile(damaged).read("xl/styles.xml").decode("utf-8")
        self.assertNotIn(_BAD_FAMILY, styles)

    # -- the defined-name defect (Coimbatore) --------------------------- #

    def test_defined_name_defect_is_detected_and_repaired(self):
        damaged = self._damaged_defined_names()
        self.assertTrue(needs_repair(str(damaged)))

        fired = repair_in_place(str(damaged))
        self.assertIn("bad_defined_names_dropped", fired)
        self.assertFalse(needs_repair(str(damaged)))

        # openpyxl raises on this defect at load time, so a clean open IS the assertion.
        wb = openpyxl.load_workbook(damaged, data_only=True)
        wb.close()

    def test_user_defined_names_and_ref_errors_are_left_alone(self):
        """The rule is narrow on purpose -- openpyxl tolerates these."""
        out = self._dir / "ref_names.xlsx"
        _rewrite_part(
            self.clean,
            out,
            "xl/workbook.xml",
            lambda x: x.replace(
                "<sheets>",
                '<definedNames><definedName name="QTY">#REF!</definedName>'
                "</definedNames><sheets>",
                1,
            ),
        )
        self.assertFalse(needs_repair(str(out)))
        self.assertEqual(repair_in_place(str(out)), [])

    # -- the repair is surgical ----------------------------------------- #

    def test_repair_preserves_every_other_zip_entry_byte_for_byte(self):
        """This is what keeps an .xlsm workbook's vbaProject.bin intact."""
        damaged = self._damaged_styles()
        original_entries = {
            i.filename: zipfile.ZipFile(damaged).read(i.filename)
            for i in zipfile.ZipFile(damaged).infolist()
        }

        repair_in_place(str(damaged))

        with zipfile.ZipFile(damaged) as zf:
            self.assertEqual(list(zf.namelist()), list(original_entries.keys()))
            for name in zf.namelist():
                if name == "xl/styles.xml":
                    continue
                with self.subTest(entry=name):
                    self.assertEqual(zf.read(name), original_entries[name])

    def test_repaired_workbook_reads_back_identical_to_the_original(self):
        """Cell data must survive the repair exactly -- the whole point."""
        expected = self._values(self.clean)

        for factory in (self._damaged_styles, self._damaged_defined_names):
            with self.subTest(defect=factory.__name__):
                damaged = factory()
                repair_in_place(str(damaged))
                self.assertEqual(self._values(damaged), expected)

    def test_boqreader_opens_a_repaired_workbook(self):
        """End-to-end: the reader every phase uses must accept the repaired file."""
        clean_reader = BoqReader(str(self.clean))
        expected_sheets = clean_reader.list_sheets()

        for factory in (self._damaged_styles, self._damaged_defined_names):
            with self.subTest(defect=factory.__name__):
                damaged = factory()
                repair_in_place(str(damaged))
                self.assertEqual(BoqReader(str(damaged)).list_sheets(), expected_sheets)

    # -- failure modes --------------------------------------------------- #

    def test_a_non_zip_file_is_reported_as_not_repairable(self):
        """A genuinely broken file belongs to the caller's error path, not here."""
        junk = self._dir / "junk.xlsx"
        junk.write_bytes(b"this is not a zip archive")
        self.assertFalse(needs_repair(str(junk)))

    def test_a_missing_file_does_not_raise(self):
        self.assertFalse(needs_repair(str(self._dir / "nope.xlsx")))
