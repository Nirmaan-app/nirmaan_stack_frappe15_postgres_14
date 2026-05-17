"""
Multi-area header pattern detection for the BoQ parser.

Examines one or two header rows and returns a MultiAreaPattern describing the
per-area column structure, or None when no pattern matches.

Three locked patterns (v5.3 §3):

  Pattern 1 — adjacent area-only labels (liberal):
      [area1] [area2] ... [TOTAL QTY?] [RATE] ...
      Single row; area cells are consecutive and not reserved keywords.
      No per-area amount columns (amount_columns=None).

  Pattern 2 — two-row merged header:
      Top row: merged cells, each spanning exactly two columns, value=area name.
      Bottom row: [QTY][AMOUNT] under each merge.

  Pattern 3 — single-row alternating label/AMOUNT pairs:
      [area1] [AMOUNT] [area2] [AMOUNT] ... [TOTAL QTY?]
      The area-name column also serves as the per-area qty column.

Priority routing:
  header_row_count == 1:  Pattern 3 → Pattern 1 (bottom) → None
  header_row_count == 2:  Pattern 2 → Pattern 3 → Pattern 1 (bottom)
                          → Pattern 1 (top, last-resort) → None

No Frappe imports — pure Python, fully testable in isolation.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Optional

from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import range_boundaries

from nirmaan_stack.services.boq_parser.reader import RawRow


# ------------------------------------------------------------------
# Compiled regexes (locked in v5.3 §3)
# ------------------------------------------------------------------

TOTAL_QTY_PATTERN = re.compile(
    r"^\s*(total|grand|sum|sub|subtotal)?\s*(qty|quantity|nos)\.?\s*$",
    re.IGNORECASE,
)

_QTY_CELL_PATTERN = re.compile(
    r"^\s*(qty|quantity|nos)\.?\s*$",
    re.IGNORECASE,
)

_AMOUNT_CELL_PATTERN = re.compile(
    r"^\s*amount\s*$",
    re.IGNORECASE,
)

# Pattern 2-rate detection accepts "RATE" and "RATES" (plural) per §9 #62 F3b fix (Phase 1.9d 2026-05-17).
_RATE_CELL_PATTERN = re.compile(
    r"^\s*rates?\s*$",
    re.IGNORECASE,
)


# ------------------------------------------------------------------
# Result dataclass
# ------------------------------------------------------------------

@dataclass
class MultiAreaPattern:
    """Detected multi-area column structure for a sheet's header row(s).

    Returned by detect_multi_area_pattern() when one of the locked patterns is
    found.  None is returned when no pattern matches.

    qty_columns and amount_columns (when not None) are parallel to areas:
      areas[i] is the area name, qty_columns[i] is its qty column (1-based),
      amount_columns[i] (if set) is its amount column.

    For Pattern 1, amount_columns is None — that pattern has per-area qty only.
    For Pattern 2-rate, rate_columns is populated (one per area, 1-based).
    detected_on_row is the 1-based row number where area names were found.

    pattern is int (1, 2, 3) for the three original patterns, or the string
    "pattern_2_rate" for the three-col-per-area Raheja shape (Phase 1.9a).
    """
    pattern: int | str
    areas: list[str]
    qty_columns: list[int]        # 1-based column indices, parallel to areas
    amount_columns: list[int] | None = None
    rate_columns: list[int] | None = None   # populated for pattern_2_rate; None otherwise
    detected_on_row: int = 0


# ------------------------------------------------------------------
# Public entry point
# ------------------------------------------------------------------

def detect_multi_area_pattern(
    bottom_header_row: RawRow,
    reserved_keywords: list[str],
    top_header_row: RawRow | None = None,
) -> MultiAreaPattern | None:
    """Examine header row(s) and return a MultiAreaPattern, or None.

    bottom_header_row: the header row (for single-row sheets) or the
        bottom header row (for two-row merged-header sheets).
    reserved_keywords: from GlobalSettings.multi_area_reserved_keywords.
        Matched case-insensitively.
    top_header_row: the row above the bottom header, if the sheet uses a
        two-row merged header (header_row_count == 2).  Pass None for
        single-row headers.

    Priority routing when top_header_row is provided (2-row mode):
        Pattern 2-rate → Pattern 2 → Pattern 3 (bottom) → Pattern 1 (bottom)
        → Pattern 1 (top, last-resort fallback) → None

    Pattern 2-rate is tried before Pattern 2 because it is a stricter variant
    (3-col-per-area vs 2-col-per-area). Without this order, a Pattern 2-rate
    sheet could match as Pattern 2, misclassifying the Rate column.

    Priority routing when top_header_row is None (1-row mode):
        Pattern 3 → Pattern 1 → None
    """
    kw_set = {kw.upper().strip() for kw in reserved_keywords}

    if top_header_row is not None:
        return (
            _try_pattern_2_rate(top_header_row, bottom_header_row, kw_set)
            or _try_pattern_2(top_header_row, bottom_header_row, kw_set)
            or _try_pattern_3(bottom_header_row, kw_set)
            or _try_pattern_1(bottom_header_row, kw_set)
            or _try_pattern_1(top_header_row, kw_set)
        )

    return _try_pattern_3(bottom_header_row, kw_set) or _try_pattern_1(bottom_header_row, kw_set)


# ------------------------------------------------------------------
# Internal helpers
# ------------------------------------------------------------------

def _sorted_cols(row: RawRow) -> list[tuple[int, str]]:
    """Return (col_index_1based, col_letter) tuples sorted left-to-right."""
    return sorted(
        ((column_index_from_string(c), c) for c in row.cells),
        key=lambda x: x[0],
    )


def _is_reserved(value: object, kw_set: set[str]) -> bool:
    """Check if a cell value is a reserved keyword (or empty).

    Empty and None cells are always reserved so patterns skip them.
    Normalizes whitespace (collapsing newlines/tabs/multiple spaces to single
    spaces) before comparison. Also tries the form with a trailing parenthetical
    stripped (e.g., 'AMOUNT (INR)' → 'AMOUNT') to handle currency-suffix
    compound variants without enumerating each one.
    """
    if value is None:
        return True
    text = str(value).upper()
    # Whitespace normalization: collapse any whitespace sequence to single space
    text = " ".join(text.split())
    if not text:
        return True
    if text in kw_set:
        return True
    # Trailing-parenthetical strip: e.g. "AMOUNT (INR)" → "AMOUNT"
    if text.endswith(")") and "(" in text:
        stripped = text[:text.rindex("(")].strip()
        if stripped and stripped in kw_set:
            return True
    return False


def _try_pattern_1(row: RawRow, kw_set: set[str]) -> MultiAreaPattern | None:
    """
    Pattern 1: consecutive non-reserved cells are treated as area names.

    Terminates at the first TOTAL_QTY_PATTERN match or at any reserved
    keyword encountered AFTER at least one area name has been collected.
    Requires at least 2 area names.
    """
    areas: list[str] = []
    qty_cols: list[int] = []

    for col_idx, col_letter in _sorted_cols(row):
        cell = row.cells[col_letter]
        val = cell.value
        text = str(val or "").strip()

        if cell.merged_range is not None and not cell.is_merged_origin:
            continue  # covered cell — value propagated from origin by reader (Part A1), not an area name

        if TOTAL_QTY_PATTERN.match(text):
            break

        if _is_reserved(val, kw_set):
            if areas:
                # A reserved cell after collecting area names ends the run.
                break
            continue

        areas.append(text)
        qty_cols.append(col_idx)

    if len(areas) >= 2:
        return MultiAreaPattern(
            pattern=1,
            areas=areas,
            qty_columns=qty_cols,
            amount_columns=None,
            detected_on_row=row.row_number,
        )
    return None


def _try_pattern_3(row: RawRow, kw_set: set[str]) -> MultiAreaPattern | None:
    """
    Pattern 3: alternating [area_name][AMOUNT] pairs.

    The area-name column doubles as the per-area qty column.
    Reserved-keyword cells between pairs are skipped without breaking.
    Requires at least 2 complete pairs.
    """
    areas: list[str] = []
    qty_cols: list[int] = []
    amount_cols: list[int] = []

    sorted_cells = _sorted_cols(row)
    i = 0
    while i < len(sorted_cells):
        col_idx, col_letter = sorted_cells[i]
        cell = row.cells[col_letter]
        val = cell.value
        text = str(val or "").strip()

        if TOTAL_QTY_PATTERN.match(text):
            break

        if _is_reserved(val, kw_set):
            i += 1
            continue

        # Non-reserved candidate — check if immediately followed by an AMOUNT cell.
        if i + 1 < len(sorted_cells):
            next_col_idx, next_col_letter = sorted_cells[i + 1]
            next_val = str(row.cells[next_col_letter].value or "").strip()
            if _AMOUNT_CELL_PATTERN.match(next_val):
                areas.append(text)
                qty_cols.append(col_idx)
                amount_cols.append(next_col_idx)
                i += 2
                continue

        # Not followed by AMOUNT — not a Pattern 3 pair; skip.
        i += 1

    if len(areas) >= 2:
        return MultiAreaPattern(
            pattern=3,
            areas=areas,
            qty_columns=qty_cols,
            amount_columns=amount_cols,
            detected_on_row=row.row_number,
        )
    return None


def _try_pattern_2_rate(
    top_row: RawRow,
    bottom_row: RawRow,
    kw_set: set[str],
) -> MultiAreaPattern | None:
    """
    Pattern 2-rate: two-row merged header where each merge spans exactly 3 columns.

    Top row: merge origins (is_merged_origin=True) with area name as value.
    Each merge spans exactly 3 columns (one row tall only).
    Bottom row: [QTY | RATE | AMOUNT] under each 3-column merge.

    Requires at least 2 qualifying merges. Results sorted left-to-right by qty column.

    For Phase 1.9a, all per-area rates default to combined_rate regardless of whether
    the bottom-row sub-label is "Supply Rate" or "Install Rate". Split-rate detection
    (mapping sub-label text to supply_rate / install_rate kinds) is deferred to a
    future iteration.
    """
    areas: list[str] = []
    qty_cols: list[int] = []
    rate_cols: list[int] = []
    amount_cols: list[int] = []

    for col_idx, col_letter in _sorted_cols(top_row):
        cell = top_row.cells[col_letter]
        if not cell.is_merged_origin or cell.merged_range is None:
            continue

        if _is_reserved(cell.value, kw_set):
            continue

        try:
            min_col, min_row_r, max_col, max_row_r = range_boundaries(cell.merged_range)
        except Exception:
            continue

        # Only handle single-row merges spanning exactly 3 columns.
        if min_row_r != max_row_r:
            continue
        if (max_col - min_col + 1) != 3:
            continue

        qty_letter = get_column_letter(min_col)
        rate_letter = get_column_letter(min_col + 1)
        amt_letter = get_column_letter(max_col)

        qty_cell = bottom_row.cells.get(qty_letter)
        rate_cell = bottom_row.cells.get(rate_letter)
        amt_cell = bottom_row.cells.get(amt_letter)

        if qty_cell is None or rate_cell is None or amt_cell is None:
            continue

        qty_text = str(qty_cell.value or "").strip()
        rate_text = str(rate_cell.value or "").strip()
        amt_text = str(amt_cell.value or "").strip()

        if not _QTY_CELL_PATTERN.match(qty_text):
            continue
        if not _RATE_CELL_PATTERN.match(rate_text):
            continue
        if not _AMOUNT_CELL_PATTERN.match(amt_text):
            continue

        areas.append(str(cell.value).strip())
        qty_cols.append(min_col)
        rate_cols.append(min_col + 1)
        amount_cols.append(max_col)

    if len(areas) >= 2:
        quads = sorted(zip(areas, qty_cols, rate_cols, amount_cols), key=lambda x: x[1])
        return MultiAreaPattern(
            pattern="pattern_2_rate",
            areas=[q[0] for q in quads],
            qty_columns=[q[1] for q in quads],
            rate_columns=[q[2] for q in quads],
            amount_columns=[q[3] for q in quads],
            detected_on_row=top_row.row_number,
        )
    return None


def _try_pattern_2(
    top_row: RawRow,
    bottom_row: RawRow,
    kw_set: set[str],
) -> MultiAreaPattern | None:
    """
    Pattern 2: two-row merged header.

    Top row: merge origins (is_merged_origin=True) whose value is an area name.
    Each merge must span exactly 2 columns.
    Bottom row: the first column under each merge must match _QTY_CELL_PATTERN
    and the second must match _AMOUNT_CELL_PATTERN.
    Requires at least 2 qualifying merges.
    Results are sorted left-to-right by qty column.
    """
    areas: list[str] = []
    qty_cols: list[int] = []
    amount_cols: list[int] = []

    for col_idx, col_letter in _sorted_cols(top_row):
        cell = top_row.cells[col_letter]
        if not cell.is_merged_origin or cell.merged_range is None:
            continue

        if _is_reserved(cell.value, kw_set):
            continue

        try:
            min_col, min_row_r, max_col, max_row_r = range_boundaries(cell.merged_range)
        except Exception:
            continue

        # Only handle single-row merges spanning exactly 2 columns.
        if min_row_r != max_row_r:
            continue
        if (max_col - min_col + 1) != 2:
            continue

        qty_letter = get_column_letter(min_col)
        amt_letter = get_column_letter(max_col)

        qty_cell = bottom_row.cells.get(qty_letter)
        amt_cell = bottom_row.cells.get(amt_letter)
        if qty_cell is None or amt_cell is None:
            continue

        qty_text = str(qty_cell.value or "").strip()
        amt_text = str(amt_cell.value or "").strip()

        if not _QTY_CELL_PATTERN.match(qty_text):
            continue
        if not _AMOUNT_CELL_PATTERN.match(amt_text):
            continue

        areas.append(str(cell.value).strip())
        qty_cols.append(min_col)
        amount_cols.append(max_col)

    if len(areas) >= 2:
        # Sort by qty column to guarantee left-to-right order.
        pairs = sorted(zip(areas, qty_cols, amount_cols), key=lambda x: x[1])
        return MultiAreaPattern(
            pattern=2,
            areas=[p[0] for p in pairs],
            qty_columns=[p[1] for p in pairs],
            amount_columns=[p[2] for p in pairs],
            detected_on_row=top_row.row_number,
        )
    return None
