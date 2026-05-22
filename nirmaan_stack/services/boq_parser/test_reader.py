# Copyright (c) 2026, Nirmaan (Stratos Infra Technologies Pvt. Ltd.) and contributors
# See license.txt

import unittest
from pathlib import Path

from nirmaan_stack.services.boq_parser.tests.fixtures.generate_synthetic import (
    generate_all,
)
from nirmaan_stack.services.boq_parser.reader import BoqReader

_FIXTURES = Path(__file__).parent / "tests" / "fixtures"


def _p(name: str) -> str:
    return str(_FIXTURES / name)


class TestBoqReader(unittest.TestCase):
    """Group 2a-B: Reader functionality."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        # Ensure all fixture files exist (idempotent)
        generate_all()

    # ------------------------------------------------------------------ #
    # Test 7 — list_sheets() exact names                                   #
    # ------------------------------------------------------------------ #

    def test_list_sheets_returns_exact_names(self):
        """list_sheets() preserves trailing whitespace and capitalisation."""
        reader = BoqReader(_p("synthetic_trailing_spaces.xlsx"))
        sheets = reader.list_sheets()
        self.assertIn("Sheet One", sheets)
        self.assertIn("Trailing  ", sheets)   # two trailing spaces

    # ------------------------------------------------------------------ #
    # Test 8 — iter_rows() row numbers and cell counts                     #
    # ------------------------------------------------------------------ #

    def test_iter_rows_yields_correct_row_numbers(self):
        """iter_rows() yields RawRow with row_number 1..6 for the simple fixture."""
        reader = BoqReader(_p("synthetic_simple.xlsx"))
        rows = list(reader.iter_rows("Sheet1", start_row=1, end_row=6))
        self.assertEqual(len(rows), 6)
        for i, row in enumerate(rows, start=1):
            self.assertEqual(row.row_number, i)

    def test_iter_rows_cells_dict_keyed_by_column_letter(self):
        """Row 1 (header row) cells are keyed by uppercase column letters."""
        reader = BoqReader(_p("synthetic_simple.xlsx"))
        rows = list(reader.iter_rows("Sheet1", start_row=1, end_row=1))
        self.assertEqual(len(rows), 1)
        header_row = rows[0]
        self.assertIn("A", header_row.cells)
        self.assertIn("B", header_row.cells)
        self.assertEqual(header_row.cells["A"].value, "Sl.No.")
        self.assertEqual(header_row.cells["B"].value, "Description")

    # ------------------------------------------------------------------ #
    # Test 9 — computed values (data_only=True)                            #
    # ------------------------------------------------------------------ #

    def test_computed_value_read_correctly(self):
        """
        F2 contains the pre-computed value 500 (written directly, not as formula).
        data_only=True workbook returns 500.
        """
        reader = BoqReader(_p("synthetic_simple.xlsx"))
        rows = list(reader.iter_rows("Sheet1", start_row=2, end_row=2))
        f2 = rows[0].get_cell("F")
        self.assertIsNotNone(f2)
        self.assertEqual(f2.value, 500)
        self.assertFalse(f2.is_formula)

    # ------------------------------------------------------------------ #
    # Test 10 — formula detection                                          #
    # ------------------------------------------------------------------ #

    def test_formula_cell_detected(self):
        """
        G2 was written as '=D2*E2' (a raw string starting with '=').
        Reader reports is_formula=True and formula='=D2*E2'.
        """
        reader = BoqReader(_p("synthetic_simple.xlsx"))
        rows = list(reader.iter_rows("Sheet1", start_row=2, end_row=2))
        g2 = rows[0].get_cell("G")
        self.assertIsNotNone(g2)
        self.assertTrue(g2.is_formula)
        self.assertEqual(g2.formula, "=D2*E2")

    def test_non_formula_cell_not_marked_as_formula(self):
        """A plain text/numeric cell must have is_formula=False and formula=None."""
        reader = BoqReader(_p("synthetic_simple.xlsx"))
        rows = list(reader.iter_rows("Sheet1", start_row=1, end_row=1))
        a1 = rows[0].get_cell("A")
        self.assertIsNotNone(a1)
        self.assertFalse(a1.is_formula)
        self.assertIsNone(a1.formula)

    # ------------------------------------------------------------------ #
    # Test 11 — bold formatting                                            #
    # ------------------------------------------------------------------ #

    def test_bold_cell_detected(self):
        """B5 in synthetic_simple has bold=True; other B cells are not bold."""
        reader = BoqReader(_p("synthetic_simple.xlsx"))
        rows = {r.row_number: r for r in reader.iter_rows("Sheet1", start_row=1, end_row=6)}

        b5 = rows[5].get_cell("B")
        self.assertIsNotNone(b5)
        self.assertTrue(b5.font_bold)

        b1 = rows[1].get_cell("B")
        self.assertIsNotNone(b1)
        self.assertFalse(b1.font_bold)

    # ------------------------------------------------------------------ #
    # Test 12 — merged ranges                                              #
    # ------------------------------------------------------------------ #

    def test_merged_origin_detected(self):
        """A3 is the top-left of merged A3:A4 — is_merged_origin=True."""
        reader = BoqReader(_p("synthetic_merged_header.xlsx"))
        rows = {r.row_number: r for r in reader.iter_rows("Data", start_row=3, end_row=4)}

        a3 = rows[3].get_cell("A")
        self.assertIsNotNone(a3)
        self.assertTrue(a3.is_merged_origin)
        self.assertEqual(a3.merged_range, "A3:A4")

    def test_merged_non_origin_not_flagged(self):
        """A4 is inside the merge but not the origin — is_merged_origin=False."""
        reader = BoqReader(_p("synthetic_merged_header.xlsx"))
        rows = {r.row_number: r for r in reader.iter_rows("Data", start_row=3, end_row=4)}

        a4 = rows[4].get_cell("A")
        self.assertIsNotNone(a4)
        self.assertFalse(a4.is_merged_origin)

    # ------------------------------------------------------------------ #
    # Test 13 — detect_header_row on simple fixture                        #
    # ------------------------------------------------------------------ #

    def test_detect_header_row_simple(self):
        """Row 1 of synthetic_simple contains header keywords — must return 1."""
        reader = BoqReader(_p("synthetic_simple.xlsx"))
        result = reader.detect_header_row("Sheet1")
        self.assertEqual(result, 1)

    # ------------------------------------------------------------------ #
    # Test 14 — detect_header_row on merged-header fixture                 #
    # ------------------------------------------------------------------ #

    def test_detect_header_row_merged(self):
        """
        Row 3 of synthetic_merged_header has 'Sl.No.', 'Description', etc.
        Rows 1-2 are title/empty rows. Must return 3.
        """
        reader = BoqReader(_p("synthetic_merged_header.xlsx"))
        result = reader.detect_header_row("Data")
        self.assertEqual(result, 3)

    # ------------------------------------------------------------------ #
    # Test 15 — detect_blank_columns                                       #
    # ------------------------------------------------------------------ #

    def test_detect_blank_columns(self):
        """
        synthetic_blank_cols has content in B, C, D; columns A and Z are blank.
        detect_blank_columns() must include A and Z in the returned set.
        """
        reader = BoqReader(_p("synthetic_blank_cols.xlsx"))
        blank = reader.detect_blank_columns("Sheet1", scan_rows=10)
        self.assertIn("A", blank)
        self.assertIn("Z", blank)
        # Content columns must not appear as blank
        self.assertNotIn("B", blank)
        self.assertNotIn("C", blank)
        self.assertNotIn("D", blank)

    # ------------------------------------------------------------------ #
    # Test 16 — empty workbook edge case                                   #
    # ------------------------------------------------------------------ #

    def test_empty_workbook_iter_rows_returns_empty(self):
        """iter_rows() on an empty sheet must yield nothing and not raise."""
        reader = BoqReader(_p("synthetic_empty.xlsx"))
        rows = list(reader.iter_rows("Empty"))
        self.assertEqual(rows, [])

    def test_empty_workbook_detect_header_returns_none(self):
        """detect_header_row() on an empty sheet must return None."""
        reader = BoqReader(_p("synthetic_empty.xlsx"))
        result = reader.detect_header_row("Empty")
        self.assertIsNone(result)

    def test_empty_workbook_no_exception(self):
        """get_sheet_dimensions() on an empty sheet returns (0, 0) without raising."""
        reader = BoqReader(_p("synthetic_empty.xlsx"))
        dims = reader.get_sheet_dimensions("Empty")
        self.assertEqual(dims, (0, 0))

    # ------------------------------------------------------------------ #
    # Additional — is_blank() on RawRow                                    #
    # ------------------------------------------------------------------ #

    def test_blank_row_detected(self):
        """Row 4 of synthetic_simple is entirely empty — is_blank() must be True."""
        reader = BoqReader(_p("synthetic_simple.xlsx"))
        rows = {r.row_number: r for r in reader.iter_rows("Sheet1", start_row=4, end_row=4)}
        self.assertTrue(rows[4].is_blank())

    def test_non_blank_row_not_blank(self):
        """Row 1 (headers) must not be reported as blank."""
        reader = BoqReader(_p("synthetic_simple.xlsx"))
        rows = {r.row_number: r for r in reader.iter_rows("Sheet1", start_row=1, end_row=1)}
        self.assertFalse(rows[1].is_blank())

    # ------------------------------------------------------------------ #
    # Additional — get_master_preamble_text                                #
    # ------------------------------------------------------------------ #

    def test_get_master_preamble_text(self):
        """get_master_preamble_text() concatenates non-empty cells by newline."""
        reader = BoqReader(_p("synthetic_simple.xlsx"))
        text = reader.get_master_preamble_text("Sheet1")
        self.assertIn("Sl.No.", text)
        self.assertIn("First item", text)

    # ------------------------------------------------------------------ #
    # Test 17 — sparse multi-area header                                   #
    # ------------------------------------------------------------------ #

    def test_detect_header_row_sparse_multi_area(self):
        """
        synthetic_sparse_header has merged area labels (rows 3-4) above a
        sparse column-header row at row 5 ("Sl.No.", "Description", "Unit", …).
        detect_header_row() must return 5.
        """
        reader = BoqReader(_p("synthetic_sparse_header.xlsx"))
        result = reader.detect_header_row("HVAC-style")
        self.assertEqual(result, 5)

    # ------------------------------------------------------------------ #
    # Test 18 — domain-vocabulary header                                   #
    # ------------------------------------------------------------------ #

    def test_detect_header_row_domain_vocab(self):
        """
        synthetic_makelist_header row 3 uses "Details of Materials" and
        "Approved Makes" — domain-specific vocabulary instead of generic names.
        detect_header_row() must still return 3.
        """
        reader = BoqReader(_p("synthetic_makelist_header.xlsx"))
        result = reader.detect_header_row("Make List")
        self.assertEqual(result, 3)

    # ------------------------------------------------------------------ #
    # Test 19 — data row with long description not picked                  #
    # ------------------------------------------------------------------ #

    def test_detect_header_row_long_description_data_row_not_picked(self):
        """
        Row 7 of synthetic_sparse_header is a data row whose description cell
        exceeds 60 characters. detect_header_row() must not return row 7; the
        correct header row (5) must be returned instead.
        """
        reader = BoqReader(_p("synthetic_sparse_header.xlsx"))
        result = reader.detect_header_row("HVAC-style")
        self.assertIsNotNone(result)
        self.assertNotEqual(result, 7)
        self.assertEqual(result, 5)


class TestMergedCellPropagation(unittest.TestCase):
    """Phase 2b.2 Part A1: covered cells now inherit value and merged_range from origin."""

    # Class-level paths assigned in setUpClass; cleared in tearDownClass.
    _temp_simple_merge = ""
    _temp_pattern2_layout = ""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        import os
        import tempfile
        import openpyxl
        from openpyxl.styles import Font

        # Fixture 1 — horizontal merge A1:B1.
        # Origin A1 = "B1 BUILDING" (bold); covered B1 has default (non-bold) formatting.
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "B1 BUILDING"
        ws["A1"].font = Font(bold=True)
        ws.merge_cells("A1:B1")
        wb.save(path)
        cls._temp_simple_merge = path

        # Fixture 2 — two-row Pattern-2-shaped layout.
        # Row 1: [A1:B1]="B1 BUILDING" | [C1:D1]="B3 BUILDING" | [E1:F1]="TOTAL"
        # Row 2: QTY | AMOUNT | QTY | AMOUNT | QTY | AMOUNT  (no merges)
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "B1 BUILDING"
        ws.merge_cells("A1:B1")
        ws["C1"] = "B3 BUILDING"
        ws.merge_cells("C1:D1")
        ws["E1"] = "TOTAL"
        ws.merge_cells("E1:F1")
        ws["A2"] = "QTY"
        ws["B2"] = "AMOUNT"
        ws["C2"] = "QTY"
        ws["D2"] = "AMOUNT"
        ws["E2"] = "QTY"
        ws["F2"] = "AMOUNT"
        wb.save(path)
        cls._temp_pattern2_layout = path

    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        import os
        for path in (cls._temp_simple_merge, cls._temp_pattern2_layout):
            if path and os.path.exists(path):
                try:
                    os.unlink(path)
                except OSError:
                    pass

    # ------------------------------------------------------------------ #
    # Test 20 — origin cell unchanged                                      #
    # ------------------------------------------------------------------ #

    def test_merged_cell_origin_unchanged(self):
        """Origin A1 of merge A1:B1 retains value, is_merged_origin=True, merged_range."""
        reader = BoqReader(self._temp_simple_merge)
        rows = list(reader.iter_rows("Sheet1", start_row=1, end_row=1))
        a1 = rows[0].get_cell("A")
        self.assertIsNotNone(a1)
        self.assertEqual(a1.value, "B1 BUILDING")
        self.assertTrue(a1.is_merged_origin)
        self.assertEqual(a1.merged_range, "A1:B1")

    # ------------------------------------------------------------------ #
    # Test 21 — covered cell inherits merged_range                         #
    # ------------------------------------------------------------------ #

    def test_merged_cell_covered_inherits_merged_range(self):
        """Covered B1 of merge A1:B1 reports merged_range='A1:B1' (was None before Part A1)."""
        reader = BoqReader(self._temp_simple_merge)
        rows = list(reader.iter_rows("Sheet1", start_row=1, end_row=1))
        b1 = rows[0].get_cell("B")
        self.assertIsNotNone(b1)
        self.assertEqual(b1.merged_range, "A1:B1")

    # ------------------------------------------------------------------ #
    # Test 22 — covered cell inherits value                                #
    # ------------------------------------------------------------------ #

    def test_merged_cell_covered_inherits_value(self):
        """Covered B1 of merge A1:B1 reports same value as origin ('B1 BUILDING')."""
        reader = BoqReader(self._temp_simple_merge)
        rows = list(reader.iter_rows("Sheet1", start_row=1, end_row=1))
        b1 = rows[0].get_cell("B")
        self.assertIsNotNone(b1)
        self.assertEqual(b1.value, "B1 BUILDING")

    # ------------------------------------------------------------------ #
    # Test 23 — covered cell is NOT the origin                             #
    # ------------------------------------------------------------------ #

    def test_merged_cell_covered_is_not_origin(self):
        """Covered B1 retains is_merged_origin=False — only the top-left is the origin."""
        reader = BoqReader(self._temp_simple_merge)
        rows = list(reader.iter_rows("Sheet1", start_row=1, end_row=1))
        b1 = rows[0].get_cell("B")
        self.assertIsNotNone(b1)
        self.assertFalse(b1.is_merged_origin)

    # ------------------------------------------------------------------ #
    # Test 24 — covered cell does NOT inherit formatting                   #
    # ------------------------------------------------------------------ #

    def test_merged_cell_covered_does_not_inherit_formatting(self):
        """Covered B1 has its own font_bold=False; bold is NOT inherited from bold origin A1."""
        reader = BoqReader(self._temp_simple_merge)
        rows = list(reader.iter_rows("Sheet1", start_row=1, end_row=1))
        a1 = rows[0].get_cell("A")
        b1 = rows[0].get_cell("B")
        self.assertTrue(a1.font_bold, "Origin A1 must be bold for this assertion to be meaningful")
        self.assertFalse(b1.font_bold, "Covered B1 must NOT inherit bold from origin")

    # ------------------------------------------------------------------ #
    # Test 25 — two-row Pattern-2-shaped merged header layout              #
    # ------------------------------------------------------------------ #

    def test_two_row_merged_header_layout(self):
        """
        Row 1 has three horizontal merges (area-name cells).
        Covered cells in row 1 share their origin's merged_range and value.
        Row 2 cells are unmerged — all have merged_range=None.
        This verifies the data shape that Part A3 multi-area detection will read.
        """
        reader = BoqReader(self._temp_pattern2_layout)
        rows = list(reader.iter_rows("Sheet1", start_row=1, end_row=2))
        self.assertEqual(len(rows), 2)
        row1, row2 = rows[0], rows[1]

        # Row 1 origins: each is_merged_origin=True, value and range set
        self.assertTrue(row1.get_cell("A").is_merged_origin)
        self.assertEqual(row1.get_cell("A").merged_range, "A1:B1")
        self.assertTrue(row1.get_cell("C").is_merged_origin)
        self.assertEqual(row1.get_cell("C").merged_range, "C1:D1")
        self.assertTrue(row1.get_cell("E").is_merged_origin)
        self.assertEqual(row1.get_cell("E").merged_range, "E1:F1")

        # Row 1 covered cells: value and range propagated from origin, not origin
        self.assertFalse(row1.get_cell("B").is_merged_origin)
        self.assertEqual(row1.get_cell("B").merged_range, "A1:B1")
        self.assertEqual(row1.get_cell("B").value, "B1 BUILDING")
        self.assertFalse(row1.get_cell("D").is_merged_origin)
        self.assertEqual(row1.get_cell("D").merged_range, "C1:D1")
        self.assertEqual(row1.get_cell("D").value, "B3 BUILDING")
        self.assertFalse(row1.get_cell("F").is_merged_origin)
        self.assertEqual(row1.get_cell("F").merged_range, "E1:F1")
        self.assertEqual(row1.get_cell("F").value, "TOTAL")

        # Row 2: unmerged — all merged_range=None
        for col in ("A", "B", "C", "D", "E", "F"):
            cell = row2.get_cell(col)
            self.assertIsNotNone(cell)
            self.assertIsNone(
                cell.merged_range,
                f"Row 2 column {col} is not merged; merged_range must be None",
            )


class TestSheetStateExposure(unittest.TestCase):
    """Phase 2c §9 #49: list_sheet_states() exposes openpyxl Worksheet.sheet_state."""

    # ------------------------------------------------------------------ #
    # Test 26 — all visible sheets                                         #
    # ------------------------------------------------------------------ #

    def test_all_visible_sheets_return_visible(self):
        """Two default sheets (never had sheet_state set) must both report 'visible'."""
        import os
        import tempfile
        import openpyxl

        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "two_visible.xlsx")
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            wb.create_sheet("Sheet One")
            wb.create_sheet("Sheet Two")
            wb.save(path)

            reader = BoqReader(path)
            states = reader.list_sheet_states()
            self.assertEqual(states, {"Sheet One": "visible", "Sheet Two": "visible"})

    # ------------------------------------------------------------------ #
    # Test 27 — hidden sheet surfaced                                      #
    # ------------------------------------------------------------------ #

    def test_hidden_sheet_state_surfaced(self):
        """Second of three sheets marked hidden must report sheet_state='hidden'."""
        import os
        import tempfile
        import openpyxl

        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "one_hidden.xlsx")
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            wb.create_sheet("Alpha")
            wb.create_sheet("Beta")
            wb.create_sheet("Gamma")
            wb["Beta"].sheet_state = "hidden"
            wb.save(path)

            reader = BoqReader(path)
            states = reader.list_sheet_states()
            self.assertEqual(set(states.keys()), {"Alpha", "Beta", "Gamma"})
            self.assertEqual(states["Beta"], "hidden")
            self.assertEqual(states["Alpha"], "visible")
            self.assertEqual(states["Gamma"], "visible")

    # ------------------------------------------------------------------ #
    # Test 28 — veryHidden sheet surfaced                                  #
    # ------------------------------------------------------------------ #

    def test_very_hidden_sheet_state_surfaced(self):
        """Sheet with sheet_state='veryHidden' must report exactly 'veryHidden' (camelCase)."""
        import os
        import tempfile
        import openpyxl

        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "very_hidden.xlsx")
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            wb.create_sheet("Visible")
            wb.create_sheet("Secret")
            wb["Secret"].sheet_state = "veryHidden"
            wb.save(path)

            reader = BoqReader(path)
            states = reader.list_sheet_states()
            self.assertEqual(states["Secret"], "veryHidden")
            self.assertEqual(states["Visible"], "visible")

    # ------------------------------------------------------------------ #
    # Test 29 — whitespace and order preserved                             #
    # ------------------------------------------------------------------ #

    def test_sheet_names_preserve_whitespace_and_order(self):
        """
        Sheet names with leading/trailing/internal whitespace are preserved exactly,
        and dict iteration order matches workbook sheet order.
        """
        import os
        import tempfile
        import openpyxl

        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "whitespace.xlsx")
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            wb.create_sheet(" Leading")
            wb.create_sheet("Trailing ")
            wb.create_sheet("Mid  Double")
            wb[" Leading"].sheet_state = "hidden"
            wb.save(path)

            reader = BoqReader(path)
            states = reader.list_sheet_states()

            # Exact keys (whitespace preserved)
            self.assertIn(" Leading", states)
            self.assertIn("Trailing ", states)
            self.assertIn("Mid  Double", states)

            # Order matches workbook order
            self.assertEqual(list(states.keys()), [" Leading", "Trailing ", "Mid  Double"])

            # Visibility
            self.assertEqual(states[" Leading"], "hidden")
            self.assertEqual(states["Trailing "], "visible")
            self.assertEqual(states["Mid  Double"], "visible")


class TestExcelErrorLiterals(unittest.TestCase):
    """Bug 13 (sec 9 #89) — Excel error literals normalised to None at reader level."""

    # ------------------------------------------------------------------
    # _is_excel_error unit tests
    # ------------------------------------------------------------------

    def test_is_excel_error_true_all_seven_literals(self):
        from nirmaan_stack.services.boq_parser.reader import _is_excel_error
        literals = ["#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#NULL!", "#N/A", "#NUM!"]
        for lit in literals:
            with self.subTest(lit=lit):
                self.assertTrue(_is_excel_error(lit))

    def test_is_excel_error_whitespace_and_case_tolerance(self):
        from nirmaan_stack.services.boq_parser.reader import _is_excel_error
        self.assertTrue(_is_excel_error("  #ref!  "))
        self.assertTrue(_is_excel_error("#value!"))
        self.assertTrue(_is_excel_error("\t#DIV/0!\n"))

    def test_is_excel_error_false_for_none_empty_numeric_normal_text(self):
        from nirmaan_stack.services.boq_parser.reader import _is_excel_error
        for val in [None, "", "   ", 0, 3.14, "hello", "#"]:
            with self.subTest(val=val):
                self.assertFalse(_is_excel_error(val))

    def test_is_excel_error_false_for_substring(self):
        """Partial match must not fire — value must equal the literal exactly (after strip/upper)."""
        from nirmaan_stack.services.boq_parser.reader import _is_excel_error
        self.assertFalse(_is_excel_error("see #REF! note"))
        self.assertFalse(_is_excel_error("Error: #VALUE!"))

    # ------------------------------------------------------------------
    # Integration: BoqReader.iter_rows() normalises error literal cells
    # ------------------------------------------------------------------

    def test_reader_normalizes_error_literal_cell_to_none(self):
        import tempfile, openpyxl
        from nirmaan_stack.services.boq_parser.reader import BoqReader

        with tempfile.TemporaryDirectory() as tmpdir:
            path = str(Path(tmpdir) / "error_test.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Normal"
            ws["B1"] = "#REF!"
            ws["C1"] = "#VALUE!"
            ws["A2"] = 42
            ws["B2"] = "#N/A"
            wb.save(path)

            reader = BoqReader(path)
            rows = list(reader.iter_rows("Sheet1"))

            row1 = rows[0]
            self.assertEqual(row1.get_cell("A").value, "Normal")
            self.assertIsNone(row1.get_cell("B").value, "#REF! should be None")
            self.assertIsNone(row1.get_cell("C").value, "#VALUE! should be None")

            row2 = rows[1]
            self.assertEqual(row2.get_cell("A").value, 42)
            self.assertIsNone(row2.get_cell("B").value, "#N/A should be None")

    def test_reader_blank_row_detection_not_confused_by_error_cells(self):
        """A row whose only non-None cell contains an error literal must be treated as blank."""
        import tempfile, openpyxl
        from nirmaan_stack.services.boq_parser.reader import BoqReader

        with tempfile.TemporaryDirectory() as tmpdir:
            path = str(Path(tmpdir) / "error_blank.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Header"
            ws["A2"] = "#DIV/0!"
            ws["A3"] = "Data"
            wb.save(path)

            reader = BoqReader(path)
            rows = list(reader.iter_rows("Sheet1"))

            self.assertEqual(len(rows), 3)
            self.assertTrue(rows[1].is_blank(), "Row with only #DIV/0! should be blank after normalisation")


if __name__ == "__main__":
    unittest.main()
