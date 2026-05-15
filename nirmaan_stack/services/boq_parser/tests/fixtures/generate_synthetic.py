"""
Generates synthetic .xlsx fixtures used by test_reader.py and test_orchestrator.py.

Run once (or call generate_all() from test setUpClass) to produce:
  - synthetic_simple.xlsx
  - synthetic_merged_header.xlsx
  - synthetic_trailing_spaces.xlsx  (for list_sheets() whitespace test)
  - synthetic_blank_cols.xlsx       (for detect_blank_columns() test)
  - synthetic_empty.xlsx            (for empty-file edge case)
  - synthetic_sparse_header.xlsx    (HVAC-style sparse multi-area header)
  - synthetic_makelist_header.xlsx  (domain-vocab header: "Details of Materials")
  - synthetic_multi_area_2row.xlsx  (Pattern 2 two-row merged header, §9 #43)

All files are written into this same directory so tests can find them
via: Path(__file__).parent / "<name>.xlsx"
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

_HERE = Path(__file__).parent


def _path(name: str) -> Path:
    return _HERE / name


# ------------------------------------------------------------------
# synthetic_simple.xlsx
# ------------------------------------------------------------------

def generate_simple() -> Path:
    """
    Single sheet "Sheet1":
    Row 1: headers — Sl.No. | Description | Unit | Qty | Rate | Amount
    Row 2: 1.0 | First item | Nos | 5 | 100 | 500   (pre-computed; also stores formula text as metadata)
    Row 3: 2.0 | Second item | Nos | 3 | 200 | 600
    Row 4: empty
    Row 5: 3.0 | Bold item | Nos | 2 | 50 | 100   — column B cell is bold
    Row 6: empty (spacer)

    NOTE: openpyxl cannot cache formula results (that requires Excel/LibreOffice
    to evaluate the workbook).  We therefore write COMPUTED VALUES directly into
    column F so data_only=True reads the correct numbers.  Row 2 column F ALSO
    stores a formula string in a *separate* cell (G2) to let us test formula
    detection without needing Excel to evaluate it.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Row 1 — headers
    ws["A1"] = "Sl.No."
    ws["B1"] = "Description"
    ws["C1"] = "Unit"
    ws["D1"] = "Qty"
    ws["E1"] = "Rate"
    ws["F1"] = "Amount"
    ws["G1"] = "Formula (test)"

    # Row 2 — normal item; F2 = pre-computed value, G2 = formula string for detection
    ws["A2"] = 1.0
    ws["B2"] = "First item"
    ws["C2"] = "Nos"
    ws["D2"] = 5
    ws["E2"] = 100
    ws["F2"] = 500          # pre-computed so data_only=True returns 500
    ws["G2"] = "=D2*E2"     # raw formula string — used to test is_formula detection

    # Row 3
    ws["A3"] = 2.0
    ws["B3"] = "Second item"
    ws["C3"] = "Nos"
    ws["D3"] = 3
    ws["E3"] = 200
    ws["F3"] = 600

    # Row 4 — empty
    # (no writes needed)

    # Row 5 — bold description
    ws["A5"] = 3.0
    ws["B5"] = "Bold item"
    ws["B5"].font = Font(bold=True)
    ws["C5"] = "Nos"
    ws["D5"] = 2
    ws["E5"] = 50
    ws["F5"] = 100

    # Row 6 — empty spacer (no writes)

    path = _path("synthetic_simple.xlsx")
    wb.save(str(path))
    return path


# ------------------------------------------------------------------
# synthetic_merged_header.xlsx
# ------------------------------------------------------------------

def generate_merged_header() -> Path:
    """
    Single sheet "Data":
    Rows 1-2: A1:F1 merged title "Project XYZ — BoQ"  (actually spans row 1)
    Row 3-4:  merged column headers
      A3:A4 = "Sl.No.", B3:B4 = "Description", C3:C4 = "Unit",
      D3:D4 = "Qty", E3:F3 = "Rate" (with E4="Supply", F4="Install")
    Row 5: data — 1.0 | Item one | Nos | 5 | 100 | 50
    Row 6: data — 2.0 | Item two | Nos | 3 | 200 | 0
    Row 7: "Total" in col A only (label-only row)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Title spanning row 1
    ws["A1"] = "Project XYZ — BoQ"
    ws.merge_cells("A1:F1")

    # Merged column headers
    ws["A3"] = "Sl.No."
    ws.merge_cells("A3:A4")

    ws["B3"] = "Description"
    ws.merge_cells("B3:B4")

    ws["C3"] = "Unit"
    ws.merge_cells("C3:C4")

    ws["D3"] = "Qty"
    ws.merge_cells("D3:D4")

    ws["E3"] = "Rate"
    ws.merge_cells("E3:F3")
    ws["E4"] = "Supply"
    ws["F4"] = "Install"

    # Data rows
    ws["A5"] = 1.0
    ws["B5"] = "Item one"
    ws["C5"] = "Nos"
    ws["D5"] = 5
    ws["E5"] = 100
    ws["F5"] = 50

    ws["A6"] = 2.0
    ws["B6"] = "Item two"
    ws["C6"] = "Nos"
    ws["D6"] = 3
    ws["E6"] = 200
    ws["F6"] = 0

    # Label-only row
    ws["A7"] = "Total"

    path = _path("synthetic_merged_header.xlsx")
    wb.save(str(path))
    return path


# ------------------------------------------------------------------
# synthetic_trailing_spaces.xlsx  (two sheets, one has trailing spaces)
# ------------------------------------------------------------------

def generate_trailing_spaces() -> Path:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet One"
    ws1["A1"] = "data"

    ws2 = wb.create_sheet("Trailing  ")   # two trailing spaces
    ws2["A1"] = "data"

    path = _path("synthetic_trailing_spaces.xlsx")
    wb.save(str(path))
    return path


# ------------------------------------------------------------------
# synthetic_blank_cols.xlsx  (cols A and Z blank, B-D have content)
# ------------------------------------------------------------------

def generate_blank_cols() -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Columns B, C, D have content in several rows
    for row in range(1, 6):
        ws.cell(row=row, column=2).value = f"B{row}"   # col B
        ws.cell(row=row, column=3).value = f"C{row}"   # col C
        ws.cell(row=row, column=4).value = f"D{row}"   # col D

    # Write empty string to Z1 so openpyxl's max_column extends to 26.
    # The cell is blank (empty string is treated as blank by detect_blank_columns).
    # Column A (index 1) stays blank throughout.
    ws["Z1"] = ""

    path = _path("synthetic_blank_cols.xlsx")
    wb.save(str(path))
    return path


# ------------------------------------------------------------------
# synthetic_empty.xlsx
# ------------------------------------------------------------------

def generate_empty() -> Path:
    wb = openpyxl.Workbook()
    wb.active.title = "Empty"
    path = _path("synthetic_empty.xlsx")
    wb.save(str(path))
    return path


# ------------------------------------------------------------------
# synthetic_sparse_header.xlsx  (HVAC-style sparse multi-area header)
# ------------------------------------------------------------------

def generate_sparse_header() -> Path:
    """
    HVAC-style sheet with merged area labels above the real header.
    Sheet "HVAC-style":
    Row 1: A1:J1 merged title "HVAC Equipment Schedule — Block A"
    Row 2: empty
    Row 3: C3:D3 merged "Supply Side", E3:F3 merged "Install Side"
           — only 2 non-empty cells after merge; fails ≥3 guard
    Row 4: C4="Qty" D4="Rate" E4="Qty" F4="Rate" G4="Amount"
           — 5 non-empty, no header keywords → score 0
    Row 5: A5="Sl.No." B5="Description" C5="Unit" G5="Total" H5="Remarks"
           — SPARSE HEADER ROW (5 non-empty; scores 4pts)
    Row 6: empty separator
    Row 7: 1.0 | <description >60 chars> | "Nos" | 2 | 1 | ... — DATA
           — last content row; B7 text is >60 chars
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HVAC-style"

    # Row 1 — merged title
    ws["A1"] = "HVAC Equipment Schedule — Block A"
    ws.merge_cells("A1:J1")

    # Row 3 — area headers (only 2 non-empty after merge → fails Guard 1)
    ws["C3"] = "Supply Side"
    ws.merge_cells("C3:D3")
    ws["E3"] = "Install Side"
    ws.merge_cells("E3:F3")

    # Row 4 — sub-labels (5 non-empty, no keywords → score 0)
    ws["C4"] = "Qty"
    ws["D4"] = "Rate"
    ws["E4"] = "Qty"
    ws["F4"] = "Rate"
    ws["G4"] = "Amount"

    # Row 5 — sparse column headers (HEADER ROW; scores 4pts via sl.no + description)
    ws["A5"] = "Sl.No."
    ws["B5"] = "Description"
    ws["C5"] = "Unit"
    ws["G5"] = "Total"
    ws["H5"] = "Remarks"

    # Row 7 — first data row; B7 is >60 chars (tests long-text guard)
    ws["A7"] = 1.0
    ws["B7"] = (
        "Supply and installation of split AC 2 Ton capacity unit "
        "with outdoor condensing unit fully integrated"
    )
    ws["C7"] = "Nos"
    ws["D7"] = 2
    ws["E7"] = 1
    ws["G7"] = 3
    ws["H7"] = "As per spec"

    path = _path("synthetic_sparse_header.xlsx")
    wb.save(str(path))
    return path


# ------------------------------------------------------------------
# synthetic_makelist_header.xlsx  (domain-vocab header)
# ------------------------------------------------------------------

def generate_makelist_header() -> Path:
    """
    Domain-vocabulary column headers: "Details of Materials", "Approved Makes".
    Sheet "Make List":
    Row 1: A1:F1 merged "Substation Equipment Make List"  (1 non-empty → fails Guard 1)
    Row 2: empty
    Row 3: "Sr.No." | "Details of Materials" | "Approved Makes" | "Unit" | "Qty" | "Rate"
           — HEADER ROW (scores 4pts: sr.no=2, details of materials=1, approved makes=1)
    Row 4: 1.0 | "Main Breaker 400A" | "ABB / Schneider" | "Nos" | 2 | 15000
    Row 5: 2.0 | "Bus Bar 60x10 copper" | "Hindalco" | "Meter" | 10 | 800
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Make List"

    # Row 1 — merged title
    ws["A1"] = "Substation Equipment Make List"
    ws.merge_cells("A1:F1")

    # Row 3 — column headers with domain vocabulary (HEADER ROW)
    ws["A3"] = "Sr.No."
    ws["B3"] = "Details of Materials"
    ws["C3"] = "Approved Makes"
    ws["D3"] = "Unit"
    ws["E3"] = "Qty"
    ws["F3"] = "Rate"

    # Row 4 — data
    ws["A4"] = 1.0
    ws["B4"] = "Main Breaker 400A"
    ws["C4"] = "ABB / Schneider"
    ws["D4"] = "Nos"
    ws["E4"] = 2
    ws["F4"] = 15000

    # Row 5 — data
    ws["A5"] = 2.0
    ws["B5"] = "Bus Bar 60x10 copper"
    ws["C5"] = "Hindalco"
    ws["D5"] = "Meter"
    ws["E5"] = 10
    ws["F5"] = 800

    path = _path("synthetic_makelist_header.xlsx")
    wb.save(str(path))
    return path


# ------------------------------------------------------------------
# synthetic_multi_area.xlsx  (Policy X + post-pass integration fixture)
# ------------------------------------------------------------------

def generate_multi_area() -> Path:
    """
    Sheet "Multi Area" — tests _apply_multi_area_post_pass():

    Header row 1:
      A=Sl.No. B=Description C=Unit D=Floor 1 E=Floor 2 F=Total Qty
      G=Rate H=Amt Floor 1 I=Amt Floor 2 J=Total Amt

    Row 2 — clean: qty sum (8) == total (8); amt sum (800) == total (800)  → no warnings
    Row 3 — qty mismatch: qty sum (8) vs total (10), diff=2 > ±1           → qty warning
    Row 4 — blank totals: F=None, J=None; per-area populated               → fallback computes totals
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Multi Area"

    # Row 1 — column headers
    ws["A1"] = "Sl.No."
    ws["B1"] = "Description"
    ws["C1"] = "Unit"
    ws["D1"] = "Floor 1"
    ws["E1"] = "Floor 2"
    ws["F1"] = "Total Qty"
    ws["G1"] = "Rate"
    ws["H1"] = "Amt Floor 1"
    ws["I1"] = "Amt Floor 2"
    ws["J1"] = "Total Amt"

    # Row 2 — good row (no warnings)
    ws["A2"] = 1.0
    ws["B2"] = "Painting works"
    ws["C2"] = "Sqm"
    ws["D2"] = 5.0    # Floor 1 qty
    ws["E2"] = 3.0    # Floor 2 qty
    ws["F2"] = 8.0    # Total Qty == sum
    ws["G2"] = 100.0
    ws["H2"] = 500.0  # Amt Floor 1
    ws["I2"] = 300.0  # Amt Floor 2
    ws["J2"] = 800.0  # Total Amt == sum

    # Row 3 — qty mismatch: sum=8, total=10, diff=2 > ±1  → qty warning
    ws["A3"] = 2.0
    ws["B3"] = "Tiling works"
    ws["C3"] = "Sqm"
    ws["D3"] = 5.0    # Floor 1 qty
    ws["E3"] = 3.0    # Floor 2 qty
    ws["F3"] = 10.0   # Total Qty != sum (mismatch)
    ws["G3"] = 200.0
    ws["H3"] = 1000.0
    ws["I3"] = 600.0
    ws["J3"] = 1600.0  # amt sum=1600 == total → no amt warning

    # Row 4 — blank totals: F=None, J=None → fallback (qty=10, amt=500, no warning)
    ws["A4"] = 3.0
    ws["B4"] = "Plumbing works"
    ws["C4"] = "m"
    ws["D4"] = 4.0    # Floor 1 qty
    ws["E4"] = 6.0    # Floor 2 qty
    # F4 intentionally blank — total blank triggers empty-total fallback
    ws["G4"] = 50.0
    ws["H4"] = 200.0
    ws["I4"] = 300.0
    # J4 intentionally blank — amount total also triggers fallback

    path = _path("synthetic_multi_area.xlsx")
    wb.save(str(path))
    return path


# ------------------------------------------------------------------
# synthetic_multi_area_2row.xlsx  (Pattern 2 — two-row merged header)
# ------------------------------------------------------------------

def generate_multi_area_2row() -> Path:
    """
    Two-row merged header sheet — exercises Pattern 2 detection via parse_boq()
    (header_row_count=2 routing, §9 #43 coverage gap).

    Row 1 (top header): area labels merged across pairs of columns.
      A1="Sl.No."  B1="Description"  C1:D1="Block A"  E1:F1="Block B"
      G1="Rate"  H1="Total"
    Row 2 (bottom header, header_row=2): column sub-labels.
      A2="Sl.No."  B2="Description"  C2="Qty"  D2="Amount"
      E2="Qty"  F2="Amount"  G2="Rate"  H2="Total"
    Row 3: empty (skipped as header_row+1 by orchestrator)
    Rows 4–6: three LINE_ITEM data rows.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Multi Area 2Row"

    # Row 1 — top header: area labels, two merged pairs
    ws["A1"] = "Sl.No."
    ws["B1"] = "Description"
    ws["C1"] = "Block A"
    ws.merge_cells("C1:D1")
    ws["E1"] = "Block B"
    ws.merge_cells("E1:F1")
    ws["G1"] = "Rate"
    ws["H1"] = "Total"

    # Row 2 — bottom header: qty/amount sub-labels (= header_row)
    ws["A2"] = "Sl.No."
    ws["B2"] = "Description"
    ws["C2"] = "Qty"
    ws["D2"] = "Amount"
    ws["E2"] = "Qty"
    ws["F2"] = "Amount"
    ws["G2"] = "Rate"
    ws["H2"] = "Total"

    # Row 3 — empty (skipped by orchestrator as header_row+1)

    # Row 4 — first data row
    ws["A4"] = 1.0
    ws["B4"] = "Electrical works"
    ws["C4"] = 5.0    # Block A qty
    ws["D4"] = 500.0  # Block A amount
    ws["E4"] = 3.0    # Block B qty
    ws["F4"] = 300.0  # Block B amount
    ws["G4"] = 100.0
    ws["H4"] = 800.0

    # Row 5
    ws["A5"] = 2.0
    ws["B5"] = "Civil works"
    ws["C5"] = 10.0
    ws["D5"] = 1000.0
    ws["E5"] = 7.0
    ws["F5"] = 700.0
    ws["G5"] = 120.0
    ws["H5"] = 1700.0

    # Row 6
    ws["A6"] = 3.0
    ws["B6"] = "HVAC works"
    ws["C6"] = 2.0
    ws["D6"] = 200.0
    ws["E6"] = 1.0
    ws["F6"] = 100.0
    ws["G6"] = 150.0
    ws["H6"] = 300.0

    path = _path("synthetic_multi_area_2row.xlsx")
    wb.save(str(path))
    return path


# ------------------------------------------------------------------
# Entry point
# ------------------------------------------------------------------

def generate_all() -> None:
    generate_simple()
    generate_merged_header()
    generate_trailing_spaces()
    generate_blank_cols()
    generate_empty()
    generate_sparse_header()
    generate_makelist_header()
    generate_multi_area()
    generate_multi_area_2row()


if __name__ == "__main__":
    generate_all()
    print("Fixtures generated in", _HERE)
