"""
Auto-guess SheetConfig column roles for the BoQ parser.

Phase 1.9h: extracted from 1.9e/1.9f diagnostic scripts into a shared module.
Root cause fixed: 1.9f Stage 1 diagnostic (chore c42eec9a) found that Raheja
Electrical at header_row_count=2 detected Pattern 2-rate correctly but all
per-area qty/rate/amount came back None because the 1.9e auto-guess explicitly
skipped area-specific roles.

Phase 1.9m: added _should_auto_promote_hrc_to_2() helper and auto-detect mode
for auto_guess_sheet_config() (pass header_row_count=None to enable). Fixes
Mode A: merged-cell two-row headers where adjacent duplicate text in the bottom
header row (flattened merged cells) triggered the singleton guard and dropped
column roles.

Phase 3 production auto-guess (wizard) will import auto_guess_sheet_config()
from here and extend it with a user-confirmation step.

No Frappe imports — pure Python, fully testable in isolation.
"""
from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl.utils import column_index_from_string, get_column_letter

from nirmaan_stack.services.boq_parser.classifier import _HEADER_KW
from nirmaan_stack.services.boq_parser.config import ColumnRole, SheetConfig
from nirmaan_stack.services.boq_parser.multi_area_detection import detect_multi_area_pattern

if TYPE_CHECKING:
    from nirmaan_stack.services.boq_parser.reader import BoqReader, RawRow

# -----------------------------------------------------------------------
# Module-level constants mirroring config._SINGLETON_ROLES
# -----------------------------------------------------------------------

_SINGLETON_ROLES: frozenset[str] = frozenset({
    "sl_no", "description", "unit", "qty_total",
    "rate_supply", "rate_install", "rate_combined",
    "amount_supply", "amount_install",
    "amount_total", "make_model", "row_notes", "reference_images",
})

_PER_AREA_ONLY_ROLES: frozenset[str] = frozenset({
    "amount_supply_by_area", "amount_install_by_area", "amount_total_by_area",
    "rate_supply_by_area", "rate_install_by_area", "rate_combined_by_area",
})


def _normalize(s: str) -> str:
    """Lowercase + collapse internal whitespace."""
    return " ".join(str(s).strip().lower().split())


def _should_auto_promote_hrc_to_2(
    bottom_row: "RawRow",
    above_row: "RawRow | None",
    below_row: "RawRow | None",
) -> bool:
    """Phase 1.9m Mode A: return True when the sheet appears to use a 2-row merged header.

    Heuristic (all three conditions must hold):
    1. bottom_row has >= 3 non-blank cells.
    2. bottom_row has at least one pair of ADJACENT spreadsheet columns
       (column-index difference == 1) with identical normalized text --- the
       flattened-merged-cell signature when a 2-row sheet is read at hrc=1.
    3. For at least one such adjacent duplicate pair, either above_row or
       below_row has DISTINCT non-blank text at those same two column positions
       (i.e., the two cells have different normalized text from each other).

    False negatives are recoverable via SheetConfig.top_header_rows_override
    (Phase 1.9d F5-b). False positives corrupt the parse --- so the heuristic
    is intentionally conservative.

    Empirical basis (Phase 1.9i §22.4 Mode A): Paytm ELEC (553 role_unassigned)
    and Paytm HVAC (170 role_unassigned) both have adjacent RATE/RATE and
    AMOUNT/AMOUNT in their detected header rows.
    """
    sorted_bottom = sorted(
        [
            (col, _normalize(str(ci.value)))
            for col, ci in bottom_row.cells.items()
            if ci.value is not None and _normalize(str(ci.value))
        ],
        key=lambda kv: column_index_from_string(kv[0]),
    )

    if len(sorted_bottom) < 3:
        return False

    dup_col_pairs: list[tuple[str, str]] = []
    for i in range(len(sorted_bottom) - 1):
        col_a, text_a = sorted_bottom[i]
        col_b, text_b = sorted_bottom[i + 1]
        if (
            column_index_from_string(col_b) - column_index_from_string(col_a) == 1
            and text_a == text_b
        ):
            dup_col_pairs.append((col_a, col_b))

    if not dup_col_pairs:
        return False

    def _row_lookup(row: "RawRow | None") -> dict[str, str]:
        if row is None:
            return {}
        return {
            col: _normalize(str(ci.value))
            for col, ci in row.cells.items()
            if ci.value is not None and _normalize(str(ci.value))
        }

    above_lookup = _row_lookup(above_row)
    below_lookup = _row_lookup(below_row)

    def _has_distinct_at(lookup: dict[str, str], col_a: str, col_b: str) -> bool:
        t_a = lookup.get(col_a)
        t_b = lookup.get(col_b)
        return t_a is not None and t_b is not None and t_a != t_b

    for col_a, col_b in dup_col_pairs:
        if _has_distinct_at(above_lookup, col_a, col_b) or _has_distinct_at(below_lookup, col_a, col_b):
            return True

    return False


# -----------------------------------------------------------------------
# Public API
# -----------------------------------------------------------------------

def auto_guess_sheet_config(
    reader: "BoqReader",
    sheet_name: str,
    header_row: int,
    header_row_count: int | None = None,
    reserved_keywords: list[str] | None = None,
) -> SheetConfig:
    """Build a SheetConfig with auto-guessed column roles.

    Pass header_row_count=None (default) to enable Mode A auto-detect: the
    function reads the row above and below header_row and calls
    _should_auto_promote_hrc_to_2() to decide between hrc=1 and hrc=2.
    Pass an explicit int (1 or 2) to bypass auto-detect.

    Two-phase logic:

    Phase 1 --- universal roles from bottom header row (always runs).
      Iterates bottom header cells left-to-right, substring-matching against
      _HEADER_KW. Singleton roles assigned at most once. Per-area-only roles
      (amount_*_by_area, rate_*_by_area) are skipped here --- they need an area name.

    Phase 2 --- per-area role assignment (only when effective_hrc >= 2 and
      detect_multi_area_pattern() returns a non-None result).
      Uses the MultiAreaPattern parallel lists (areas, qty_columns,
      amount_columns, rate_columns) to assign per-area roles directly:
        qty column under area  → ColumnRole(role="qty", area=<area>)
        amount column under area → ColumnRole(role="amount_total_by_area", area=<area>)
        rate column under area   → ColumnRole(role="rate_combined_by_area", area=<area>)
      Per-area assignments override any universal singleton assigned in Phase 1
      for the same column. area_dimensions is populated with all detected areas.

    Patterns covered in Phase 2:
      Pattern 2       --- qty + amount per area
      Pattern 2-rate  --- qty + rate + amount per area
      Pattern 3       --- qty + amount per area (alternating)
      Pattern 1       --- qty only per area (area name = qty column)
    """
    if reserved_keywords is None:
        reserved_keywords = []

    bottom_rows = list(reader.iter_rows(sheet_name, start_row=header_row, end_row=header_row))
    if not bottom_rows:
        effective_hrc = header_row_count if header_row_count is not None else 1
        return SheetConfig(
            sheet_name=sheet_name,
            header_row=header_row,
            header_row_count=effective_hrc,  # type: ignore[arg-type]
        )

    bottom_raw = bottom_rows[0]

    # Phase 1.9m Mode A: resolve effective header_row_count
    if header_row_count is None:
        above_rows = (
            list(reader.iter_rows(sheet_name, start_row=header_row - 1, end_row=header_row - 1))
            if header_row > 1 else []
        )
        below_rows = list(reader.iter_rows(sheet_name, start_row=header_row + 1, end_row=header_row + 1))
        if _should_auto_promote_hrc_to_2(
            bottom_raw,
            above_rows[0] if above_rows else None,
            below_rows[0] if below_rows else None,
        ):
            effective_hrc = 2
        else:
            effective_hrc = 1
    else:
        effective_hrc = header_row_count

    column_role_map: dict[str, ColumnRole] = {}
    assigned_singletons: set[str] = set()

    # ------------------------------------------------------------------
    # Phase 1: universal roles from bottom header row
    # ------------------------------------------------------------------
    sorted_cells = sorted(
        bottom_raw.cells.items(),
        key=lambda kv: column_index_from_string(kv[0]),
    )

    for col_letter, ci in sorted_cells:
        if ci.value is None:
            continue
        cell_text = _normalize(str(ci.value))
        if not cell_text:
            continue

        # Phase 1.9l Mode D --- longest matched keyword wins. Tie-break by iteration order.
        # Fixes: "Supply Rate" mis-labeled as rate_combined (1.9i Mode D, target 8).
        matched_role: str | None = None
        best_kw_len: int = -1
        for role_key, kw_set in _HEADER_KW.items():
            matched_kws = [kw for kw in kw_set if kw in cell_text]
            if not matched_kws:
                continue
            longest_match = max(len(kw) for kw in matched_kws)
            if longest_match > best_kw_len:
                best_kw_len = longest_match
                matched_role = role_key

        if matched_role is None:
            continue
        if matched_role in _SINGLETON_ROLES and matched_role in assigned_singletons:
            continue
        if matched_role in _PER_AREA_ONLY_ROLES:
            continue

        column_role_map[col_letter] = ColumnRole(role=matched_role)  # type: ignore[arg-type]
        if matched_role in _SINGLETON_ROLES:
            assigned_singletons.add(matched_role)

    # ------------------------------------------------------------------
    # Phase 2: pattern detection + per-area role assignment
    # ------------------------------------------------------------------
    area_dimensions: list[str] = []

    if effective_hrc >= 2 and header_row >= 2:
        top_rows = list(reader.iter_rows(sheet_name, start_row=header_row - 1, end_row=header_row - 1))
        if top_rows:
            mp = detect_multi_area_pattern(
                bottom_raw,
                reserved_keywords,
                top_header_row=top_rows[0],
            )
            if mp is not None:
                area_dimensions = list(mp.areas)
                for i, area_name in enumerate(mp.areas):
                    qty_col = get_column_letter(mp.qty_columns[i])
                    column_role_map[qty_col] = ColumnRole(role="qty", area=area_name)

                    if mp.amount_columns is not None:
                        amt_col = get_column_letter(mp.amount_columns[i])
                        # Detected single per-area amount column = the combined/total amount
                        # (field-set Slice 2a): role amount_total_by_area (mirrors how a detected
                        # per-area rate column maps to rate_combined_by_area).
                        column_role_map[amt_col] = ColumnRole(role="amount_total_by_area", area=area_name)

                    if mp.rate_columns is not None:
                        rate_col = get_column_letter(mp.rate_columns[i])
                        column_role_map[rate_col] = ColumnRole(
                            role="rate_combined_by_area", area=area_name
                        )

    return SheetConfig(
        sheet_name=sheet_name,
        header_row=header_row,
        header_row_count=effective_hrc,  # type: ignore[arg-type]
        column_role_map=column_role_map,
        area_dimensions=area_dimensions,
    )
