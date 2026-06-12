# Copyright (c) 2026, Nirmaan (Stratos Infra Technologies Pvt. Ltd.) and contributors
# See license.txt

"""
Tests for Bug 18 -- reader merge-propagation blanking for text-role columns.

When a BoQ author merges a section-banner cell across many columns (e.g. A41:G41),
openpyxl propagates the origin text into every covered cell. Without the fix, the
classifier sees text in sl_no, description, AND unit columns and misclassifies the
row as LINE_ITEM. The fix suppresses the propagated value in covered cells whose
column role is in text_role_columns; only the merge origin retains the text.

Test groups:
  A. TestBug18SyntheticMergeBanner  (5) -- unit tests on synthetic xlsx fixtures
  B. TestBug18SafronIntegration     (2) -- real-fixture tests on safron r41
"""
from __future__ import annotations

import os
import tempfile
import unittest
from pathlib import Path

import openpyxl

import nirmaan_stack.services.boq_parser.reader as reader_mod
from nirmaan_stack.services.boq_parser.reader import BoqReader
from nirmaan_stack.services.boq_parser.classifier import (
    RowClassification,
    classify_row,
)
from nirmaan_stack.services.boq_parser.config import (
    ColumnRole,
    GlobalSettings,
    SheetConfig,
)
from nirmaan_stack.services.boq_parser.orchestrator import TEXT_ROLE_ROLES

_FIXTURES = Path(__file__).parent / "tests" / "fixtures"
_SAFRON_FIXTURE = _FIXTURES / "safron_hvac_2026-04-11.xlsx"
_SAFRON_SHEET = "Low Side Works R2 11.4.26"
_SAFRON_BANNER_ROW = 41
_SAFRON_BANNER_TEXT = "PART- 2 INSULATION"

_GS = GlobalSettings()


# ==================================================================
# A. Synthetic unit tests (5)
# ==================================================================

class TestBug18SyntheticMergeBanner(unittest.TestCase):
    """
    Bug 18 -- synthetic merge banner tests.

    Fixture layout: A1:C1 merged with 'SECTION HEADER' (A=sl_no origin,
    B=description covered, C=unit covered). D1 and E1 are unmerged numeric
    cells representing qty/rate (non-text-role).
    """

    def setUp(self):
        fd, self._path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "SECTION HEADER"
        ws.merge_cells("A1:C1")
        ws["D1"] = 5.0    # qty column (not text-role)
        ws["E1"] = 100.0  # rate column (not text-role)
        wb.save(self._path)

    def tearDown(self):
        if os.path.exists(self._path):
            os.unlink(self._path)

    def test_merge_origin_retains_value(self):
        """Origin A1 retains 'SECTION HEADER' even with text_role_columns active."""
        reader = BoqReader(self._path)
        rows = list(reader.iter_rows("Sheet1", start_row=1, end_row=1,
                                     text_role_columns={"A", "B", "C"}))
        a1 = rows[0].get_cell("A")
        self.assertIsNotNone(a1)
        self.assertEqual(a1.value, "SECTION HEADER",
                         "Origin A1 must retain its value; Bug 18 only blanks covered peers")

    def test_covered_text_role_cells_are_blank(self):
        """Covered B1 and C1 (text-role columns) report None after Bug 18 suppression."""
        reader = BoqReader(self._path)
        rows = list(reader.iter_rows("Sheet1", start_row=1, end_row=1,
                                     text_role_columns={"A", "B", "C"}))
        row = rows[0]
        b1 = row.get_cell("B")
        c1 = row.get_cell("C")
        self.assertIsNone(b1.value,
                          "Covered B1 in text-role column must be None after Bug 18 fix")
        self.assertIsNone(c1.value,
                          "Covered C1 in text-role column must be None after Bug 18 fix")

    def test_covered_non_text_role_cells_preserved(self):
        """
        Covered cells outside text_role_columns must NOT be suppressed.
        This is the must-not-regress test: area-header merges over qty/amount
        columns (the legitimate multi-area propagation pattern) must pass through.
        """
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Area 1"
        ws.merge_cells("A1:B1")  # area label; A and B are qty/amount roles (not text-role)
        wb.save(path)
        try:
            reader = BoqReader(path)
            # text_role_columns does NOT include A or B
            rows = list(reader.iter_rows("Sheet1", start_row=1, end_row=1,
                                         text_role_columns={"C"}))
            b1 = rows[0].get_cell("B")
            self.assertEqual(b1.value, "Area 1",
                             "Covered B1 outside text_role_columns must preserve propagated value")
        finally:
            os.unlink(path)

    def test_toggle_off_propagated_values_flow_through(self):
        """With BUG_18_MERGE_PROPAGATION_BLANK_ENABLED=False, covered text-role cells get value."""
        try:
            reader_mod.BUG_18_MERGE_PROPAGATION_BLANK_ENABLED = False
            reader = BoqReader(self._path)
            rows = list(reader.iter_rows("Sheet1", start_row=1, end_row=1,
                                         text_role_columns={"A", "B", "C"}))
            b1 = rows[0].get_cell("B")
            self.assertEqual(b1.value, "SECTION HEADER",
                             "Toggle=False: covered B1 must propagate origin value through")
        finally:
            reader_mod.BUG_18_MERGE_PROPAGATION_BLANK_ENABLED = True

    def test_no_text_role_columns_param_backward_compat(self):
        """Without text_role_columns param, covered cells still propagate value (backward compat)."""
        reader = BoqReader(self._path)
        rows = list(reader.iter_rows("Sheet1", start_row=1, end_row=1))
        b1 = rows[0].get_cell("B")
        self.assertEqual(b1.value, "SECTION HEADER",
                         "Without text_role_columns, covered cells must still propagate value")


# ==================================================================
# B. Real-fixture integration: safron r41 (2 tests)
# ==================================================================

class TestBug18SafronIntegration(unittest.TestCase):
    """
    Bug 18 integration tests on safron_hvac_2026-04-11.xlsx.

    Sheet: "Low Side Works R2 11.4.26"
    Row 41: "PART- 2 INSULATION" merged across A41:G41.
    Pre-fix: classified as LINE_ITEM (propagated text in unit column).
    Post-fix: classified as NOTE (description and unit columns blanked).
    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls._fixture_exists = _SAFRON_FIXTURE.exists()
        if not cls._fixture_exists:
            return

        cls.reader = BoqReader(str(_SAFRON_FIXTURE))

        # Minimal SheetConfig with standard HVAC column layout.
        # A=sl_no, B=description, C=unit are text-role (will be in text_role_columns).
        # D=qty, E=rate_supply, F=rate_install, G=amount_supply are numeric-role.
        cls.sheet_config = SheetConfig(
            sheet_name=_SAFRON_SHEET,
            header_row=5,
            column_role_map={
                "A": ColumnRole(role="sl_no"),
                "B": ColumnRole(role="description"),
                "C": ColumnRole(role="unit"),
                "D": ColumnRole(role="qty"),
                "E": ColumnRole(role="rate_supply"),
                "F": ColumnRole(role="rate_install"),
                "G": ColumnRole(role="amount_supply"),
            },
        )

        # Compute text_role_columns the same way the orchestrator does
        cls.text_role_columns = {
            col for col, cr in cls.sheet_config.column_role_map.items()
            if cr.role in TEXT_ROLE_ROLES
        }  # {"A", "B", "C"}

        cls.rows_by_number = {
            rr.row_number: rr
            for rr in cls.reader.iter_rows(
                _SAFRON_SHEET,
                text_role_columns=cls.text_role_columns,
            )
        }

    def setUp(self):
        if not self._fixture_exists:
            self.skipTest(f"Safron fixture not found: {_SAFRON_FIXTURE}")
        if _SAFRON_BANNER_ROW not in self.rows_by_number:
            self.skipTest(
                f"Row {_SAFRON_BANNER_ROW} not found in safron sheet '{_SAFRON_SHEET}'"
            )

    def test_safron_r41_covered_description_cell_blank(self):
        """
        After Bug 18 fix, covered B41 (description, text-role) must be None.
        Origin A41 must retain 'PART- 2 INSULATION'.
        Pre-fix baseline: B41.value = 'PART- 2 INSULATION' (propagated).
        """
        row = self.rows_by_number[_SAFRON_BANNER_ROW]
        a41 = row.get_cell("A")
        b41 = row.get_cell("B")
        self.assertEqual(
            a41.value, _SAFRON_BANNER_TEXT,
            f"Origin A41 must retain banner text '{_SAFRON_BANNER_TEXT}'",
        )
        self.assertIsNone(
            b41.value,
            f"Covered B41 (description) must be None after Bug 18; "
            f"row 41 banner is '{_SAFRON_BANNER_TEXT}'",
        )

    def test_safron_r41_classifies_as_note_not_line_item(self):
        """
        Row 41 ('PART- 2 INSULATION') must classify as NOTE after Bug 18 fix.
        Pre-fix baseline: LINE_ITEM (propagated text fills unit column, unit
        passes the junk check, Bug 16 Clause 2 does not fire).
        Post-fix path: description=None, unit=None -> Bug 16 Clause 2 fires -> NOTE.
        """
        row = self.rows_by_number[_SAFRON_BANNER_ROW]
        result = classify_row(row, self.sheet_config, _GS)
        self.assertNotEqual(
            result.classification, RowClassification.LINE_ITEM,
            f"Row {_SAFRON_BANNER_ROW} '{_SAFRON_BANNER_TEXT}' must NOT be LINE_ITEM "
            f"after Bug 18 fix; got {result.classification}",
        )
        self.assertEqual(
            result.classification, RowClassification.NOTE,
            f"Row {_SAFRON_BANNER_ROW} '{_SAFRON_BANNER_TEXT}' must classify as NOTE "
            f"after Bug 18 fix; got {result.classification}",
        )


if __name__ == "__main__":
    unittest.main()
