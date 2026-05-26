"""
BoqReader — clean structured access to Excel workbooks for the BoQ parser.

Wraps openpyxl and provides:
  - list_sheets()           — exact sheet names
  - get_sheet_dimensions()  — (last_row, last_col), content-based
  - iter_rows()             — lazy RawRow iterator
  - detect_header_row()     — heuristic header finder
  - detect_blank_columns()  — set of entirely-blank column letters
  - get_master_preamble_text() — concatenated text from a preamble sheet

No Frappe imports; fully testable in isolation.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Any, Iterator

import openpyxl
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------
# Bug 17: numeric-to-display-string helper for text-role columns
# ------------------------------------------------------------------

# Matches the decimal-precision specifier inside an Excel number format string.
# Handles "0.00", "#,##0.00", "0.00%", "_ * #,##0.00_ ;...", etc.
# Scientific-notation formats ("0.00E+00") contain "E" — callers guard against
# those separately if needed (BoQ sl_no fields never use scientific notation).
_PRECISION_FMT_RE = re.compile(r"0\.(0+)")


def _format_numeric_as_displayed(value: Any, number_format: str | None) -> Any:
    """
    Return the string a BoQ author would have seen on screen for a numeric cell.

    Applies only to text-role columns (sl_no, description, unit, make_model,
    append_to_notes). Numeric-role columns keep raw float for downstream math.

    Rules:
      None  → None    (blank-cell semantics preserved; do NOT return "")
      bool  → str()   (bool is int subclass in Python; guard before int check)
      str   → pass-through
      int/float:
        If number_format contains "0.<zeros>" precision specifier, apply
        format(value, ".Nf") where N = len(zeros).
        Otherwise (including "General"): str(round(value, 10)) — removes the
        15th-16th digit IEEE 754 noise typical of =A+0.1 chain formulas.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, str):
        return value
    if isinstance(value, (int, float)):
        fmt = number_format or "General"
        m = _PRECISION_FMT_RE.search(fmt)
        if m:
            n_decimals = len(m.group(1))
            return format(value, f".{n_decimals}f")
        # No decimal-precision specifier: strip IEEE 754 noise via round-to-10
        return str(round(value, 10))
    return str(value)

# ------------------------------------------------------------------
# Data structures
# ------------------------------------------------------------------

@dataclass
class CellInfo:
    value: Any                  # computed value (data_only=True read)
    formula: str | None         # formula text if cell is a formula, else None
    is_formula: bool            # True when value comes from a formula
    is_merged_origin: bool      # True if top-left of a merged range
    merged_range: str | None    # e.g. "A2:B3", or None
    font_bold: bool             # detected bold formatting
    fill_color_rgb: str | None  # hex RGB e.g. "FFFFAA00", or None
    indent: int                 # alignment indent level (0+)


@dataclass
class RawRow:
    row_number: int                          # 1-indexed
    cells: dict[str, CellInfo] = field(default_factory=dict)

    def get_cell(self, col_letter: str) -> CellInfo | None:
        return self.cells.get(col_letter.upper())

    def is_blank(self) -> bool:
        for ci in self.cells.values():
            if ci.value is not None and str(ci.value).strip() != "":
                return False
        return True


# ------------------------------------------------------------------
# Header-row heuristic keywords (weighted)
# ------------------------------------------------------------------

# Strong match: +2 pts per cell (unambiguous column-header vocabulary)
_STRONG_HEADER_KEYWORDS = {
    "sl.no", "s.no", "sno", "sr.no", "description", "item description",
}

# Medium match: +1 pt per cell (domain-specific but less definitive)
_MEDIUM_HEADER_KEYWORDS = {
    "item", "material", "materials", "details of materials",
    "make/model", "approved make", "approved makes",
}

_COL_LETTER_RE = re.compile(r"^[A-Z]+$")


# ------------------------------------------------------------------
# Excel error literal detection (Bug 13, sec 9 #89)
# ------------------------------------------------------------------

EXCEL_ERROR_LITERALS: frozenset[str] = frozenset({
    "#REF!", "#VALUE!", "#NAME?", "#DIV/0!",
    "#NULL!", "#N/A", "#NUM!",
})


def _is_excel_error(value: Any) -> bool:
    """Return True if value is one of Excel's seven error literal strings.
    Whitespace-tolerant, case-insensitive comparison."""
    return isinstance(value, str) and value.strip().upper() in EXCEL_ERROR_LITERALS


# ------------------------------------------------------------------
# BoqReader
# ------------------------------------------------------------------

class BoqReader:
    """
    Opens an .xlsx workbook in both data_only=True and data_only=False modes
    simultaneously to capture computed values AND formula strings.

    Use read_only=False (default) so cell formatting attributes are available.
    """

    def __init__(self, file_path: str) -> None:
        self._path = file_path
        # data_only=True — computed cell values (cached by Excel/LibreOffice)
        self._wb_values = openpyxl.load_workbook(
            file_path, data_only=True, read_only=False
        )
        # data_only=False — raw formula strings
        self._wb_formulas = openpyxl.load_workbook(
            file_path, data_only=False, read_only=False
        )

        # Pre-build merged-range lookup per sheet:
        # sheet_name → { "A1" → "A1:B3", ... } for each top-left cell
        self._merged_origins: dict[str, dict[str, str]] = {}
        for sheet_name in self._wb_values.sheetnames:
            ws = self._wb_values[sheet_name]
            origins: dict[str, str] = {}
            for rng in ws.merged_cells.ranges:
                tl_col = get_column_letter(rng.min_col)
                tl_key = f"{tl_col}{rng.min_row}"
                origins[tl_key] = str(rng)
            self._merged_origins[sheet_name] = origins

    # ------------------------------------------------------------------ #
    # Public API                                                           #
    # ------------------------------------------------------------------ #

    def list_sheets(self) -> list[str]:
        """Return sheet names exactly as openpyxl reports them."""
        return list(self._wb_values.sheetnames)

    def list_sheet_states(self) -> dict[str, str]:
        """
        Return a mapping from sheet name to sheet visibility state.

        Sheet visibility values come straight from openpyxl's `Worksheet.sheet_state`:
        one of 'visible', 'hidden', or 'veryHidden'. Sheet names preserve exact
        whitespace and casing, matching `list_sheets()`.

        Used by the Phase 3 wizard to default hidden / veryHidden sheets to the
        'skip' disposition. Pure pass-through — no caching, no transformation.
        """
        return {ws.title: ws.sheet_state for ws in self._wb_values.worksheets}

    def get_sheet_dimensions(self, sheet_name: str) -> tuple[int, int]:
        """
        Return (last_row, last_col) — 1-indexed, content-based.

        Walks rows to find the actual last row with any non-empty cell because
        openpyxl's ws.max_row can be unreliable for files edited by non-Excel
        tools (it reports blank trailing rows that were previously written to).
        """
        ws = self._wb_values[sheet_name]
        last_row = 0
        last_col = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != "":
                    if cell.row > last_row:
                        last_row = cell.row
                    if cell.column > last_col:
                        last_col = cell.column
        return last_row, last_col

    def iter_rows(
        self,
        sheet_name: str,
        start_row: int = 1,
        end_row: int | None = None,
        text_role_columns: set[str] | None = None,
    ) -> Iterator[RawRow]:
        """
        Yield one RawRow per row, lazy.

        Reads computed values from the data_only workbook and formula strings
        from the non-data_only workbook, then merges them per cell.

        Covered cells (inside a merged range but not the top-left origin) receive
        the origin's value, formula text, and merged_range string.  Formatting
        fields (font_bold, fill_color_rgb, indent) always reflect the covered
        cell's own data — they are never inherited from the origin.

        text_role_columns: optional set of column letters (e.g. {"A", "C"}) for
        which numeric values are formatted as the author saw them on screen, using
        the cell's number_format. Pass None (default) for unchanged behavior.
        See _format_numeric_as_displayed() for the conversion rules (Bug 17).
        """
        ws_val = self._wb_values[sheet_name]
        ws_fml = self._wb_formulas[sheet_name]
        origins = self._merged_origins.get(sheet_name, {})

        # Build a per-invocation lookup for covered cells (inside a merge but
        # not the origin).  Key: (row, col) integer tuple.
        # Value: (range_str, origin_value, origin_formula_text, origin_is_formula,
        #         origin_number_format)
        covered_lookup: dict[tuple[int, int], tuple[str, Any, str | None, bool, str | None]] = {}
        for rng in ws_val.merged_cells.ranges:
            origin_row, origin_col = rng.min_row, rng.min_col
            origin_cell_val = ws_val.cell(origin_row, origin_col)
            origin_cell_fml = ws_fml.cell(origin_row, origin_col)
            origin_value = origin_cell_val.value
            origin_raw_fml = origin_cell_fml.value
            origin_is_formula = (
                isinstance(origin_raw_fml, str) and origin_raw_fml.startswith("=")
            )
            origin_formula_text = origin_raw_fml if origin_is_formula else None
            origin_number_format: str | None = getattr(origin_cell_val, "number_format", None)
            range_str = str(rng)
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    if r == origin_row and c == origin_col:
                        continue  # origin handled by existing logic below
                    covered_lookup[(r, c)] = (
                        range_str,
                        origin_value,
                        origin_formula_text,
                        origin_is_formula,
                        origin_number_format,
                    )

        if end_row is None:
            # Use content-based last-row detection so empty sheets yield nothing
            # instead of returning a row of all-None cells (openpyxl's max_row
            # defaults to 1 for empty sheets after save/reload).
            content_last_row, _ = self.get_sheet_dimensions(sheet_name)
            if content_last_row == 0:
                return
            end_row = content_last_row

        for row_cells in ws_val.iter_rows(min_row=start_row, max_row=end_row):
            row_num = row_cells[0].row
            cells: dict[str, CellInfo] = {}

            for cell_val in row_cells:
                col_letter = get_column_letter(cell_val.column)
                cell_key = f"{col_letter}{row_num}"
                coord = (row_num, cell_val.column)

                covered = covered_lookup.get(coord)
                if covered is not None:
                    # Covered cell: propagate value, formula, and range from origin.
                    range_str, origin_value, origin_formula_text, origin_is_formula, origin_number_format = covered
                    computed_value = origin_value
                    formula_text = origin_formula_text
                    is_formula = origin_is_formula
                    is_origin = False
                    merged_range_val = range_str
                    cell_number_format = origin_number_format
                else:
                    # Non-covered cell (origin or unmerged): existing behavior.
                    cell_fml = ws_fml[cell_key]
                    raw_formula = cell_fml.value
                    is_formula = isinstance(raw_formula, str) and raw_formula.startswith("=")
                    formula_text = raw_formula if is_formula else None
                    computed_value = cell_val.value
                    is_origin = cell_key in origins
                    merged_range_val = origins.get(cell_key)
                    cell_number_format = getattr(cell_val, "number_format", None)

                # Normalize Excel error literals to None (Bug 13, sec 9 #89)
                if _is_excel_error(computed_value):
                    computed_value = None

                # Bug 17: format numeric values as displayed for text-role columns
                if text_role_columns is not None and col_letter in text_role_columns:
                    computed_value = _format_numeric_as_displayed(
                        computed_value, cell_number_format
                    )

                # Formatting — always the covered cell's own, never inherited from origin.
                font_bold = bool(cell_val.font and cell_val.font.bold)
                fill_rgb = _extract_fill_rgb(cell_val)
                indent = _extract_indent(cell_val)

                cells[col_letter] = CellInfo(
                    value=computed_value,
                    formula=formula_text,
                    is_formula=is_formula,
                    is_merged_origin=is_origin,
                    merged_range=merged_range_val,
                    font_bold=font_bold,
                    fill_color_rgb=fill_rgb,
                    indent=indent,
                )

            yield RawRow(row_number=row_num, cells=cells)

    def detect_header_row(
        self, sheet_name: str, scan_top_n: int = 15
    ) -> int | None:
        """
        Scan the first scan_top_n rows and return the row number that best
        matches a BoQ column-header row.

        Scoring (weighted, per cell — each cell counted at most once):
          Strong keywords (e.g. "sl.no", "description"): +2
          Medium keywords (e.g. "item", "materials"):     +1

        Row-shape guards (row skipped when any fail):
          1. ≥ 3 non-empty cells — filters merged-title and label-only rows
          2. ≤ 1 cell with text > 60 chars — filters narrative description rows
          3. Not the last content row — avoids isolated data rows at sheet end

        Returns None if no row scores > 1 (requires ≥ 2 weighted points).
        Tiebreak: earliest row wins.
        """
        ws = self._wb_values[sheet_name]
        last_content_row, _ = self.get_sheet_dimensions(sheet_name)

        best_row: int | None = None
        best_score = 1  # require score strictly > 1

        for row_cells in ws.iter_rows(max_row=scan_top_n):
            if not row_cells:
                continue
            row_num = row_cells[0].row

            # Guard 3 — skip last content row (likely a totals / data row)
            if row_num == last_content_row:
                continue

            # Collect non-empty cell texts
            cell_texts: list[str] = []
            long_text_count = 0
            for cell in row_cells:
                if cell.value is None:
                    continue
                text = str(cell.value).strip()
                if not text:
                    continue
                cell_texts.append(text)
                if len(text) > 60:
                    long_text_count += 1

            # Guard 1 — at least 3 non-empty cells
            if len(cell_texts) < 3:
                continue

            # Guard 2 — at most 1 cell with text > 60 chars
            if long_text_count > 1:
                continue

            # Weighted scoring
            score = 0
            for text in cell_texts:
                text_lower = text.lower()
                matched = False
                for kw in _STRONG_HEADER_KEYWORDS:
                    if kw in text_lower:
                        score += 2
                        matched = True
                        break
                if not matched:
                    for kw in _MEDIUM_HEADER_KEYWORDS:
                        if kw in text_lower:
                            score += 1
                            break

            if score > best_score:
                best_score = score
                best_row = row_num

        return best_row

    def detect_blank_columns(
        self, sheet_name: str, scan_rows: int = 50
    ) -> set[str]:
        """
        Return the set of column letters (e.g. {"A", "Z"}) whose every cell
        in rows 1..scan_rows is None or empty string.

        Only inspects columns up to the workbook's max_column for that sheet.
        """
        ws = self._wb_values[sheet_name]
        max_col = ws.max_column or 0
        if max_col == 0:
            return set()

        non_blank: set[str] = set()
        for row_cells in ws.iter_rows(max_row=scan_rows, max_col=max_col):
            for cell in row_cells:
                if cell.value is not None and str(cell.value).strip() != "":
                    non_blank.add(get_column_letter(cell.column))

        all_cols = {get_column_letter(c) for c in range(1, max_col + 1)}
        return all_cols - non_blank

    def get_master_preamble_text(self, sheet_name: str) -> str:
        """
        Concatenate all non-empty cell text from the sheet, row-major order,
        separated by newlines.
        """
        ws = self._wb_values[sheet_name]
        parts: list[str] = []
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                if cell.value is not None:
                    text = str(cell.value).strip()
                    if text:
                        parts.append(text)
        return "\n".join(parts)


# ------------------------------------------------------------------
# Internal helpers
# ------------------------------------------------------------------

def _extract_fill_rgb(cell: Any) -> str | None:
    """
    Return the solid fill foreground color as a hex RGB string, or None.

    Theme colors and pattern fills that aren't solid return None.
    """
    try:
        fill = cell.fill
        if fill is None or fill.fill_type != "solid":
            return None
        rgb = fill.fgColor.rgb
        # Theme colors return an integer 0 or non-string types
        if not isinstance(rgb, str) or len(rgb) < 6:
            return None
        return rgb
    except Exception:
        return None


def _extract_indent(cell: Any) -> int:
    """Return the cell's alignment indent level (0 if not set)."""
    try:
        indent = cell.alignment.indent
        return int(indent) if indent else 0
    except Exception:
        return 0
