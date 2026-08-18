# Copyright (c) 2026, Nirmaan (Stratos Infra Technologies Pvt. Ltd.) and contributors
# For license information, please see license.txt

"""Bank-outflow statement parser (Bulk Import Outflow, slice S1).

PURE MODULE -- no `frappe`, no database, no request context, no filesystem. It takes the uploaded
bytes and returns a `ParseResult`; persistence is the caller's job.

TWO CONTRACTS THAT LOOK LIKE DETAILS AND ARE NOT:

1. **The parser NEVER filters.** A FAILED, PENDING or REVERSED transfer is parsed and returned
   exactly like a successful one, carrying its real status. Downstream auto-skips it, and that skip
   is VISIBLE on the row. The reason is concrete: a failed transfer still carries a bank reference
   number, and its successful retry appears as a separate row in the same file -- in a real 19-row
   statement, `Sri Sai Roadlines` Rs 22,000 FAILED at 18:17:59 with reference 620918791146 and
   succeeded at 19:23 with reference 620919871893, and BOTH would otherwise match the same payment.
   A parser that dropped the failure silently would leave nobody able to see why the numbers moved.

2. **`charges_amount` sums EVERY row; `gross_amount` sums only successful ones.** The asymmetry is
   deliberate. A charge is money the bank took whatever the transfer's outcome, so excluding failed
   rows from it would understate the debit. The beneficiary amount of a failed transfer, by
   contrast, never left the account. This is the same reason the batch total can never equal the sum
   of its rows: gateway charge plus tax belongs to no settlement target at all.

MONEY IS `Decimal` THROUGHOUT. These figures are compared for exact equality against stored amounts
and then differenced; binary floating point makes both unreliable at the paisa level. The conversion
to `float` happens once, at the persistence boundary.

Adding a source is a new entry in `_ADAPTERS` -- a column map plus its required set. No other code
in this module is source-aware.

⚠️ CASHBOOK NEEDED TWO GENERAL CAPABILITIES, AND BOTH ARE DELIBERATELY GENERAL RATHER THAN
CASHBOOK-SHAPED. The parked note that adding a second source was "one adapter entry" was optimistic;
it was one adapter entry plus these two, each of which is a rule about statements in general:

1. **A field may be mapped to SEVERAL columns.** A map value may be one header or a tuple of them,
   whose non-empty values join with `_MULTI_COLUMN_JOIN`. Cashbook splits its free text across
   `Remark` and `Note` -- 7 rows in a 115-row sample carry a `Note` saying more than the `Remark`
   does ("Pay to BharatPe Merchant" / "VR mall site") -- and both feed the same matcher, so they
   have to arrive as one string. The join is visible, so the two halves stay legible; the file
   itself is attached to the batch, so the exact cells stay recoverable.

2. **A row carrying NO transfer id, NO amount and NO status is file furniture, not a transfer.**
   Statements end in a totals block -- Cashbook's is six rows reading "Opening VA Balance",
   "Ending Total Balance" and so on down the date column. Those were being reported as six
   "has no Transfer Id" warnings on every single import, which is exactly how a warning list stops
   being read. A row with no id but a real amount or status is still warned about, because that one
   is a genuine data problem: see `_build_row`.

⚠️ SOURCE AND FORMAT ARE DIFFERENT AXES -- see below. Neither capability above is a format concern.

FORMAT IS SNIFFED FROM THE BYTES, NOT DECLARED (Q10, slice V3). `.csv` and `.xlsx` both arrive here
as bytes and are told apart by the ZIP magic number that starts every xlsx -- a CSV cannot begin
with it. The caller does not say which it has, and neither does the accountant: "the sheet format
stops being something they have to think about" was the point of the ruling. A file whose extension
disagrees with its contents therefore still parses correctly, which is the common shape when
someone renames an export.

⚠️ SOURCE AND FORMAT ARE DIFFERENT AXES. `source` is WHOSE statement this is (Cashfree, one day
Cashbook) and selects the column map; format is how the bytes are encoded. A Cashfree export is the
same statement whether saved as .csv or .xlsx, so adding a format must never mean adding a source.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal

from nirmaan_stack.services.outflow_import.duplicates import row_identity
from nirmaan_stack.services.outflow_import.normalize import (
    normalize_account,
    normalize_amount,
    normalize_reference,
)

__all__ = [
    "RawRow",
    "ParseResult",
    "StatementFormatError",
    "parse_statement",
    "SUPPORTED_SOURCES",
    "BANK_SUCCESS_STATUS",
    "BANK_TERMINAL_STATUSES",
    "is_success_status",
    "is_terminal_status",
]

# The one word the gateway uses for money that actually left the account. Everything else -- FAILED,
# PENDING, REVERSED -- is a transfer that did not happen.
BANK_SUCCESS_STATUS = "SUCCESS"

# The statuses that are the LAST WORD on a transfer -- its story is over, whichever way it went.
#
# ⚠️ THIS IS A DIFFERENT QUESTION FROM `is_success_status`, AND CONFLATING THEM COSTS MONEY IN ONE
# DIRECTION AND CLARITY IN THE OTHER. "Did money move?" decides whether a row can SETTLE anything.
# "Is this the final account of the transfer?" decides whether a stored row may be treated as an
# IMPORT -- see `candidates.find_earlier_batches_for_rows`. A transfer still QUEUED is neither: it
# settles nothing today AND tomorrow's export will say something new about it, so freezing it as
# "already imported" strands the money permanently. A FAILED transfer also settles nothing, but it
# IS final -- a retry gets a new transfer id, so that row will never say anything else -- and it must
# keep counting as a duplicate, or a re-uploaded statement stops being recognised as fully imported
# and stages a batch with nothing in it to action.
#
# ⚠️ ANYTHING NOT LISTED HERE IS TREATED AS STILL IN FLIGHT, AND THE DEFAULT LEANS THAT WAY ON
# PURPOSE. The two mistakes are not symmetric: calling an in-flight status final loses real money
# with no trace and no way back (a skipped row is frozen against re-matching), while calling a final
# status in-flight costs a re-staged row that is skipped anyway. Only one of those is recoverable,
# so an unrecognised status must fall on the recoverable side.
BANK_TERMINAL_STATUSES = frozenset({BANK_SUCCESS_STATUS, "FAILED", "REJECTED", "REVERSED"})


def is_success_status(status_raw: str | None) -> bool:
    """Did this transfer actually move money?

    ⚠️ THE SINGLE DEFINITION OF "SUCCESSFUL", AND IT HAD TO BECOME ONE. It was a property on `RawRow`
    and nothing else needed it -- until the import summary had to exclude failed transfers from every
    figure it reports, which happens long after parsing, against rows read back out of the database.
    A second copy of `== "SUCCESS"` in that query is the shape where one side later learns about
    `REVERSED` and the other does not, and the two disagree about the same statement.

    `review.get_import_summary` binds `BANK_SUCCESS_STATUS` into its `GROUP BY` for exactly this
    reason: the SQL and this function compare against the same literal, from here.
    """
    return (status_raw or "").strip().upper() == BANK_SUCCESS_STATUS


def is_terminal_status(status_raw: str | None) -> bool:
    """Is this the LAST thing the statement will ever say about this transfer?

    True for a transfer that succeeded AND for one that definitively did not. False while it is
    still in flight, and false for a status we do not recognise -- see `BANK_TERMINAL_STATUSES` for
    why the unknown case leans that way.

    Normalises identically to `is_success_status`, so the two can never disagree about the same
    cell, and the SQL that binds this set uses `UPPER(BTRIM(...))` to match.
    """
    return (status_raw or "").strip().upper() in BANK_TERMINAL_STATUSES


class StatementFormatError(ValueError):
    """The uploaded file is not a statement this source knows how to read.

    Raised only for a whole-file problem -- an unreadable encoding, an empty file, or a header that
    is missing a column the matcher depends on. A bad value in a single row is never this; it is a
    row-level finding the caller reports.
    """


@dataclass(frozen=True)
class RawRow:
    """One transfer, verbatim, plus the two identity forms the matcher needs.

    Every raw field is kept exactly as the statement wrote it. The derived fields sit BESIDE the raw
    ones rather than replacing them, so the original is always recoverable and a re-match can never
    destroy evidence.
    """

    row_number: int
    transfer_id: str
    reference_id: str
    added_on: datetime | None
    amount: Decimal
    status_raw: str
    beneficiary_name: str
    beneficiary_id: str
    bank_account: str
    ifsc: str
    remarks: str
    bank_reference_no: str
    service_charge: Decimal
    service_tax: Decimal
    added_by_raw: str
    normalized_account: str
    normalized_reference: str
    row_kind: str = ""
    """What KIND of movement the statement says this is, verbatim, or "" where it does not say.

    Cashbook mixes spends with wallet top-ups and bank loads in one file and tells them apart in a
    `Type` column; only `Wallet Spend` is money leaving on someone's behalf. Cashfree's export is
    transfers only, so it has no such column and this stays blank there.

    ⚠️ THE PARSER STILL NEVER FILTERS -- this field only RECORDS what the statement said. Deciding
    that a top-up is not importable is a downstream judgement, and it stays downstream so the skip
    is visible on a staged row rather than being an absence nobody can account for. Same contract
    the FAILED transfer above already follows, for the same reason.

    It carries a default because it is the one field a statement may genuinely not have, and every
    caller that builds a `RawRow` by hand predates it.
    """

    @property
    def is_success(self) -> bool:
        return is_success_status(self.status_raw)

    @property
    def added_on_date(self) -> date | None:
        return self.added_on.date() if self.added_on else None


@dataclass(frozen=True)
class ParseResult:
    """Everything the caller needs to stage a batch, and nothing it has to recompute."""

    source: str
    rows: tuple[RawRow, ...]
    period_from: date | None
    period_to: date | None
    gross_amount: Decimal
    charges_amount: Decimal
    duplicate_transfer_ids: tuple[str, ...]
    warnings: tuple[str, ...]

    @property
    def success_count(self) -> int:
        return sum(1 for row in self.rows if row.is_success)


# --- source adapters ---------------------------------------------------------------------------
#
# A statement may carry columns we ignore entirely (VPA, Acknowledged, Mode, Status Code, Payment
# Instrument ID, Last checked at, Status Description, Extended UTR). Unknown columns are fine;
# a MISSING required column is not, because the matcher would then silently lose a signal.

_CASHFREE_COLUMNS = {
    "transfer_id": "Transfer Id",
    "reference_id": "Reference Id",
    "added_on": "Added On",
    "amount": "Amount",
    "status_raw": "Status",
    "beneficiary_name": "Beneficiary Name",
    "beneficiary_id": "Beneficiary Id",
    "bank_account": "Bank Account",
    "ifsc": "IFSC",
    "remarks": "Remarks",
    "bank_reference_no": "Bank Reference No",
    "service_charge": "Service Charge",
    "service_tax": "Service Tax",
    "added_by_raw": "Added by",
}

# Only the columns without which a row cannot be identified, matched or reported.
# `Beneficiary Id` is absent on purpose: it is stored nowhere in this database, so it can never
# resolve to anything and is retained for provenance only.
_CASHFREE_REQUIRED = frozenset(
    {"Transfer Id", "Added On", "Amount", "Status", "Beneficiary Name", "Bank Reference No"}
)

# Cashbook is a PETTY-CASH WALLET statement, not a bank transfer export, and the difference shows
# in what is absent: no bank reference, no beneficiary account, no IFSC, no gateway charges. Match
# tiers 0 and 1 read exactly those fields, so they can find nothing here -- which is the mechanical
# reason this source gets its own downstream path rather than a widened matcher.
#
# ⚠️ `Date` IS TAKEN WITHOUT `Time`, WHICH SITS BESIDE IT (owner ruling). `%d/%m/%Y` is already in
# `_DATETIME_FORMATS`, so the cell parses to midnight and the clock time is dropped. Two spends to
# one payee on one day therefore share a timestamp and are told apart by their transfer ids, which
# are unique -- verified across a 115-row sample.
_CASHBOOK_COLUMNS = {
    "transfer_id": "Txn Id",
    "added_on": "Date",
    "amount": "Debit",
    "status_raw": "Payment Status",
    "beneficiary_name": "To",
    # Who spent it. On a wallet statement this is a real person holding the card, and it becomes the
    # created expense's `payment_by` -- so the record says who spent the money rather than who
    # imported the file.
    "added_by_raw": "From",
    # See capability 1 in the module docstring. `Note` is genuinely optional and often blank.
    "remarks": ("Remark", "Note"),
    "row_kind": "Type",
}

# ⚠️ `Remark` IS REQUIRED, AND IT IS THE ONE ENTRY HERE THAT IS NOT ABOUT IDENTIFYING A ROW. It is
# the ONLY signal this source carries for choosing a project or an expense type -- there is no
# account number to resolve and no reference to look up. A file missing it would parse perfectly and
# then book every single row to a fallback, which is the silent-loss-of-a-signal case the note above
# `_CASHFREE_REQUIRED` exists to prevent, in its most complete form.
#
# `Debit` is required as a COLUMN, not as a value: it is blank on every top-up row by design.
_CASHBOOK_REQUIRED = frozenset(
    {"Txn Id", "Date", "Type", "Debit", "Payment Status", "To", "Remark"}
)

_ADAPTERS = {
    "Cashfree": (_CASHFREE_COLUMNS, _CASHFREE_REQUIRED),
    "Cashbook": (_CASHBOOK_COLUMNS, _CASHBOOK_REQUIRED),
}

# Visible on purpose: a reader of a joined remark can see where one cell ended and the next began,
# and it matches the separator `settle._default_description` already composes with.
_MULTI_COLUMN_JOIN = " - "

SUPPORTED_SOURCES = tuple(sorted(_ADAPTERS))

_DATETIME_FORMATS = (
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%dT%H:%M",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
    "%d-%m-%Y %H:%M:%S",
    "%d/%m/%Y %H:%M:%S",
    "%d-%m-%Y",
    "%d/%m/%Y",
)


def parse_statement(content: bytes, source: str = "Cashfree") -> ParseResult:
    """Parse an uploaded statement into rows plus the batch-level figures.

    Raises `StatementFormatError` for a whole-file problem. Row-level oddities become warnings.
    """
    if source not in _ADAPTERS:
        raise StatementFormatError(
            f"Unsupported outflow source {source!r}. Supported: {', '.join(SUPPORTED_SOURCES)}."
        )
    column_map, required = _ADAPTERS[source]

    fieldnames, records = _read_records(content)
    headers = [h.strip() for h in fieldnames]
    if not headers:
        raise StatementFormatError("The uploaded statement has no header row.")

    missing = sorted(required - set(headers))
    if missing:
        raise StatementFormatError(
            f"This does not look like a {source} statement. Missing column(s): {', '.join(missing)}."
        )

    header_lookup = {h.strip(): h for h in fieldnames}
    rows: list[RawRow] = []
    warnings: list[str] = []

    for position, record in enumerate(records, start=1):
        row = _build_row(position, record, column_map, header_lookup, warnings)
        if row is None:
            continue
        rows.append(row)

    if not rows:
        raise StatementFormatError("The uploaded statement contains no transaction rows.")

    duplicates = _duplicate_transfer_ids(rows)
    if duplicates:
        warnings.append(
            f"{len(duplicates)} transfer id(s) appear more than once in this file: "
            + ", ".join(duplicates[:5])
            + ("..." if len(duplicates) > 5 else "")
        )

    dates = [row.added_on_date for row in rows if row.added_on_date]
    # Charges across EVERY row, gross across successful ones only -- see the module docstring.
    gross = sum((row.amount for row in rows if row.is_success), Decimal("0"))
    charges = sum((row.service_charge + row.service_tax for row in rows), Decimal("0"))

    return ParseResult(
        source=source,
        rows=tuple(rows),
        period_from=min(dates) if dates else None,
        period_to=max(dates) if dates else None,
        gross_amount=gross,
        charges_amount=charges,
        duplicate_transfer_ids=duplicates,
        warnings=tuple(warnings),
    )


def _build_row(
    position: int,
    record: dict,
    column_map: dict,
    header_lookup: dict,
    warnings: list[str],
) -> RawRow | None:
    def cell(header: str) -> str:
        value = record.get(header_lookup.get(header, header))
        return "" if value is None else str(value)

    def raw(field: str) -> str:
        """This field's value, joining several columns into one where the map names several.

        A tuple joins only the parts that carry something, so a blank optional column leaves no
        dangling separator and a row with one of the two reads exactly as it would have if the
        other column did not exist.
        """
        header = column_map.get(field)
        if header is None:
            return ""
        if isinstance(header, tuple):
            parts = [part for part in (cell(name).strip() for name in header) if part]
            return _MULTI_COLUMN_JOIN.join(parts)
        return cell(header)

    transfer_id = raw("transfer_id").strip()
    if not transfer_id:
        # ⚠️ TWO DIFFERENT THINGS ARRIVE HERE AND ONLY ONE IS WORTH A WARNING. A row with no id but
        # a real amount or status is a TRANSFER WE CANNOT IDENTIFY -- a genuine defect in the export
        # and exactly what this warning was written for. A row with no id, no amount and no status
        # is the totals block at the foot of the sheet: it was never a transfer, so reporting it
        # says nothing and, at six lines per import, trains people to skip the warning list
        # entirely. See capability 2 in the module docstring.
        if raw("amount").strip() or raw("status_raw").strip():
            warnings.append(f"Row {position} has no Transfer Id and was not staged.")
        return None

    added_on = _parse_datetime(raw("added_on"))
    if added_on is None:
        warnings.append(f"Row {position} ({transfer_id}) has an unreadable Added On value.")

    bank_account = raw("bank_account").strip()
    bank_reference_no = raw("bank_reference_no").strip()

    return RawRow(
        row_number=position,
        transfer_id=transfer_id,
        reference_id=raw("reference_id").strip(),
        added_on=added_on,
        amount=normalize_amount(raw("amount")),
        status_raw=raw("status_raw").strip(),
        beneficiary_name=raw("beneficiary_name").strip(),
        beneficiary_id=raw("beneficiary_id").strip(),
        bank_account=bank_account,
        ifsc=raw("ifsc").strip(),
        # Remarks is kept VERBATIM -- no strip, no truncation. It is stored in a Text column
        # precisely so a long remark survives; as Data it is varchar(140) and Frappe throws
        # CharacterLengthExceededError rather than truncating.
        remarks=raw("remarks"),
        bank_reference_no=bank_reference_no,
        service_charge=normalize_amount(raw("service_charge")),
        service_tax=normalize_amount(raw("service_tax")),
        added_by_raw=raw("added_by_raw").strip(),
        normalized_account=normalize_account(bank_account),
        normalized_reference=normalize_reference(bank_reference_no),
        row_kind=raw("row_kind").strip(),
    )


def _read_records(content: bytes) -> tuple[list[str], list[dict]]:
    """Turn statement bytes into (header names, records) whichever format they are in.

    ONE seam for both formats, so everything downstream -- the required-column check, the column
    map, `_build_row` -- is format-blind and stays that way. A format is a way of encoding a table;
    it must not become a second parser.
    """
    if _is_xlsx(content):
        return _read_xlsx(content)
    return _read_csv(content)


def _is_xlsx(content: bytes) -> bool:
    """An .xlsx is a ZIP archive, so it starts with the ZIP local-file-header magic.

    Sniffing beats trusting the extension: a renamed export is common, and the failure mode of
    guessing wrong is a baffling "missing column" error rather than an honest one.
    """
    return isinstance(content, (bytes, bytearray)) and bytes(content[:4]) == b"PK\x03\x04"


def _read_csv(content: bytes) -> tuple[list[str], list[dict]]:
    text = _decode(content)
    if not text.strip():
        raise StatementFormatError("The uploaded statement is empty.")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader.fieldnames or []), list(reader)


def _read_xlsx(content: bytes) -> tuple[list[str], list[dict]]:
    """Read the FIRST worksheet of an .xlsx as a header row plus records.

    ⚠️ `data_only=True` returns a formula cell's CACHED VALUE. A workbook saved by a tool that never
    calculated would hand us `None` there -- which surfaces as an empty field and a row-level
    warning, not a wrong number. Reading the formula TEXT instead would be far worse: it would
    parse as a garbage amount.

    ⚠️ Values arrive TYPED here and as text from a CSV -- a date cell is a `datetime`, an amount a
    `float`. Everything is stringified so `_build_row` sees exactly what it sees from a CSV and
    there is one set of coercion rules, not two. `str(datetime)` yields "YYYY-MM-DD HH:MM:SS",
    which `_parse_datetime` already accepts; that is why the format list carries it.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover -- openpyxl ships with Frappe
        raise StatementFormatError(
            "This server cannot read .xlsx statements. Save the sheet as .csv and upload that."
        ) from exc

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise StatementFormatError(
            "The uploaded file looks like a spreadsheet but could not be opened."
        ) from exc

    try:
        if not workbook.sheetnames:
            raise StatementFormatError("The uploaded workbook has no sheets.")
        sheet = workbook[workbook.sheetnames[0]]

        row_iter = sheet.iter_rows(values_only=True)
        header_cells = next(row_iter, None)
        if header_cells is None:
            raise StatementFormatError("The uploaded statement has no header row.")
        # Trailing empty header cells are what Excel leaves behind after a column is cleared.
        fieldnames = [("" if cell is None else str(cell)).strip() for cell in header_cells]
        while fieldnames and not fieldnames[-1]:
            fieldnames.pop()
        if not any(fieldnames):
            raise StatementFormatError("The uploaded statement has no header row.")

        records: list[dict] = []
        for cells in row_iter:
            # Excel reports a used range that routinely outruns the data; a wholly empty tuple is
            # padding, not a transfer, and staging it would manufacture a row nobody exported.
            if cells is None or all(cell is None or str(cell).strip() == "" for cell in cells):
                continue
            record = {}
            for index, name in enumerate(fieldnames):
                if not name:
                    continue
                value = cells[index] if index < len(cells) else None
                record[name] = "" if value is None else str(value)
            records.append(record)
    finally:
        workbook.close()

    return fieldnames, records


def _duplicate_transfer_ids(rows: list[RawRow]) -> tuple[str, ...]:
    """Transfer ids that appear MORE THAN ONCE AS THE SAME TRANSFER within this one statement.

    ⚠️ IT REPEATS ON IDENTITY BUT REPORTS THE ID (slice D3). The two halves are deliberate:

    * Repeats are counted on `(transfer_id, amount, date)`, because `_stage_batch` marks its
      `duplicate_in_file` rows on exactly that key. Leaving this on `transfer_id` alone would let
      the preview call two rows repeated while staging called them distinct -- the same file,
      two answers.
    * What comes BACK is still the transfer id, because `duplicate_transfer_ids` is an API payload
      field typed `string[]` on the client. Returning tuples would be a wire change for a list
      nothing currently renders.

    A consequence worth being explicit about: one id carried twice at two DIFFERENT amounts is no
    longer reported here, because those are now two different transfers that happen to share an id.
    """
    seen: set = set()
    repeated: list[str] = []
    for row in rows:
        identity = row_identity(row.transfer_id, row.amount, row.added_on_date)
        if identity in seen and row.transfer_id not in repeated:
            repeated.append(row.transfer_id)
        seen.add(identity)
    return tuple(repeated)


def _decode(content: bytes) -> str:
    """Decode statement bytes, preferring UTF-8 and falling back rather than failing.

    `utf-8-sig` first because a spreadsheet-exported CSV routinely carries a BOM, which would
    otherwise become part of the first header name and break the required-column check with a
    baffling message.
    """
    if isinstance(content, str):
        return content
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise StatementFormatError("The uploaded statement could not be decoded as text.")


def _parse_datetime(value: str) -> datetime | None:
    text = (value or "").strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        pass
    for fmt in _DATETIME_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None
