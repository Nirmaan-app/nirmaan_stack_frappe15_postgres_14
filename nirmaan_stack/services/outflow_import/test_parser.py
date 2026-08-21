# Copyright (c) 2026, Nirmaan (Stratos Infra Technologies Pvt. Ltd.) and contributors
# See license.txt

"""Tests for nirmaan_stack.services.outflow_import.parser.

FIXTURES ARE SYNTHETIC ON PURPOSE. This repository is public, so no real statement is committed --
real beneficiary names, bank account numbers and IFSC codes would be a data leak. `cashfree_sample.csv`
is fabricated to reproduce the STRUCTURAL properties of two real exports (43 rows and 19 rows), and
each row that exists to reproduce one is labelled in the fixture generator and named in the test
that depends on it.

The two contracts these tests defend hardest, because both are easy to "tidy away":
  * the parser never filters -- a FAILED transfer is staged, not dropped
  * charges sum across EVERY row while gross sums only successful ones
"""

import unittest
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from nirmaan_stack.services.outflow_import.parser import (
    SUPPORTED_SOURCES,
    ParseResult,
    StatementFormatError,
    parse_statement,
)

FIXTURES = Path(__file__).parent / "tests" / "fixtures"


def _load(name: str) -> bytes:
    return (FIXTURES / name).read_bytes()


def _sample() -> ParseResult:
    return parse_statement(_load("cashfree_sample.csv"), source="Cashfree")


def _by_transfer(result: ParseResult, suffix: str):
    for row in result.rows:
        if row.transfer_id.endswith(suffix):
            return row
    raise AssertionError(f"no staged row ending {suffix!r}")


def _cashbook() -> ParseResult:
    return parse_statement(_load("cashbook_sample.csv"), source="Cashbook")


class TestParseShape(unittest.TestCase):
    def test_cashfree_is_a_supported_source(self):
        self.assertIn("Cashfree", SUPPORTED_SOURCES)

    def test_rows_are_staged_and_the_blank_transfer_id_row_is_not(self):
        result = _sample()
        # 12 data rows in the fixture; the one with a blank Transfer Id cannot be identified and is
        # dropped with a warning rather than staged under an empty key.
        self.assertEqual(len(result.rows), 11)
        self.assertTrue(all(row.transfer_id for row in result.rows))

    def test_row_numbers_reflect_file_position(self):
        result = _sample()
        self.assertEqual(result.rows[0].row_number, 1)
        self.assertEqual([r.row_number for r in result.rows], sorted(r.row_number for r in result.rows))

    def test_period_is_derived_from_the_rows(self):
        result = _sample()
        self.assertEqual(result.period_from, date(2026, 7, 28))
        self.assertEqual(result.period_to, date(2026, 7, 28))

    def test_datetime_is_parsed_with_full_precision(self):
        row = _by_transfer(_sample(), "0001")
        self.assertEqual(row.added_on, datetime(2026, 7, 28, 17, 5, 18))
        self.assertEqual(row.added_on_date, date(2026, 7, 28))


class TestNeverFilters(unittest.TestCase):
    """The parser stages every identifiable row, whatever its outcome."""

    def test_failed_transfer_is_staged_not_dropped(self):
        row = _by_transfer(_sample(), "0002")
        self.assertEqual(row.status_raw, "FAILED")
        self.assertFalse(row.is_success)

    def test_failed_transfer_still_carries_a_bank_reference(self):
        # This is why filtering matters downstream: a failed transfer has a reference that would
        # match a payment just as well as a real one.
        row = _by_transfer(_sample(), "0002")
        self.assertEqual(row.bank_reference_no, "900000000002")
        self.assertEqual(row.normalized_reference, "900000000002")

    def test_failed_transfer_and_its_retry_are_distinct_rows_with_distinct_references(self):
        # Mirrors the live pair: the same Rs 22,000 to the same beneficiary, failed at 18:17:59 and
        # succeeded at 19:23. Both would otherwise match the same payment.
        failed = _by_transfer(_sample(), "0002")
        retry = _by_transfer(_sample(), "0003")
        self.assertEqual(failed.amount, retry.amount)
        self.assertEqual(failed.beneficiary_name, retry.beneficiary_name)
        self.assertNotEqual(failed.transfer_id, retry.transfer_id)
        self.assertNotEqual(failed.bank_reference_no, retry.bank_reference_no)
        self.assertTrue(retry.is_success)

    def test_success_count_excludes_the_failure(self):
        result = _sample()
        self.assertEqual(result.success_count, len(result.rows) - 1)


class TestAmounts(unittest.TestCase):
    def test_gross_sums_successful_rows_only(self):
        result = _sample()
        expected = sum((r.amount for r in result.rows if r.is_success), Decimal("0"))
        self.assertEqual(result.gross_amount, expected)
        # the failed Rs 22,000 must not be in there
        self.assertNotIn(Decimal("22000.0"), [result.gross_amount])

    def test_charges_sum_every_row_including_the_failure(self):
        # Deliberate asymmetry: a charge is money the bank took whatever the outcome, so excluding
        # failed rows would understate the debit.
        result = _sample()
        expected = sum((r.service_charge + r.service_tax for r in result.rows), Decimal("0"))
        self.assertEqual(result.charges_amount, expected)

    def test_batch_total_never_equals_the_sum_of_rows(self):
        # The gateway charge belongs to no settlement target; this is why a reconciled batch still
        # does not tie to the bank statement without booking the charges separately.
        result = _sample()
        self.assertGreater(result.charges_amount, Decimal("0"))

    def test_thousands_separator_in_an_amount(self):
        row = _by_transfer(_sample(), "0009")
        self.assertEqual(row.amount, Decimal("1234.50"))

    def test_amounts_are_decimal_not_float(self):
        for row in _sample().rows:
            self.assertIsInstance(row.amount, Decimal)
            self.assertIsInstance(row.service_charge, Decimal)


class TestDerivedIdentityForms(unittest.TestCase):
    def test_leading_zero_account_is_normalized_beside_the_verbatim_value(self):
        row = _by_transfer(_sample(), "0006")
        self.assertEqual(row.bank_account, "0042345678904")       # verbatim, untouched
        self.assertEqual(row.normalized_account, "42345678904")   # comparable

    def test_padded_bank_reference_is_normalized_beside_the_verbatim_value(self):
        row = _by_transfer(_sample(), "0007")
        self.assertEqual(row.bank_reference_no, "900000000007")   # stripped at field level
        self.assertEqual(row.normalized_reference, "900000000007")

    def test_raw_values_are_preserved_so_a_rematch_cannot_destroy_evidence(self):
        row = _by_transfer(_sample(), "0006")
        self.assertNotEqual(row.bank_account, row.normalized_account)


class TestRemarks(unittest.TestCase):
    def test_long_remark_survives_untruncated(self):
        # Stored in a Text column precisely so this survives; as Data it is varchar(140) and Frappe
        # THROWS rather than truncating.
        row = _by_transfer(_sample(), "0008")
        self.assertGreater(len(row.remarks), 140)

    def test_remark_is_kept_verbatim(self):
        row = _by_transfer(_sample(), "0001")
        self.assertEqual(row.remarks, "Sample Project materials")


class TestWarnings(unittest.TestCase):
    def test_blank_transfer_id_is_reported(self):
        result = _sample()
        self.assertTrue(any("no Transfer Id" in w for w in result.warnings))

    def test_unreadable_date_is_staged_with_a_warning_not_dropped(self):
        result = _sample()
        row = _by_transfer(result, "0010")
        self.assertIsNone(row.added_on)
        self.assertTrue(any("Added On" in w for w in result.warnings))

    def test_in_file_duplicate_transfer_id_is_reported(self):
        result = _sample()
        self.assertIn("TID0000000000000000000000000001", result.duplicate_transfer_ids)
        self.assertTrue(any("more than once" in w for w in result.warnings))

    def test_duplicate_rows_are_still_staged(self):
        # Reporting the duplicate is the parser's job; deciding what to do about it is not.
        result = _sample()
        matching = [r for r in result.rows if r.transfer_id == "TID0000000000000000000000000001"]
        self.assertEqual(len(matching), 2)


class TestWholeFileFailures(unittest.TestCase):
    def test_unknown_source_is_refused(self):
        with self.assertRaises(StatementFormatError):
            parse_statement(_load("cashfree_sample.csv"), source="Nonesuch")

    def test_wrong_header_is_refused_and_names_the_missing_columns(self):
        with self.assertRaises(StatementFormatError) as ctx:
            parse_statement(_load("cashfree_bad_header.csv"), source="Cashfree")
        self.assertIn("Transfer Id", str(ctx.exception))

    def test_empty_file_is_refused(self):
        with self.assertRaises(StatementFormatError):
            parse_statement(_load("cashfree_empty.csv"), source="Cashfree")

    def test_header_only_file_is_refused(self):
        header = _load("cashfree_sample.csv").split(b"\n")[0]
        with self.assertRaises(StatementFormatError):
            parse_statement(header, source="Cashfree")

    def test_utf8_bom_is_tolerated(self):
        # A spreadsheet-exported CSV routinely carries a BOM, which would otherwise become part of
        # the first header name and fail the required-column check with a baffling message.
        with_bom = b"\xef\xbb\xbf" + _load("cashfree_sample.csv")
        self.assertEqual(len(parse_statement(with_bom).rows), 11)

    def test_unknown_extra_columns_are_ignored(self):
        # Real statements carry VPA, Acknowledged, Mode, Status Code and more; none are required.
        result = _sample()
        self.assertTrue(result.rows)


class TestXlsx(unittest.TestCase):
    """.xlsx alongside .csv (owner ruling Q10, slice V3).

    `cashfree_sample.xlsx` is the same statement as `cashfree_sample.csv`, with dates as DATETIME
    cells and amounts as FLOATS but identity fields left as TEXT. That mix is the point of the
    fixture -- see the leading-zero test below.

    ⚠️ IT IS NOT, HOWEVER, "SAVED THE WAY A REAL EXPORT SAVES IT" -- THAT CLAIM WAS HERE AND IT WAS
    FALSE. It was generated by openpyxl, which writes an HONEST `<dimension>`. A real Cashfree
    export writes `<dimension ref="A1"/>`, and every real .xlsx upload therefore failed from the
    day this format shipped until 2026-08-21 while this suite stayed green. That is the shape of
    defect a fixture built by the same library it is testing will always miss.
    `cashfree_bad_dimension.xlsx` is the twin with ONLY that lie introduced -- see
    `TestAWorkbookThatLiesAboutItsSize`.
    """

    def test_it_parses_to_exactly_the_same_rows_as_its_csv_twin(self):
        """THE GATE FOR THIS SLICE. Format is an encoding, not a dialect: the same statement must
        produce the same rows whichever way it was saved, or every downstream rule silently has two
        behaviours."""
        from_csv = parse_statement(_load("cashfree_sample.csv"), source="Cashfree")
        from_xlsx = parse_statement(_load("cashfree_sample.xlsx"), source="Cashfree")
        self.assertEqual(from_csv.rows, from_xlsx.rows)

    def test_the_batch_level_figures_agree_too(self):
        from_csv = parse_statement(_load("cashfree_sample.csv"), source="Cashfree")
        from_xlsx = parse_statement(_load("cashfree_sample.xlsx"), source="Cashfree")
        # Compared as Decimals, not as text: 57727.50 and 57727.5 are the same money and different
        # strings, and a repr comparison here would fail for no reason that matters.
        self.assertEqual(from_csv.gross_amount, from_xlsx.gross_amount)
        self.assertEqual(from_csv.charges_amount, from_xlsx.charges_amount)
        self.assertEqual(from_csv.period_from, from_xlsx.period_from)
        self.assertEqual(from_csv.period_to, from_xlsx.period_to)
        self.assertEqual(from_csv.duplicate_transfer_ids, from_xlsx.duplicate_transfer_ids)

    def test_a_declared_dimension_of_one_cell_does_not_truncate_the_read(self):
        """THE REAL-WORLD SHAPE. `cashfree_bad_dimension.xlsx` is byte-identical to the twin except
        that its sheet declares `<dimension ref="A1"/>`, exactly as a Cashfree Transfers export
        does. Under `read_only=True` openpyxl believes that declaration and clips every row to
        column A -- so the statement lost 21 of its 22 columns and the required-column check
        reported five of six missing, while `Added On` (column A) was found. Bounded `iter_rows`
        is what overrides it."""
        honest = parse_statement(_load("cashfree_sample.xlsx"), source="Cashfree")
        lying = parse_statement(_load("cashfree_bad_dimension.xlsx"), source="Cashfree")
        self.assertEqual(honest.rows, lying.rows)

    def test_the_lying_workbook_agrees_with_the_csv_too(self):
        """Through the whole chain, not just against its xlsx twin -- the csv is the reference."""
        from_csv = parse_statement(_load("cashfree_sample.csv"), source="Cashfree")
        lying = parse_statement(_load("cashfree_bad_dimension.xlsx"), source="Cashfree")
        self.assertEqual(from_csv.rows, lying.rows)
        self.assertEqual(from_csv.gross_amount, lying.gross_amount)
        self.assertEqual(from_csv.period_from, lying.period_from)
        self.assertEqual(from_csv.period_to, lying.period_to)

    def test_the_column_scan_ceiling_refuses_rather_than_short_reads(self):
        """⚠️ `_MAX_SCAN_COLUMNS` IS A REAL CEILING, unlike the row one, so it must never be reached
        silently. A statement wider than the scan would otherwise be parsed with its tail chopped
        off -- which reads as a successful import of a statement nobody exported."""
        import io as _io

        from openpyxl import Workbook

        from nirmaan_stack.services.outflow_import.parser import _MAX_SCAN_COLUMNS

        workbook = Workbook()
        sheet = workbook.active
        sheet.append([f"Column {i}" for i in range(_MAX_SCAN_COLUMNS + 5)])
        sheet.append(["x"] * (_MAX_SCAN_COLUMNS + 5))
        buffer = _io.BytesIO()
        workbook.save(buffer)

        with self.assertRaises(StatementFormatError) as caught:
            parse_statement(buffer.getvalue(), source="Cashfree")
        self.assertIn(str(_MAX_SCAN_COLUMNS), str(caught.exception))

    def test_the_row_ceiling_is_excels_own_limit_so_it_can_never_truncate(self):
        """It is NOT a performance guard -- it exists only because `iter_rows` needs a bound, and
        it is set where no workbook Excel can open could ever reach it."""
        from nirmaan_stack.services.outflow_import.parser import _MAX_SCAN_ROWS

        self.assertEqual(_MAX_SCAN_ROWS, 1_048_576)

    def test_the_format_is_sniffed_from_the_bytes_not_the_name(self):
        """No filename reaches the parser at all, so a renamed export still works. The failure mode
        of trusting an extension is a 'missing column' error on a perfectly good file."""
        result = parse_statement(_load("cashfree_sample.xlsx"), source="Cashfree")
        self.assertEqual(len(result.rows), 11)

    def test_a_leading_zero_bank_account_survives(self):
        """⚠️ THE ONE THAT BREAKS SILENTLY. Account `0042345678904` written to a NUMBER cell comes
        back as 42345678904 -- a valid-looking account that belongs to nobody. The fixture keeps
        identity fields as text for exactly this reason, and this test is what would catch a future
        change that started coercing them."""
        result = parse_statement(_load("cashfree_sample.xlsx"), source="Cashfree")
        accounts = {row.bank_account for row in result.rows}
        self.assertIn("0042345678904", accounts)

    def test_a_typed_datetime_cell_reads_as_the_same_instant_as_its_text_twin(self):
        from_csv = parse_statement(_load("cashfree_sample.csv"), source="Cashfree")
        from_xlsx = parse_statement(_load("cashfree_sample.xlsx"), source="Cashfree")
        self.assertEqual(from_csv.rows[0].added_on, datetime(2026, 7, 28, 17, 5, 18))
        self.assertEqual(from_csv.rows[0].added_on, from_xlsx.rows[0].added_on)

    def test_an_unreadable_date_stays_unreadable_rather_than_becoming_today(self):
        """The fixture carries a deliberate 'not-a-date'. It must warn, not silently substitute --
        a fabricated date would place the transfer in the wrong period."""
        result = parse_statement(_load("cashfree_sample.xlsx"), source="Cashfree")
        undated = [row for row in result.rows if row.added_on is None]
        self.assertTrue(undated)

    def test_a_file_that_is_not_a_workbook_but_starts_like_one_is_refused_clearly(self):
        """A truncated or corrupt upload still carries the ZIP magic. The message has to say the
        file could not be opened, not report a missing column."""
        with self.assertRaises(StatementFormatError) as caught:
            parse_statement(b"PK\x03\x04 and then nothing useful", source="Cashfree")
        self.assertIn("could not be opened", str(caught.exception))

    def test_an_xlsx_missing_a_required_column_fails_the_same_way_a_csv_does(self):
        """The required-column check is downstream of the format seam, so it must be reached
        identically by both. If .xlsx skipped it, a wrong workbook would stage garbage rows."""
        from openpyxl import Workbook

        import io as _io

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Transfer Id", "Amount"])
        sheet.append(["T1", 100])
        buffer = _io.BytesIO()
        workbook.save(buffer)

        with self.assertRaises(StatementFormatError) as caught:
            parse_statement(buffer.getvalue(), source="Cashfree")
        self.assertIn("Missing column", str(caught.exception))


class TestCashbookSource(unittest.TestCase):
    """The petty-cash wallet statement (slice 1).

    `cashbook_sample.csv` is fabricated on the same terms as the Cashfree fixture, and reproduces
    every structural shape found in a real 137-row export: a spend whose Note says more than its
    Remark, a spend with no free text at all, two spends identical but for their ids, a refused
    spend, a spend with no debit figure, three movements that are not spends, one transfer missing
    its id, and the six-row totals block the sheet ends with.
    """

    def test_cashbook_is_a_supported_source(self):
        self.assertIn("Cashbook", SUPPORTED_SOURCES)

    def test_every_row_carrying_a_transfer_id_is_staged(self):
        # 17 rows in the fixture: 10 identifiable movements, 1 unidentifiable, 6 totals lines.
        result = _cashbook()
        self.assertEqual(len(result.rows), 10)
        self.assertTrue(all(row.transfer_id for row in result.rows))

    def test_the_kind_of_movement_is_recorded_verbatim(self):
        result = _cashbook()
        self.assertEqual(_by_transfer(result, "AAAAAA").row_kind, "Wallet Spend")
        self.assertEqual(_by_transfer(result, "900001-0").row_kind, "VA → Wallet")
        self.assertEqual(_by_transfer(result, "VALOAD-9000002").row_kind, "Bank → VA")
        self.assertEqual(_by_transfer(result, "PTM9000003").row_kind, "Wallet Credit")

    def test_a_top_up_is_staged_rather_than_filtered_out(self):
        """The parser never filters -- deciding a top-up is not importable is downstream's job.

        This is the same contract the FAILED Cashfree transfer defends, and it matters more here:
        a top-up is not an error, it is simply not a spend, and the staged row is what makes the
        eventual skip visible instead of an absence nobody can account for.
        """
        result = _cashbook()
        self.assertIsNotNone(_by_transfer(result, "900001-0"))
        self.assertIsNotNone(_by_transfer(result, "VALOAD-9000002"))

    def test_a_refused_spend_is_staged_and_keeps_its_status(self):
        row = _by_transfer(_cashbook(), "FFFFFF")
        self.assertEqual(row.status_raw, "FAILED")
        self.assertFalse(row.is_success)


class TestCashbookRemarkJoin(unittest.TestCase):
    """Remark and Note arrive as one string, because both feed the same matcher."""

    def test_a_note_saying_more_than_its_remark_is_not_lost(self):
        row = _by_transfer(_cashbook(), "BBBBBB")
        self.assertEqual(row.remarks, "Pay to merchant - Printout charges beta site")

    def test_a_blank_note_leaves_no_dangling_separator(self):
        row = _by_transfer(_cashbook(), "AAAAAA")
        self.assertEqual(row.remarks, "Transport charges alpha project")

    def test_a_row_with_neither_reads_as_empty_not_as_a_separator(self):
        row = _by_transfer(_cashbook(), "CCCCCC")
        self.assertEqual(row.remarks, "")


class TestCashbookTrailerRows(unittest.TestCase):
    """The totals block must cost NOTHING, and an unidentifiable transfer must still be reported.

    ⚠️ THESE TWO TESTS ARE A PAIR AND NEITHER IS SAFE ALONE. Suppressing the totals block by
    dropping the warning outright would pass the first and break the second, and losing that
    warning means a transfer we cannot identify disappears in silence.
    """

    def test_the_totals_block_produces_no_warnings(self):
        result = _cashbook()
        self.assertEqual(
            [w for w in result.warnings if "no Transfer Id" in w],
            ["Row 11 has no Transfer Id and was not staged."],
        )

    def test_a_transfer_with_an_amount_but_no_id_is_still_reported(self):
        result = _cashbook()
        self.assertTrue(any("Row 11" in w for w in result.warnings))

    def test_no_totals_line_is_staged_as_a_transfer(self):
        result = _cashbook()
        self.assertFalse(any("Balance" in (row.beneficiary_name or "") for row in result.rows))


class TestCashbookFigures(unittest.TestCase):
    def test_the_date_is_read_without_the_time_beside_it(self):
        """Owner ruling: `Date` is taken, `Time` is discarded, so a spend lands at midnight."""
        row = _by_transfer(_cashbook(), "AAAAAA")
        self.assertEqual(row.added_on, datetime(2026, 8, 1, 0, 0))
        self.assertEqual(row.added_on_date, date(2026, 8, 1))

    def test_period_spans_the_statement(self):
        result = _cashbook()
        self.assertEqual(result.period_from, date(2026, 8, 1))
        self.assertEqual(result.period_to, date(2026, 8, 4))

    def test_a_wallet_statement_carries_no_charges(self):
        self.assertEqual(_cashbook().charges_amount, Decimal("0"))

    def test_gross_sums_successful_debits_and_the_refused_spend_is_not_in_it(self):
        # 180 + 70 + 6000 + 250 + 250; the FAILED 400 never left, and a credit is not a debit.
        self.assertEqual(_cashbook().gross_amount, Decimal("6750"))

    def test_a_spend_with_no_debit_figure_reads_as_zero_rather_than_failing_the_file(self):
        row = _by_transfer(_cashbook(), "GGGGGG")
        self.assertEqual(row.amount, Decimal("0"))
        self.assertTrue(row.is_success)

    def test_two_spends_alike_but_for_their_ids_are_not_called_duplicates(self):
        """Identity is `(id, amount, date)`, and the ids differ -- so these are two real spends.

        Measured on a real export: one such pair exists in 115 rows, two porter payments minutes
        apart. Keying the check on anything coarser would silently merge them.
        """
        result = _cashbook()
        self.assertEqual(result.duplicate_transfer_ids, ())

    def test_who_spent_it_is_captured_separately_from_who_was_paid(self):
        row = _by_transfer(_cashbook(), "AAAAAA")
        self.assertEqual(row.added_by_raw, "Asha Menon")
        self.assertEqual(row.beneficiary_name, "Testvendor Alpha")


class TestCashbookWholeFileFailures(unittest.TestCase):
    def test_a_statement_without_a_remark_column_is_refused(self):
        """Remark is the ONLY signal this source carries for choosing a project or a type.

        Without it the file would parse perfectly and book every row to a fallback -- a silent
        loss, which is precisely what the required-column check exists to prevent.
        """
        with self.assertRaises(StatementFormatError) as caught:
            parse_statement(_load("cashbook_no_remark.csv"), source="Cashbook")
        self.assertIn("Remark", str(caught.exception))

    def test_a_cashfree_statement_is_refused_as_cashbook(self):
        with self.assertRaises(StatementFormatError) as caught:
            parse_statement(_load("cashfree_sample.csv"), source="Cashbook")
        self.assertIn("Missing column", str(caught.exception))


class TestCashbookXlsx(unittest.TestCase):
    def test_it_parses_to_exactly_the_same_rows_as_its_csv_twin(self):
        from_csv = _cashbook()
        from_xlsx = parse_statement(_load("cashbook_sample.xlsx"), source="Cashbook")
        self.assertEqual(
            [(r.transfer_id, r.amount, r.remarks, r.row_kind) for r in from_csv.rows],
            [(r.transfer_id, r.amount, r.remarks, r.row_kind) for r in from_xlsx.rows],
        )

    def test_the_totals_block_is_dropped_from_the_workbook_too(self):
        result = parse_statement(_load("cashbook_sample.xlsx"), source="Cashbook")
        self.assertEqual(len(result.rows), 10)
        self.assertEqual(len([w for w in result.warnings if "no Transfer Id" in w]), 1)


class TestCashfreeIsUnaffected(unittest.TestCase):
    """Slice 1 touched shared code. Cashfree must not have moved.

    ⚠️ THE TRAILER RULE IS THE RISK HERE, and it is narrow by design: it suppresses a warning only
    where a row has NO id, NO amount and NO status. The Cashfree fixture's unidentifiable row
    carries an amount of 400 and a status of SUCCESS, so it stays reported -- which is the whole
    reason the rule tests content rather than just emptiness.
    """

    def test_the_blank_transfer_id_row_is_still_warned_about(self):
        result = _sample()
        self.assertTrue(any("no Transfer Id" in w for w in result.warnings))

    def test_row_kind_is_blank_on_a_source_that_does_not_declare_one(self):
        self.assertTrue(all(row.row_kind == "" for row in _sample().rows))

    def test_a_single_column_remark_is_unchanged_by_the_multi_column_map(self):
        """A map naming ONE column must still read exactly that cell, unjoined and unstripped.

        `raw()` gained a tuple branch; this pins that the string branch beneath it did not change
        behaviour. It is the whole-fixture form of the assertion rather than a spot check, because
        a join defect would be uniform and a single row could miss it.
        """
        self.assertEqual(
            [row.remarks for row in _sample().rows],
            [
                "Sample Project materials",
                "Sample Project Transportation Services",
                "Sample Project Transportation Services",
                "Sample Project HVAC Services",
                "Sample Project HVAC Services",
                "Sample Project materials",
                "Sample Project materials",
                "Sample Project materials supplied against multiple indents raised during the "
                "month including consumables, fasteners and sundry site items delivered in "
                "several tranches across the north and south blocks",
                "Sample Project Room Rent",
                "Sample Project materials",
                "Sample Project materials",
            ],
        )


if __name__ == "__main__":
    unittest.main()
