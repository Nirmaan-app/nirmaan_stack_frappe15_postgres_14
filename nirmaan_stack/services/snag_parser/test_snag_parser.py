"""Unit tests for the PURE Snag List parser.

Plain `unittest.TestCase`, NOT `FrappeTestCase`: this package has no frappe
dependency, so the suite runs with no bench context, no site and no database.

    # in-container / anywhere frappe's deps import cleanly:
    python3 -m unittest nirmaan_stack.services.snag_parser.test_snag_parser -v

    # on a host where `nirmaan_stack/__init__.py` cannot import firebase_admin,
    # run it from the services directory instead (same code, shorter path):
    cd nirmaan_stack/services && python3 -m unittest snag_parser.test_snag_parser -v

The fixture is a copy of the real consultant file `Food Box MEP Snags list.xlsx`
(plan §2). Every number below is MEASURED from it, not assumed.
"""

from __future__ import annotations

import os
import tempfile
import unittest

from openpyxl import Workbook

from openpyxl import load_workbook

from .guess import guess_mapping
from .parser import parse_grid, parse_sheet
from .reader import columns_for_header_row, inspect_sheet, inspect_workbook, read_grid

FIXTURE = os.path.join(os.path.dirname(__file__), "fixtures", "food_box_mep_snags.xlsx")

#: The mapping the wizard pre-selects for the fixture's Sheet1.
FOOD_BOX_MAPPING = {"area": "B", "category": "C", "description": "D", "remarks": "G"}


def _reasons(preview):
    """{source_row: skipped_reason} for every SKIPPED row of the merged list."""
    return {r["source_row"]: r["skipped_reason"] for r in preview["rows"] if r["skipped_reason"]}


def _accepted(preview):
    """The accepted rows of the merged list, in Excel row order."""
    return [r for r in preview["rows"] if r["skipped_reason"] is None]


def _skipped(preview):
    return [r for r in preview["rows"] if r["skipped_reason"] is not None]


def _write_workbook(rows, sheet_title="Sheet1"):
    """Write `rows` (list of lists) to a temp .xlsx and return its path."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    for row in rows:
        ws.append(row)
    handle, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(handle)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# reader.inspect_workbook against the real file
# ---------------------------------------------------------------------------


class TestInspectFoodBoxWorkbook(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.sheets = inspect_workbook(FIXTURE)

    def test_three_sheets_only_the_first_has_content(self):
        self.assertEqual([s["name"] for s in self.sheets], ["Sheet1", "Sheet2", "Sheet3"])
        self.assertFalse(self.sheets[0]["is_empty"])
        self.assertTrue(self.sheets[1]["is_empty"])
        self.assertTrue(self.sheets[2]["is_empty"])

    def test_empty_sheets_carry_no_header_columns_or_guess(self):
        for sheet in self.sheets[1:]:
            self.assertEqual(sheet["row_count"], 0)
            self.assertIsNone(sheet["header_row"])
            self.assertEqual(sheet["columns"], [])
            self.assertIsNone(sheet["mapping_guess"])

    def test_header_row_is_seven(self):
        """Rows 1-6 are the title / "Total Snags:124" block, NOT the header."""
        self.assertEqual(self.sheets[0]["header_row"], 7)

    def test_row_count_is_the_real_last_used_row(self):
        self.assertEqual(self.sheets[0]["row_count"], 149)

    def test_columns_are_the_header_row_cells_a_to_g(self):
        self.assertEqual(
            self.sheets[0]["columns"],
            [
                {"letter": "A", "label": "S.No"},
                {"letter": "B", "label": "Area / Location"},
                {"letter": "C", "label": "Category"},
                {"letter": "D", "label": "Snag Description"},
                {"letter": "E", "label": "Risk Level"},
                {"letter": "F", "label": "Status"},
                {"letter": "G", "label": "Remarks"},
            ],
        )

    def test_mapping_guess(self):
        self.assertEqual(self.sheets[0]["mapping_guess"], FOOD_BOX_MAPPING)


# ---------------------------------------------------------------------------
# parser.parse_sheet against the real file
# ---------------------------------------------------------------------------


class TestParseFoodBoxSheet1(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.preview = parse_sheet(FIXTURE, "Sheet1", FOOD_BOX_MAPPING)
        cls.rows = _accepted(cls.preview)
        cls.reasons = _reasons(cls.preview)

    def test_exactly_124_rows(self):
        """THE assertion: the file's own title row says "Total Snags:124".

        A naive "map three columns and ingest" reads 142.
        """
        self.assertEqual(len(self.rows), 124)
        self.assertEqual(self.preview["accepted_count"], 124)
        self.assertEqual(self.preview["skipped_count"], 25)

    def test_the_124_split_into_the_two_blocks(self):
        block1 = [r["source_row"] for r in self.rows if r["source_row"] < 50]
        block2 = [r["source_row"] for r in self.rows if r["source_row"] >= 50]
        self.assertEqual(len(block1), 37)
        self.assertEqual((min(block1), max(block1)), (8, 44))
        self.assertEqual(len(block2), 87)
        self.assertEqual((min(block2), max(block2)), (55, 141))

    def test_every_row_of_the_sheet_is_accounted_for(self):
        """ONE merged list: accepted + skipped == every row of the sheet, in
        ascending Excel order, no gaps and no duplicates."""
        self.assertEqual(len(self.preview["rows"]), 149)
        self.assertEqual(
            self.preview["accepted_count"] + self.preview["skipped_count"], 149
        )
        numbers = [r["source_row"] for r in self.preview["rows"]]
        self.assertEqual(numbers, list(range(1, 150)))

    def test_accepted_and_skipped_are_interleaved_not_two_blocks(self):
        """Row 45 (blank) sits BETWEEN accepted rows 44 and 55 in the list."""
        flags = [r["skipped_reason"] is None for r in self.preview["rows"]]
        self.assertEqual(flags[43], True)   # row 44 accepted
        self.assertEqual(flags[44], False)  # row 45 blank
        self.assertEqual(flags[54], True)   # row 55 accepted

    def test_row_54_is_the_repeated_header(self):
        self.assertEqual(self.reasons[54], "repeated_header")

    def test_rows_46_to_49_are_the_risk_summary_block(self):
        for row in (46, 47, 48, 49):
            self.assertEqual(self.reasons[row], "summary_block", "row %d" % row)

    def test_rows_143_to_146_are_the_fls_risk_summary_block(self):
        for row in (143, 144, 145, 146):
            self.assertEqual(self.reasons[row], "summary_block", "row %d" % row)

    def test_blank_rows(self):
        for row in (45, 50, 53):
            self.assertEqual(self.reasons[row], "blank", "row %d" % row)
            self.assertEqual(
                [r["preview_text"] for r in self.preview["rows"] if r["source_row"] == row],
                [""],
            )

    def test_rows_above_the_header_are_excluded_by_POSITION(self):
        """Rows 1-6 are the title / "Total Snags:124" block.

        They used to fall out INCIDENTALLY (their Description cell happens to be
        empty, so `no_description` / `blank`). The header row is now a real rule:
        every row above it is out of the data region whatever its cells say.
        """
        for row in (1, 2, 3, 4, 5, 6):
            self.assertEqual(self.reasons[row], "above_header", "row %d" % row)

    def test_the_header_row_itself_keeps_repeated_header(self):
        self.assertEqual(self.reasons[7], "repeated_header")

    def test_the_resolved_header_row_rides_the_response(self):
        self.assertEqual(self.preview["header_row"], 7)

    def test_preview_text_is_the_first_non_empty_cell_truncated(self):
        by_row = {r["source_row"]: r["preview_text"] for r in _skipped(self.preview)}
        self.assertEqual(by_row[1], "VR BENGALURU (VRB MALL) - FOOD BOX AREA SNAG LIST")
        self.assertEqual(by_row[46], "RISK SUMMARY")
        self.assertEqual(by_row[54], "S.No")
        for text in by_row.values():
            self.assertLessEqual(len(text), 80)

    def test_distinct_categories(self):
        """MEASURED: 10 distinct categories (the plan doc's "11" was an estimate)."""
        categories = self.preview["distinct_categories"]
        self.assertEqual(len(categories), 10)
        self.assertEqual(categories[0], {"value": "Fire & Life Safety", "count": 79})
        self.assertEqual(sum(c["count"] for c in categories), 124)

    def test_distinct_areas(self):
        """MEASURED: 17 distinct areas (the plan doc's "19" was an estimate)."""
        areas = self.preview["distinct_areas"]
        self.assertEqual(len(areas), 17)
        self.assertEqual(areas[0], {"value": "Food Box - General", "count": 74})
        self.assertEqual(sum(a["count"] for a in areas), 124)

    def test_distinct_values_sorted_by_count_desc_then_value_asc(self):
        for bucket in (self.preview["distinct_areas"], self.preview["distinct_categories"]):
            keys = [(-b["count"], b["value"]) for b in bucket]
            self.assertEqual(keys, sorted(keys))
            self.assertNotIn("", [b["value"] for b in bucket])

    def test_values_are_verbatim_never_normalised(self):
        by_row = {r["source_row"]: r for r in self.rows}
        self.assertEqual(by_row[10]["area"], "Opp. McD Shop")
        self.assertEqual(by_row[10]["category"], "HVAC - Chilled Water")
        self.assertEqual(
            by_row[8]["description"],
            "Detector cables are hanging and not properly laid. Proper conduit routing is required.",
        )
        # No value carries edge whitespace, and internal spacing is untouched.
        for row in self.rows:
            for field in ("area", "category", "description", "remark"):
                self.assertEqual(row[field], row[field].strip())

    def test_row_shape_matches_the_wire_contract(self):
        """EVERY row — accepted or skipped — carries the SAME fields.

        `ParsedSnagRow` minus `is_duplicate`, which the API layer folds in.
        """
        expected = {
            "source_row",
            "area",
            "category",
            "description",
            "remark",
            "skipped_reason",
            "tickable",
            "preview_text",
        }
        for row in self.preview["rows"]:
            self.assertEqual(set(row), expected, "row %d" % row["source_row"])
            self.assertIn(
                row["skipped_reason"],
                {None, "blank", "repeated_header", "summary_block", "no_description", "above_header"},
            )
        self.assertEqual(self.preview["sheet_name"], "Sheet1")

    def test_the_remark_key_is_singular_and_source_remarks_is_gone(self):
        """ADR-0018: the row VALUE is `remark`; `remarks` is a MAPPING key
        holding an Excel column LETTER."""
        for row in self.preview["rows"]:
            self.assertIn("remark", row)
            self.assertNotIn("source_remarks", row)
            self.assertNotIn("remarks", row)

    def test_preview_text_is_empty_on_an_accepted_row(self):
        for row in self.rows:
            self.assertEqual(row["preview_text"], "")

    def test_tickable_is_true_for_every_accepted_row(self):
        for row in self.rows:
            self.assertTrue(row["tickable"], "row %d" % row["source_row"])

    def test_a_row_with_no_description_is_never_tickable(self):
        """It could not become a Snag — description is the one required field."""
        no_desc = [r for r in self.preview["rows"] if not r["description"].strip()]
        self.assertEqual(len(no_desc), 23)
        for row in no_desc:
            self.assertFalse(row["tickable"], "row %d" % row["source_row"])
        # The two repeated-header rows DO carry a description cell ("Snag
        # Description"), so they stay tickable — a parser mistake stays
        # recoverable by hand.
        tickable_skips = [r["source_row"] for r in _skipped(self.preview) if r["tickable"]]
        self.assertEqual(tickable_skips, [7, 54])

    def test_a_skipped_row_carries_the_same_values_as_an_accepted_one(self):
        """The R2.1 fix: a re-ticked skipped row must arrive with its data."""
        row54 = next(r for r in self.preview["rows"] if r["source_row"] == 54)
        self.assertEqual(row54["skipped_reason"], "repeated_header")
        self.assertEqual(row54["area"], "Area / Location")
        self.assertEqual(row54["category"], "Category")
        self.assertEqual(row54["description"], "Snag Description")
        self.assertEqual(row54["remark"], "Remarks")


# ---------------------------------------------------------------------------
# guess.guess_mapping — synthetic
# ---------------------------------------------------------------------------


class TestGuessMapping(unittest.TestCase):
    @staticmethod
    def _columns(*labels):
        letters = "ABCDEFGHIJ"
        return [{"letter": letters[i], "label": label} for i, label in enumerate(labels)]

    def test_returns_none_when_there_is_no_description_column(self):
        self.assertIsNone(
            guess_mapping(self._columns("S.No", "Area / Location", "Category", "Remarks"))
        )

    def test_returns_none_for_no_columns(self):
        self.assertIsNone(guess_mapping([]))

    def test_matches_case_insensitively(self):
        self.assertEqual(
            guess_mapping(self._columns("AREA", "category", "DeScRiPtIoN", "remarks")),
            {"area": "A", "category": "B", "description": "C", "remarks": "D"},
        )

    def test_synonyms(self):
        self.assertEqual(
            guess_mapping(self._columns("Zone", "Discipline", "Observation", "Action")),
            {"area": "A", "category": "B", "description": "C", "remarks": "D"},
        )

    def test_most_specific_wins(self):
        # "Snag Description" beats the bare "Snag" column, wherever it sits.
        mapping = guess_mapping(self._columns("Snag", "Snag Description", "Area"))
        self.assertEqual(mapping["description"], "B")
        self.assertEqual(mapping["area"], "C")

    def test_a_letter_is_never_claimed_twice(self):
        # One column could serve description, category and remarks; description
        # claims it first and the others come back None.
        mapping = guess_mapping(self._columns("Defect / Comment / System"))
        self.assertEqual(mapping["description"], "A")
        self.assertIsNone(mapping["category"])
        self.assertIsNone(mapping["remarks"])
        letters = [v for v in mapping.values() if v]
        self.assertEqual(len(letters), len(set(letters)))

    def test_unmapped_roles_are_none(self):
        self.assertEqual(
            guess_mapping(self._columns("S.No", "Snag Description", "Risk Level")),
            {"area": None, "category": None, "description": "B", "remarks": None},
        )


# ---------------------------------------------------------------------------
# Synthetic workbooks
# ---------------------------------------------------------------------------


class TestSyntheticWorkbooks(unittest.TestCase):
    def setUp(self):
        self._paths = []

    def tearDown(self):
        for path in self._paths:
            try:
                os.remove(path)
            except OSError:
                pass

    def _workbook(self, rows, sheet_title="Sheet1"):
        path = _write_workbook(rows, sheet_title)
        self._paths.append(path)
        return path

    def test_a_sheet_with_no_header_row(self):
        path = self._workbook([["just"], ["some"], ["prose"]])
        sheet = inspect_workbook(path)[0]
        self.assertFalse(sheet["is_empty"])
        self.assertIsNone(sheet["header_row"])
        self.assertEqual(sheet["columns"], [])
        self.assertIsNone(sheet["mapping_guess"])

    def test_a_totally_empty_sheet(self):
        path = self._workbook([])
        sheet = inspect_workbook(path)[0]
        self.assertTrue(sheet["is_empty"])
        self.assertEqual(sheet["row_count"], 0)

    def test_description_only_mapping_yields_empty_strings(self):
        path = self._workbook(
            [
                ["Snag Description"],
                ["Loose cable tray"],
                [None],
                ["Missing sprinkler cap"],
            ]
        )
        preview = parse_sheet(path, "Sheet1", {"area": None, "category": None, "description": "A", "remarks": None})
        self.assertEqual(preview["accepted_count"], 2)
        for row in _accepted(preview):
            self.assertEqual(row["area"], "")
            self.assertEqual(row["category"], "")
            self.assertEqual(row["remark"], "")
        self.assertEqual(preview["distinct_areas"], [])
        self.assertEqual(preview["distinct_categories"], [])
        self.assertEqual(_reasons(preview)[3], "blank")
        # With only ONE column mapped the >= 2 rule is unreachable, so the
        # threshold falls to that one cell — the header row is still a header.
        self.assertEqual(_reasons(preview)[1], "repeated_header")

    def test_two_mapped_columns_still_need_two_header_cells(self):
        """The >= 2 rule is UNCHANGED as soon as two columns are mapped."""
        path = self._workbook(
            [
                ["Area / Location", "Snag Description"],
                ["Category", "Loose cable tray"],  # 1 header label only -> data
            ]
        )
        preview = parse_sheet(
            path, "Sheet1", {"area": "A", "category": None, "description": "B", "remarks": None}
        )
        self.assertEqual(_reasons(preview)[1], "repeated_header")
        self.assertEqual([r["source_row"] for r in _accepted(preview)], [2])

    def test_a_mapped_column_beyond_the_sheet_reads_as_empty(self):
        path = self._workbook([["Snag Description"], ["Chipped paint"]])
        preview = parse_sheet(path, "Sheet1", {"area": "Z", "category": None, "description": "A", "remarks": None})
        self.assertEqual(_accepted(preview)[0]["area"], "")

    def test_missing_sheet_raises(self):
        path = self._workbook([["Snag Description"], ["x"]])
        with self.assertRaises(ValueError):
            parse_sheet(path, "Nope", FOOD_BOX_MAPPING)


# ---------------------------------------------------------------------------
# parse_grid — the pure classifier, no file at all
# ---------------------------------------------------------------------------


class TestParseGrid(unittest.TestCase):
    MAPPING = {"area": "A", "category": "B", "description": "C", "remarks": "D"}

    def test_description_mapping_is_required(self):
        for mapping in ({}, {"description": None}, {"description": ""}, None):
            with self.assertRaises(ValueError):
                parse_grid("S", [["a", "b", "c"]], mapping)

    def test_a_bad_column_letter_raises(self):
        with self.assertRaises(ValueError):
            parse_grid("S", [["a"]], {"description": "not-a-letter"})

    def test_summary_block_runs_to_the_next_blank_row(self):
        grid = [
            ["Area / Location", "Category", "Snag Description", "Remarks"],
            ["Lobby", "Electrical", "Loose socket", ""],
            ["RISK SUMMARY", "", "", ""],
            ["High", "3", "", ""],
            ["Low", "1", "", ""],
            ["", "", "", ""],
            ["Lobby", "Electrical", "Cracked conduit", ""],
        ]
        preview = parse_grid("S", grid, self.MAPPING)
        reasons = _reasons(preview)
        self.assertEqual([r["source_row"] for r in _accepted(preview)], [2, 7])
        self.assertEqual(reasons[1], "repeated_header")
        for row in (3, 4, 5):
            self.assertEqual(reasons[row], "summary_block")
        self.assertEqual(reasons[6], "blank")

    def test_summary_block_also_stops_at_the_next_header_row(self):
        # Row 1 is the header, so the summary block below it is inside the data
        # region and terminates on the REPEATED header at row 4. (Before the
        # header row became a positional rule this grid opened with the summary
        # block itself; rows above the header are now excluded by POSITION, which
        # would have masked what this test is about.)
        grid = [
            ["Area / Location", "Category", "Snag Description", "Remarks"],
            ["Total Snags", "", "", ""],
            ["High", "3", "", ""],
            ["Area / Location", "Category", "Snag Description", "Remarks"],
            ["Lobby", "Electrical", "Loose socket", ""],
        ]
        preview = parse_grid("S", grid, self.MAPPING)
        reasons = _reasons(preview)
        self.assertEqual(reasons[1], "repeated_header")
        self.assertEqual(reasons[2], "summary_block")
        self.assertEqual(reasons[3], "summary_block")
        self.assertEqual(reasons[4], "repeated_header")
        self.assertEqual([r["source_row"] for r in _accepted(preview)], [5])

    def test_a_data_row_whose_serial_is_a_number_never_opens_a_block(self):
        grid = [
            ["1", "Lobby", "Total station room cabling is loose", ""],
        ]
        preview = parse_grid("S", grid, {"area": "B", "category": None, "description": "C", "remarks": None})
        self.assertEqual(preview["accepted_count"], 1)

# ---------------------------------------------------------------------------
# header_row — the explicit override, and the positional data region
# ---------------------------------------------------------------------------


class TestHeaderRowOverride(unittest.TestCase):
    """The header row is a RULE now, not a coincidence.

    Before this, `parser.py` never called `find_header_row`: rows above the
    header fell out only because their Description cell happened to be empty.
    """

    @classmethod
    def setUpClass(cls):
        cls.auto = parse_sheet(FIXTURE, "Sheet1", FOOD_BOX_MAPPING)
        cls.explicit = parse_sheet(FIXTURE, "Sheet1", FOOD_BOX_MAPPING, header_row=7)
        cls.wrong = parse_sheet(FIXTURE, "Sheet1", FOOD_BOX_MAPPING, header_row=54)

    def test_an_override_that_AGREES_with_the_guess_is_a_no_op(self):
        """THE assertion for this parameter: `header_row=7` must return exactly
        what auto-detection returns — same 124 rows, byte for byte."""
        self.assertEqual(self.explicit, self.auto)
        self.assertEqual(self.explicit["accepted_count"], 124)
        self.assertEqual(
            [r["source_row"] for r in _accepted(self.explicit)],
            [r["source_row"] for r in _accepted(self.auto)],
        )

    def test_rows_above_an_explicit_header_row_are_above_header(self):
        reasons = _reasons(self.explicit)
        for row in (1, 2, 3, 4, 5, 6):
            self.assertEqual(reasons[row], "above_header", "row %d" % row)
        self.assertEqual(reasons[7], "repeated_header")
        self.assertIsNone(
            next(r for r in self.explicit["rows"] if r["source_row"] == 8)["skipped_reason"]
        )

    def test_a_WRONG_override_really_reaches_classification(self):
        """`header_row=54` is the fixture's REPEATED header. Honoured literally:
        the whole first data block (8-44) falls out of the data region."""
        self.assertEqual(self.wrong["header_row"], 54)
        self.assertEqual(self.wrong["accepted_count"], 87)
        self.assertEqual(self.wrong["skipped_count"], 62)
        accepted = [r["source_row"] for r in _accepted(self.wrong)]
        self.assertEqual((min(accepted), max(accepted)), (55, 141))
        self.assertLess(self.wrong["accepted_count"], self.auto["accepted_count"])

        reasons = _reasons(self.wrong)
        for row in range(8, 45):
            self.assertEqual(reasons[row], "above_header", "row %d" % row)
        self.assertEqual(reasons[54], "repeated_header")
        # Every one of rows 1-53 is out by position, whatever its content was.
        self.assertEqual(
            sum(1 for v in reasons.values() if v == "above_header"), 53
        )

    def test_a_wrong_override_still_accounts_for_every_row(self):
        numbers = [r["source_row"] for r in self.wrong["rows"]]
        self.assertEqual(numbers, list(range(1, 150)))

    def test_distinct_values_follow_the_narrowed_data_region(self):
        """They count ACCEPTED rows only — so a wrong override changes them."""
        self.assertEqual(sum(a["count"] for a in self.wrong["distinct_areas"]), 87)
        self.assertEqual(sum(c["count"] for c in self.wrong["distinct_categories"]), 87)

    def test_no_header_row_anywhere_means_no_positional_filter(self):
        """A grid with nothing header-looking classifies on content alone,
        exactly as it did before this parameter existed."""
        grid = [
            ["Lobby", "Electrical", "Loose socket", ""],
            ["Lobby", "Electrical", "Cracked conduit", ""],
        ]
        preview = parse_grid("S", grid, {"area": "A", "category": "B", "description": "C", "remarks": "D"})
        self.assertIsNone(preview["header_row"])
        self.assertEqual([r["source_row"] for r in _accepted(preview)], [1, 2])

    def test_an_override_is_honoured_on_a_row_row_is_header_would_reject(self):
        grid = [
            ["Lobby", "Electrical", "Loose socket", ""],
            ["Lobby", "Electrical", "Cracked conduit", ""],
            ["Lobby", "Electrical", "Chipped paint", ""],
        ]
        preview = parse_grid(
            "S", grid, {"area": "A", "category": "B", "description": "C", "remarks": "D"}, header_row=2
        )
        self.assertEqual(preview["header_row"], 2)
        self.assertEqual(_reasons(preview), {1: "above_header", 2: "repeated_header"})
        self.assertEqual([r["source_row"] for r in _accepted(preview)], [3])


# ---------------------------------------------------------------------------
# reader.columns_for_header_row / inspect_sheet(header_row=...)
# ---------------------------------------------------------------------------


class TestColumnsForAnArbitraryHeaderRow(unittest.TestCase):
    def setUp(self):
        self._paths = []

    def tearDown(self):
        for path in self._paths:
            try:
                os.remove(path)
            except OSError:
                pass

    def _grid(self, rows):
        path = _write_workbook(rows)
        self._paths.append(path)
        wb = load_workbook(path, data_only=True)
        try:
            return read_grid(wb["Sheet1"]), path
        finally:
            wb.close()

    #: Nothing here reads as a header label, so auto-detection finds nothing —
    #: the sheet the wizard could not map by hand until now.
    HEADERLESS = [
        ["Place", "Team", "Issue seen"],
        ["Lobby", "Electrical", "Loose socket"],
        ["Lobby", "Electrical", "Cracked conduit"],
    ]

    def test_a_headerless_sheet_still_yields_a_real_column_list(self):
        grid, path = self._grid(self.HEADERLESS)
        sheet = inspect_workbook(path)[0]
        self.assertIsNone(sheet["header_row"])
        self.assertEqual(sheet["columns"], [])  # auto-detect: nothing to show

        self.assertEqual(
            columns_for_header_row(grid, 1),
            [
                {"letter": "A", "label": "Place"},
                {"letter": "B", "label": "Team"},
                {"letter": "C", "label": "Issue seen"},
            ],
        )

    def test_inspect_sheet_honours_an_explicit_header_row(self):
        _, path = self._grid(self.HEADERLESS)
        wb = load_workbook(path, data_only=True)
        try:
            sheet = inspect_sheet(wb["Sheet1"], header_row=1)
        finally:
            wb.close()
        self.assertEqual(sheet["header_row"], 1)
        self.assertEqual([c["label"] for c in sheet["columns"]], ["Place", "Team", "Issue seen"])
        # No column reads as a description, so there is still nothing to guess.
        self.assertIsNone(sheet["mapping_guess"])

    def test_columns_for_a_DATA_row_are_that_row_cells(self):
        grid, _ = self._grid(self.HEADERLESS)
        self.assertEqual(
            [c["label"] for c in columns_for_header_row(grid, 2)],
            ["Lobby", "Electrical", "Loose socket"],
        )

    def test_a_row_outside_the_grid_yields_letters_with_blank_labels(self):
        grid, _ = self._grid(self.HEADERLESS)
        self.assertEqual(
            columns_for_header_row(grid, 99),
            [
                {"letter": "A", "label": ""},
                {"letter": "B", "label": ""},
                {"letter": "C", "label": ""},
            ],
        )

    def test_none_and_an_empty_grid_yield_nothing(self):
        grid, _ = self._grid(self.HEADERLESS)
        self.assertEqual(columns_for_header_row(grid, None), [])
        self.assertEqual(columns_for_header_row([], 1), [])

    def test_the_fixture_header_row_matches_inspect_workbook(self):
        wb = load_workbook(FIXTURE, data_only=True)
        try:
            grid = read_grid(wb["Sheet1"])
            forced = inspect_sheet(wb["Sheet1"], header_row=7)
        finally:
            wb.close()
        auto = inspect_workbook(FIXTURE)[0]
        self.assertEqual(columns_for_header_row(grid, 7), auto["columns"])
        self.assertEqual(forced, auto)


if __name__ == "__main__":
    unittest.main()
