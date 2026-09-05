"""Row-level parse of ONE Snag List worksheet (step 2 — parse preview).

PURE: openpyxl + stdlib only. Returns the `ParsePreviewResponse` shape MINUS
`is_duplicate` / `duplicate_count` — this module knows nothing about the
database, so the API layer folds those in.

`rows` is ONE merged list: accepted and skipped rows interleaved in Excel row
order, each carrying the SAME fields (`skipped_reason` is None on an accepted
row). There is deliberately no second `skipped` list — a skipped row the user
re-ticks must arrive carrying its values, or it is silently dropped at import
(plan § R2.1).

THE SOURCE FILE IS A REPORT, NOT A TABLE (plan §2): one worksheet holds a title
block, two data blocks, a REPEATED header row between them and two hand-kept
RISK SUMMARY tallies. A naive "map three columns and ingest" reads 142 rows
where the file's own title says 124. Every row is therefore classified, and a
skipped row is skipped WITH A STATED REASON so the wizard can show it.
"""

from __future__ import annotations

import re

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from .guess import is_known_header_label
from .reader import all_header_rows, find_header_row, read_grid

#: A row whose FIRST non-empty cell reads like this opens a summary/tally block.
SUMMARY_RE = re.compile(r"summary|total|risk\s*summary", re.IGNORECASE)

#: How many mapped cells must equal a header label for the row to BE a header.
MIN_REPEATED_HEADER_CELLS = 2

#: `preview_text` is a glance, not the row.
PREVIEW_TEXT_MAX_LEN = 80

SKIP_BLANK = "blank"
SKIP_REPEATED_HEADER = "repeated_header"
SKIP_SUMMARY_BLOCK = "summary_block"
SKIP_NO_DESCRIPTION = "no_description"
#: A row sitting ABOVE the header row: outside the data region by POSITION, not
#: by content. The header row ITSELF keeps `SKIP_REPEATED_HEADER`.
SKIP_ABOVE_HEADER = "above_header"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _col_index(letter) -> "int | None":
    """0-based grid index for an Excel column letter; None when unmapped."""
    if not letter:
        return None
    try:
        return column_index_from_string(str(letter).strip().upper()) - 1
    except (ValueError, TypeError):
        return None


def _at(cells, idx) -> str:
    if idx is None or idx < 0 or idx >= len(cells):
        return ""
    return cells[idx]


def first_non_empty(cells) -> str:
    for text in cells:
        if text:
            return text
    return ""


def preview_text(cells) -> str:
    text = first_non_empty(cells)
    return text[:PREVIEW_TEXT_MAX_LEN]


def summary_block_rows(grid, header_rows) -> "set[int]":
    """1-based rows sitting inside a summary / tally block.

    A block OPENS on a row whose first non-empty cell matches `SUMMARY_RE` and
    RUNS until the next blank row or the next header row (neither of which is
    part of it).
    """
    rows: "set[int]" = set()
    in_block = False
    for idx, cells in enumerate(grid):
        row = idx + 1
        blank = not any(cells)
        if in_block:
            if blank or row in header_rows:
                in_block = False
            else:
                rows.add(row)
                continue
        if blank or row in header_rows:
            continue
        if SUMMARY_RE.search(first_non_empty(cells)):
            in_block = True
            rows.add(row)
    return rows


def is_repeated_header(cells, mapped_indexes) -> bool:
    """True when the mapped cells of this row are header labels, not data.

    The rule is >= 2 mapped cells holding a header label. When FEWER than two
    columns are mapped at all (a description-only mapping is legal — area /
    category / remarks are optional) that threshold is unreachable, and the
    header row would ingest itself as a snag reading "Snag Description". The
    requirement then falls to every mapped cell. With two or more mapped
    columns this is exactly the >= 2 rule, unchanged.
    """
    if not mapped_indexes:
        return False
    needed = min(MIN_REPEATED_HEADER_CELLS, len(mapped_indexes))
    hits = 0
    for idx in mapped_indexes:
        text = _at(cells, idx)
        if text and is_known_header_label(text):
            hits += 1
    return hits >= needed


def _distinct(values) -> "list[dict]":
    """Distinct non-blank values, count DESC then value ASC."""
    counts: "dict[str, int]" = {}
    for value in values:
        if not value:
            continue
        counts[value] = counts.get(value, 0) + 1
    return [
        {"value": value, "count": count}
        for value, count in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))
    ]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def _row(source_row, area, category, description, remark, reason, preview, serial="") -> dict:
    """ONE merged preview row — the `ParsedSnagRow` shape minus `is_duplicate`
    (which the API layer folds in; this module knows nothing about the DB).

    `skipped_reason` is None for an accepted row. `tickable` is False ONLY when
    the description is empty: such a row can NEVER become a Snag, because
    description is the one required field — offering a tick that the import must
    then refuse is exactly the silent drop this merged shape exists to prevent.
    """
    return {
        "source_row": source_row,
        "area": area,
        "category": category,
        "description": description,
        # NOTE: `remark` (SINGULAR) is the row's VALUE — the text read out of the
        # mapped column. It lands in `ProjectSnag.remark`. Do NOT confuse it with
        # the MAPPING key `remarks` (plural), which is an Excel COLUMN LETTER —
        # see the `remarks_i` lookup a few lines below, in `parse_grid`. ADR-0018.
        "remark": remark,
        #: The consultant's OWN S.No, VERBATIM, "" when the sheet has no such column
        #: or the cell is blank. The import supplies a number for those rows; this
        #: layer never invents one, because a parser row must report what is there.
        "serial": serial,
        "skipped_reason": reason,
        "tickable": bool(description.strip()),
        "preview_text": preview,
    }


def parse_grid(sheet_name: str, grid, mapping, header_row: "int | None" = None) -> dict:
    """Classify every row of an already-read grid. See `parse_sheet`."""
    description_letter = (mapping or {}).get("description")
    if not description_letter:
        raise ValueError("mapping['description'] is required — row detection keys on it")

    desc_i = _col_index(description_letter)
    if desc_i is None:
        raise ValueError("mapping['description'] is not a valid Excel column letter: %r" % (description_letter,))

    area_i = _col_index(mapping.get("area"))
    category_i = _col_index(mapping.get("category"))
    # MAPPING KEY `remarks` (PLURAL) -> an Excel COLUMN LETTER. The row VALUE it
    # produces is called `remark` (SINGULAR) — see `_row` below. ADR-0018.
    remarks_i = _col_index(mapping.get("remarks"))
    serial_i = _col_index(mapping.get("serial"))
    # ⚠️ `serial_i` is NOT in `mapped_indexes`. That list drives `is_repeated_header`,
    # so widening it would change how EVERY existing sheet classifies its rows -- for a
    # column that carries no snag content. The S.No is read, never used to judge a row.
    mapped_indexes = [i for i in (area_i, category_i, desc_i, remarks_i) if i is not None]

    # The header row bounds the DATA REGION by POSITION. An explicit `header_row`
    # is honoured literally, even if `row_is_header` would reject it; None means
    # auto-detect. When neither yields a row there is no positional filter at all
    # and every row classifies on content alone, exactly as before.
    if header_row is None:
        header_row = find_header_row(grid)

    header_rows = all_header_rows(grid)
    block_rows = summary_block_rows(grid, header_rows)

    rows: "list[dict]" = []
    accepted = 0

    for idx, cells in enumerate(grid):
        row = idx + 1
        area = _at(cells, area_i)
        category = _at(cells, category_i)
        description = _at(cells, desc_i)
        remark = _at(cells, remarks_i)
        serial = _at(cells, serial_i)

        def skip(reason, preview=None):
            rows.append(
                _row(
                    row,
                    area,
                    category,
                    description,
                    remark,
                    reason,
                    preview_text(cells) if preview is None else preview,
                    serial,
                )
            )

        # POSITION first: at or above the header row is outside the data region,
        # whatever the cells say.
        if header_row is not None and row <= header_row:
            skip(SKIP_REPEATED_HEADER if row == header_row else SKIP_ABOVE_HEADER)
            continue

        if not any(cells):
            skip(SKIP_BLANK, "")
            continue
        if is_repeated_header(cells, mapped_indexes):
            skip(SKIP_REPEATED_HEADER)
            continue
        if row in block_rows:
            skip(SKIP_SUMMARY_BLOCK)
            continue
        if not description:
            skip(SKIP_NO_DESCRIPTION)
            continue

        rows.append(_row(row, area, category, description, remark, None, "", serial))
        accepted += 1

    accepted_rows = [r for r in rows if r["skipped_reason"] is None]

    return {
        "sheet_name": sheet_name,
        # The header row the parse ACTUALLY used — the caller's override, or the
        # auto-guess when it sent none. None = no positional filter was applied.
        "header_row": header_row,
        # ONE list: accepted and skipped INTERLEAVED, in Excel row order.
        "rows": rows,
        "accepted_count": accepted,
        "skipped_count": len(rows) - accepted,
        # Distinct values count ACCEPTED rows ONLY.
        "distinct_areas": _distinct(r["area"] for r in accepted_rows),
        "distinct_categories": _distinct(r["category"] for r in accepted_rows),
    }


def parse_sheet(path: str, sheet_name: str, mapping, header_row: "int | None" = None) -> dict:
    """Parse one worksheet of the workbook at `path` with `mapping`.

    `mapping["description"]` is REQUIRED; area / category / remarks may be None,
    in which case those values come back as "". Values are trimmed but never
    lowercased or normalised (ADR-0016: stored verbatim).

    `header_row` (1-based) overrides auto-detection; every row AT or ABOVE it is
    excluded from the data region with a stated reason.
    """
    wb = load_workbook(path, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError("sheet not found in workbook: %r" % (sheet_name,))
        grid = read_grid(wb[sheet_name])
    finally:
        wb.close()
    return parse_grid(sheet_name, grid, mapping, header_row=header_row)
