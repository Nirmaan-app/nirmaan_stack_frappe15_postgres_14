"""Workbook inspection for the Snag List import wizard (step 1 — upload).

PURE: openpyxl + stdlib only. Takes a filesystem PATH and returns plain dicts
matching the `WorkbookSheet` TS interface in
`frontend/src/pages/SnagList/types.ts`. No frappe import anywhere.
"""

from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .guess import cell_text, guess_mapping, looks_like_header_label

#: How far down a sheet we look for the header row.
HEADER_SCAN_ROWS = 30

#: A header cell holds a short label, never a paragraph.
MAX_HEADER_LABEL_LEN = 60

#: A header row carries at least this many short text cells.
MIN_HEADER_CELLS = 3


# ---------------------------------------------------------------------------
# Sheet geometry
# ---------------------------------------------------------------------------


def sheet_extent(ws) -> "tuple[int, int]":
    """The sheet's REAL last used (row, col) — openpyxl's max_row/max_column
    count rows that only carry formatting. (0, 0) for an empty sheet."""
    last_row = 0
    last_col = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is None:
                continue
            if not cell_text(cell.value):
                continue
            if cell.row > last_row:
                last_row = cell.row
            if cell.column > last_col:
                last_col = cell.column
    return last_row, last_col


def read_grid(ws) -> "list[list[str]]":
    """The sheet as trimmed text, row-major, 0-indexed. `grid[r][c]` is Excel
    row r+1, column c+1. Trailing empty rows/columns are dropped."""
    last_row, last_col = sheet_extent(ws)
    if not last_row or not last_col:
        return []
    grid = [["" for _ in range(last_col)] for _ in range(last_row)]
    for row in ws.iter_rows(min_row=1, max_row=last_row, max_col=last_col):
        for cell in row:
            text = cell_text(cell.value)
            if text:
                grid[cell.row - 1][cell.column - 1] = text
    return grid


# ---------------------------------------------------------------------------
# Header-row detection
# ---------------------------------------------------------------------------


def row_is_header(cells) -> bool:
    """A row is a header row when >= 3 of its cells are short non-empty text
    AND at least one of them reads as a known header word."""
    labels = [c for c in cells if c and len(c) <= MAX_HEADER_LABEL_LEN]
    if len(labels) < MIN_HEADER_CELLS:
        return False
    return any(looks_like_header_label(label) for label in labels)


def find_header_row(grid, scan_rows: int = HEADER_SCAN_ROWS) -> "int | None":
    """1-based Excel row of the FIRST header-looking row, or None."""
    for idx, cells in enumerate(grid[:scan_rows]):
        if row_is_header(cells):
            return idx + 1
    return None


def all_header_rows(grid) -> "set[int]":
    """Every header-looking row in the sheet (1-based). Used by the parser to
    terminate a summary block and to recognise a repeated header."""
    return {i + 1 for i, cells in enumerate(grid) if row_is_header(cells)}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def columns_for_header_row(grid, header_row) -> "list[dict]":
    """The `WorkbookColumn[]` list for an ARBITRARY row of `grid`.

    `header_row` is 1-based; None (or a row outside the grid) yields []. This
    works for ANY row the user names — including a sheet where `find_header_row`
    found nothing, which is the only way such a sheet can be mapped by hand.
    """
    if not grid or header_row is None:
        return []
    last_col = max(len(r) for r in grid)
    if not last_col:
        return []
    header_cells = grid[header_row - 1] if 1 <= header_row <= len(grid) else []
    return [
        {
            "letter": get_column_letter(c + 1),
            "label": header_cells[c] if c < len(header_cells) else "",
        }
        for c in range(last_col)
    ]


def inspect_sheet(ws, header_row: "int | None" = None) -> dict:
    """One `WorkbookSheet` dict for an open openpyxl worksheet.

    `header_row` (1-based) OVERRIDES auto-detection and is honoured LITERALLY —
    a user may name a row `row_is_header` would reject. None = auto-detect,
    exactly as before.
    """
    grid = read_grid(ws)
    if not grid:
        return {
            "name": ws.title,
            "is_empty": True,
            "row_count": 0,
            "header_row": None,
            "columns": [],
            "mapping_guess": None,
        }

    if header_row is None:
        header_row = find_header_row(grid)

    columns = columns_for_header_row(grid, header_row)

    return {
        "name": ws.title,
        "is_empty": False,
        "row_count": len(grid),
        "header_row": header_row,
        "columns": columns,
        "mapping_guess": guess_mapping(columns) if columns else None,
    }


def inspect_workbook(path: str) -> "list[dict]":
    """Inspect every worksheet in the workbook at `path`.

    Returns one dict per sheet, in workbook order, matching the
    `WorkbookSheet` TS interface.
    """
    wb = load_workbook(path, data_only=True)
    try:
        return [inspect_sheet(ws) for ws in wb.worksheets]
    finally:
        wb.close()
